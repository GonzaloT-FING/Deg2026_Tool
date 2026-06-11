"""Degradation (.DTA) -> Excel (.xlsx) exporter and V vs t viewer."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date as date_cls, datetime, time as time_cls, timedelta
from pathlib import Path
from math import ceil, floor, log10
import re

import tkinter as tk
import matplotlib.dates as mdates
from tkinter import ttk, messagebox, filedialog
from matplotlib.backends.backend_agg import FigureCanvasAgg
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.ticker import LinearLocator, LogFormatterSciNotation, LogLocator, MaxNLocator, StrMethodFormatter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from pipelines import ocp_pip as ocp
from pipelines.activ_pip import ACTIV_CYCLE_GRADIENTS, _build_scrollable_cycle_selector, _cycle_gradient_color
from plot_defaults import PlotFontDefaults, apply_x_tick_label_padding, make_legend_draggable, resolve_plot_font_defaults
from i18n import normalize_language, translate
from ui_layout import create_resizable_plot_layout, create_scrollable_controls


META_FIELDS = [
    ("TITLE", "Tecnica"),
    ("DATE", "Fecha"),
    ("TIME", "Hora"),
    ("ISTEP1", "Corriente"),
    ("TSTEP1", "Duracion"),
    ("SAMPLETIME", "Tiempo de muestreo"),
    ("AREA", "Area"),
]

DATA_EXPORT = [
    ("Pt", "Pt"),
    ("T", "time"),
    ("Vf", "Voltaje"),
    ("Im", "Corriente"),
    ("Temp", "Temperatura"),
]

DEG_FILE_RE = re.compile(
    r"^Degradacion_galvanostatica_19A_60C_#(?P<stage>\d+)\.DTA$",
    re.IGNORECASE,
)
DEG_OCP_FILE_RE = re.compile(
    r"^Degradacion_OCP_#(?P<stage>\d+)\.DTA$",
    re.IGNORECASE,
)

LINESTYLE_OPTIONS = ["none", "-", "--", ":", "-."]
SMOOTHING_ALGORITHMS = ["Median filter", "Rolling average"]
DEG_STAGE_SELECTOR_HEIGHT = 140

DEG_PLOT_COLORS = {
    "voltage": "#1f5f99",
    "temperature": "#cf9a32",
    "dvdt": "#1f5f99",
    "range": "#a6bddb",
}

DEG_STAGE_GRADIENTS = {
    "voltage": ACTIV_CYCLE_GRADIENTS["voltage"],
    "temperature": ACTIV_CYCLE_GRADIENTS["temperature"],
    "dvdt": ACTIV_CYCLE_GRADIENTS["voltage"],
}

SECONDS_PER_MINUTE = 60.0
SECONDS_PER_HOUR = 3600.0
DATE_AXIS_OPTION = "fecha"
TIME_UNIT_OPTIONS = ["s", "min", "h", DATE_AXIS_OPTION]
DATETIME_DISPLAY_FORMAT = "%Y-%m-%d %H:%M:%S"
DATE_INPUT_FORMATS = (
    "%d/%m/%Y",
    "%d/%m/%y",
    "%m/%d/%Y",
    "%m/%d/%y",
    "%Y-%m-%d",
)
TIME_INPUT_FORMATS = (
    "%H:%M:%S",
    "%H:%M",
    "%I:%M:%S %p",
    "%I:%M %p",
)
DATETIME_INPUT_FORMATS = (
    DATETIME_DISPLAY_FORMAT,
    "%Y-%m-%d %H:%M",
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M",
    "%d/%m/%y %H:%M:%S",
    "%d/%m/%y %H:%M",
    "%m/%d/%Y %H:%M:%S",
    "%m/%d/%Y %H:%M",
    "%m/%d/%y %H:%M:%S",
    "%m/%d/%y %H:%M",
)
TimeAxisValue = float | datetime


@dataclass(frozen=True)
class DegFile:
    path: Path
    stage: int


@dataclass
class ParsedDTA:
    meta_values: dict[str, str]
    meta_units: dict[str, str]
    header: list[str]
    units: list[str]
    rows: list[list[str]]


def to_float(val: str) -> float | None:
    s = val.strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _drop_leading_blank(parts: list[str]) -> list[str]:
    if parts and parts[0] == "":
        return parts[1:]
    return parts


def _extract_parenthesized_unit(text: str) -> str:
    matches = re.findall(r"\(([^()]*)\)", text)
    if matches:
        return matches[-1].strip()
    return ""


def _extract_meta_unit(key: str, description: str) -> str:
    unit = _extract_parenthesized_unit(description)
    if unit:
        return unit

    fallback_units = {
        "TITLE": "",
        "DATE": "",
        "TIME": "",
        "ISTEP1": "A",
        "TSTEP1": "s",
        "SAMPLETIME": "s",
        "AREA": "cm^2",
    }
    return fallback_units.get(key, "")


def parse_gamry_dta(path: Path) -> ParsedDTA:
    text = path.read_text(encoding="latin-1", errors="replace")
    lines = text.splitlines()

    meta_values: dict[str, str] = {}
    meta_units: dict[str, str] = {}
    header: list[str] | None = None
    units: list[str] = []
    rows: list[list[str]] = []
    table_started = False

    for line in lines:
        if not table_started:
            if line.startswith("CURVE") and "TABLE" in line:
                table_started = True
                continue

            if not line.strip():
                continue

            parts = line.split("\t")
            if len(parts) >= 3 and parts[0].strip():
                key = parts[0].strip()
                value = parts[2].strip()
                description = " ".join(p.strip() for p in parts[3:] if p.strip())
                meta_values[key] = value
                meta_units[key] = _extract_meta_unit(key, description)
            continue

        if not line.strip():
            continue

        parts = _drop_leading_blank([p.strip() for p in line.rstrip("\r\n").split("\t")])
        if not parts:
            continue

        first = parts[0]
        if header is None:
            if first == "Pt":
                header = parts
            continue

        if not units:
            if first == "#":
                units = parts
            continue

        if re.fullmatch(r"-?\d+", first):
            rows.append(parts)

    if header is None:
        raise ValueError(f"No data header found in {path.name} (expected 'Pt ...')")

    return ParsedDTA(
        meta_values=meta_values,
        meta_units=meta_units,
        header=header,
        units=units,
        rows=rows,
    )


def find_deg_files(input_dir: Path) -> list[DegFile]:
    files: list[DegFile] = []
    for path in sorted(Path(input_dir).glob("*.DTA")):
        match = DEG_FILE_RE.match(path.name)
        if not match:
            continue
        files.append(DegFile(path=path, stage=int(match.group("stage"))))
    return sorted(files, key=lambda item: item.stage)


def find_deg_ocp_files(input_dir: Path) -> list[DegFile]:
    files: list[DegFile] = []
    for path in sorted(Path(input_dir).glob("*.DTA")):
        match = DEG_OCP_FILE_RE.match(path.name)
        if not match:
            continue
        files.append(DegFile(path=path, stage=int(match.group("stage"))))
    return sorted(files, key=lambda item: item.stage)


def _column_index(parsed: ParsedDTA, column_name: str) -> int | None:
    try:
        return parsed.header.index(column_name)
    except ValueError:
        return None


def _required_numeric_series(parsed: ParsedDTA, x_name: str, y_name: str) -> tuple[list[float], list[float]]:
    x_idx = _column_index(parsed, x_name)
    y_idx = _column_index(parsed, y_name)
    if x_idx is None or y_idx is None:
        raise ValueError(f"Faltan columnas requeridas: {x_name}, {y_name}")

    x_values: list[float] = []
    y_values: list[float] = []

    for row in parsed.rows:
        if x_idx >= len(row) or y_idx >= len(row):
            continue
        x_val = to_float(row[x_idx])
        y_val = to_float(row[y_idx])
        if x_val is None or y_val is None:
            continue
        x_values.append(x_val)
        y_values.append(y_val)

    if not x_values or not y_values:
        raise ValueError("No se encontraron datos numericos validos para graficar.")

    return x_values, y_values


def _optional_float(value: str | None) -> float | None:
    if value is None:
        return None
    text = value.strip()
    if not text:
        return None
    return float(text.replace(",", "."))


def _positive_float(text: str, name: str) -> float:
    value = text.strip().replace(",", ".")
    if not value:
        raise ValueError(f"{name} no puede estar vacio.")
    num = float(value)
    if num <= 0:
        raise ValueError(f"{name} debe ser mayor que 0.")
    return num


def _is_date_axis(time_unit: str) -> bool:
    return time_unit == DATE_AXIS_OPTION


def _parse_date_text(text: str) -> date_cls:
    normalized = text.strip()
    if not normalized:
        raise ValueError("DATE esta vacio.")

    for date_format in DATE_INPUT_FORMATS:
        try:
            return datetime.strptime(normalized, date_format).date()
        except ValueError:
            continue

    raise ValueError(f"DATE no tiene un formato valido: {normalized!r}")


def _parse_time_text(text: str) -> time_cls:
    normalized = text.strip()
    if not normalized:
        raise ValueError("TIME esta vacio.")

    for time_format in TIME_INPUT_FORMATS:
        try:
            return datetime.strptime(normalized, time_format).time()
        except ValueError:
            continue

    raise ValueError(f"TIME no tiene un formato valido: {normalized!r}")


def _parse_datetime_text(text: str) -> datetime:
    normalized = text.strip()
    if not normalized:
        raise ValueError("La fecha no puede estar vacia.")

    iso_candidate = normalized.replace("T", " ")
    try:
        return datetime.fromisoformat(iso_candidate)
    except ValueError:
        pass

    for datetime_format in DATETIME_INPUT_FORMATS:
        try:
            return datetime.strptime(normalized, datetime_format)
        except ValueError:
            continue

    try:
        parsed_date = _parse_date_text(normalized)
    except ValueError as exc:
        raise ValueError(
            "La fecha debe tener formato YYYY-MM-DD HH:MM[:SS] o DD/MM/YYYY HH:MM[:SS]."
        ) from exc

    return datetime.combine(parsed_date, time_cls.min)


def _start_datetime(parsed: ParsedDTA, source_name: str) -> datetime:
    raw_date = parsed.meta_values.get("DATE", "")
    raw_time = parsed.meta_values.get("TIME", "")

    if not raw_date.strip():
        raise ValueError(f"{source_name}: falta DATE para construir el eje de fecha.")
    if not raw_time.strip():
        raise ValueError(f"{source_name}: falta TIME para construir el eje de fecha.")

    try:
        return datetime.combine(_parse_date_text(raw_date), _parse_time_text(raw_time))
    except ValueError as exc:
        raise ValueError(f"{source_name}: {exc}") from exc


def _reference_start_datetime(parsed_items: list[tuple[DegFile, ParsedDTA]]) -> datetime:
    if not parsed_items:
        raise ValueError("No hay archivos de degradacion para construir el eje de fecha.")

    return min(_start_datetime(parsed, deg_file.path.name) for deg_file, parsed in parsed_items)


def _format_datetime_value(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime(DATETIME_DISPLAY_FORMAT)


def _format_time_limit(value: TimeAxisValue | None, time_unit: str) -> str:
    if value is None:
        return ""
    if _is_date_axis(time_unit):
        if not isinstance(value, datetime):
            raise ValueError("Los limites del eje de fecha deben ser fechas validas.")
        return _format_datetime_value(value)
    return _format_limit_value(float(value), _time_unit_decimals(time_unit))


def _parse_time_limit_text(value_text: str, time_unit: str) -> TimeAxisValue | None:
    if not value_text.strip():
        return None
    if _is_date_axis(time_unit):
        return _parse_datetime_text(value_text)
    return _optional_float(value_text)


def _time_limit_to_seconds(value: TimeAxisValue, time_unit: str, reference_start: datetime) -> float:
    if _is_date_axis(time_unit):
        if not isinstance(value, datetime):
            raise ValueError("Los limites del eje de fecha deben ser fechas validas.")
        return (value - reference_start).total_seconds()
    return float(value) * _time_unit_scale(time_unit)


def _seconds_to_time_limit(value_seconds: float, time_unit: str, reference_start: datetime) -> TimeAxisValue:
    if _is_date_axis(time_unit):
        return reference_start + timedelta(seconds=value_seconds)
    return value_seconds / _time_unit_scale(time_unit)


def _convert_time_limit_text(
    value_text: str,
    from_unit: str,
    to_unit: str,
    reference_start: datetime,
) -> str:
    if from_unit == to_unit or not value_text.strip():
        return value_text

    parsed_value = _parse_time_limit_text(value_text, from_unit)
    if parsed_value is None:
        return value_text

    seconds = _time_limit_to_seconds(parsed_value, from_unit, reference_start)
    converted_value = _seconds_to_time_limit(seconds, to_unit, reference_start)
    return _format_time_limit(converted_value, to_unit)


def _plot_time_values(
    parsed: ParsedDTA,
    elapsed_seconds: list[float],
    time_unit: str,
    source_name: str,
) -> list[TimeAxisValue]:
    if _is_date_axis(time_unit):
        start_dt = _start_datetime(parsed, source_name)
        return [start_dt + timedelta(seconds=value) for value in elapsed_seconds]

    time_scale = _time_unit_scale(time_unit)
    return [value / time_scale for value in elapsed_seconds]


def _padded_datetime_limits(
    values: list[datetime],
    rel_pad: float = 0.05,
) -> tuple[datetime | None, datetime | None]:
    if not values:
        return None, None

    vmin = min(values)
    vmax = max(values)
    span_seconds = (vmax - vmin).total_seconds()
    pad_seconds = 1.0 if span_seconds == 0 else span_seconds * rel_pad
    pad = timedelta(seconds=pad_seconds)
    return vmin - pad, vmax + pad


def _apply_time_axis_format(ax, time_unit: str, tick_count: int) -> None:
    if _is_date_axis(time_unit):
        tick_target = max(2, int(tick_count))
        locator = mdates.AutoDateLocator(
            minticks=max(2, min(4, tick_target)),
            maxticks=tick_target,
        )
        ax.xaxis.set_major_locator(locator)
        ax.xaxis.set_major_formatter(mdates.ConciseDateFormatter(locator))
        return

    ax.xaxis.set_major_locator(MaxNLocator(nbins=max(2, int(tick_count))))
    ax.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))


def _time_unit_scale(time_unit: str) -> float:
    if time_unit == "min":
        return SECONDS_PER_MINUTE
    if time_unit == "h":
        return SECONDS_PER_HOUR
    return 1.0


def _time_unit_decimals(time_unit: str) -> int:
    if time_unit == "h":
        return 2
    if time_unit == "min":
        return 1
    return 1


def _dv_dt_scale(time_unit: str) -> float:
    return 1e6 * _time_unit_scale(time_unit)


def _dv_dt_unit_label(time_unit: str) -> str:
    return f"μV/{time_unit}" if time_unit != "s" else "μV/s"


def _mpl_linestyle(value: str) -> str:
    return "None" if value == "none" else value


def _linear_fit_slope(x_vals: list[float], y_vals: list[float]) -> float | None:
    if len(x_vals) < 2:
        return None
    sum_x = sum(x_vals)
    sum_y = sum(y_vals)
    sum_x2 = sum(x * x for x in x_vals)
    sum_xy = sum(x * y for x, y in zip(x_vals, y_vals))
    n = float(len(x_vals))
    denom = (n * sum_x2) - (sum_x * sum_x)
    if denom == 0:
        return None
    return (n * sum_xy - sum_x * sum_y) / denom


def _linear_fit_params(x_vals: list[float], y_vals: list[float]) -> tuple[float, float] | None:
    if len(x_vals) < 2:
        return None
    slope = _linear_fit_slope(x_vals, y_vals)
    if slope is None:
        return None
    n = float(len(x_vals))
    intercept = (sum(y_vals) - slope * sum(x_vals)) / n
    return slope, intercept


def _edge_slope(x_vals: list[float], y_vals: list[float]) -> float | None:
    if len(x_vals) < 2:
        return None
    dt = x_vals[-1] - x_vals[0]
    if dt == 0:
        return None
    return (y_vals[-1] - y_vals[0]) / dt


def _filtered_time_voltage(
    time_seconds: list[float],
    voltages: list[float],
    t_min_seconds: float | None,
    t_max_seconds: float | None,
) -> tuple[list[float], list[float]]:
    pairs = [
        (t, v)
        for t, v in zip(time_seconds, voltages)
        if (t_min_seconds is None or t >= t_min_seconds)
        and (t_max_seconds is None or t <= t_max_seconds)
    ]
    if not pairs:
        return [], []
    pairs.sort(key=lambda item: item[0])
    out_t = [item[0] for item in pairs]
    out_v = [item[1] for item in pairs]
    return out_t, out_v


def compute_degradation_rate(
    parsed: ParsedDTA,
    source_name: str,
    time_unit: str,
    reference_start: datetime | None,
    t_min: TimeAxisValue | None,
    t_max: TimeAxisValue | None,
    use_linear_fit: bool,
) -> tuple[float | None, int]:
    time_values, voltage_values = _required_numeric_series(parsed, "T", "Vf")

    if _is_date_axis(time_unit):
        if reference_start is None:
            raise ValueError("Falta referencia de tiempo para el eje de fecha.")
        stage_start = _start_datetime(parsed, source_name)
        offset = (stage_start - reference_start).total_seconds()
        time_seconds = [offset + value for value in time_values]
        t_min_seconds = (
            _time_limit_to_seconds(t_min, time_unit, reference_start) if t_min is not None else None
        )
        t_max_seconds = (
            _time_limit_to_seconds(t_max, time_unit, reference_start) if t_max is not None else None
        )
    else:
        time_seconds = list(time_values)
        t_min_seconds = (
            float(t_min) * _time_unit_scale(time_unit) if t_min is not None else None
        )
        t_max_seconds = (
            float(t_max) * _time_unit_scale(time_unit) if t_max is not None else None
        )

    filt_t, filt_v = _filtered_time_voltage(time_seconds, voltage_values, t_min_seconds, t_max_seconds)
    if len(filt_t) < 2:
        return None, len(filt_t)

    slope = _linear_fit_slope(filt_t, filt_v) if use_linear_fit else _edge_slope(filt_t, filt_v)
    return slope, len(filt_t)


def _deg_language(language: str | None = None) -> str:
    return normalize_language(language)


def _safe_filename_part(text: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", text.strip())
    return cleaned.strip("._") or "Deg"


def _format_report_value(value: object, digits: int = 6) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.{digits}g}"
    return str(value)


def build_deg_report_metadata(
    parsed_items: list[tuple[DegFile, ParsedDTA]],
    language: str | None = None,
) -> list[tuple[str, object, str]]:
    language = _deg_language(language)
    if not parsed_items:
        return []

    _first_file, first_parsed = parsed_items[0]
    label_keys = {
        "TITLE": "technique",
        "DATE": "date",
        "TIME": "start_time",
        "ISTEP1": "current",
        "TSTEP1": "step_duration",
        "SAMPLETIME": "sampling_time",
        "AREA": "area",
    }
    rows: list[tuple[str, object, str]] = [
        (translate("deg_stage_count", language), len(parsed_items), ""),
    ]

    for key, fallback_label in META_FIELDS:
        label = translate(label_keys.get(key, fallback_label), language)
        raw_value = first_parsed.meta_values.get(key, "")
        value = to_float(raw_value)
        rows.append((label, _format_report_value(value if value is not None else raw_value), first_parsed.meta_units.get(key, "")))
    return rows


def build_deg_report_indicators(
    parsed_items: list[tuple[DegFile, ParsedDTA]],
    language: str | None = None,
) -> list[tuple[str, object, str]]:
    language = _deg_language(language)
    slopes_uv_h: list[float] = []
    voltage_deltas: list[float] = []
    temperature_deltas: list[float] = []
    temperatures: list[float] = []

    for _deg_file, parsed in parsed_items:
        try:
            t_vals, v_vals = _required_numeric_series(parsed, "T", "Vf")
        except ValueError:
            continue
        slope = _edge_slope(t_vals, v_vals)
        if slope is not None:
            slopes_uv_h.append(slope * 1e6 * SECONDS_PER_HOUR)
        if v_vals:
            voltage_deltas.append(max(v_vals) - min(v_vals))

        try:
            _temp_t, temp_vals = _required_numeric_series(parsed, "T", "Temp")
        except ValueError:
            temp_vals = []
        if temp_vals:
            temperature_deltas.append(max(temp_vals) - min(temp_vals))
            temperatures.extend(temp_vals)

    rows: list[tuple[str, object, str]] = []
    if slopes_uv_h:
        rows.append((translate("deg_average_slope", language), _format_report_value(sum(slopes_uv_h) / len(slopes_uv_h)), "µV/h"))
    if voltage_deltas:
        rows.append((translate("deg_max_delta_voltage", language), _format_report_value(max(voltage_deltas)), "V"))
    if temperature_deltas:
        rows.append((translate("deg_max_delta_temperature", language), _format_report_value(max(temperature_deltas)), "C"))
    if temperatures:
        rows.append((translate("deg_max_temperature", language), _format_report_value(max(temperatures), digits=3), "C"))
    return rows


def _add_report_table(
    ax,
    title: str,
    rows: list[tuple[str, object, str]],
    bbox: list[float],
    language: str = "es",
) -> None:
    ax.text(
        bbox[0],
        bbox[1] + bbox[3] + 0.025,
        title,
        fontsize=13,
        fontweight="bold",
        ha="left",
        va="bottom",
        transform=ax.transAxes,
    )
    table = ax.table(
        cellText=[[_format_report_value(name), _format_report_value(value), _format_report_value(unit)] for name, value, unit in rows],
        colLabels=[translate("name", language), translate("value", language), translate("unit", language)],
        cellLoc="left",
        colLoc="left",
        colWidths=[bbox[2] * 0.58, bbox[2] * 0.27, bbox[2] * 0.15],
        bbox=bbox,
    )
    table.auto_set_font_size(False)
    table.set_fontsize(8.0)
    for (row_idx, _col_idx), cell in table.get_celld().items():
        cell.set_edgecolor("#b8c0c8")
        cell.set_linewidth(0.5)
        if row_idx == 0:
            cell.set_facecolor("#edf2f7")
            cell.set_text_props(weight="bold")
        else:
            cell.set_facecolor("white")


def export_deg_v_vs_t_report_pdf(
    parsed_items: list[tuple[DegFile, ParsedDTA]],
    output_path: Path,
    *,
    language: str = "es",
    font_defaults: PlotFontDefaults | None = None,
) -> Path:
    if not parsed_items:
        raise ValueError("Debe seleccionar al menos una etapa para generar el reporte.")

    language = _deg_language(language)
    font_defaults = resolve_plot_font_defaults(font_defaults)
    font_values = font_defaults.as_strings()
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    limits = compute_autofit_v_vs_t_limits(parsed_items, show_temperature=True, time_unit="h")
    plot_kwargs = {
        "show_temperature": True,
        "voltage_linestyle": "-",
        "temperature_linestyle": "--",
        "time_unit": "h",
        "x_tick_count": 6,
        "y_tick_count": 6,
        "plot_title": translate("deg_report_plot_title", language),
        "show_title": True,
        "title_fontsize": float(font_values["title"]),
        "tick_fontsize": float(font_values["tick"]),
        "label_fontsize": float(font_values["label"]),
        "legend_fontsize": float(font_values["legend"]),
        "line_width": 1.5,
        "t_min": _parse_time_limit_text(limits["t_min"], "h"),
        "t_max": _parse_time_limit_text(limits["t_max"], "h"),
        "v_min": _optional_float(limits["v_min"]),
        "v_max": _optional_float(limits["v_max"]),
        "temp_min": _optional_float(limits["temp_min"]),
        "temp_max": _optional_float(limits["temp_max"]),
        "show_fit_line": False,
        "show_fit_range": False,
    }

    metadata_rows = build_deg_report_metadata(parsed_items, language=language)
    indicator_rows = build_deg_report_indicators(parsed_items, language=language)

    with PdfPages(output_path) as pdf:
        plot_fig = Figure(figsize=(9.5, 6.2), dpi=150)
        FigureCanvasAgg(plot_fig)
        has_plot = draw_v_vs_t_on_figure(plot_fig, parsed_items=parsed_items, **plot_kwargs)
        if not has_plot:
            raise ValueError("No hay datos suficientes para generar el grafico del reporte.")
        pdf.savefig(plot_fig, bbox_inches="tight")

        table_fig = Figure(figsize=(8.5, 11.0), dpi=150)
        ax = table_fig.add_subplot(111)
        ax.axis("off")
        table_fig.text(
            0.05,
            0.965,
            translate("deg_report_title", language),
            fontsize=18,
            fontweight="bold",
            ha="left",
            va="top",
        )
        table_fig.text(
            0.05,
            0.935,
            translate("deg_report_subtitle", language),
            fontsize=10,
            ha="left",
            va="top",
            color="#4a5568",
        )
        _add_report_table(ax, translate("metadata", language), metadata_rows, [0.05, 0.53, 0.90, 0.34], language=language)
        _add_report_table(ax, translate("indicators", language), indicator_rows, [0.05, 0.08, 0.90, 0.36], language=language)
        pdf.savefig(table_fig, bbox_inches="tight")

    return output_path


def _build_scrollable_controls(parent) -> ttk.Frame:
    _outer, inner = create_scrollable_controls(parent, reset_y_on_resize=True)
    return inner


def _selected_stage_items(
    parsed_items: list[tuple[DegFile, ParsedDTA]],
    stage_vars: dict[int, tk.BooleanVar],
) -> list[tuple[DegFile, ParsedDTA]]:
    return [
        (deg_file, parsed)
        for deg_file, parsed in parsed_items
        if stage_vars.get(deg_file.stage) is not None and stage_vars[deg_file.stage].get()
    ]


def _format_limit_value(value: float | None, decimals: int = 1) -> str:
    if value is None:
        return ""
    return f"{value:.{decimals}f}"


def _round_down_dec(value: float, decimals: int = 1) -> float:
    scale = 10 ** decimals
    return floor(value * scale) / scale


def _round_up_dec(value: float, decimals: int = 1) -> float:
    scale = 10 ** decimals
    return ceil(value * scale) / scale


def _padded_limits(
    values: list[float],
    rel_pad: float = 0.05,
    decimals: int = 1,
) -> tuple[float | None, float | None]:
    if not values:
        return None, None

    vmin = min(values)
    vmax = max(values)

    if vmin == vmax:
        pad = max(abs(vmin) * rel_pad, 1e-6)
    else:
        pad = (vmax - vmin) * rel_pad

    lo = _round_down_dec(vmin - pad, decimals)
    hi = _round_up_dec(vmax + pad, decimals)
    return lo, hi


def _tight_limits(values: list[float], decimals: int = 1) -> tuple[float | None, float | None]:
    if not values:
        return None, None
    lo = _round_down_dec(min(values), decimals)
    hi = _round_up_dec(max(values), decimals)
    if lo == hi:
        return _padded_limits(values, decimals=decimals)
    return lo, hi


def _auto_decimals(
    values: list[float],
    default: int = 1,
    min_decimals: int = 1,
    max_decimals: int = 5,
) -> int:
    if not values:
        return max(default, min_decimals)
    span = max(values) - min(values)
    if span <= 0:
        ref = max(abs(value) for value in values)
        span = ref * 0.02 if ref > 0 else 1.0
    if span <= 0:
        return max(default, min_decimals)
    magnitude = floor(log10(span))
    decimals = 1 - magnitude
    decimals = max(min_decimals, min(max_decimals, int(decimals)))
    return max(default, decimals)


def _positive_dv_dt_rows(rows: list[dict[str, float]]) -> list[dict[str, float]]:
    return [row for row in rows if row["dVdt"] > 0]


def _scale_compatible_dv_dt_rows(
    rows: list[dict[str, float]],
    logarithmic_y: bool,
) -> list[dict[str, float]]:
    return _positive_dv_dt_rows(rows) if logarithmic_y else rows


def _format_log_limit_value(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:.6g}"


def _format_adaptive_limit_value(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:.6g}"


def _adaptive_linear_limits(values: list[float]) -> tuple[float | None, float | None]:
    if not values:
        return None, None

    vmin = min(values)
    vmax = max(values)

    if vmin == vmax:
        pad = max(abs(vmin) * 0.05, 1e-9)
    else:
        pad = (vmax - vmin) * 0.05

    lo = vmin - pad
    hi = vmax + pad

    if lo == hi:
        hi = lo + max(abs(lo) * 0.05, 1e-9)

    return lo, hi


def _log_axis_limits(values: list[float]) -> tuple[float | None, float | None]:
    positive_values = [value for value in values if value > 0]
    if not positive_values:
        return None, None

    vmin = min(positive_values)
    vmax = max(positive_values)

    lower = 10 ** floor(log10(vmin))
    upper = 10 ** ceil(log10(vmax))

    if lower == upper:
        upper *= 10.0

    return lower, upper


def _normalize_median_window(window_size: int) -> int:
    size = max(1, int(window_size))
    if size % 2 == 0:
        size += 1
    return size


def _median_filter(values: list[float], window_size: int) -> list[float]:
    if not values:
        return []

    size = _normalize_median_window(window_size)
    if size == 1:
        return list(values)

    radius = size // 2
    filtered: list[float] = []

    for idx in range(len(values)):
        start = max(0, idx - radius)
        end = min(len(values), idx + radius + 1)
        window = sorted(values[start:end])
        filtered.append(window[len(window) // 2])

    return filtered


def _normalize_average_window(window_size: int) -> int:
    return max(1, int(window_size))


def _rolling_average_filter(values: list[float], window_size: int) -> list[float]:
    if not values:
        return []

    size = _normalize_average_window(window_size)
    if size == 1:
        return list(values)

    radius_left = (size - 1) // 2
    radius_right = size // 2
    filtered: list[float] = []

    for idx in range(len(values)):
        start = max(0, idx - radius_left)
        end = min(len(values), idx + radius_right + 1)
        window = values[start:end]
        filtered.append(sum(window) / len(window))

    return filtered


def apply_smoothing(values: list[float], algorithm: str, window_size: int) -> list[float]:
    if algorithm == "Median filter":
        return _median_filter(values, window_size)
    if algorithm == "Rolling average":
        return _rolling_average_filter(values, window_size)
    raise ValueError(f"Algoritmo de suavizado no soportado: {algorithm}")


def apply_x_edge_ticks(ax, x_min: float | None, x_max: float | None, tick_count: int) -> None:
    if x_min is None and x_max is None:
        return
    ax.xaxis.set_major_locator(LinearLocator(max(2, int(tick_count))))
    ax.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))


def compute_default_v_vs_t_limits(
    parsed_items: list[tuple[DegFile, ParsedDTA]],
    time_unit: str = "s",
) -> dict[str, str]:
    time_values: list[float] = []
    datetime_values: list[datetime] = []
    voltage_values: list[float] = []
    temperature_values: list[float] = []

    for deg_file, parsed in parsed_items:
        t_vals, v_vals = _required_numeric_series(parsed, "T", "Vf")
        if _is_date_axis(time_unit):
            datetime_values.extend(_plot_time_values(parsed, t_vals, time_unit, deg_file.path.name))
        else:
            time_values.extend(t_vals)
        voltage_values.extend(v_vals)
        try:
            _temp_time, temp_vals = _required_numeric_series(parsed, "T", "Temp")
            temperature_values.extend(temp_vals)
        except ValueError:
            pass

    if (not time_values and not datetime_values) or not voltage_values:
        return {"t_min": "", "t_max": "", "v_min": "", "v_max": "", "temp_min": "", "temp_max": ""}

    if _is_date_axis(time_unit):
        _t_min, t_max = _padded_datetime_limits(datetime_values)
        t_min = min(datetime_values)
    else:
        time_scale = _time_unit_scale(time_unit)
        time_decimals = _time_unit_decimals(time_unit)
        scaled_time_values = [value / time_scale for value in time_values]
        _t_min, t_max = _padded_limits(scaled_time_values, decimals=time_decimals)
        t_min = 0.0
    voltage_decimals = _auto_decimals(voltage_values)
    temp_decimals = _auto_decimals(temperature_values) if temperature_values else 1
    v_min, v_max = _tight_limits(voltage_values, decimals=voltage_decimals)
    temp_min, temp_max = _tight_limits(temperature_values, decimals=temp_decimals) if temperature_values else (None, None)

    return {
        "t_min": _format_time_limit(t_min, time_unit),
        "t_max": _format_time_limit(t_max, time_unit),
        "v_min": _format_limit_value(v_min, voltage_decimals),
        "v_max": _format_limit_value(v_max, voltage_decimals),
        "temp_min": _format_limit_value(temp_min, temp_decimals) if temp_min is not None else "",
        "temp_max": _format_limit_value(temp_max, temp_decimals) if temp_max is not None else "",
    }


def compute_autofit_v_vs_t_limits(
    parsed_items: list[tuple[DegFile, ParsedDTA]],
    show_temperature: bool,
    time_unit: str = "s",
) -> dict[str, str]:
    time_values: list[float] = []
    datetime_values: list[datetime] = []
    voltage_values: list[float] = []
    temperature_values: list[float] = []

    for deg_file, parsed in parsed_items:
        t_vals, v_vals = _required_numeric_series(parsed, "T", "Vf")
        if _is_date_axis(time_unit):
            datetime_values.extend(_plot_time_values(parsed, t_vals, time_unit, deg_file.path.name))
        else:
            time_values.extend(t_vals)
        voltage_values.extend(v_vals)
        if show_temperature:
            try:
                _temp_time, temp_vals = _required_numeric_series(parsed, "T", "Temp")
                temperature_values.extend(temp_vals)
            except ValueError:
                pass

    if (not time_values and not datetime_values) or not voltage_values:
        raise ValueError("No hay datos validos para ajustar los ejes.")

    if _is_date_axis(time_unit):
        _t_min, t_max = _padded_datetime_limits(datetime_values)
        t_min = min(datetime_values)
    else:
        time_scale = _time_unit_scale(time_unit)
        time_decimals = _time_unit_decimals(time_unit)
        scaled_time_values = [value / time_scale for value in time_values]
        _t_min, t_max = _padded_limits(scaled_time_values, decimals=time_decimals)
        t_min = 0.0
    voltage_decimals = _auto_decimals(voltage_values)
    v_min, v_max = _tight_limits(voltage_values, decimals=voltage_decimals)

    out = {
        "t_min": _format_time_limit(t_min, time_unit),
        "t_max": _format_time_limit(t_max, time_unit),
        "v_min": _format_limit_value(v_min, voltage_decimals),
        "v_max": _format_limit_value(v_max, voltage_decimals),
        "temp_min": "",
        "temp_max": "",
    }
    if show_temperature and temperature_values:
        temp_decimals = _auto_decimals(temperature_values)
        temp_min, temp_max = _tight_limits(temperature_values, decimals=temp_decimals)
        out["temp_min"] = _format_limit_value(temp_min, temp_decimals)
        out["temp_max"] = _format_limit_value(temp_max, temp_decimals)
    return out


def build_dv_dt_rows(
    parsed: ParsedDTA,
    time_unit: str = "s",
    smoothing_algorithm: str = "Median filter",
    smoothing_window: int = 1,
    source_name: str = "",
) -> list[dict[str, float | datetime]]:
    time_values, voltage_values = _required_numeric_series(parsed, "T", "Vf")
    dvdt_scale = _dv_dt_scale(time_unit)

    elapsed_seconds: list[float] = []
    dvdt_values: list[float] = []

    for idx in range(len(time_values) - 1):
        t0 = time_values[idx]
        t1 = time_values[idx + 1]
        v0 = voltage_values[idx]
        v1 = voltage_values[idx + 1]

        dt = t1 - t0
        if dt <= 0:
            continue

        elapsed_seconds.append((t0 + t1) / 2.0)
        dvdt_values.append(((v1 - v0) / dt) * dvdt_scale)

    smoothed_dvdt = apply_smoothing(dvdt_values, smoothing_algorithm, smoothing_window)
    plot_time_values = _plot_time_values(parsed, elapsed_seconds, time_unit, source_name or "archivo")
    dvdt_rows = [
        {"time": time_value, "dVdt": dvdt_value}
        for time_value, dvdt_value in zip(plot_time_values, smoothed_dvdt)
    ]

    if not dvdt_rows:
        raise ValueError("No hay suficientes datos validos para calcular dV/dt.")

    return dvdt_rows


def compute_default_dv_dt_limits(
    parsed_items: list[tuple[DegFile, ParsedDTA]],
    time_unit: str = "s",
    smoothing_algorithm: str = "Median filter",
    smoothing_window: int = 1,
    logarithmic_y: bool = True,
    decimals: int = 3,
) -> dict[str, str]:
    time_values: list[TimeAxisValue] = []
    dvdt_values: list[float] = []

    for deg_file, parsed in parsed_items:
        rows = _scale_compatible_dv_dt_rows(
            build_dv_dt_rows(
                parsed,
                time_unit=time_unit,
                smoothing_algorithm=smoothing_algorithm,
                smoothing_window=smoothing_window,
                source_name=deg_file.path.name,
            ),
            logarithmic_y=logarithmic_y,
        )
        time_values.extend(row["time"] for row in rows)
        dvdt_values.extend(row["dVdt"] for row in rows)

    if not time_values or not dvdt_values:
        return {
            "t_min": "",
            "t_max": "",
            "dvdt_min": "",
            "dvdt_max": "",
        }

    if _is_date_axis(time_unit):
        datetime_time_values = [value for value in time_values if isinstance(value, datetime)]
        _t_min, t_max = _padded_datetime_limits(datetime_time_values)
        t_min = min(datetime_time_values)
    else:
        _t_min, t_max = _padded_limits([float(value) for value in time_values], decimals=_time_unit_decimals(time_unit))
        t_min = 0.0
    if logarithmic_y:
        dvdt_min, dvdt_max = _log_axis_limits(dvdt_values)
    else:
        dvdt_min, dvdt_max = _adaptive_linear_limits(dvdt_values)

    return {
        "t_min": _format_time_limit(t_min, time_unit),
        "t_max": _format_time_limit(t_max, time_unit),
        "dvdt_min": _format_log_limit_value(dvdt_min) if logarithmic_y else _format_adaptive_limit_value(dvdt_min),
        "dvdt_max": _format_log_limit_value(dvdt_max) if logarithmic_y else _format_adaptive_limit_value(dvdt_max),
    }


def compute_autofit_dv_dt_limits(
    parsed_items: list[tuple[DegFile, ParsedDTA]],
    time_unit: str = "s",
    smoothing_algorithm: str = "Median filter",
    smoothing_window: int = 1,
    logarithmic_y: bool = True,
    locked_t_min: TimeAxisValue | None = None,
    locked_t_max: TimeAxisValue | None = None,
    locked_dvdt_min: float | None = None,
    locked_dvdt_max: float | None = None,
    decimals: int = 3,
) -> dict[str, str]:
    rows: list[dict[str, float | datetime]] = []

    for deg_file, parsed in parsed_items:
        rows.extend(
            _scale_compatible_dv_dt_rows(
                build_dv_dt_rows(
                    parsed,
                    time_unit=time_unit,
                    smoothing_algorithm=smoothing_algorithm,
                    smoothing_window=smoothing_window,
                    source_name=deg_file.path.name,
                ),
                logarithmic_y=logarithmic_y,
            )
        )

    filtered_rows: list[dict[str, float | datetime]] = []
    for row in rows:
        if locked_t_min is not None and row["time"] < locked_t_min:
            continue
        if locked_t_max is not None and row["time"] > locked_t_max:
            continue
        if locked_dvdt_min is not None and row["dVdt"] < locked_dvdt_min:
            continue
        if locked_dvdt_max is not None and row["dVdt"] > locked_dvdt_max:
            continue
        filtered_rows.append(row)

    if not filtered_rows:
        raise ValueError("No hay datos validos dentro de los limites bloqueados para ajustar los ejes.")

    time_values = [row["time"] for row in filtered_rows]
    dvdt_values = [row["dVdt"] for row in filtered_rows]

    if _is_date_axis(time_unit):
        datetime_time_values = [value for value in time_values if isinstance(value, datetime)]
        _t_min, t_max = _padded_datetime_limits(datetime_time_values)
        t_min = min(datetime_time_values)
    else:
        _t_min, t_max = _padded_limits([float(value) for value in time_values], decimals=_time_unit_decimals(time_unit))
        t_min = 0.0
    if logarithmic_y:
        dvdt_min, dvdt_max = _log_axis_limits(dvdt_values)
    else:
        dvdt_min, dvdt_max = _padded_limits(dvdt_values, decimals=decimals)

    return {
        "t_min": _format_time_limit(t_min, time_unit),
        "t_max": _format_time_limit(t_max, time_unit),
        "dvdt_min": _format_log_limit_value(dvdt_min) if logarithmic_y else _format_adaptive_limit_value(dvdt_min),
        "dvdt_max": _format_log_limit_value(dvdt_max) if logarithmic_y else _format_adaptive_limit_value(dvdt_max),
    }


def draw_dv_dt_on_figure(
    fig: Figure,
    parsed_items: list[tuple[DegFile, ParsedDTA]],
    dvdt_linestyle: str,
    time_unit: str = "s",
    smoothing_algorithm: str = "Median filter",
    smoothing_window: int = 1,
    logarithmic_y: bool = True,
    x_tick_count: int = 6,
    y_tick_count: int = 6,
    t_min: TimeAxisValue | None = None,
    t_max: TimeAxisValue | None = None,
    dvdt_min: float | None = None,
    dvdt_max: float | None = None,
    plot_title: str = "",
    show_title: bool = True,
    title_fontsize: float = 14,
    tick_fontsize: float = 10,
    label_fontsize: float = 11,
    legend_fontsize: float = 10,
    line_width: float = 1.5,
) -> bool:
    fig.clear()

    if not parsed_items:
        return False

    x_tick_count = max(2, int(x_tick_count))
    y_tick_count = max(2, int(y_tick_count))
    mpl_linestyle = _mpl_linestyle(dvdt_linestyle)

    if mpl_linestyle == "None":
        return False

    ax = fig.add_subplot(111)

    def _line_kwargs(color: str) -> dict:
        return {
            "color": color,
            "linestyle": mpl_linestyle,
            "linewidth": line_width,
        }

    for idx, (deg_file, parsed) in enumerate(parsed_items):
        delta_rows = _scale_compatible_dv_dt_rows(
            build_dv_dt_rows(
                parsed,
                time_unit=time_unit,
                smoothing_algorithm=smoothing_algorithm,
                smoothing_window=smoothing_window,
                source_name=deg_file.path.name,
            ),
            logarithmic_y=logarithmic_y,
        )
        if not delta_rows:
            continue
        color = _cycle_gradient_color(DEG_STAGE_GRADIENTS["dvdt"], idx, len(parsed_items))
        ax.plot(
            [row["time"] for row in delta_rows],
            [row["dVdt"] for row in delta_rows],
            label=f"Etapa #{deg_file.stage} dV/dt",
            **_line_kwargs(color),
        )

    handles, labels = ax.get_legend_handles_labels()
    if not handles:
        fig.clear()
        return False

    default_title = "dV/dt vs t - Degradacion galvanostatica"
    final_title = plot_title.strip() if plot_title.strip() else default_title

    x_label = "Fecha / hora" if _is_date_axis(time_unit) else f"Tiempo [{time_unit}]"
    dvdt_label_unit = "s" if _is_date_axis(time_unit) else time_unit
    ax.set_xlabel(x_label, fontsize=label_fontsize)
    ax.set_ylabel(f"dV/dt [{_dv_dt_unit_label(dvdt_label_unit)}]", fontsize=label_fontsize)
    ax.set_title(final_title if show_title else "", fontsize=title_fontsize)
    ax.grid(True, alpha=0.25)
    ax.tick_params(axis="both", labelsize=tick_fontsize)
    apply_x_tick_label_padding(ax, tick_fontsize)
    _apply_time_axis_format(ax, time_unit, x_tick_count)

    if t_min is not None or t_max is not None:
        ax.set_xlim(left=t_min, right=t_max)
        if not _is_date_axis(time_unit):
            apply_x_edge_ticks(ax, t_min, t_max, x_tick_count)

    if logarithmic_y:
        if dvdt_min is not None and dvdt_min <= 0:
            raise ValueError("dV/dt min debe ser mayor que 0 para usar escala logaritmica.")
        if dvdt_max is not None and dvdt_max <= 0:
            raise ValueError("dV/dt max debe ser mayor que 0 para usar escala logaritmica.")

        ax.set_yscale("log")
        if dvdt_min is not None or dvdt_max is not None:
            ax.set_ylim(bottom=dvdt_min, top=dvdt_max)

        ax.yaxis.set_major_locator(LogLocator(base=10.0, numticks=max(2, y_tick_count)))
        ax.yaxis.set_minor_locator(LogLocator(base=10.0, subs=tuple(range(2, 10))))
        ax.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
        ax.yaxis.set_minor_formatter(LogFormatterSciNotation(base=10.0, labelOnlyBase=True))
        ax.grid(True, which="minor", axis="y", alpha=0.25)
    else:
        if dvdt_min is not None or dvdt_max is not None:
            ax.set_ylim(bottom=dvdt_min, top=dvdt_max)
            ax.yaxis.set_major_locator(LinearLocator(y_tick_count))
        else:
            ax.yaxis.set_major_locator(MaxNLocator(nbins=y_tick_count))
        ax.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))

    make_legend_draggable(ax.legend(handles, labels, fontsize=legend_fontsize))
    fig.tight_layout()
    return True


def draw_v_vs_t_on_figure(
    fig: Figure,
    parsed_items: list[tuple[DegFile, ParsedDTA]],
    show_temperature: bool,
    voltage_linestyle: str = "-",
    temperature_linestyle: str = "--",
    time_unit: str = "s",
    x_tick_count: int = 6,
    y_tick_count: int = 6,
    plot_title: str = "",
    show_title: bool = True,
    title_fontsize: float = 14,
    tick_fontsize: float = 10,
    label_fontsize: float = 11,
    legend_fontsize: float = 10,
    line_width: float = 1.5,
    t_min: TimeAxisValue | None = None,
    t_max: TimeAxisValue | None = None,
    v_min: float | None = None,
    v_max: float | None = None,
    temp_min: float | None = None,
    temp_max: float | None = None,
    fit_t_min: TimeAxisValue | None = None,
    fit_t_max: TimeAxisValue | None = None,
    fit_use_linear: bool = False,
    show_fit_line: bool = True,
    show_fit_range: bool = False,
    reference_start: datetime | None = None,
) -> bool:
    fig.clear()

    if not parsed_items:
        return False

    has_voltage = _mpl_linestyle(voltage_linestyle) != "None"
    has_temperature = show_temperature and _mpl_linestyle(temperature_linestyle) != "None"
    if not has_voltage and not has_temperature:
        return False

    ax = fig.add_subplot(111)
    ax_temp = ax.twinx() if has_temperature else None

    for idx, (deg_file, parsed) in enumerate(parsed_items):
        t_vals, v_vals = _required_numeric_series(parsed, "T", "Vf")
        plot_t_vals = _plot_time_values(parsed, t_vals, time_unit, deg_file.path.name)
        voltage_color = _cycle_gradient_color(DEG_STAGE_GRADIENTS["voltage"], idx, len(parsed_items))
        temperature_color = _cycle_gradient_color(DEG_STAGE_GRADIENTS["temperature"], idx, len(parsed_items))
        if has_voltage:
            ax.plot(
                plot_t_vals,
                v_vals,
                color=voltage_color,
                linewidth=line_width,
                linestyle=_mpl_linestyle(voltage_linestyle),
                label=f"Etapa #{deg_file.stage} V",
            )

            if show_fit_line:
                if _is_date_axis(time_unit):
                    if reference_start is None:
                        raise ValueError("Falta referencia de tiempo para el eje de fecha.")
                    stage_start = _start_datetime(parsed, deg_file.path.name)
                    offset = (stage_start - reference_start).total_seconds()
                    time_seconds = [offset + value for value in t_vals]
                    fit_min_sec = (
                        _time_limit_to_seconds(fit_t_min, time_unit, reference_start)
                        if fit_t_min is not None
                        else None
                    )
                    fit_max_sec = (
                        _time_limit_to_seconds(fit_t_max, time_unit, reference_start)
                        if fit_t_max is not None
                        else None
                    )
                else:
                    time_seconds = list(t_vals)
                    fit_min_sec = (
                        float(fit_t_min) * _time_unit_scale(time_unit) if fit_t_min is not None else None
                    )
                    fit_max_sec = (
                        float(fit_t_max) * _time_unit_scale(time_unit) if fit_t_max is not None else None
                    )

                filt_t, filt_v = _filtered_time_voltage(time_seconds, v_vals, fit_min_sec, fit_max_sec)
                if len(filt_t) >= 2:
                    if fit_use_linear:
                        params = _linear_fit_params(filt_t, filt_v)
                        if params is not None:
                            slope, intercept = params
                        else:
                            slope = None
                            intercept = None
                    else:
                        slope = _edge_slope(filt_t, filt_v)
                        intercept = filt_v[0] - slope * filt_t[0] if slope is not None else None

                    if slope is not None and intercept is not None:
                        t0 = filt_t[0]
                        t1 = filt_t[-1]
                        y0 = slope * t0 + intercept
                        y1 = slope * t1 + intercept
                        if _is_date_axis(time_unit):
                            plot_fit_t = [
                                _seconds_to_time_limit(t0, time_unit, reference_start),
                                _seconds_to_time_limit(t1, time_unit, reference_start),
                            ]
                        else:
                            scale = _time_unit_scale(time_unit)
                            plot_fit_t = [t0 / scale, t1 / scale]
                        ax.plot(
                            plot_fit_t,
                            [y0, y1],
                            color=voltage_color,
                            linewidth=max(1.0, line_width),
                            linestyle="--",
                            alpha=0.6,
                        )
        if has_temperature and ax_temp is not None:
            temp_time, temp_vals = _required_numeric_series(parsed, "T", "Temp")
            plot_temp_time = _plot_time_values(parsed, temp_time, time_unit, deg_file.path.name)
            ax_temp.plot(
                plot_temp_time,
                temp_vals,
                color=temperature_color,
                linewidth=line_width,
                linestyle=_mpl_linestyle(temperature_linestyle),
                label=f"Etapa #{deg_file.stage} T",
            )

    default_title = "V vs t - Degradacion galvanostatica"
    final_title = plot_title.strip() if plot_title.strip() else default_title

    x_label = "Fecha / hora" if _is_date_axis(time_unit) else f"Tiempo [{time_unit}]"
    ax.set_xlabel(x_label, fontsize=label_fontsize)
    if has_voltage:
        ax.set_ylabel("Voltaje [V]", fontsize=label_fontsize)
    ax.set_title(final_title if show_title else "", fontsize=title_fontsize)
    ax.grid(True, alpha=0.25)
    ax.tick_params(axis="both", labelsize=tick_fontsize)
    apply_x_tick_label_padding(ax, tick_fontsize)
    _apply_time_axis_format(ax, time_unit, x_tick_count)
    ax.yaxis.set_major_locator(MaxNLocator(nbins=max(2, int(y_tick_count))))
    ax.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))

    if t_min is not None or t_max is not None:
        ax.set_xlim(left=t_min, right=t_max)
        if not _is_date_axis(time_unit):
            apply_x_edge_ticks(ax, t_min, t_max, x_tick_count)
    if has_voltage and (v_min is not None or v_max is not None):
        ax.set_ylim(bottom=v_min, top=v_max)
        ax.yaxis.set_major_locator(LinearLocator(max(2, int(y_tick_count))))
        ax.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))

    if show_fit_range and fit_t_min is not None and fit_t_max is not None:
        ax.axvspan(
            fit_t_min,
            fit_t_max,
            color=DEG_PLOT_COLORS["range"],
            alpha=0.18,
            zorder=0,
        )

    handles, labels = ax.get_legend_handles_labels()
    if ax_temp is not None:
        ax_temp.tick_params(axis="y", labelsize=tick_fontsize)
        ax_temp.set_ylabel("Temperatura [°C]", fontsize=label_fontsize)
        if temp_min is not None or temp_max is not None:
            ax_temp.set_ylim(bottom=temp_min, top=temp_max)
            ax_temp.yaxis.set_major_locator(LinearLocator(max(2, int(y_tick_count))))
        else:
            ax_temp.yaxis.set_major_locator(MaxNLocator(nbins=max(2, int(y_tick_count))))
        ax_temp.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
        h2, l2 = ax_temp.get_legend_handles_labels()
        handles += h2
        labels += l2

    if has_voltage:
        ax.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))

    if handles:
        make_legend_draggable(ax.legend(handles, labels, fontsize=legend_fontsize))
    fig.tight_layout()
    return True


def open_v_vs_t_window(
    input_dir: Path,
    font_defaults: PlotFontDefaults | None = None,
    language: str = "es",
) -> None:
    language = _deg_language(language)
    deg_files = find_deg_files(Path(input_dir))
    if not deg_files:
        raise ValueError("No se encontraron archivos de degradacion validos.")

    font_defaults = resolve_plot_font_defaults(font_defaults)
    font_default_values = font_defaults.as_strings()
    parsed_items = [(deg_file, parse_gamry_dta(deg_file.path)) for deg_file in deg_files]
    stage_vars = {deg_file.stage: tk.BooleanVar(value=True) for deg_file in deg_files}
    default_limits = compute_default_v_vs_t_limits(parsed_items, time_unit="s")

    win = tk.Toplevel()
    win.title("Deg - V vs t")
    win.geometry("1200x720")

    controls_host, plot_outer = create_resizable_plot_layout(win, sidebar_width=320)
    controls_frame = _build_scrollable_controls(controls_host)

    toolbar_frame = ttk.Frame(plot_outer)
    toolbar_frame.pack(side="top", fill="x")

    canvas_frame = ttk.Frame(plot_outer)
    canvas_frame.pack(side="top", fill="both", expand=True)

    fig = Figure(figsize=(8.8, 5.4), dpi=100)

    canvas = FigureCanvasTkAgg(fig, master=canvas_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    toolbar = NavigationToolbar2Tk(canvas, toolbar_frame, pack_toolbar=False)
    toolbar.update()
    toolbar.pack(side="left", fill="x")

    status_var = tk.StringVar(value="Listo.")
    temperature_var = tk.BooleanVar(value=False)
    time_unit_var = tk.StringVar(value="s")
    voltage_line_var = tk.StringVar(value="-")
    temperature_line_var = tk.StringVar(value="--")
    x_tick_count_var = tk.IntVar(value=6)
    y_tick_count_var = tk.IntVar(value=6)
    plot_title_var = tk.StringVar(value="")
    show_title_var = tk.BooleanVar(value=True)
    title_fontsize_var = tk.StringVar(value=font_default_values["title"])
    tick_fontsize_var = tk.StringVar(value=font_default_values["tick"])
    label_fontsize_var = tk.StringVar(value=font_default_values["label"])
    legend_fontsize_var = tk.StringVar(value=font_default_values["legend"])
    line_width_var = tk.StringVar(value="1.5")
    t_min_var = tk.StringVar(value=default_limits["t_min"])
    t_max_var = tk.StringVar(value=default_limits["t_max"])
    v_min_var = tk.StringVar(value=default_limits["v_min"])
    v_max_var = tk.StringVar(value=default_limits["v_max"])
    temp_min_var = tk.StringVar(value=default_limits["temp_min"])
    temp_max_var = tk.StringVar(value=default_limits["temp_max"])
    slope_t_min_var = tk.StringVar(value="")
    slope_t_max_var = tk.StringVar(value="")
    slope_fit_var = tk.BooleanVar(value=False)
    show_fit_var = tk.BooleanVar(value=False)
    show_range_var = tk.BooleanVar(value=True)
    slope_indicator_var = tk.StringVar(value="—")

    initial_state = {
        "temperature": False,
        "time_unit": "s",
        "voltage_line": "-",
        "temperature_line": "--",
        "x_tick_count": 6,
        "y_tick_count": 6,
        "plot_title": "",
        "show_title": True,
        "title_fontsize": font_default_values["title"],
        "tick_fontsize": font_default_values["tick"],
        "label_fontsize": font_default_values["label"],
        "legend_fontsize": font_default_values["legend"],
        "line_width": "1.5",
        "t_min": default_limits["t_min"],
        "t_max": default_limits["t_max"],
        "v_min": default_limits["v_min"],
        "v_max": default_limits["v_max"],
        "temp_min": default_limits["temp_min"],
        "temp_max": default_limits["temp_max"],
        "slope_t_min": "",
        "slope_t_max": "",
        "slope_fit": False,
        "show_fit": False,
        "show_range": True,
    }

    plot_job = {"id": None}
    current_time_unit = {"value": time_unit_var.get()}
    suspend_events = {"value": False}
    reference_start_cache: dict[str, datetime | None] = {"value": None}

    def _reference_start() -> datetime:
        if reference_start_cache["value"] is None:
            reference_start_cache["value"] = _reference_start_datetime(parsed_items)
        return reference_start_cache["value"]

    def _collect_limits():
        return dict(
            t_min=_parse_time_limit_text(t_min_var.get(), time_unit_var.get()),
            t_max=_parse_time_limit_text(t_max_var.get(), time_unit_var.get()),
            v_min=_optional_float(v_min_var.get()),
            v_max=_optional_float(v_max_var.get()),
            temp_min=_optional_float(temp_min_var.get()),
            temp_max=_optional_float(temp_max_var.get()),
        )

    def _collect_slope_limits() -> tuple[TimeAxisValue | None, TimeAxisValue | None]:
        return (
            _parse_time_limit_text(slope_t_min_var.get(), time_unit_var.get()),
            _parse_time_limit_text(slope_t_max_var.get(), time_unit_var.get()),
        )

    def _update_degradation_indicator(visible_items: list[tuple[DegFile, ParsedDTA]]) -> None:
        if not visible_items:
            slope_indicator_var.set("Seleccione al menos una etapa.")
            return

        try:
            t_min, t_max = _collect_slope_limits()
        except ValueError as exc:
            slope_indicator_var.set(f"Error: {exc}")
            return
        use_fit = slope_fit_var.get()
        try:
            ref_start = _reference_start() if _is_date_axis(time_unit_var.get()) else None
        except ValueError as exc:
            slope_indicator_var.set(f"Error: {exc}")
            return
        lines: list[str] = []

        for deg_file, parsed in visible_items:
            try:
                slope, _count = compute_degradation_rate(
                    parsed=parsed,
                    source_name=deg_file.path.name,
                    time_unit=time_unit_var.get(),
                    reference_start=ref_start,
                    t_min=t_min,
                    t_max=t_max,
                    use_linear_fit=use_fit,
                )
            except ValueError as exc:
                lines.append(f"Etapa #{deg_file.stage}: Error ({exc})")
                continue

            if slope is None:
                lines.append(f"Etapa #{deg_file.stage}: Sin datos")
                continue

            slope_uv_hour = slope * 1e6 * SECONDS_PER_HOUR
            slope_text = _format_adaptive_limit_value(slope_uv_hour)
            lines.append(f"Etapa #{deg_file.stage}: {slope_text} µV/h")


        slope_indicator_var.set("\n".join(lines) if lines else "Sin datos.")

    def _plot():
        plot_job["id"] = None
        visible_items = _selected_stage_items(parsed_items, stage_vars)
        if not visible_items:
            fig.clear()
            canvas.draw_idle()
            status_var.set("No se muestra grafico: seleccione al menos una etapa.")
            _update_degradation_indicator([])
            return
        try:
            fit_t_min, fit_t_max = _collect_slope_limits()
        except ValueError:
            fit_t_min, fit_t_max = None, None
        show_fit_range = show_range_var.get() and fit_t_min is not None and fit_t_max is not None
        try:
            has_plot = draw_v_vs_t_on_figure(
                fig=fig,
                parsed_items=visible_items,
                show_temperature=temperature_var.get(),
                voltage_linestyle=voltage_line_var.get(),
                temperature_linestyle=temperature_line_var.get(),
                time_unit=time_unit_var.get(),
                x_tick_count=x_tick_count_var.get(),
                y_tick_count=y_tick_count_var.get(),
                plot_title=plot_title_var.get(),
                show_title=show_title_var.get(),
                title_fontsize=_positive_float(title_fontsize_var.get(), "Tamaño del título"),
                tick_fontsize=_positive_float(tick_fontsize_var.get(), "Tamaño de ticks"),
                label_fontsize=_positive_float(label_fontsize_var.get(), "Tamaño de etiquetas"),
                legend_fontsize=_positive_float(legend_fontsize_var.get(), "Tamaño de leyenda"),
                line_width=_positive_float(line_width_var.get(), "Grosor de línea"),
                fit_t_min=fit_t_min,
                fit_t_max=fit_t_max,
                fit_use_linear=slope_fit_var.get(),
                show_fit_line=show_fit_var.get(),
                show_fit_range=show_fit_range,
                reference_start=_reference_start() if _is_date_axis(time_unit_var.get()) else None,
                **_collect_limits(),
            )
        except ValueError as exc:
            fig.clear()
            canvas.draw_idle()
            status_var.set(f"Error: {exc}")
            _update_degradation_indicator(visible_items)
            return

        if not has_plot:
            fig.clear()
            canvas.draw_idle()
            status_var.set("No se muestra grafico: active al menos una serie.")
            _update_degradation_indicator(visible_items)
            return

        canvas.draw_idle()
        status_var.set("Grafico actualizado.")
        _update_degradation_indicator(visible_items)

    def _schedule_plot(*_args):
        if suspend_events["value"]:
            return
        if plot_job["id"] is not None:
            win.after_cancel(plot_job["id"])
        plot_job["id"] = win.after(20, _plot)

    def _autofit():
        visible_items = _selected_stage_items(parsed_items, stage_vars)
        if not visible_items:
            status_var.set("Autoescala no disponible: seleccione al menos una etapa.")
            return
        try:
            fitted = compute_autofit_v_vs_t_limits(
                visible_items,
                temperature_var.get(),
                time_unit=time_unit_var.get(),
            )
        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            return

        suspend_events["value"] = True
        try:
            t_min_var.set(fitted["t_min"])
            t_max_var.set(fitted["t_max"])
            v_min_var.set(fitted["v_min"])
            v_max_var.set(fitted["v_max"])
            temp_min_var.set(fitted["temp_min"])
            temp_max_var.set(fitted["temp_max"])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Autoescala aplicada.")

    def _on_time_unit_changed(*_args):
        if suspend_events["value"]:
            return

        old_unit = current_time_unit["value"]
        new_unit = time_unit_var.get()
        old_t_min = t_min_var.get()
        old_t_max = t_max_var.get()
        old_slope_t_min = slope_t_min_var.get()
        old_slope_t_max = slope_t_max_var.get()

        suspend_events["value"] = True
        try:
            if t_min_var.get().strip():
                t_min_var.set(_convert_time_limit_text(t_min_var.get(), old_unit, new_unit, _reference_start()))
            if t_max_var.get().strip():
                t_max_var.set(_convert_time_limit_text(t_max_var.get(), old_unit, new_unit, _reference_start()))
            if slope_t_min_var.get().strip():
                slope_t_min_var.set(
                    _convert_time_limit_text(slope_t_min_var.get(), old_unit, new_unit, _reference_start())
                )
            if slope_t_max_var.get().strip():
                slope_t_max_var.set(
                    _convert_time_limit_text(slope_t_max_var.get(), old_unit, new_unit, _reference_start())
                )
            current_time_unit["value"] = new_unit
        except ValueError as exc:
            t_min_var.set(old_t_min)
            t_max_var.set(old_t_max)
            slope_t_min_var.set(old_slope_t_min)
            slope_t_max_var.set(old_slope_t_max)
            current_time_unit["value"] = old_unit
            time_unit_var.set(old_unit)
            status_var.set(f"Error: {exc}")
            return
        finally:
            suspend_events["value"] = False

        _schedule_plot()

    def _reset():
        suspend_events["value"] = True
        try:
            temperature_var.set(initial_state["temperature"])
            current_time_unit["value"] = initial_state["time_unit"]
            time_unit_var.set(initial_state["time_unit"])
            voltage_line_var.set(initial_state["voltage_line"])
            temperature_line_var.set(initial_state["temperature_line"])
            x_tick_count_var.set(initial_state["x_tick_count"])
            y_tick_count_var.set(initial_state["y_tick_count"])
            plot_title_var.set(initial_state["plot_title"])
            show_title_var.set(initial_state["show_title"])
            title_fontsize_var.set(initial_state["title_fontsize"])
            tick_fontsize_var.set(initial_state["tick_fontsize"])
            label_fontsize_var.set(initial_state["label_fontsize"])
            legend_fontsize_var.set(initial_state["legend_fontsize"])
            line_width_var.set(initial_state["line_width"])
            t_min_var.set(initial_state["t_min"])
            t_max_var.set(initial_state["t_max"])
            v_min_var.set(initial_state["v_min"])
            v_max_var.set(initial_state["v_max"])
            temp_min_var.set(initial_state["temp_min"])
            temp_max_var.set(initial_state["temp_max"])
            slope_t_min_var.set(initial_state["slope_t_min"])
            slope_t_max_var.set(initial_state["slope_t_max"])
            slope_fit_var.set(initial_state["slope_fit"])
            show_fit_var.set(initial_state["show_fit"])
            show_range_var.set(initial_state["show_range"])
            for var in stage_vars.values():
                var.set(True)
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Valores restaurados.")

    def _export_report() -> None:
        visible_items = _selected_stage_items(parsed_items, stage_vars)
        if not visible_items:
            status_var.set("Reporte no disponible: seleccione al menos una etapa.")
            return
        try:
            stage_text = f"{visible_items[0][0].stage}-{visible_items[-1][0].stage}"
            default_name = f"Deg_Report_stages_{_safe_filename_part(stage_text)}.pdf"
            path_text = filedialog.asksaveasfilename(
                parent=win,
                title=translate("save_deg_report", language),
                initialfile=default_name,
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            )
            if not path_text:
                status_var.set(translate("export_cancelled", language))
                return
            exported_path = export_deg_v_vs_t_report_pdf(
                visible_items,
                Path(path_text),
                language=language,
                font_defaults=font_defaults,
            )
        except Exception as exc:
            status_var.set(f"Error al exportar reporte: {type(exc).__name__}: {exc}")
            messagebox.showerror("Deg Report", str(exc), parent=win)
            return

        status_var.set(f"{translate('report_exported', language)}: {exported_path}")

    def _open_composer_placeholder():
        existing = getattr(win, "_composer_win", None)
        if existing is not None and existing.winfo_exists():
            existing.lift()
            existing.focus_force()
            return

        comp = tk.Toplevel(win)
        win._composer_win = comp  # type: ignore[attr-defined]
        comp.title("Componer (Deg)")
        comp.geometry("520x260")

        body = ttk.Frame(comp, padding=16)
        body.pack(fill="both", expand=True)

        ttk.Label(
            body,
            text="Componer provisional",
            font=("TkDefaultFont", 11, "bold"),
        ).pack(anchor="w", pady=(0, 10))

        ttk.Label(
            body,
            text=(
                "Este espacio se reservará para componer curvas de degradación "
                "y superponer distintas etapas en un único gráfico."
            ),
            wraplength=460,
            justify="left",
        ).pack(anchor="w", pady=(0, 10))

        ttk.Label(
            body,
            text=(
                f"Etapas detectadas actualmente: "
                f"{', '.join(f'#{item.stage}' for item in deg_files)}"
            ),
            wraplength=460,
            justify="left",
        ).pack(anchor="w", pady=(0, 14))

        ttk.Button(body, text="Cerrar", command=comp.destroy).pack(anchor="e")

        def _on_close():
            comp.destroy()
            try:
                delattr(win, "_composer_win")
            except Exception:
                pass

        comp.protocol("WM_DELETE_WINDOW", _on_close)

    ttk.Label(
        controls_frame,
        text=f"Etapas detectadas: {', '.join(f'#{item.stage}' for item in deg_files)}",
        justify="left",
        wraplength=260,
    ).pack(anchor="w", pady=(0, 10))

    stages_box = ttk.LabelFrame(controls_frame, text="Etapas")
    stages_box.pack(fill="x", pady=5)

    series_box = ttk.LabelFrame(controls_frame, text="Series")
    series_box.pack(fill="x", pady=5)

    style_box = ttk.LabelFrame(controls_frame, text="Estilo")
    style_box.pack(fill="x", pady=5)

    text_box = ttk.LabelFrame(controls_frame, text="Texto / tamaños")
    text_box.pack(fill="x", pady=5)

    degradation_box = ttk.LabelFrame(controls_frame, text="Degradación")
    degradation_box.pack(fill="x", pady=5)

    limits_box = ttk.LabelFrame(controls_frame, text="Límites de ejes")
    limits_box.pack(fill="x", pady=5)

    stages_inner, bind_stage_scroll = _build_scrollable_cycle_selector(
        stages_box,
        height=DEG_STAGE_SELECTOR_HEIGHT,
    )
    for index, deg_file in enumerate(deg_files):
        stage_toggle = ttk.Checkbutton(
            stages_inner,
            text=f"Etapa #{deg_file.stage}",
            variable=stage_vars[deg_file.stage],
            command=_schedule_plot,
        )
        pady = (0, 2) if index < len(deg_files) - 1 else 0
        stage_toggle.pack(anchor="w", padx=4, pady=pady)
        bind_stage_scroll(stage_toggle)

    ttk.Checkbutton(
        series_box,
        text="Temperatura",
        variable=temperature_var,
        command=_schedule_plot,
    ).pack(anchor="w", padx=8, pady=4)
    ttk.Label(series_box, text="Unidad de tiempo").pack(anchor="w", padx=8, pady=(8, 2))
    time_unit_combo = ttk.Combobox(
        series_box,
        textvariable=time_unit_var,
        values=TIME_UNIT_OPTIONS,
        state="readonly",
        width=8,
    )
    time_unit_combo.pack(anchor="w", padx=8, pady=(0, 4))

    ttk.Label(style_box, text="Línea de voltaje").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    voltage_line_combo = ttk.Combobox(
        style_box,
        textvariable=voltage_line_var,
        values=LINESTYLE_OPTIONS,
        state="readonly",
        width=10,
    )
    voltage_line_combo.grid(row=0, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Línea de temperatura").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    temperature_line_combo = ttk.Combobox(
        style_box,
        textvariable=temperature_line_var,
        values=LINESTYLE_OPTIONS,
        state="readonly",
        width=10,
    )
    temperature_line_combo.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Título").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    title_entry = ttk.Entry(text_box, textvariable=plot_title_var, width=28)
    title_entry.grid(row=0, column=1, sticky="we", padx=8, pady=3)
    ttk.Checkbutton(text_box, text="Mostrar titulo", variable=show_title_var, command=_schedule_plot).grid(
        row=7, column=0, columnspan=2, sticky="w", padx=8, pady=3
    )

    ttk.Label(text_box, text="Tamaño del título").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    title_size_spin = tk.Spinbox(text_box, from_=6, to=30, increment=0.5, textvariable=title_fontsize_var, width=8)
    title_size_spin.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de ticks").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    tick_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=tick_fontsize_var, width=8)
    tick_size_spin.grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de etiquetas").grid(row=3, column=0, sticky="w", padx=8, pady=3)
    label_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=label_fontsize_var, width=8)
    label_size_spin.grid(row=3, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de leyenda").grid(row=4, column=0, sticky="w", padx=8, pady=3)
    legend_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=legend_fontsize_var, width=8)
    legend_size_spin.grid(row=4, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Grosor de línea").grid(row=5, column=0, sticky="w", padx=8, pady=3)
    line_width_spin = tk.Spinbox(text_box, from_=0.5, to=5.0, increment=0.1, textvariable=line_width_var, width=8)
    line_width_spin.grid(row=5, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(degradation_box, text="t min").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    slope_t_min_entry = ttk.Entry(degradation_box, textvariable=slope_t_min_var, width=16)
    slope_t_min_entry.grid(row=0, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(degradation_box, text="t max").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    slope_t_max_entry = ttk.Entry(degradation_box, textvariable=slope_t_max_var, width=16)
    slope_t_max_entry.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    slope_fit_check = ttk.Checkbutton(
        degradation_box,
        text="Ajuste lineal",
        variable=slope_fit_var,
        command=_schedule_plot,
    )
    slope_fit_check.grid(row=2, column=0, columnspan=2, sticky="w", padx=8, pady=(3, 3))

    show_fit_check = ttk.Checkbutton(
        degradation_box,
        text="Mostrar ajuste",
        variable=show_fit_var,
        command=_schedule_plot,
    )
    show_fit_check.grid(row=3, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 6))

    show_range_check = ttk.Checkbutton(
        degradation_box,
        text="Mostrar rango",
        variable=show_range_var,
        command=_schedule_plot,
    )
    show_range_check.grid(row=4, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 6))

    ttk.Label(degradation_box, text="dV/dt promedio [µV/h]").grid(
        row=5,
        column=0,
        columnspan=2,
        sticky="w",
        padx=8,
        pady=(6, 2),
    )
    slope_indicator_label = ttk.Label(
        degradation_box,
        textvariable=slope_indicator_var,
        justify="left",
        wraplength=240,
    )
    slope_indicator_label.grid(row=6, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 6))

    limit_specs = [
        ("t min", t_min_var),
        ("t max", t_max_var),
        ("V min", v_min_var),
        ("V max", v_max_var),
        ("T min", temp_min_var),
        ("T max", temp_max_var),
    ]

    for row_idx, (label, var) in enumerate(limit_specs):
        ttk.Label(limits_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        entry = ttk.Entry(limits_box, textvariable=var, width=22)
        entry.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        entry.bind("<Return>", _schedule_plot)
        entry.bind("<KP_Enter>", _schedule_plot)
        entry.bind("<FocusOut>", _schedule_plot)

    for entry in (slope_t_min_entry, slope_t_max_entry):
        entry.bind("<Return>", _schedule_plot)
        entry.bind("<KP_Enter>", _schedule_plot)
        entry.bind("<FocusOut>", _schedule_plot)

    ttk.Label(limits_box, text="x-Ticks").grid(row=len(limit_specs), column=0, sticky="w", padx=8, pady=3)
    x_tick_spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=x_tick_count_var, width=8)
    x_tick_spin.grid(row=len(limit_specs), column=1, sticky="w", padx=8, pady=3)

    ttk.Label(limits_box, text="y-Ticks").grid(row=len(limit_specs) + 1, column=0, sticky="w", padx=8, pady=3)
    y_tick_spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=y_tick_count_var, width=8)
    y_tick_spin.grid(row=len(limit_specs) + 1, column=1, sticky="w", padx=8, pady=3)

    for combo in (voltage_line_combo, temperature_line_combo):
        combo.bind("<<ComboboxSelected>>", _schedule_plot)
    time_unit_combo.bind("<<ComboboxSelected>>", _on_time_unit_changed)
    for spin in (x_tick_spin, y_tick_spin):
        spin.bind("<Return>", _schedule_plot)
        spin.bind("<FocusOut>", _schedule_plot)
        spin.config(command=_schedule_plot)

    for widget in (
        title_entry,
        title_size_spin,
        tick_size_spin,
        label_size_spin,
        legend_size_spin,
        line_width_spin,
    ):
        widget.bind("<Return>", _schedule_plot)
        widget.bind("<KP_Enter>", _schedule_plot)
        widget.bind("<FocusOut>", _schedule_plot)

    for spin in (
        title_size_spin,
        tick_size_spin,
        label_size_spin,
        legend_size_spin,
        line_width_spin,
    ):
        spin.config(command=_schedule_plot)

    ttk.Label(
        controls_frame,
        textvariable=status_var,
        wraplength=240,
        justify="left",
    ).pack(anchor="w", fill="x", pady=(10, 10))

    buttons_frame = ttk.Frame(controls_frame)
    buttons_frame.pack(fill="x", pady=(5, 0))

    ttk.Button(buttons_frame, text="Restablecer", command=_reset).pack(side="left", padx=(0, 6))
    ttk.Button(buttons_frame, text="Componer", command=_open_composer_placeholder).pack(side="left", padx=(0, 6))
    ttk.Button(buttons_frame, text="Autoescala", command=_autofit).pack(side="left")
    ttk.Button(buttons_frame, text=translate("pdf_report", language), command=_export_report).pack(side="left", padx=(6, 0))

    _plot()


def open_dv_dt_window(input_dir: Path, font_defaults: PlotFontDefaults | None = None) -> None:
    deg_files = find_deg_files(Path(input_dir))
    if not deg_files:
        raise ValueError("No se encontraron archivos de degradacion validos.")

    font_defaults = resolve_plot_font_defaults(font_defaults)
    font_default_values = font_defaults.as_strings()
    parsed_items = [(deg_file, parse_gamry_dta(deg_file.path)) for deg_file in deg_files]
    stage_vars = {deg_file.stage: tk.BooleanVar(value=True) for deg_file in deg_files}
    default_limits = compute_default_dv_dt_limits(
        parsed_items,
        time_unit="s",
        smoothing_algorithm="Median filter",
        smoothing_window=1,
        logarithmic_y=True,
    )

    win = tk.Toplevel()
    win.title("Deg - dV/dt")
    win.geometry("1200x720")

    controls_host, plot_outer = create_resizable_plot_layout(win, sidebar_width=320)
    controls_frame = _build_scrollable_controls(controls_host)

    toolbar_frame = ttk.Frame(plot_outer)
    toolbar_frame.pack(side="top", fill="x")

    canvas_frame = ttk.Frame(plot_outer)
    canvas_frame.pack(side="top", fill="both", expand=True)

    fig = Figure(figsize=(8.8, 5.4), dpi=100)

    canvas = FigureCanvasTkAgg(fig, master=canvas_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    toolbar = NavigationToolbar2Tk(canvas, toolbar_frame, pack_toolbar=False)
    toolbar.update()
    toolbar.pack(side="left", fill="x")

    status_var = tk.StringVar(value="Listo.")
    time_unit_var = tk.StringVar(value="s")
    logarithmic_y_var = tk.BooleanVar(value=True)
    dvdt_line_var = tk.StringVar(value="-")
    smoothing_algorithm_var = tk.StringVar(value="Median filter")
    smoothing_window_var = tk.IntVar(value=1)
    x_tick_count_var = tk.IntVar(value=6)
    y_tick_count_var = tk.IntVar(value=6)
    plot_title_var = tk.StringVar(value="")
    show_title_var = tk.BooleanVar(value=True)
    title_fontsize_var = tk.StringVar(value=font_default_values["title"])
    tick_fontsize_var = tk.StringVar(value=font_default_values["tick"])
    label_fontsize_var = tk.StringVar(value=font_default_values["label"])
    legend_fontsize_var = tk.StringVar(value=font_default_values["legend"])
    line_width_var = tk.StringVar(value="1.5")
    t_min_var = tk.StringVar(value=default_limits["t_min"])
    t_max_var = tk.StringVar(value=default_limits["t_max"])
    dvdt_min_var = tk.StringVar(value=default_limits["dvdt_min"])
    dvdt_max_var = tk.StringVar(value=default_limits["dvdt_max"])
    t_min_lock_var = tk.BooleanVar(value=False)
    t_max_lock_var = tk.BooleanVar(value=False)
    dvdt_min_lock_var = tk.BooleanVar(value=False)
    dvdt_max_lock_var = tk.BooleanVar(value=False)

    initial_state = {
        "time_unit": "s",
        "logarithmic_y": True,
        "dvdt_line": "-",
        "smoothing_algorithm": "Median filter",
        "smoothing_window": 1,
        "x_tick_count": 6,
        "y_tick_count": 6,
        "plot_title": "",
        "show_title": True,
        "title_fontsize": font_default_values["title"],
        "tick_fontsize": font_default_values["tick"],
        "label_fontsize": font_default_values["label"],
        "legend_fontsize": font_default_values["legend"],
        "line_width": "1.5",
        "t_min": default_limits["t_min"],
        "t_max": default_limits["t_max"],
        "dvdt_min": default_limits["dvdt_min"],
        "dvdt_max": default_limits["dvdt_max"],
        "t_min_lock": False,
        "t_max_lock": False,
        "dvdt_min_lock": False,
        "dvdt_max_lock": False,
    }

    plot_job = {"id": None}
    current_time_unit = {"value": time_unit_var.get()}
    suspend_events = {"value": False}
    limit_entries: dict[str, ttk.Entry] = {}
    reference_start_cache: dict[str, datetime | None] = {"value": None}

    def _reference_start() -> datetime:
        if reference_start_cache["value"] is None:
            reference_start_cache["value"] = _reference_start_datetime(parsed_items)
        return reference_start_cache["value"]

    def _smoothing_config() -> dict[str, object]:
        return {
            "smoothing_algorithm": smoothing_algorithm_var.get(),
            "smoothing_window": smoothing_window_var.get(),
        }

    def _axis_mode_config() -> dict[str, object]:
        return {
            "logarithmic_y": logarithmic_y_var.get(),
        }

    def _collect_limits():
        return dict(
            t_min=_parse_time_limit_text(t_min_var.get(), time_unit_var.get()),
            t_max=_parse_time_limit_text(t_max_var.get(), time_unit_var.get()),
            dvdt_min=_optional_float(dvdt_min_var.get()),
            dvdt_max=_optional_float(dvdt_max_var.get()),
        )

    def _apply_lock_states() -> None:
        states = {
            "t_min": t_min_lock_var.get(),
            "t_max": t_max_lock_var.get(),
            "dvdt_min": dvdt_min_lock_var.get(),
            "dvdt_max": dvdt_max_lock_var.get(),
        }
        for key, entry in limit_entries.items():
            entry.state(["disabled"] if states[key] else ["!disabled"])

    def _convert_dvdt_limit_text(value_text: str, from_unit: str, to_unit: str) -> str:
        value = _optional_float(value_text)
        if value is None or from_unit == to_unit:
            return value_text
        from_dvdt_unit = "s" if _is_date_axis(from_unit) else from_unit
        to_dvdt_unit = "s" if _is_date_axis(to_unit) else to_unit
        base_v_per_s = value / _dv_dt_scale(from_dvdt_unit)
        converted = base_v_per_s * _dv_dt_scale(to_dvdt_unit)
        return _format_log_limit_value(converted) if logarithmic_y_var.get() else _format_adaptive_limit_value(converted)

    def _plot():
        plot_job["id"] = None
        visible_items = _selected_stage_items(parsed_items, stage_vars)
        if not visible_items:
            fig.clear()
            canvas.draw_idle()
            status_var.set("No se muestra grafico: seleccione al menos una etapa.")
            return
        try:
            has_plot = draw_dv_dt_on_figure(
                fig=fig,
                parsed_items=visible_items,
                dvdt_linestyle=dvdt_line_var.get(),
                time_unit=time_unit_var.get(),
                x_tick_count=x_tick_count_var.get(),
                y_tick_count=y_tick_count_var.get(),
                plot_title=plot_title_var.get(),
                show_title=show_title_var.get(),
                title_fontsize=_positive_float(title_fontsize_var.get(), "Tamaño del título"),
                tick_fontsize=_positive_float(tick_fontsize_var.get(), "Tamaño de ticks"),
                label_fontsize=_positive_float(label_fontsize_var.get(), "Tamaño de etiquetas"),
                legend_fontsize=_positive_float(legend_fontsize_var.get(), "Tamaño de leyenda"),
                line_width=_positive_float(line_width_var.get(), "Grosor de línea"),
                **_smoothing_config(),
                **_axis_mode_config(),
                **_collect_limits(),
            )
        except ValueError as exc:
            fig.clear()
            canvas.draw_idle()
            status_var.set(f"Error: {exc}")
            return

        if not has_plot:
            fig.clear()
            canvas.draw_idle()
            status_var.set("No se muestra grafico: active una linea visible.")
            return

        canvas.draw_idle()
        status_var.set("Grafico actualizado.")

    def _schedule_plot(*_args):
        if suspend_events["value"]:
            return
        if plot_job["id"] is not None:
            win.after_cancel(plot_job["id"])
        plot_job["id"] = win.after(20, _plot)

    def _autofit():
        visible_items = _selected_stage_items(parsed_items, stage_vars)
        if not visible_items:
            status_var.set("Autoescala no disponible: seleccione al menos una etapa.")
            return
        try:
            fitted = compute_autofit_dv_dt_limits(
                visible_items,
                time_unit=time_unit_var.get(),
                **_smoothing_config(),
                **_axis_mode_config(),
                locked_t_min=_parse_time_limit_text(t_min_var.get(), time_unit_var.get()) if t_min_lock_var.get() else None,
                locked_t_max=_parse_time_limit_text(t_max_var.get(), time_unit_var.get()) if t_max_lock_var.get() else None,
                locked_dvdt_min=_optional_float(dvdt_min_var.get()) if dvdt_min_lock_var.get() else None,
                locked_dvdt_max=_optional_float(dvdt_max_var.get()) if dvdt_max_lock_var.get() else None,
            )
        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            return

        suspend_events["value"] = True
        try:
            if not t_min_lock_var.get():
                t_min_var.set(fitted["t_min"])
            if not t_max_lock_var.get():
                t_max_var.set(fitted["t_max"])
            if not dvdt_min_lock_var.get():
                dvdt_min_var.set(fitted["dvdt_min"])
            if not dvdt_max_lock_var.get():
                dvdt_max_var.set(fitted["dvdt_max"])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Autoescala aplicada.")

    def _on_time_unit_changed(*_args):
        if suspend_events["value"]:
            return

        old_unit = current_time_unit["value"]
        new_unit = time_unit_var.get()
        old_t_min = t_min_var.get()
        old_t_max = t_max_var.get()
        old_dvdt_min = dvdt_min_var.get()
        old_dvdt_max = dvdt_max_var.get()

        suspend_events["value"] = True
        try:
            if not t_min_lock_var.get() and t_min_var.get().strip():
                t_min_var.set(_convert_time_limit_text(t_min_var.get(), old_unit, new_unit, _reference_start()))
            if not t_max_lock_var.get() and t_max_var.get().strip():
                t_max_var.set(_convert_time_limit_text(t_max_var.get(), old_unit, new_unit, _reference_start()))
            if not dvdt_min_lock_var.get() and dvdt_min_var.get().strip():
                dvdt_min_var.set(_convert_dvdt_limit_text(dvdt_min_var.get(), old_unit, new_unit))
            if not dvdt_max_lock_var.get() and dvdt_max_var.get().strip():
                dvdt_max_var.set(_convert_dvdt_limit_text(dvdt_max_var.get(), old_unit, new_unit))
            current_time_unit["value"] = new_unit
        except ValueError as exc:
            t_min_var.set(old_t_min)
            t_max_var.set(old_t_max)
            dvdt_min_var.set(old_dvdt_min)
            dvdt_max_var.set(old_dvdt_max)
            current_time_unit["value"] = old_unit
            time_unit_var.set(old_unit)
            status_var.set(f"Error: {exc}")
            return
        finally:
            suspend_events["value"] = False

        _schedule_plot()

    def _reset():
        suspend_events["value"] = True
        try:
            current_time_unit["value"] = initial_state["time_unit"]
            time_unit_var.set(initial_state["time_unit"])
            logarithmic_y_var.set(initial_state["logarithmic_y"])
            dvdt_line_var.set(initial_state["dvdt_line"])
            smoothing_algorithm_var.set(initial_state["smoothing_algorithm"])
            smoothing_window_var.set(initial_state["smoothing_window"])
            x_tick_count_var.set(initial_state["x_tick_count"])
            y_tick_count_var.set(initial_state["y_tick_count"])
            plot_title_var.set(initial_state["plot_title"])
            show_title_var.set(initial_state["show_title"])
            title_fontsize_var.set(initial_state["title_fontsize"])
            tick_fontsize_var.set(initial_state["tick_fontsize"])
            label_fontsize_var.set(initial_state["label_fontsize"])
            legend_fontsize_var.set(initial_state["legend_fontsize"])
            line_width_var.set(initial_state["line_width"])
            t_min_var.set(initial_state["t_min"])
            t_max_var.set(initial_state["t_max"])
            dvdt_min_var.set(initial_state["dvdt_min"])
            dvdt_max_var.set(initial_state["dvdt_max"])
            t_min_lock_var.set(initial_state["t_min_lock"])
            t_max_lock_var.set(initial_state["t_max_lock"])
            dvdt_min_lock_var.set(initial_state["dvdt_min_lock"])
            dvdt_max_lock_var.set(initial_state["dvdt_max_lock"])
            for var in stage_vars.values():
                var.set(True)
            _apply_lock_states()
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Valores restaurados.")

    ttk.Label(
        controls_frame,
        text=f"Etapas detectadas: {', '.join(f'#{item.stage}' for item in deg_files)}",
        justify="left",
        wraplength=260,
    ).pack(anchor="w", pady=(0, 10))

    stages_box = ttk.LabelFrame(controls_frame, text="Etapas")
    stages_box.pack(fill="x", pady=5)

    series_box = ttk.LabelFrame(controls_frame, text="Series")
    series_box.pack(fill="x", pady=5)

    style_box = ttk.LabelFrame(controls_frame, text="Estilo")
    style_box.pack(fill="x", pady=5)

    smoothing_box = ttk.LabelFrame(controls_frame, text="Suavizado")
    smoothing_box.pack(fill="x", pady=5)

    text_box = ttk.LabelFrame(controls_frame, text="Texto / tamaños")
    text_box.pack(fill="x", pady=5)

    limits_box = ttk.LabelFrame(controls_frame, text="Límites de ejes")
    limits_box.pack(fill="x", pady=5)

    stages_inner, bind_stage_scroll = _build_scrollable_cycle_selector(
        stages_box,
        height=DEG_STAGE_SELECTOR_HEIGHT,
    )
    for index, deg_file in enumerate(deg_files):
        stage_toggle = ttk.Checkbutton(
            stages_inner,
            text=f"Etapa #{deg_file.stage}",
            variable=stage_vars[deg_file.stage],
            command=_schedule_plot,
        )
        pady = (0, 2) if index < len(deg_files) - 1 else 0
        stage_toggle.pack(anchor="w", padx=4, pady=pady)
        bind_stage_scroll(stage_toggle)

    ttk.Label(series_box, text="Unidad de tiempo").pack(anchor="w", padx=8, pady=(4, 2))
    time_unit_combo = ttk.Combobox(
        series_box,
        textvariable=time_unit_var,
        values=TIME_UNIT_OPTIONS,
        state="readonly",
        width=8,
    )
    time_unit_combo.pack(anchor="w", padx=8, pady=(0, 4))
    ttk.Checkbutton(
        series_box,
        text="Logarithmic y-axis",
        variable=logarithmic_y_var,
        command=_schedule_plot,
    ).pack(anchor="w", padx=8, pady=(4, 4))

    ttk.Label(style_box, text="Línea de dV/dt").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    dvdt_line_combo = ttk.Combobox(
        style_box,
        textvariable=dvdt_line_var,
        values=LINESTYLE_OPTIONS,
        state="readonly",
        width=10,
    )
    dvdt_line_combo.grid(row=0, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(smoothing_box, text="Algoritmo").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    smoothing_algo_combo = ttk.Combobox(
        smoothing_box,
        textvariable=smoothing_algorithm_var,
        values=SMOOTHING_ALGORITHMS,
        state="readonly",
        width=16,
    )
    smoothing_algo_combo.grid(row=0, column=1, sticky="w", padx=8, pady=3)

    smoothing_window_label = ttk.Label(smoothing_box, text="Ventana de mediana")
    smoothing_window_label.grid(row=1, column=0, sticky="w", padx=8, pady=3)
    smoothing_window_spin = tk.Spinbox(
        smoothing_box,
        from_=1,
        to=500,
        increment=1,
        textvariable=smoothing_window_var,
        width=8,
    )
    smoothing_window_spin.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    def _update_smoothing_ui(*_args):
        algorithm = smoothing_algorithm_var.get()
        if algorithm == "Rolling average":
            smoothing_window_label.config(text="Average window")
        else:
            smoothing_window_label.config(text="Ventana de mediana")
        _schedule_plot()

    ttk.Label(text_box, text="Título").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    title_entry = ttk.Entry(text_box, textvariable=plot_title_var, width=28)
    title_entry.grid(row=0, column=1, sticky="we", padx=8, pady=3)
    ttk.Checkbutton(text_box, text="Mostrar titulo", variable=show_title_var, command=_schedule_plot).grid(
        row=6, column=0, columnspan=2, sticky="w", padx=8, pady=3
    )

    ttk.Label(text_box, text="Tamaño del título").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    title_size_spin = tk.Spinbox(text_box, from_=6, to=30, increment=0.5, textvariable=title_fontsize_var, width=8)
    title_size_spin.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de ticks").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    tick_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=tick_fontsize_var, width=8)
    tick_size_spin.grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de etiquetas").grid(row=3, column=0, sticky="w", padx=8, pady=3)
    label_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=label_fontsize_var, width=8)
    label_size_spin.grid(row=3, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de leyenda").grid(row=4, column=0, sticky="w", padx=8, pady=3)
    legend_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=legend_fontsize_var, width=8)
    legend_size_spin.grid(row=4, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Grosor de línea").grid(row=5, column=0, sticky="w", padx=8, pady=3)
    line_width_spin = tk.Spinbox(text_box, from_=0.0, to=10.0, increment=0.1, textvariable=line_width_var, width=8)
    line_width_spin.grid(row=5, column=1, sticky="w", padx=8, pady=3)

    limit_specs = [
        ("t min", "t_min", t_min_var, t_min_lock_var),
        ("t max", "t_max", t_max_var, t_max_lock_var),
        ("dV/dt min", "dvdt_min", dvdt_min_var, dvdt_min_lock_var),
        ("dV/dt max", "dvdt_max", dvdt_max_var, dvdt_max_lock_var),
    ]

    for row_idx, (label, key, var, lock_var) in enumerate(limit_specs):
        ttk.Label(limits_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        entry = ttk.Entry(limits_box, textvariable=var, width=22)
        entry.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        entry.bind("<Return>", _schedule_plot)
        entry.bind("<KP_Enter>", _schedule_plot)
        entry.bind("<FocusOut>", _schedule_plot)
        limit_entries[key] = entry
        ttk.Checkbutton(
            limits_box,
            text="Lock",
            variable=lock_var,
            command=_apply_lock_states,
        ).grid(row=row_idx, column=2, sticky="w", padx=8, pady=3)

    ttk.Label(limits_box, text="x-Ticks").grid(row=len(limit_specs), column=0, sticky="w", padx=8, pady=3)
    x_tick_spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=x_tick_count_var, width=8)
    x_tick_spin.grid(row=len(limit_specs), column=1, sticky="w", padx=8, pady=3)

    ttk.Label(limits_box, text="y-Ticks").grid(row=len(limit_specs) + 1, column=0, sticky="w", padx=8, pady=3)
    y_tick_spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=y_tick_count_var, width=8)
    y_tick_spin.grid(row=len(limit_specs) + 1, column=1, sticky="w", padx=8, pady=3)

    dvdt_line_combo.bind("<<ComboboxSelected>>", _schedule_plot)
    time_unit_combo.bind("<<ComboboxSelected>>", _on_time_unit_changed)
    smoothing_algo_combo.bind("<<ComboboxSelected>>", _update_smoothing_ui)

    for spin in (smoothing_window_spin, x_tick_spin, y_tick_spin):
        spin.bind("<Return>", _schedule_plot)
        spin.bind("<FocusOut>", _schedule_plot)
        spin.config(command=_schedule_plot)

    for widget in (
        title_entry,
        title_size_spin,
        tick_size_spin,
        label_size_spin,
        legend_size_spin,
        line_width_spin,
    ):
        widget.bind("<Return>", _schedule_plot)
        widget.bind("<KP_Enter>", _schedule_plot)
        widget.bind("<FocusOut>", _schedule_plot)

    for spin in (
        title_size_spin,
        tick_size_spin,
        label_size_spin,
        legend_size_spin,
        line_width_spin,
    ):
        spin.config(command=_schedule_plot)

    ttk.Label(
        controls_frame,
        textvariable=status_var,
        wraplength=240,
        justify="left",
    ).pack(anchor="w", fill="x", pady=(10, 10))

    buttons_frame = ttk.Frame(controls_frame)
    buttons_frame.pack(fill="x", pady=(5, 0))

    ttk.Button(buttons_frame, text="Restablecer", command=_reset).pack(side="left", padx=(0, 6))
    ttk.Button(buttons_frame, text="Autoescala", command=_autofit).pack(side="left")

    _apply_lock_states()
    _update_smoothing_ui()
    _plot()


def open_ocp_window(
    input_dir: Path,
    font_defaults: PlotFontDefaults | None = None,
    language: str = "es",
) -> None:
    language = _deg_language(language)
    ocp_files = find_deg_ocp_files(Path(input_dir))
    if not ocp_files:
        raise ValueError("No se encontraron archivos Degradacion_OCP validos.")

    win = tk.Toplevel()
    win.title("Deg - OCP")
    win.geometry("1200x720")

    notebook = ttk.Notebook(win)
    notebook.pack(fill="both", expand=True)

    for deg_file in ocp_files:
        tab_frame = ttk.Frame(notebook)
        notebook.add(tab_frame, text=f"Etapa #{deg_file.stage}")
        ocp._build_v_vs_t_tab(tab_frame, deg_file.path, font_defaults=font_defaults, language=language)


def _write_metadata_sheet(ws, parsed: ParsedDTA) -> None:
    ws["A1"] = "Campo"
    ws["B1"] = "Valor"
    ws["C1"] = "Unidad"

    for ref in ("A1", "B1", "C1"):
        ws[ref].font = Font(bold=True)

    for row_idx, (key, label) in enumerate(META_FIELDS, start=2):
        raw_value = parsed.meta_values.get(key, "")
        raw_unit = parsed.meta_units.get(key, "")
        ws.cell(row=row_idx, column=1, value=label)

        num = to_float(raw_value)
        ws.cell(row=row_idx, column=2, value=num if num is not None else raw_value)
        ws.cell(row=row_idx, column=3, value=raw_unit)

    ws.freeze_panes = "A2"


def _write_data_sheet(ws, parsed: ParsedDTA) -> None:
    col_idx = {name: idx for idx, name in enumerate(parsed.header)}
    selected_source_cols = [name for name, _label in DATA_EXPORT if name in col_idx]

    for col_num, source_name in enumerate(selected_source_cols, start=1):
        label = dict(DATA_EXPORT)[source_name]
        cell = ws.cell(row=1, column=col_num, value=label)
        cell.font = Font(bold=True)

    for col_num, source_name in enumerate(selected_source_cols, start=1):
        source_index = col_idx[source_name]
        unit_value = ""
        if parsed.units and source_index < len(parsed.units):
            unit_value = parsed.units[source_index]
            if unit_value == "#":
                unit_value = ""
        ws.cell(row=2, column=col_num, value=unit_value)

    for row_num, raw_parts in enumerate(parsed.rows, start=3):
        for col_num, source_name in enumerate(selected_source_cols, start=1):
            source_index = col_idx[source_name]
            raw_value = raw_parts[source_index] if source_index < len(raw_parts) else ""
            num = to_float(raw_value)
            ws.cell(row=row_num, column=col_num, value=num if num is not None else raw_value)

    ws.freeze_panes = "A3"


def _auto_format_sheet(ws) -> None:
    for col_num in range(1, ws.max_column + 1):
        max_len = 0
        for row_num in range(1, min(ws.max_row, 100) + 1):
            value = ws.cell(row=row_num, column=col_num).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_num)].width = min(max(10, max_len + 2), 45)

    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top")


def export_file(source_path: Path, output_path: Path) -> None:
    parsed = parse_gamry_dta(source_path)
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "Metadata"
    _write_metadata_sheet(ws_meta, parsed)

    ws_data = wb.create_sheet("Data")
    _write_data_sheet(ws_data, parsed)

    for ws in (ws_meta, ws_data):
        _auto_format_sheet(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def export_folder(input_dir: Path, output_dir: Path) -> list[Path]:
    exported: list[Path] = []
    for deg_file in [*find_deg_files(input_dir), *find_deg_ocp_files(input_dir)]:
        out_path = output_dir / f"{deg_file.path.stem}.xlsx"
        export_file(deg_file.path, out_path)
        exported.append(out_path)
    return exported


def _show_deg_stub(title: str) -> None:
    messagebox.showinfo("Deg", f"{title} aun no esta implementado.")


def run_pipeline(
    input_dir: Path,
    output_dir: Path,
    selected_options: list[str] | None = None,
    font_defaults: PlotFontDefaults | None = None,
    language: str = "es",
) -> list[Path]:
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)

    exported_files = export_folder(input_dir, output_dir)
    if not exported_files:
        return []

    chosen = set(selected_options or [])

    if "V vs t" in chosen:
        open_v_vs_t_window(input_dir, font_defaults=font_defaults, language=language)

    if "dV/dt" in chosen:
        open_dv_dt_window(input_dir, font_defaults=font_defaults)

    if "OCP" in chosen:
        open_ocp_window(input_dir, font_defaults=font_defaults, language=language)

    return exported_files
