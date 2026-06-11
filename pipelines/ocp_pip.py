"""Open circuit potential (.DTA) -> Excel (.xlsx) exporter for Gamry OCP files."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from math import floor, log10

import tkinter as tk
from tkinter import ttk, filedialog
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.ticker import MaxNLocator

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from i18n import normalize_language, translate
from plot_defaults import PlotFontDefaults, apply_x_tick_label_padding, make_legend_draggable, resolve_plot_font_defaults
from ui_layout import create_resizable_plot_layout, create_scrollable_controls


META_FIELDS = [
    ("TITLE", "Técnica"),
    ("DATE", "Fecha"),
    ("TIME", "Hora"),
    ("TIMEOUT", "Duración"),
    ("SAMPLETIME", "tiempo de muestreo"),
    ("STABILITY", "estabilización"),
    ("AREA", "Area"),
]

DATA_EXPORT = [
    ("Pt", "Pt"),
    ("T", "tiempo"),
    ("Vf", "Vf"),
    ("Vm", "Vm"),
    ("Ach", "Ach"),
    ("Temp", "Temperatura"),
]

OCP_FILE_RE = re.compile(r"^OCP.*\.DTA$", re.IGNORECASE)

OCP_PLOT_COLORS = {
    "voltage": "#06a8c2",
    "temperature": "#cf9a32",
}

LINESTYLE_OPTIONS = ["none", "-", "--", ":", "-."]

DVDT_UNIT_OPTIONS = ["Auto", "V/s", "mV/s", "µV/s"]
DVDT_UNIT_SCALES = {
    "V/s": 1.0,
    "mV/s": 1e3,
    "µV/s": 1e6,
}

DELTA_V_UNIT_OPTIONS = ["Auto", "V", "mV", "µV"]
DELTA_V_UNIT_SCALES = {
    "V": 1.0,
    "mV": 1e3,
    "µV": 1e6,
}


@dataclass
class ParsedDTA:
    meta_values: dict[str, str]
    meta_units: dict[str, str]
    header: list[str]
    units: list[str]
    rows: list[list[str]]


def to_float(val: str) -> float | None:
    """Convert Gamry-style numbers (decimal comma) to float."""
    s = val.strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _drop_leading_blank(parts: list[str]) -> list[str]:
    if parts and parts[0] == "":
        return parts[1:]
    return parts


def _extract_parenthesized_unit(text: str) -> str:
    matches = re.findall(r"\(([^()]*)\)", text)
    if matches:
        return matches[-1].strip()
    return ""


def _extract_meta_unit(key: str, description: str) -> str:
    unit = _extract_parenthesized_unit(description)
    if unit:
        return unit

    fallback_units = {
        "TITLE": "",
        "DATE": "",
        "TIME": "",
        "TIMEOUT": "s",
        "SAMPLETIME": "s",
        "STABILITY": "mV/s",
        "AREA": "cm^2",
    }
    return fallback_units.get(key, "")


def parse_gamry_dta(path: Path) -> ParsedDTA:
    """Parse one Gamry .DTA file containing a CURVE table."""
    text = path.read_text(encoding="latin-1", errors="replace")
    lines = text.splitlines()

    meta_values: dict[str, str] = {}
    meta_units: dict[str, str] = {}
    header: list[str] | None = None
    units: list[str] = []
    rows: list[list[str]] = []

    table_started = False

    for line in lines:
        if not table_started:
            if line.startswith("CURVE") and "TABLE" in line:
                table_started = True
                continue

            if not line.strip():
                continue

            parts = line.split("\t")
            if len(parts) >= 3 and parts[0].strip():
                key = parts[0].strip()
                value = parts[2].strip()
                description = parts[-1].strip() if len(parts) >= 4 else ""

                meta_values[key] = value
                meta_units[key] = _extract_meta_unit(key, description)

            continue

        if not line.strip():
            continue

        parts = _drop_leading_blank(line.rstrip("\r\n").split("\t"))
        if not parts:
            continue

        if header is None:
            if parts[0] == "Pt":
                header = parts
            continue

        if parts[0] == "#":
            units = parts
            continue

        if re.fullmatch(r"-?\d+", parts[0]):
            rows.append(parts)

    if header is None:
        raise ValueError(f"No data header found in {path.name} (expected 'Pt ...')")

    return ParsedDTA(
        meta_values=meta_values,
        meta_units=meta_units,
        header=header,
        units=units,
        rows=rows,
    )


def _write_metadata_sheet(ws, parsed: ParsedDTA) -> None:
    ws["A1"] = "Campo"
    ws["B1"] = "Valor"
    ws["C1"] = "Unidad"

    for ref in ("A1", "B1", "C1"):
        ws[ref].font = Font(bold=True)

    numeric_meta_keys = {"TIMEOUT", "SAMPLETIME", "STABILITY", "AREA"}

    for row_idx, (key, label) in enumerate(META_FIELDS, start=2):
        raw_value = parsed.meta_values.get(key, "")
        raw_unit = parsed.meta_units.get(key, "")

        ws.cell(row=row_idx, column=1, value=label)

        if key in numeric_meta_keys:
            num = to_float(raw_value)
            ws.cell(row=row_idx, column=2, value=num if num is not None else raw_value)
        else:
            ws.cell(row=row_idx, column=2, value=raw_value)

        ws.cell(row=row_idx, column=3, value=raw_unit)

    ws.freeze_panes = "A2"


def _write_data_sheet(ws, parsed: ParsedDTA) -> None:
    col_idx = {name: idx for idx, name in enumerate(parsed.header)}
    selected_source_cols = [name for name, _label in DATA_EXPORT if name in col_idx]

    for col_num, source_name in enumerate(selected_source_cols, start=1):
        label = dict(DATA_EXPORT)[source_name]
        cell = ws.cell(row=1, column=col_num, value=label)
        cell.font = Font(bold=True)

    for col_num, source_name in enumerate(selected_source_cols, start=1):
        source_index = col_idx[source_name]
        unit_value = ""
        if parsed.units and source_index < len(parsed.units):
            unit_value = parsed.units[source_index]
            if unit_value == "#":
                unit_value = ""
        ws.cell(row=2, column=col_num, value=unit_value)

    for row_num, raw_parts in enumerate(parsed.rows, start=3):
        for col_num, source_name in enumerate(selected_source_cols, start=1):
            source_index = col_idx[source_name]
            raw_value = raw_parts[source_index] if source_index < len(raw_parts) else ""

            num = to_float(raw_value)
            ws.cell(row=row_num, column=col_num, value=num if num is not None else raw_value)

    ws.freeze_panes = "A3"


def _auto_format_sheet(ws) -> None:
    for col_num in range(1, ws.max_column + 1):
        max_len = 0
        for row_num in range(1, min(ws.max_row, 100) + 1):
            value = ws.cell(row=row_num, column=col_num).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_num)].width = min(max(10, max_len + 2), 45)

    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top")


def _column_index(parsed: ParsedDTA, column_name: str) -> int | None:
    try:
        return parsed.header.index(column_name)
    except ValueError:
        return None


def _required_numeric_series(parsed: ParsedDTA, x_col: str, y_col: str) -> tuple[list[float], list[float]]:
    x_idx = _column_index(parsed, x_col)
    y_idx = _column_index(parsed, y_col)

    if x_idx is None or y_idx is None:
        raise ValueError(f"No se encontraron las columnas requeridas: {x_col}, {y_col}.")

    x_values: list[float] = []
    y_values: list[float] = []

    for row in parsed.rows:
        if x_idx >= len(row) or y_idx >= len(row):
            continue

        x_val = to_float(row[x_idx])
        y_val = to_float(row[y_idx])
        if x_val is None or y_val is None:
            continue

        x_values.append(x_val)
        y_values.append(y_val)

    if not x_values or not y_values:
        raise ValueError(f"No hay datos numÃ©ricos vÃ¡lidos para {y_col} vs {x_col}.")

    return x_values, y_values


def _optional_float(value: str | None) -> float | None:
    if value is None:
        return None
    text = value.strip()
    if not text:
        return None
    return float(text.replace(",", "."))


def _positive_float(text: str, name: str) -> float:
    value = text.strip().replace(",", ".")
    if not value:
        raise ValueError(f"{name} no puede estar vacÃ­o.")
    num = float(value)
    if num <= 0:
        raise ValueError(f"{name} debe ser mayor que 0.")
    return num


def _round_down_dec(value: float, decimals: int = 1) -> float:
    scale = 10 ** decimals
    return int(value * scale) / scale if value >= 0 else -int((-value * scale) + 0.9999999999) / scale


def _round_up_dec(value: float, decimals: int = 1) -> float:
    scale = 10 ** decimals
    return -int((-value * scale)) / scale if value <= 0 else int((value * scale) + 0.9999999999) / scale


def _format_limit_value(value: float | None, decimals: int = 1) -> str:
    if value is None:
        return ""
    return f"{value:.{decimals}f}"


def _padded_limits(
    values: list[float],
    rel_pad: float = 0.05,
    decimals: int = 1,
) -> tuple[float | None, float | None]:
    if not values:
        return None, None

    vmin = min(values)
    vmax = max(values)

    if vmin == vmax:
        pad = max(abs(vmin) * rel_pad, 1e-6)
    else:
        pad = (vmax - vmin) * rel_pad

    lo = _round_down_dec(vmin - pad, decimals)
    hi = _round_up_dec(vmax + pad, decimals)
    return lo, hi


def _tight_limits(values: list[float], decimals: int = 1) -> tuple[float | None, float | None]:
    if not values:
        return None, None

    vmin = min(values)
    vmax = max(values)

    lo = _round_down_dec(vmin, decimals)
    hi = _round_up_dec(vmax, decimals)
    if lo == hi:
        hi = lo + 1.0

    return lo, hi


def _auto_decimals(
    values: list[float],
    default: int = 1,
    min_decimals: int = 1,
    max_decimals: int = 4,
) -> int:
    if not values:
        return default

    vmin = min(values)
    vmax = max(values)
    span = abs(vmax - vmin)
    if span <= 0:
        return max(default, min_decimals)

    magnitude = floor(log10(span))
    decimals = 1 - magnitude
    decimals = max(min_decimals, min(max_decimals, int(decimals)))
    return max(default, decimals)


def _auto_dvdt_unit(values: list[float]) -> str:
    if not values:
        return "V/s"
    max_abs = max(abs(value) for value in values)
    if max_abs >= 1:
        return "V/s"
    if max_abs >= 1e-3:
        return "mV/s"
    return "µV/s"


def _resolve_dvdt_unit(values: list[float], mode: str) -> tuple[str, float]:
    if mode == "Auto":
        unit = _auto_dvdt_unit(values)
    else:
        unit = mode
    return unit, DVDT_UNIT_SCALES[unit]


def _auto_delta_v_unit(values: list[float]) -> str:
    if not values:
        return "V"
    max_abs = max(abs(value) for value in values)
    if max_abs >= 1:
        return "V"
    if max_abs >= 1e-3:
        return "mV"
    return "µV"


def _resolve_delta_v_unit(values: list[float], mode: str) -> tuple[str, float]:
    if mode == "Auto":
        unit = _auto_delta_v_unit(values)
    else:
        unit = mode
    return unit, DELTA_V_UNIT_SCALES[unit]


def _linspace_ticks(lo: float, hi: float, count: int) -> list[float]:
    if count <= 1:
        return [lo]
    if lo == hi:
        return [lo for _ in range(count)]
    step = (hi - lo) / (count - 1)
    return [lo + step * idx for idx in range(count)]


def _mpl_linestyle(value: str) -> str:
    return "None" if value == "none" else value


def _pick_legend_corner(series_specs: list[tuple[list[float], list[float], tuple[float, float], tuple[float, float]]]) -> str:
    counts = {
        "upper left": 0,
        "upper right": 0,
        "lower left": 0,
        "lower right": 0,
    }

    for x_values, y_values, xlim, ylim in series_specs:
        if not x_values or not y_values:
            continue

        x_min, x_max = xlim
        y_min, y_max = ylim
        if x_max == x_min or y_max == y_min:
            continue

        step = max(1, len(x_values) // 500)
        for x_val, y_val in zip(x_values[::step], y_values[::step]):
            x_norm = (x_val - x_min) / (x_max - x_min)
            y_norm = (y_val - y_min) / (y_max - y_min)
            if x_norm <= 0.5 and y_norm >= 0.5:
                counts["upper left"] += 1
            elif x_norm > 0.5 and y_norm >= 0.5:
                counts["upper right"] += 1
            elif x_norm <= 0.5 and y_norm < 0.5:
                counts["lower left"] += 1
            else:
                counts["lower right"] += 1

    return min(counts, key=counts.get)


def _build_scrollable_controls(parent) -> ttk.Frame:
    _outer, controls_frame = create_scrollable_controls(parent)
    return controls_frame


def _ocp_language(language: str | None = None) -> str:
    return normalize_language(language)


def _safe_filename_part(text: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", text.strip())
    return cleaned.strip("._") or "OCP"


def _format_report_value(value: object, digits: int = 6) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.{digits}g}"
    return str(value)


def compute_default_v_vs_t_limits(parsed: ParsedDTA, decimals: int = 1) -> dict[str, str]:
    time_values, voltage_values = _required_numeric_series(parsed, "T", "Vf")
    _temp_time_values, temp_values = _required_numeric_series(parsed, "T", "Temp")

    t_min, t_max = _padded_limits(time_values, decimals=decimals)
    v_decimals = _auto_decimals(voltage_values, default=decimals)
    temp_decimals = _auto_decimals(temp_values, default=decimals)
    v_min, v_max = _tight_limits(voltage_values, decimals=v_decimals)
    temp_min, temp_max = _tight_limits(temp_values, decimals=temp_decimals)

    return {
        "t_min": _format_limit_value(t_min, decimals),
        "t_max": _format_limit_value(t_max, decimals),
        "v_min": _format_limit_value(v_min, v_decimals),
        "v_max": _format_limit_value(v_max, v_decimals),
        "temp_min": _format_limit_value(temp_min, temp_decimals),
        "temp_max": _format_limit_value(temp_max, temp_decimals),
    }


def compute_autofit_v_vs_t_limits(parsed: ParsedDTA, show_temperature: bool, decimals: int = 1) -> dict[str, str]:
    limits = compute_default_v_vs_t_limits(parsed, decimals=decimals)
    if not show_temperature:
        limits["temp_min"] = ""
        limits["temp_max"] = ""
    return limits


def build_ocp_report_metadata(
    parsed: ParsedDTA,
    source_name: str,
    language: str | None = None,
) -> list[tuple[str, object, str]]:
    language = _ocp_language(language)
    label_keys = {
        "TITLE": "technique",
        "DATE": "date",
        "TIME": "start_time",
        "TIMEOUT": "duration",
        "SAMPLETIME": "sampling_time",
        "STABILITY": "stabilization",
        "AREA": "area",
    }
    numeric_meta_keys = {"TIMEOUT", "SAMPLETIME", "STABILITY", "AREA"}
    rows: list[tuple[str, object, str]] = [(translate("source_file", language), source_name, "")]
    for key, fallback_label in META_FIELDS:
        raw_value = parsed.meta_values.get(key, "")
        value = to_float(raw_value) if key in numeric_meta_keys else None
        rows.append(
            (
                translate(label_keys.get(key, fallback_label), language),
                _format_report_value(value if value is not None else raw_value),
                parsed.meta_units.get(key, ""),
            )
        )
    return rows


def build_ocp_v_vs_t_report_indicators(
    parsed: ParsedDTA,
    language: str | None = None,
) -> list[tuple[str, object, str]]:
    language = _ocp_language(language)
    time_values, voltage_values = _required_numeric_series(parsed, "T", "Vf")
    rows: list[tuple[str, object, str]] = []
    if time_values:
        rows.append((translate("total_time", language), _format_report_value(max(time_values) - min(time_values)), "s"))
    if voltage_values:
        rows.extend(
            [
                (translate("initial_voltage", language), _format_report_value(voltage_values[0]), "V"),
                (translate("final_voltage", language), _format_report_value(voltage_values[-1]), "V"),
                (translate("delta_voltage", language), _format_report_value(voltage_values[-1] - voltage_values[0]), "V"),
                (translate("minimum_voltage", language), _format_report_value(min(voltage_values)), "V"),
                (translate("maximum_voltage", language), _format_report_value(max(voltage_values)), "V"),
            ]
        )

    try:
        _temp_time_values, temp_values = _required_numeric_series(parsed, "T", "Temp")
    except ValueError:
        temp_values = []
    if temp_values:
        rows.extend(
            [
                (translate("average_temperature", language), _format_report_value(sum(temp_values) / len(temp_values), digits=3), "C"),
                (translate("minimum_temperature", language), _format_report_value(min(temp_values), digits=3), "C"),
                (translate("maximum_temperature", language), _format_report_value(max(temp_values), digits=3), "C"),
            ]
        )
    return rows


def _add_report_table(
    ax,
    title: str,
    rows: list[tuple[str, object, str]],
    bbox: list[float],
    language: str = "es",
) -> None:
    ax.text(
        bbox[0],
        bbox[1] + bbox[3] + 0.025,
        title,
        fontsize=13,
        fontweight="bold",
        ha="left",
        va="bottom",
        transform=ax.transAxes,
    )
    table = ax.table(
        cellText=[[_format_report_value(name), _format_report_value(value), _format_report_value(unit)] for name, value, unit in rows],
        colLabels=[translate("name", language), translate("value", language), translate("unit", language)],
        cellLoc="left",
        colLoc="left",
        colWidths=[bbox[2] * 0.58, bbox[2] * 0.27, bbox[2] * 0.15],
        bbox=bbox,
    )
    table.auto_set_font_size(False)
    table.set_fontsize(8.0)
    for (row_idx, _col_idx), cell in table.get_celld().items():
        cell.set_edgecolor("#b8c0c8")
        cell.set_linewidth(0.5)
        if row_idx == 0:
            cell.set_facecolor("#edf2f7")
            cell.set_text_props(weight="bold")
        else:
            cell.set_facecolor("white")


def export_ocp_v_vs_t_report_pdf(
    parsed: ParsedDTA,
    output_path: Path,
    *,
    source_name: str,
    language: str = "es",
    font_defaults: PlotFontDefaults | None = None,
) -> Path:
    language = _ocp_language(language)
    font_defaults = resolve_plot_font_defaults(font_defaults)
    font_values = font_defaults.as_strings()
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    limits = compute_autofit_v_vs_t_limits(parsed, show_temperature=True)
    plot_kwargs = {
        "show_temperature": True,
        "voltage_linestyle": "-",
        "temperature_linestyle": "--",
        "tick_count": 6,
        "plot_title": translate("ocp_v_vs_t_report_plot_title", language, file=source_name),
        "show_title": True,
        "title_fontsize": float(font_values["title"]),
        "tick_fontsize": float(font_values["tick"]),
        "label_fontsize": float(font_values["label"]),
        "legend_fontsize": float(font_values["legend"]),
        "line_width": 1.5,
        "t_min": _optional_float(limits["t_min"]),
        "t_max": _optional_float(limits["t_max"]),
        "v_min": _optional_float(limits["v_min"]),
        "v_max": _optional_float(limits["v_max"]),
        "temp_min": _optional_float(limits["temp_min"]),
        "temp_max": _optional_float(limits["temp_max"]),
        "language": language,
    }

    metadata_rows = build_ocp_report_metadata(parsed, source_name, language=language)
    indicator_rows = build_ocp_v_vs_t_report_indicators(parsed, language=language)

    with PdfPages(output_path) as pdf:
        plot_fig = Figure(figsize=(9.5, 6.2), dpi=150)
        FigureCanvasAgg(plot_fig)
        has_plot = draw_v_vs_t_on_figure(plot_fig, parsed=parsed, source_name=source_name, **plot_kwargs)
        if not has_plot:
            raise ValueError("No hay datos suficientes para generar el grafico del reporte.")
        pdf.savefig(plot_fig, bbox_inches="tight")

        table_fig = Figure(figsize=(8.5, 11.0), dpi=150)
        ax = table_fig.add_subplot(111)
        ax.axis("off")
        table_fig.text(
            0.05,
            0.965,
            translate("ocp_v_vs_t_report_title", language, file=source_name),
            fontsize=18,
            fontweight="bold",
            ha="left",
            va="top",
        )
        table_fig.text(
            0.05,
            0.935,
            translate("ocp_v_vs_t_report_subtitle", language),
            fontsize=10,
            ha="left",
            va="top",
            color="#4a5568",
        )
        _add_report_table(ax, translate("metadata", language), metadata_rows, [0.05, 0.53, 0.90, 0.34], language=language)
        _add_report_table(ax, translate("indicators", language), indicator_rows, [0.05, 0.08, 0.90, 0.36], language=language)
        pdf.savefig(table_fig, bbox_inches="tight")

    return output_path


def build_delta_v_data(parsed: ParsedDTA) -> dict[str, list[float]]:
    time_values, voltage_values = _required_numeric_series(parsed, "T", "Vf")

    dvdt_time: list[float] = []
    dvdt_values: list[float] = []
    start_times: list[float] = []
    end_times: list[float] = []
    delta_v_values: list[float] = []

    for idx in range(len(time_values) - 1):
        t0 = time_values[idx]
        t1 = time_values[idx + 1]
        v0 = voltage_values[idx]
        v1 = voltage_values[idx + 1]

        dt = t1 - t0
        if dt <= 0:
            continue

        dv = v1 - v0
        dvdt_time.append((t0 + t1) / 2.0)
        dvdt_values.append(dv / dt)
        start_times.append(t0)
        end_times.append(t1)
        delta_v_values.append(dv)

    if not dvdt_time:
        raise ValueError("No hay suficientes datos vÃ¡lidos para calcular dV/dt.")

    return {
        "raw_time": time_values,
        "raw_voltage": voltage_values,
        "dvdt_time": dvdt_time,
        "dvdt_values": dvdt_values,
        "start_times": start_times,
        "end_times": end_times,
        "delta_v_values": delta_v_values,
    }


def compute_default_delta_v_limits(parsed: ParsedDTA, decimals: int = 3, dvdt_scale: float = 1.0) -> dict[str, str]:
    delta_data = build_delta_v_data(parsed)
    t_min, t_max = _padded_limits(delta_data["dvdt_time"], decimals=decimals)
    dvdt_scaled = [value * dvdt_scale for value in delta_data["dvdt_values"]]
    dvdt_decimals = _auto_decimals(dvdt_scaled, default=decimals, max_decimals=6)
    dvdt_min, dvdt_max = _tight_limits(dvdt_scaled, decimals=dvdt_decimals)
    return {
        "t_min": _format_limit_value(t_min, decimals),
        "t_max": _format_limit_value(t_max, decimals),
        "dvdt_min": _format_limit_value(dvdt_min, dvdt_decimals),
        "dvdt_max": _format_limit_value(dvdt_max, dvdt_decimals),
    }


def compute_autofit_delta_v_limits(
    parsed: ParsedDTA,
    t_min: float | None = None,
    t_max: float | None = None,
    decimals: int = 3,
    dvdt_scale: float = 1.0,
) -> dict[str, str]:
    delta_data = build_delta_v_data(parsed)
    filtered_time: list[float] = []
    filtered_values: list[float] = []

    for time_value, dvdt_value in zip(delta_data["dvdt_time"], delta_data["dvdt_values"]):
        if t_min is not None and time_value < t_min:
            continue
        if t_max is not None and time_value > t_max:
            continue
        filtered_time.append(time_value)
        filtered_values.append(dvdt_value)

    if not filtered_time:
        raise ValueError("No hay datos vÃ¡lidos de dV/dt en el rango de tiempo seleccionado.")

    t_min_fit, t_max_fit = _padded_limits(filtered_time, decimals=decimals)
    filtered_scaled = [value * dvdt_scale for value in filtered_values]
    dvdt_decimals = _auto_decimals(filtered_scaled, default=decimals, max_decimals=6)
    dvdt_min_fit, dvdt_max_fit = _tight_limits(filtered_scaled, decimals=dvdt_decimals)

    return {
        "t_min": _format_limit_value(t_min_fit, decimals),
        "t_max": _format_limit_value(t_max_fit, decimals),
        "dvdt_min": _format_limit_value(dvdt_min_fit, dvdt_decimals),
        "dvdt_max": _format_limit_value(dvdt_max_fit, dvdt_decimals),
    }


def compute_delta_v_indicators(
    parsed: ParsedDTA,
    t_min: float | None = None,
    t_max: float | None = None,
) -> dict[str, float]:
    delta_data = build_delta_v_data(parsed)

    selected_rates: list[float] = []
    selected_start: float | None = None
    selected_end: float | None = None
    last_dvdt: float | None = None

    for start_t, end_t, dvdt_value in zip(
        delta_data["start_times"],
        delta_data["end_times"],
        delta_data["dvdt_values"],
    ):
        mid_t = (start_t + end_t) / 2.0
        if t_min is not None and mid_t < t_min:
            continue
        if t_max is not None and mid_t > t_max:
            continue

        selected_rates.append(dvdt_value)
        if selected_start is None:
            selected_start = start_t
        selected_end = end_t
        last_dvdt = dvdt_value

    if not selected_rates or selected_start is None or selected_end is None or last_dvdt is None:
        raise ValueError("No hay datos vÃ¡lidos de dV/dt en el rango de tiempo seleccionado.")

    raw_time = delta_data["raw_time"]
    raw_voltage = delta_data["raw_voltage"]

    start_index = min(range(len(raw_time)), key=lambda idx: abs(raw_time[idx] - selected_start))
    end_index = min(range(len(raw_time)), key=lambda idx: abs(raw_time[idx] - selected_end))

    if end_index < start_index:
        start_index, end_index = end_index, start_index

    total_delta_v = raw_voltage[end_index] - raw_voltage[start_index]

    return {
        "avg_dvdt": sum(selected_rates) / len(selected_rates),
        "last_dvdt": last_dvdt,
        "total_delta_v": total_delta_v,
    }


def draw_v_vs_t_on_figure(
    fig: Figure,
    parsed: ParsedDTA,
    source_name: str,
    show_temperature: bool,
    voltage_linestyle: str = "-",
    temperature_linestyle: str = "--",
    tick_count: int = 6,
    plot_title: str = "",
    show_title: bool = True,
    title_fontsize: float = 14,
    tick_fontsize: float = 10,
    label_fontsize: float = 11,
    legend_fontsize: float = 10,
    line_width: float = 1.5,
    t_min: float | None = None,
    t_max: float | None = None,
    v_min: float | None = None,
    v_max: float | None = None,
    temp_min: float | None = None,
    temp_max: float | None = None,
    language: str = "es",
) -> bool:
    fig.clear()
    language = _ocp_language(language)

    time_values, voltage_values = _required_numeric_series(parsed, "T", "Vf")
    if not _mpl_linestyle(voltage_linestyle) == "None":
        has_voltage = True
    else:
        has_voltage = False

    if not has_voltage and not show_temperature:
        return False

    ax_main = fig.add_subplot(111)
    ax_temp = None

    if has_voltage:
        ax_main.plot(
            time_values,
            voltage_values,
            color=OCP_PLOT_COLORS["voltage"],
            linewidth=line_width,
            linestyle=_mpl_linestyle(voltage_linestyle),
            label=translate("voltage", language),
        )

    ax_main.set_xlabel(translate("time_seconds", language), fontsize=label_fontsize)
    ax_main.set_ylabel(translate("voltage_v", language), color=OCP_PLOT_COLORS["voltage"], fontsize=label_fontsize)
    ax_main.tick_params(axis="x", labelsize=tick_fontsize)
    apply_x_tick_label_padding(ax_main, tick_fontsize)
    ax_main.tick_params(axis="y", labelcolor=OCP_PLOT_COLORS["voltage"], labelsize=tick_fontsize)
    ax_main.grid(True, alpha=0.25)
    ax_main.xaxis.set_major_locator(MaxNLocator(nbins=max(2, int(tick_count))))
    ax_main.yaxis.set_major_locator(MaxNLocator(nbins=max(2, int(tick_count))))

    if t_min is not None or t_max is not None:
        ax_main.set_xlim(left=t_min, right=t_max)
    if v_min is not None or v_max is not None:
        ax_main.set_ylim(bottom=v_min, top=v_max)

    temp_time_values: list[float] = []
    temp_values: list[float] = []

    if show_temperature:
        temp_time_values, temp_values = _required_numeric_series(parsed, "T", "Temp")
        ax_temp = ax_main.twinx()
        ax_temp.plot(
            temp_time_values,
            temp_values,
            color=OCP_PLOT_COLORS["temperature"],
            linewidth=line_width,
            linestyle=_mpl_linestyle(temperature_linestyle),
            label=translate("temperature", language),
        )
        ax_temp.set_ylabel(translate("temperature_c", language), color=OCP_PLOT_COLORS["temperature"], fontsize=label_fontsize)
        ax_temp.tick_params(axis="y", labelcolor=OCP_PLOT_COLORS["temperature"], labelsize=tick_fontsize)
        ax_temp.yaxis.set_major_locator(MaxNLocator(nbins=max(2, int(tick_count))))
        if temp_min is not None or temp_max is not None:
            ax_temp.set_ylim(bottom=temp_min, top=temp_max)

    y_tick_count = max(2, int(tick_count))
    v_lo, v_hi = ax_main.get_ylim()
    ax_main.set_yticks(_linspace_ticks(v_lo, v_hi, y_tick_count))
    if ax_temp is not None:
        t_lo, t_hi = ax_temp.get_ylim()
        ax_temp.set_yticks(_linspace_ticks(t_lo, t_hi, y_tick_count))

    default_title = parsed.meta_values.get("TITLE", "").strip() or source_name
    final_title = plot_title.strip() if plot_title.strip() else f"V vs t - {default_title}"
    ax_main.set_title(final_title if show_title else "", fontsize=title_fontsize)

    handles = []
    labels = []
    main_handles, main_labels = ax_main.get_legend_handles_labels()
    handles.extend(main_handles)
    labels.extend(main_labels)
    if ax_temp is not None:
        temp_handles, temp_labels = ax_temp.get_legend_handles_labels()
        handles.extend(temp_handles)
        labels.extend(temp_labels)

    if handles:
        series_specs = []
        if has_voltage:
            series_specs.append((time_values, voltage_values, ax_main.get_xlim(), ax_main.get_ylim()))
        if show_temperature and ax_temp is not None:
            series_specs.append((temp_time_values, temp_values, ax_main.get_xlim(), ax_temp.get_ylim()))
        legend_loc = _pick_legend_corner(series_specs) if series_specs else "best"
        make_legend_draggable(ax_main.legend(handles, labels, loc=legend_loc, fontsize=legend_fontsize))

    fig.tight_layout()
    return True


def _build_v_vs_t_tab(
    parent: tk.Widget,
    source_path: Path,
    font_defaults: PlotFontDefaults | None = None,
    language: str = "es",
) -> None:
    language = _ocp_language(language)
    root = parent.winfo_toplevel()
    parsed = parse_gamry_dta(source_path)
    default_limits = compute_default_v_vs_t_limits(parsed)
    font_defaults = resolve_plot_font_defaults(font_defaults)
    font_default_values = font_defaults.as_strings()


    controls_host, plot_outer = create_resizable_plot_layout(parent, sidebar_width=320)
    controls_frame = _build_scrollable_controls(controls_host)

    toolbar_frame = ttk.Frame(plot_outer)
    toolbar_frame.pack(side="top", fill="x")

    canvas_frame = ttk.Frame(plot_outer)
    canvas_frame.pack(side="top", fill="both", expand=True)

    fig = Figure(figsize=(8.8, 5.4), dpi=100)

    canvas = FigureCanvasTkAgg(fig, master=canvas_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    toolbar = NavigationToolbar2Tk(canvas, toolbar_frame, pack_toolbar=False)
    toolbar.update()
    toolbar.pack(side="left", fill="x")

    status_var = tk.StringVar(value=translate("ready", language))
    temperature_var = tk.BooleanVar(value=False)
    voltage_line_var = tk.StringVar(value="-")
    temperature_line_var = tk.StringVar(value="--")
    tick_count_var = tk.IntVar(value=6)
    plot_title_var = tk.StringVar(value="")
    show_title_var = tk.BooleanVar(value=True)
    title_fontsize_var = tk.StringVar(value=font_default_values["title"])
    tick_fontsize_var = tk.StringVar(value=font_default_values["tick"])
    label_fontsize_var = tk.StringVar(value=font_default_values["label"])
    legend_fontsize_var = tk.StringVar(value=font_default_values["legend"])
    line_width_var = tk.StringVar(value="1.5")
    t_min_var = tk.StringVar(value=default_limits["t_min"])
    t_max_var = tk.StringVar(value=default_limits["t_max"])
    v_min_var = tk.StringVar(value=default_limits["v_min"])
    v_max_var = tk.StringVar(value=default_limits["v_max"])
    temp_min_var = tk.StringVar(value="")
    temp_max_var = tk.StringVar(value="")

    initial_state = {
        "temperature": False,
        "voltage_line": "-",
        "temperature_line": "--",
        "tick_count": 6,
        "plot_title": "",
        "show_title": True,
        "title_fontsize": font_default_values["title"],
        "tick_fontsize": font_default_values["tick"],
        "label_fontsize": font_default_values["label"],
        "legend_fontsize": font_default_values["legend"],
        "line_width": "1.5",
        "t_min": default_limits["t_min"],
        "t_max": default_limits["t_max"],
        "v_min": default_limits["v_min"],
        "v_max": default_limits["v_max"],
        "temp_min": "",
        "temp_max": "",
    }

    plot_job = {"id": None}
    suspend_events = {"value": False}

    def _collect_limits():
        return dict(
            t_min=_optional_float(t_min_var.get()),
            t_max=_optional_float(t_max_var.get()),
            v_min=_optional_float(v_min_var.get()),
            v_max=_optional_float(v_max_var.get()),
            temp_min=_optional_float(temp_min_var.get()),
            temp_max=_optional_float(temp_max_var.get()),
        )

    def _plot():
        try:
            has_plot = draw_v_vs_t_on_figure(
                fig=fig,
                parsed=parsed,
                source_name=source_path.stem,
                show_temperature=temperature_var.get(),
                voltage_linestyle=voltage_line_var.get(),
                temperature_linestyle=temperature_line_var.get(),
                tick_count=tick_count_var.get(),
                plot_title=plot_title_var.get(),
                show_title=show_title_var.get(),
                title_fontsize=_positive_float(title_fontsize_var.get(), translate("title_font_size", language)),
                tick_fontsize=_positive_float(tick_fontsize_var.get(), translate("tick_font_size", language)),
                label_fontsize=_positive_float(label_fontsize_var.get(), translate("label_font_size", language)),
                legend_fontsize=_positive_float(legend_fontsize_var.get(), translate("legend_font_size", language)),
                line_width=_positive_float(line_width_var.get(), translate("line_width", language)),
                language=language,
                **_collect_limits(),
            )
        except ValueError as exc:
            fig.clear()
            canvas.draw_idle()
            status_var.set(f"Error: {exc}")
            return

        if not has_plot:
            fig.clear()
            canvas.draw_idle()
            status_var.set(translate("graph_no_visible_series", language))
            return

        canvas.draw_idle()
        status_var.set(translate("graph_updated", language))

    def _schedule_plot(*_args):
        if suspend_events["value"]:
            return
        if plot_job["id"] is not None:
            root.after_cancel(plot_job["id"])
        plot_job["id"] = root.after(20, _plot)

    def _autofit():
        try:
            fitted = compute_autofit_v_vs_t_limits(parsed, show_temperature=temperature_var.get())
        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            return

        suspend_events["value"] = True
        try:
            t_min_var.set(fitted["t_min"])
            t_max_var.set(fitted["t_max"])
            v_min_var.set(fitted["v_min"])
            v_max_var.set(fitted["v_max"])
            temp_min_var.set(fitted["temp_min"])
            temp_max_var.set(fitted["temp_max"])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set(translate("autofit_applied", language))

    def _reset():
        suspend_events["value"] = True
        try:
            temperature_var.set(initial_state["temperature"])
            voltage_line_var.set(initial_state["voltage_line"])
            temperature_line_var.set(initial_state["temperature_line"])
            tick_count_var.set(initial_state["tick_count"])
            plot_title_var.set(initial_state["plot_title"])
            show_title_var.set(initial_state["show_title"])
            title_fontsize_var.set(initial_state["title_fontsize"])
            tick_fontsize_var.set(initial_state["tick_fontsize"])
            label_fontsize_var.set(initial_state["label_fontsize"])
            legend_fontsize_var.set(initial_state["legend_fontsize"])
            line_width_var.set(initial_state["line_width"])
            t_min_var.set(initial_state["t_min"])
            t_max_var.set(initial_state["t_max"])
            v_min_var.set(initial_state["v_min"])
            v_max_var.set(initial_state["v_max"])
            temp_min_var.set(initial_state["temp_min"])
            temp_max_var.set(initial_state["temp_max"])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set(translate("values_reset", language))

    def _current_font_defaults() -> PlotFontDefaults:
        return PlotFontDefaults(
            title=_positive_float(title_fontsize_var.get(), translate("title_font_size", language)),
            tick=_positive_float(tick_fontsize_var.get(), translate("tick_font_size", language)),
            label=_positive_float(label_fontsize_var.get(), translate("label_font_size", language)),
            legend=_positive_float(legend_fontsize_var.get(), translate("legend_font_size", language)),
        )

    def _export_report() -> None:
        default_name = f"OCP_V_vs_t_Report_{_safe_filename_part(source_path.stem)}.pdf"
        path_text = filedialog.asksaveasfilename(
            parent=root,
            title=translate("save_ocp_v_vs_t_report", language),
            initialfile=default_name,
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if not path_text:
            status_var.set(translate("export_cancelled", language))
            return
        try:
            exported_path = export_ocp_v_vs_t_report_pdf(
                parsed,
                Path(path_text),
                source_name=source_path.stem,
                language=language,
                font_defaults=_current_font_defaults(),
            )
        except Exception as exc:
            status_var.set(f"{translate('report_export_failed', language)}: {type(exc).__name__}: {exc}")
            return

        status_var.set(f"{translate('report_exported', language)}: {exported_path}")

    ttk.Label(
        controls_frame,
        text=f"{translate('detected_file', language)}:\n{source_path.name}",
        justify="left",
    ).pack(anchor="w", pady=(0, 10))

    series_box = ttk.LabelFrame(controls_frame, text=translate("series", language))
    series_box.pack(fill="x", pady=5)

    style_box = ttk.LabelFrame(controls_frame, text=translate("style", language))
    style_box.pack(fill="x", pady=5)

    text_box = ttk.LabelFrame(controls_frame, text=translate("text_sizes", language))
    text_box.pack(fill="x", pady=5)

    limits_box = ttk.LabelFrame(controls_frame, text=translate("axis_limits", language))
    limits_box.pack(fill="x", pady=5)

    ttk.Checkbutton(
        series_box,
        text=translate("temperature", language),
        variable=temperature_var,
        command=_schedule_plot,
    ).pack(anchor="w", padx=8, pady=4)

    ttk.Label(style_box, text=translate("voltage_line", language)).grid(row=0, column=0, sticky="w", padx=8, pady=3)
    voltage_line_combo = ttk.Combobox(
        style_box,
        textvariable=voltage_line_var,
        values=LINESTYLE_OPTIONS,
        state="readonly",
        width=10,
    )
    voltage_line_combo.grid(row=0, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text=translate("temperature_line", language)).grid(row=1, column=0, sticky="w", padx=8, pady=3)
    temperature_line_combo = ttk.Combobox(
        style_box,
        textvariable=temperature_line_var,
        values=LINESTYLE_OPTIONS,
        state="readonly",
        width=10,
    )
    temperature_line_combo.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text=translate("ticks", language)).grid(row=2, column=0, sticky="w", padx=8, pady=3)
    tick_spin = tk.Spinbox(style_box, from_=2, to=10, textvariable=tick_count_var, width=8)
    tick_spin.grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text=translate("title", language)).grid(row=0, column=0, sticky="w", padx=8, pady=3)
    title_entry = ttk.Entry(text_box, textvariable=plot_title_var, width=28)
    title_entry.grid(row=0, column=1, sticky="we", padx=8, pady=3)
    ttk.Checkbutton(text_box, text=translate("show_title", language), variable=show_title_var, command=_schedule_plot).grid(
        row=6, column=0, columnspan=2, sticky="w", padx=8, pady=3
    )

    ttk.Label(text_box, text=translate("title_font_size", language)).grid(row=1, column=0, sticky="w", padx=8, pady=3)
    title_size_spin = tk.Spinbox(
        text_box,
        from_=6,
        to=30,
        increment=0.5,
        textvariable=title_fontsize_var,
        width=8,
    )
    title_size_spin.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text=translate("tick_font_size", language)).grid(row=2, column=0, sticky="w", padx=8, pady=3)
    tick_size_spin = tk.Spinbox(
        text_box,
        from_=6,
        to=24,
        increment=0.5,
        textvariable=tick_fontsize_var,
        width=8,
    )
    tick_size_spin.grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text=translate("label_font_size", language)).grid(row=3, column=0, sticky="w", padx=8, pady=3)
    label_size_spin = tk.Spinbox(
        text_box,
        from_=6,
        to=24,
        increment=0.5,
        textvariable=label_fontsize_var,
        width=8,
    )
    label_size_spin.grid(row=3, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text=translate("legend_font_size", language)).grid(row=4, column=0, sticky="w", padx=8, pady=3)
    legend_size_spin = tk.Spinbox(
        text_box,
        from_=6,
        to=24,
        increment=0.5,
        textvariable=legend_fontsize_var,
        width=8,
    )
    legend_size_spin.grid(row=4, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text=translate("line_width", language)).grid(row=5, column=0, sticky="w", padx=8, pady=3)
    line_width_spin = tk.Spinbox(
        text_box,
        from_=0.5,
        to=5.0,
        increment=0.1,
        textvariable=line_width_var,
        width=8,
    )
    line_width_spin.grid(row=5, column=1, sticky="w", padx=8, pady=3)

    limit_specs = [
        ("t min", t_min_var),
        ("t max", t_max_var),
        ("V min", v_min_var),
        ("V max", v_max_var),
        ("Temp min", temp_min_var),
        ("Temp max", temp_max_var),
    ]

    for row_idx, (label, var) in enumerate(limit_specs):
        ttk.Label(limits_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        entry = ttk.Entry(limits_box, textvariable=var, width=12)
        entry.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        entry.bind("<Return>", _schedule_plot)
        entry.bind("<KP_Enter>", _schedule_plot)
        entry.bind("<FocusOut>", _schedule_plot)

    for combo in (voltage_line_combo, temperature_line_combo):
        combo.bind("<<ComboboxSelected>>", _schedule_plot)

    tick_spin.bind("<Return>", _schedule_plot)
    tick_spin.bind("<FocusOut>", _schedule_plot)
    tick_spin.config(command=_schedule_plot)

    for widget in (
        title_entry,
        title_size_spin,
        tick_size_spin,
        label_size_spin,
        legend_size_spin,
        line_width_spin,
    ):
        widget.bind("<Return>", _schedule_plot)
        widget.bind("<KP_Enter>", _schedule_plot)
        widget.bind("<FocusOut>", _schedule_plot)

    for spin in (
        title_size_spin,
        tick_size_spin,
        label_size_spin,
        legend_size_spin,
        line_width_spin,
    ):
        spin.config(command=_schedule_plot)

    ttk.Label(
        controls_frame,
        textvariable=status_var,
        wraplength=240,
        justify="left",
    ).pack(anchor="w", fill="x", pady=(10, 10))

    buttons_frame = ttk.Frame(controls_frame)
    buttons_frame.pack(fill="x", pady=(5, 0))

    ttk.Button(buttons_frame, text=translate("reset", language), command=_reset).pack(side="left", padx=(0, 6))
    ttk.Button(buttons_frame, text=translate("autofit", language), command=_autofit).pack(side="left")
    ttk.Button(buttons_frame, text=translate("pdf_report", language), command=_export_report).pack(side="left", padx=(6, 0))

    _plot()


def open_delta_v_window(input_dir: Path, font_defaults: PlotFontDefaults | None = None) -> None:
    ocp_files = find_ocp_files(Path(input_dir))
    if not ocp_files:
        raise ValueError("No se encontraron archivos OCP válidos.")

    win = tk.Toplevel()
    win.title("OCP - DeltaV")
    win.geometry("1200x720")

    notebook = ttk.Notebook(win)
    notebook.pack(fill="both", expand=True)

    for source_path in ocp_files:
        tab_frame = ttk.Frame(notebook)
        notebook.add(tab_frame, text=source_path.stem)
        _build_delta_v_tab(tab_frame, source_path, font_defaults=font_defaults)


def open_v_vs_t_window(
    input_dir: Path,
    font_defaults: PlotFontDefaults | None = None,
    language: str = "es",
) -> None:
    language = _ocp_language(language)
    ocp_files = find_ocp_files(Path(input_dir))
    if not ocp_files:
        raise ValueError("No se encontraron archivos OCP válidos.")

    win = tk.Toplevel()
    win.title("OCP - V vs t")
    win.geometry("1200x720")


    notebook = ttk.Notebook(win)
    notebook.pack(fill="both", expand=True)

    for source_path in ocp_files:
        tab_frame = ttk.Frame(notebook)
        notebook.add(tab_frame, text=source_path.stem)
        _build_v_vs_t_tab(tab_frame, source_path, font_defaults=font_defaults, language=language)


def draw_delta_v_on_figure(
    fig: Figure,
    parsed: ParsedDTA,
    source_name: str,
    show_voltage: bool = False,
    show_temperature: bool = False,
    dvdt_linestyle: str = "-",
    voltage_linestyle: str = "--",
    temperature_linestyle: str = "--",
    dvdt_scale: float = 1.0,
    dvdt_unit: str = "V/s",
    tick_count: int = 6,
    plot_title: str = "",
    show_title: bool = True,
    title_fontsize: float = 14,
    tick_fontsize: float = 10,
    label_fontsize: float = 11,
    legend_fontsize: float = 10,
    line_width: float = 1.5,
    t_min: float | None = None,
    t_max: float | None = None,
    dvdt_min: float | None = None,
    dvdt_max: float | None = None,
    v_min: float | None = None,
    v_max: float | None = None,
) -> bool:
    fig.clear()

    linestyle = _mpl_linestyle(dvdt_linestyle)
    voltage_style = _mpl_linestyle(voltage_linestyle)
    temp_style = _mpl_linestyle(temperature_linestyle)
    if linestyle == "None" and not show_voltage and not show_temperature:
        return False

    delta_data = build_delta_v_data(parsed)
    raw_time = delta_data["raw_time"]
    raw_voltage = delta_data["raw_voltage"]
    dvdt_scaled = [value * dvdt_scale for value in delta_data["dvdt_values"]]
    ax_main = fig.add_subplot(111)
    ax_voltage = None
    ax_temp = None
    handles = []
    labels = []

    if linestyle != "None":
        ax_main.plot(
            delta_data["dvdt_time"],
            dvdt_scaled,
            color=OCP_PLOT_COLORS["voltage"],
            linewidth=line_width,
            linestyle=linestyle,
            label="dV/dt",
        )

    if show_voltage and voltage_style != "None":
        ax_voltage = ax_main.twinx()
        ax_voltage.plot(
            raw_time,
            raw_voltage,
            color=OCP_PLOT_COLORS["temperature"],
            linewidth=line_width,
            linestyle=voltage_style,
            label="Voltaje",
        )
        ax_voltage.set_ylabel("Voltaje [V]", color=OCP_PLOT_COLORS["temperature"], fontsize=label_fontsize)
        ax_voltage.tick_params(axis="y", labelsize=tick_fontsize, labelcolor=OCP_PLOT_COLORS["temperature"])
        ax_voltage.yaxis.set_major_locator(MaxNLocator(nbins=max(2, int(tick_count))))
        if v_min is not None or v_max is not None:
            ax_voltage.set_ylim(bottom=v_min, top=v_max)

    if show_temperature and temp_style != "None":
        temp_time, temp_values = _required_numeric_series(parsed, "T", "Temp")
        if ax_voltage is None:
            ax_temp = ax_main.twinx()
        else:
            ax_temp = ax_main.twinx()
            ax_temp.spines["right"].set_position(("outward", 60))
            ax_temp.spines["left"].set_visible(False)
            ax_temp.yaxis.tick_right()
            ax_temp.yaxis.set_label_position("right")
        ax_temp.plot(
            temp_time,
            temp_values,
            color=OCP_PLOT_COLORS["temperature"],
            linewidth=line_width,
            linestyle=temp_style,
            label="Temperatura",
        )
        ax_temp.set_ylabel("Temperatura [°C]", color=OCP_PLOT_COLORS["temperature"], fontsize=label_fontsize)
        ax_temp.tick_params(axis="y", labelsize=tick_fontsize, labelcolor=OCP_PLOT_COLORS["temperature"])
        ax_temp.yaxis.set_major_locator(MaxNLocator(nbins=max(2, int(tick_count))))

    ax_main.set_xlabel("Tiempo [s]", fontsize=label_fontsize)
    ax_main.set_ylabel(f"dV/dt [{dvdt_unit}]", color=OCP_PLOT_COLORS["voltage"], fontsize=label_fontsize)
    ax_main.tick_params(axis="x", labelsize=tick_fontsize)
    apply_x_tick_label_padding(ax_main, tick_fontsize)
    ax_main.tick_params(axis="y", labelsize=tick_fontsize, labelcolor=OCP_PLOT_COLORS["voltage"])
    ax_main.grid(True, alpha=0.25)
    ax_main.xaxis.set_major_locator(MaxNLocator(nbins=max(2, int(tick_count))))
    ax_main.yaxis.set_major_locator(MaxNLocator(nbins=max(2, int(tick_count))))

    if t_min is not None or t_max is not None:
        ax_main.set_xlim(left=t_min, right=t_max)
        if ax_voltage is not None:
            ax_voltage.set_xlim(left=t_min, right=t_max)
        if ax_temp is not None:
            ax_temp.set_xlim(left=t_min, right=t_max)
    if dvdt_min is not None or dvdt_max is not None:
        ax_main.set_ylim(bottom=dvdt_min, top=dvdt_max)

    y_tick_count = max(2, int(tick_count))
    dvdt_lo, dvdt_hi = ax_main.get_ylim()
    ax_main.set_yticks(_linspace_ticks(dvdt_lo, dvdt_hi, y_tick_count))
    if ax_voltage is not None:
        v_lo, v_hi = ax_voltage.get_ylim()
        ax_voltage.set_yticks(_linspace_ticks(v_lo, v_hi, y_tick_count))
    if ax_temp is not None:
        t_lo, t_hi = ax_temp.get_ylim()
        ax_temp.set_yticks(_linspace_ticks(t_lo, t_hi, y_tick_count))

    default_title = parsed.meta_values.get("TITLE", "").strip() or source_name
    final_title = plot_title.strip() if plot_title.strip() else f"DeltaV - {default_title}"
    ax_main.set_title(final_title if show_title else "", fontsize=title_fontsize)

    main_handles, main_labels = ax_main.get_legend_handles_labels()
    handles.extend(main_handles)
    labels.extend(main_labels)

    if ax_voltage is not None:
        voltage_handles, voltage_labels = ax_voltage.get_legend_handles_labels()
        handles.extend(voltage_handles)
        labels.extend(voltage_labels)
    if ax_temp is not None:
        temp_handles, temp_labels = ax_temp.get_legend_handles_labels()
        handles.extend(temp_handles)
        labels.extend(temp_labels)

    if handles:
        series_specs = []
        if linestyle != "None":
            series_specs.append((delta_data["dvdt_time"], dvdt_scaled, ax_main.get_xlim(), ax_main.get_ylim()))
        if ax_voltage is not None:
            series_specs.append((raw_time, raw_voltage, ax_main.get_xlim(), ax_voltage.get_ylim()))
        if ax_temp is not None:
            temp_time, temp_values = _required_numeric_series(parsed, "T", "Temp")
            series_specs.append((temp_time, temp_values, ax_main.get_xlim(), ax_temp.get_ylim()))
        legend_loc = _pick_legend_corner(series_specs) if series_specs else "best"
        make_legend_draggable(ax_main.legend(handles, labels, loc=legend_loc, fontsize=legend_fontsize))

    fig.tight_layout()
    return True


def _build_delta_v_tab(
    parent: tk.Widget,
    source_path: Path,
    font_defaults: PlotFontDefaults | None = None,
) -> None:
    root = parent.winfo_toplevel()
    parsed = parse_gamry_dta(source_path)
    delta_data = build_delta_v_data(parsed)
    default_unit, default_scale = _resolve_dvdt_unit(delta_data["dvdt_values"], "Auto")
    default_limits = compute_default_delta_v_limits(parsed, dvdt_scale=default_scale)
    font_defaults = resolve_plot_font_defaults(font_defaults)
    font_default_values = font_defaults.as_strings()


    controls_host, plot_outer = create_resizable_plot_layout(parent, sidebar_width=320)
    controls_frame = _build_scrollable_controls(controls_host)

    toolbar_frame = ttk.Frame(plot_outer)
    toolbar_frame.pack(side="top", fill="x")

    canvas_frame = ttk.Frame(plot_outer)
    canvas_frame.pack(side="top", fill="both", expand=True)

    fig = Figure(figsize=(8.8, 5.4), dpi=100)

    canvas = FigureCanvasTkAgg(fig, master=canvas_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    toolbar = NavigationToolbar2Tk(canvas, toolbar_frame, pack_toolbar=False)
    toolbar.update()
    toolbar.pack(side="left", fill="x")

    status_var = tk.StringVar(value="Listo.")
    avg_dvdt_var = tk.StringVar(value="")
    last_dvdt_var = tk.StringVar(value="")
    total_delta_v_var = tk.StringVar(value="")

    show_voltage_var = tk.BooleanVar(value=False)
    show_temperature_var = tk.BooleanVar(value=False)
    dvdt_line_var = tk.StringVar(value="-")
    voltage_line_var = tk.StringVar(value="--")
    temperature_line_var = tk.StringVar(value="--")
    dvdt_unit_var = tk.StringVar(value="Auto")
    delta_v_unit_var = tk.StringVar(value="Auto")
    tick_count_var = tk.IntVar(value=6)
    plot_title_var = tk.StringVar(value="")
    show_title_var = tk.BooleanVar(value=True)
    title_fontsize_var = tk.StringVar(value=font_default_values["title"])
    tick_fontsize_var = tk.StringVar(value=font_default_values["tick"])
    label_fontsize_var = tk.StringVar(value=font_default_values["label"])
    legend_fontsize_var = tk.StringVar(value=font_default_values["legend"])
    line_width_var = tk.StringVar(value="1.5")
    t_min_var = tk.StringVar(value=default_limits["t_min"])
    t_max_var = tk.StringVar(value=default_limits["t_max"])
    dvdt_min_var = tk.StringVar(value=default_limits["dvdt_min"])
    dvdt_max_var = tk.StringVar(value=default_limits["dvdt_max"])
    v_default_limits = compute_default_v_vs_t_limits(parsed)
    v_min_var = tk.StringVar(value="")
    v_max_var = tk.StringVar(value="")

    initial_state = {
        "show_voltage": False,
        "show_temperature": False,
        "dvdt_line": "-",
        "voltage_line": "--",
        "temperature_line": "--",
        "dvdt_unit": "Auto",
        "delta_v_unit": "Auto",
        "tick_count": 6,
        "plot_title": "",
        "show_title": True,
        "title_fontsize": font_default_values["title"],
        "tick_fontsize": font_default_values["tick"],
        "label_fontsize": font_default_values["label"],
        "legend_fontsize": font_default_values["legend"],
        "line_width": "1.5",
        "t_min": default_limits["t_min"],
        "t_max": default_limits["t_max"],
        "dvdt_min": default_limits["dvdt_min"],
        "dvdt_max": default_limits["dvdt_max"],
        "v_min": "",
        "v_max": "",
    }

    plot_job = {"id": None}
    suspend_events = {"value": False}

    def _collect_limits():
        return dict(
            t_min=_optional_float(t_min_var.get()),
            t_max=_optional_float(t_max_var.get()),
            dvdt_min=_optional_float(dvdt_min_var.get()),
            dvdt_max=_optional_float(dvdt_max_var.get()),
            v_min=_optional_float(v_min_var.get()),
            v_max=_optional_float(v_max_var.get()),
        )

    current_dvdt_scale = {"value": default_scale, "unit": default_unit}

    def _filtered_dvdt_values():
        t_min = _optional_float(t_min_var.get())
        t_max = _optional_float(t_max_var.get())
        values: list[float] = []
        for time_value, dvdt_value in zip(delta_data["dvdt_time"], delta_data["dvdt_values"]):
            if t_min is not None and time_value < t_min:
                continue
            if t_max is not None and time_value > t_max:
                continue
            values.append(dvdt_value)
        return values

    def _update_indicators():
        indicators = compute_delta_v_indicators(
            parsed,
            t_min=_optional_float(t_min_var.get()),
            t_max=_optional_float(t_max_var.get()),
        )
        scale = current_dvdt_scale["value"]
        unit = current_dvdt_scale["unit"]
        avg_dvdt_var.set(f"{indicators['avg_dvdt'] * scale:.6g} {unit}")
        last_dvdt_var.set(f"{indicators['last_dvdt'] * scale:.6g} {unit}")
        delta_v_unit, delta_v_scale = _resolve_delta_v_unit(
            [indicators["total_delta_v"]],
            delta_v_unit_var.get(),
        )
        total_delta_v_var.set(f"{indicators['total_delta_v'] * delta_v_scale:.6g} {delta_v_unit}")

    dvdt_min_label_var = tk.StringVar(value=f"dV/dt min ({default_unit})")
    dvdt_max_label_var = tk.StringVar(value=f"dV/dt max ({default_unit})")

    def _plot():
        try:
            unit, scale = _resolve_dvdt_unit(_filtered_dvdt_values(), dvdt_unit_var.get())
            current_dvdt_scale["value"] = scale
            current_dvdt_scale["unit"] = unit
            dvdt_min_label_var.set(f"dV/dt min ({unit})")
            dvdt_max_label_var.set(f"dV/dt max ({unit})")
            has_plot = draw_delta_v_on_figure(
                fig=fig,
                parsed=parsed,
                source_name=source_path.stem,
                show_voltage=show_voltage_var.get(),
                show_temperature=show_temperature_var.get(),
                dvdt_linestyle=dvdt_line_var.get(),
                voltage_linestyle=voltage_line_var.get(),
                temperature_linestyle=temperature_line_var.get(),
                dvdt_scale=scale,
                dvdt_unit=unit,
                tick_count=tick_count_var.get(),
                plot_title=plot_title_var.get(),
                show_title=show_title_var.get(),
                title_fontsize=_positive_float(title_fontsize_var.get(), "Tamaño del título"),
                tick_fontsize=_positive_float(tick_fontsize_var.get(), "Tamaño de ticks"),
                label_fontsize=_positive_float(label_fontsize_var.get(), "Tamaño de etiquetas"),
                legend_fontsize=_positive_float(legend_fontsize_var.get(), "Tamaño de leyenda"),
                line_width=_positive_float(line_width_var.get(), "Grosor de línea"),
                **_collect_limits(),
            )
            _update_indicators()
        except ValueError as exc:
            fig.clear()
            canvas.draw_idle()
            avg_dvdt_var.set("")
            last_dvdt_var.set("")
            total_delta_v_var.set("")
            status_var.set(f"Error: {exc}")
            return

        if not has_plot:
            fig.clear()
            canvas.draw_idle()
            status_var.set("No se muestra gráfico: active una lí­nea visible.")
            return

        canvas.draw_idle()
        status_var.set("Gráfico actualizado.")

    def _schedule_plot(*_args):
        if suspend_events["value"]:
            return
        if plot_job["id"] is not None:
            root.after_cancel(plot_job["id"])
        plot_job["id"] = root.after(20, _plot)

    def _on_unit_change(_event=None):
        old_scale = current_dvdt_scale["value"]
        _unit, new_scale = _resolve_dvdt_unit(_filtered_dvdt_values(), dvdt_unit_var.get())

        suspend_events["value"] = True
        try:
            for var in (dvdt_min_var, dvdt_max_var):
                current = _optional_float(var.get())
                if current is None:
                    continue
                converted = current * old_scale / new_scale
                var.set(f"{converted:.6g}")
        finally:
            suspend_events["value"] = False

        _schedule_plot()

    def _autofit():
        try:
            unit, scale = _resolve_dvdt_unit(_filtered_dvdt_values(), dvdt_unit_var.get())
            fitted = compute_autofit_delta_v_limits(
                parsed,
                t_min=_optional_float(t_min_var.get()),
                t_max=_optional_float(t_max_var.get()),
                dvdt_scale=scale,
            )
        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            return

        suspend_events["value"] = True
        try:
            t_min_var.set(fitted["t_min"])
            t_max_var.set(fitted["t_max"])
            dvdt_min_var.set(fitted["dvdt_min"])
            dvdt_max_var.set(fitted["dvdt_max"])
            if show_voltage_var.get():
                v_min_var.set(v_default_limits["v_min"])
                v_max_var.set(v_default_limits["v_max"])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Autoescala aplicada.")

    def _reset():
        suspend_events["value"] = True
        try:
            show_voltage_var.set(initial_state["show_voltage"])
            show_temperature_var.set(initial_state["show_temperature"])
            dvdt_line_var.set(initial_state["dvdt_line"])
            voltage_line_var.set(initial_state["voltage_line"])
            temperature_line_var.set(initial_state["temperature_line"])
            dvdt_unit_var.set(initial_state["dvdt_unit"])
            delta_v_unit_var.set(initial_state["delta_v_unit"])
            tick_count_var.set(initial_state["tick_count"])
            plot_title_var.set(initial_state["plot_title"])
            show_title_var.set(initial_state["show_title"])
            title_fontsize_var.set(initial_state["title_fontsize"])
            tick_fontsize_var.set(initial_state["tick_fontsize"])
            label_fontsize_var.set(initial_state["label_fontsize"])
            legend_fontsize_var.set(initial_state["legend_fontsize"])
            line_width_var.set(initial_state["line_width"])
            t_min_var.set(initial_state["t_min"])
            t_max_var.set(initial_state["t_max"])
            dvdt_min_var.set(initial_state["dvdt_min"])
            dvdt_max_var.set(initial_state["dvdt_max"])
            v_min_var.set(initial_state["v_min"])
            v_max_var.set(initial_state["v_max"])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Valores restaurados.")

    ttk.Label(
        controls_frame,
        text=f"Archivo detectado:\n{source_path.name}",
        justify="left",
    ).pack(anchor="w", pady=(0, 10))

    series_box = ttk.LabelFrame(controls_frame, text="Series")
    series_box.pack(fill="x", pady=5)

    indicators_box = ttk.LabelFrame(controls_frame, text="Indicadores")
    indicators_box.pack(fill="x", pady=5)

    style_box = ttk.LabelFrame(controls_frame, text="Estilo")
    style_box.pack(fill="x", pady=5)

    text_box = ttk.LabelFrame(controls_frame, text="Texto / tamaños")
    text_box.pack(fill="x", pady=5)

    limits_box = ttk.LabelFrame(controls_frame, text="Límites de ejes")
    limits_box.pack(fill="x", pady=5)

    ttk.Checkbutton(
        series_box,
        text="Voltaje",
        variable=show_voltage_var,
        command=_schedule_plot,
    ).pack(anchor="w", padx=8, pady=4)
    ttk.Checkbutton(
        series_box,
        text="Temperatura",
        variable=show_temperature_var,
        command=_schedule_plot,
    ).pack(anchor="w", padx=8, pady=4)
    ttk.Label(series_box, text="Unidades dV/dt").pack(anchor="w", padx=8, pady=(6, 2))
    dvdt_unit_combo = ttk.Combobox(
        series_box,
        textvariable=dvdt_unit_var,
        values=DVDT_UNIT_OPTIONS,
        state="readonly",
        width=10,
    )
    dvdt_unit_combo.pack(anchor="w", padx=8, pady=(0, 4))

    ttk.Label(indicators_box, text="Promedio dV/dt").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    ttk.Label(indicators_box, textvariable=avg_dvdt_var).grid(row=0, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(indicators_box, text="Último dV/dt").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    ttk.Label(indicators_box, textvariable=last_dvdt_var).grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(indicators_box, text="Delta V total").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    ttk.Label(indicators_box, textvariable=total_delta_v_var).grid(row=2, column=1, sticky="w", padx=8, pady=3)
    ttk.Label(indicators_box, text="Unidades Delta V").grid(row=3, column=0, sticky="w", padx=8, pady=3)
    delta_v_unit_combo = ttk.Combobox(
        indicators_box,
        textvariable=delta_v_unit_var,
        values=DELTA_V_UNIT_OPTIONS,
        state="readonly",
        width=10,
    )
    delta_v_unit_combo.grid(row=3, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Línea de dV/dt").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    dvdt_line_combo = ttk.Combobox(
        style_box,
        textvariable=dvdt_line_var,
        values=LINESTYLE_OPTIONS,
        state="readonly",
        width=10,
    )
    dvdt_line_combo.grid(row=0, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Línea de voltaje").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    voltage_line_combo = ttk.Combobox(
        style_box,
        textvariable=voltage_line_var,
        values=LINESTYLE_OPTIONS,
        state="readonly",
        width=10,
    )
    voltage_line_combo.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Línea de temperatura").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    temperature_line_combo = ttk.Combobox(
        style_box,
        textvariable=temperature_line_var,
        values=LINESTYLE_OPTIONS,
        state="readonly",
        width=10,
    )
    temperature_line_combo.grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Ticks").grid(row=3, column=0, sticky="w", padx=8, pady=3)
    tick_spin = tk.Spinbox(style_box, from_=2, to=10, textvariable=tick_count_var, width=8)
    tick_spin.grid(row=3, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Título").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    title_entry = ttk.Entry(text_box, textvariable=plot_title_var, width=28)
    title_entry.grid(row=0, column=1, sticky="we", padx=8, pady=3)
    ttk.Checkbutton(text_box, text="Mostrar titulo", variable=show_title_var, command=_schedule_plot).grid(
        row=6, column=0, columnspan=2, sticky="w", padx=8, pady=3
    )

    ttk.Label(text_box, text="Tamaño del título").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    title_size_spin = tk.Spinbox(text_box, from_=6, to=30, increment=0.5, textvariable=title_fontsize_var, width=8)
    title_size_spin.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de ticks").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    tick_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=tick_fontsize_var, width=8)
    tick_size_spin.grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de etiquetas").grid(row=3, column=0, sticky="w", padx=8, pady=3)
    label_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=label_fontsize_var, width=8)
    label_size_spin.grid(row=3, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de leyenda").grid(row=4, column=0, sticky="w", padx=8, pady=3)
    legend_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=legend_fontsize_var, width=8)
    legend_size_spin.grid(row=4, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Grosor de línea").grid(row=5, column=0, sticky="w", padx=8, pady=3)
    line_width_spin = tk.Spinbox(text_box, from_=0.5, to=5.0, increment=0.1, textvariable=line_width_var, width=8)
    line_width_spin.grid(row=5, column=1, sticky="w", padx=8, pady=3)

    limit_specs = [
        ("t min", t_min_var),
        ("t max", t_max_var),
        (dvdt_min_label_var, dvdt_min_var),
        (dvdt_max_label_var, dvdt_max_var),
        ("V min", v_min_var),
        ("V max", v_max_var),
    ]

    for row_idx, (label, var) in enumerate(limit_specs):
        if isinstance(label, tk.StringVar):
            ttk.Label(limits_box, textvariable=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        else:
            ttk.Label(limits_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        entry = ttk.Entry(limits_box, textvariable=var, width=12)
        entry.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        entry.bind("<Return>", _schedule_plot)
        entry.bind("<KP_Enter>", _schedule_plot)
        entry.bind("<FocusOut>", _schedule_plot)

    for combo in (dvdt_line_combo, voltage_line_combo, temperature_line_combo, dvdt_unit_combo, delta_v_unit_combo):
        if combo is dvdt_unit_combo:
            combo.bind("<<ComboboxSelected>>", _on_unit_change)
        else:
            combo.bind("<<ComboboxSelected>>", _schedule_plot)

    tick_spin.bind("<Return>", _schedule_plot)
    tick_spin.bind("<FocusOut>", _schedule_plot)
    tick_spin.config(command=_schedule_plot)

    for widget in (
        title_entry,
        title_size_spin,
        tick_size_spin,
        label_size_spin,
        legend_size_spin,
        line_width_spin,
    ):
        widget.bind("<Return>", _schedule_plot)
        widget.bind("<KP_Enter>", _schedule_plot)
        widget.bind("<FocusOut>", _schedule_plot)

    for spin in (
        title_size_spin,
        tick_size_spin,
        label_size_spin,
        legend_size_spin,
        line_width_spin,
    ):
        spin.config(command=_schedule_plot)

    ttk.Label(
        controls_frame,
        textvariable=status_var,
        wraplength=240,
        justify="left",
    ).pack(anchor="w", fill="x", pady=(10, 10))

    buttons_frame = ttk.Frame(controls_frame)
    buttons_frame.pack(fill="x", pady=(5, 0))

    ttk.Button(buttons_frame, text="Restablecer", command=_reset).pack(side="left", padx=(0, 6))
    ttk.Button(buttons_frame, text="Autoescala", command=_autofit).pack(side="left")

    _plot()


def export_ocp_file(path: Path, out_path: Path) -> None:
    parsed = parse_gamry_dta(path)

    wb = Workbook()
    wb.remove(wb.active)

    ws_meta = wb.create_sheet("Metadata")
    _write_metadata_sheet(ws_meta, parsed)

    ws_data = wb.create_sheet("Data")
    _write_data_sheet(ws_data, parsed)

    for ws in (ws_meta, ws_data):
        _auto_format_sheet(ws)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def find_ocp_files(input_dir: Path) -> list[Path]:
    return sorted(
        [
            p for p in input_dir.iterdir()
            if p.is_file() and p.suffix.lower() == ".dta" and OCP_FILE_RE.match(p.name)
        ]
    )


def export_folder(input_dir: Path, output_dir: Path) -> list[Path]:
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    dta_files = find_ocp_files(input_dir)
    if not dta_files:
        return []

    exported_files: list[Path] = []
    for dta_file in dta_files:
        out_path = output_dir / f"{dta_file.stem}.xlsx"
        export_ocp_file(dta_file, out_path)
        exported_files.append(out_path)

    return exported_files


def run_pipeline(
    input_dir: Path,
    output_dir: Path,
    selected_options: list[str] | None = None,
    font_defaults: PlotFontDefaults | None = None,
    language: str = "es",
) -> list[Path]:
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)

    exported_files = export_folder(input_dir, output_dir)
    if not exported_files:
        return []

    chosen = set(selected_options or [])

    if "V vs t" in chosen:
        open_v_vs_t_window(input_dir, font_defaults=font_defaults, language=language)

    if "DeltaV" in chosen:
        open_delta_v_window(input_dir, font_defaults=font_defaults)

    return exported_files


if __name__ == "__main__":
    repo_dir = Path(__file__).resolve().parents[1]
    input_dir = repo_dir / "data"
    output_dir = repo_dir / "outputs"

    exported = export_folder(input_dir, output_dir)
    if exported:
        print("Archivos exportados:")
        for path in exported:
            print(f" - {path}")
    else:
        print("No se encontraron archivos OCP para exportar.")
