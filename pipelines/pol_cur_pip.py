"""Polarization curve (.DTA) -> Excel (.xlsx) exporter for Gamry chronopotentiometry.

What this first version does:
  - Finds all .DTA files whose filename starts with 'Curva_Polarizacion_'
  - Separates Asc and Dsc files
  - Sorts each curve by the LAST number after '#'
  - Reconstructs the full ascending and descending polarization curves
  - Exports ONE .xlsx per curve with three sheets:
        1) Metadata  -> Campo / Valor / Unidad
        2) Asc       -> headers row, units row, then numeric data
        3) Dsc       -> headers row, units row, then numeric data

This version focuses on importing, gathering, parsing, concatenating, and exporting.
No further processing/plotting is implemented yet.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from collections import defaultdict
import re

from math import floor, ceil, log10, isfinite

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.backends.backend_pdf import PdfPages

from matplotlib.ticker import LinearLocator, LogLocator, LogFormatterSciNotation, MaxNLocator, FormatStrFormatter, StrMethodFormatter
from math import floor, ceil

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from i18n import normalize_language, translate
from plot_defaults import PlotFontDefaults, apply_x_tick_label_padding, make_legend_draggable, resolve_plot_font_defaults
from ui_layout import create_resizable_plot_layout, create_scrollable_controls


# ---------------------------------------------------------------------------
# Export labels
# ---------------------------------------------------------------------------
META_ROWS_ORDER = [
    "Técnica",
    "Fecha",
    "Hora",
    "Duración del paso",
    "Rango I",
    "Paso I",
    "Tiempo de muestreo",
    "Área",
]

DATA_EXPORT = [
    ("Pt", "Pt", ""),
    ("T", "time", "s"),
    ("Vf", "Voltaje", "V"),
    ("Im", "Corriente", "A"),
    ("Sig", "Sig", "V"),
    ("Ach", "Ach", "V"),
    ("Temp", "Temperatura", "ºC"),
]

FILE_RE = re.compile(
    r"^Curva_Polarizacion_(?P<direction>Asc|Dsc)_(?P<description>.+?)_#(?P<curve_id>\d+)_#(?P<file_index>\d+)\.DTA$",
    re.IGNORECASE,
)

MARKER_OPTIONS = ["none", "^", "v", "o", "s", "d", "x", "+"]
LINESTYLE_OPTIONS = ["none", "-", "--", ":", "-."]
SMOOTHING_ALGORITHMS = ["Median filter", "Rolling average"]

PC_PLOT_COLORS = {
    "asc_voltage": "#06a8c2",
    "dsc_voltage": "#2b3d8c",
    "asc_temperature": "#cf9a32",
    "dsc_temperature": "#ab3030",
    "asc_current": "#2f9e44",
    "dsc_current": "#1b5e20",
}

SECONDS_PER_MINUTE = 60.0
SECONDS_PER_HOUR = 3600.0
TIME_UNIT_OPTIONS = ["s", "min", "h"]
PC_LANGUAGE = "es"
PC_HIGH_CURRENT_TARGET_A = 18.0
PC_HIGH_CURRENT_FIT_POINTS = 5


def _pc_language(language: str | None = None) -> str:
    return normalize_language(language or PC_LANGUAGE)


# ---------------------------------------------------------------------------
# Data containers
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class PolarizationFile:
    path: Path
    direction: str
    description: str
    curve_id: int
    file_index: int


@dataclass
class ParsedDTA:
    meta_values: dict[str, str]
    meta_units: dict[str, str]
    header: list[str]
    units: list[str]
    rows: list[list[str]]


@dataclass
class CurveBundle:
    description: str
    curve_id: int
    asc_files: list[PolarizationFile]
    dsc_files: list[PolarizationFile]


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------
def to_float(val: str) -> float | None:
    """Convert Gamry-style numbers (decimal comma) to float."""
    s = val.strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None

BACKUP_SUFFIX_RE = re.compile(r"\s+\([^)]*\)$")

SINGLE_INDEX_RE = re.compile(
    r"^Curva_Polarizacion_(?P<direction>Asc|Dsc)_(?P<description>.+?)_#(?P<file_index>\d+)\.DTA$",
    re.IGNORECASE,
)

def _normalize_filename_for_parser(path: Path, default_curve_id: int = 1) -> str:
    name = path.name

    # First remove trailing backup signature, if present:
    # "Curva_Polarizacion_Dsc_33s_2x_#145 (2026_03_10 19_56_50 UTC).DTA"
    # -> "Curva_Polarizacion_Dsc_33s_2x_#145.DTA"
    stem = path.stem
    stem = BACKUP_SUFFIX_RE.sub("", stem)
    cleaned_name = f"{stem}{path.suffix}"

    # If it already matches the normal format, keep it unchanged
    if FILE_RE.match(cleaned_name):
        return cleaned_name

    # If it is the backup format with only one #number, convert it to the
    # standard two-number format expected by FILE_RE
    match = SINGLE_INDEX_RE.match(cleaned_name)
    if match:
        return (
            f"Curva_Polarizacion_"
            f"{match.group('direction').title()}_"
            f"{match.group('description')}"
            f"_#{default_curve_id}"
            f"_#{match.group('file_index')}"
            f"{path.suffix}"
        )

    return cleaned_name

def _drop_leading_blank(parts: list[str]) -> list[str]:
    if parts and parts[0] == "":
        return parts[1:]
    return parts

def apply_temperature_axis_scaling(
    ax_temp,
    temp_values: list[float],
    tick_count: int,
    t_min: float | None = None,
    t_max: float | None = None,
) -> None:
    if not temp_values:
        return

    if t_min is not None or t_max is not None:
        current_lo, current_hi = ax_temp.get_ylim()
        t_lo = current_lo if t_min is None else t_min
        t_hi = current_hi if t_max is None else t_max
    else:
        t_lo = floor(min(temp_values))
        t_hi = ceil(max(temp_values))

    if t_lo == t_hi:
        t_hi = t_lo + 1

    tick_count = max(2, int(tick_count))

    ax_temp.set_ylim(t_lo, t_hi)
    ax_temp.yaxis.set_major_locator(LinearLocator(tick_count))
    ax_temp.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))

def apply_current_axis_scaling(
    ax_current,
    current_values: list[float],
    tick_count: int,
    i_min: float | None = None,
    i_max: float | None = None,
) -> None:
    if not current_values:
        return

    i_lo = 0.0 if i_min is None else i_min
    i_hi = ceil(max(current_values)) if i_max is None else i_max

    if i_hi <= i_lo:
        i_hi = i_lo + 1.0

    tick_count = max(2, int(tick_count))

    ax_current.set_ylim(i_lo, i_hi)
    ax_current.yaxis.set_major_locator(LinearLocator(tick_count))
    ax_current.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))


def apply_x_edge_ticks(ax, x_min: float | None, x_max: float | None, tick_count: int) -> None:
    if x_min is None and x_max is None:
        return

    tick_count = max(2, int(tick_count))
    ax.xaxis.set_major_locator(LinearLocator(tick_count))
    ax.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))


def _extract_parenthesized_unit(text: str) -> str:
    matches = re.findall(r"\(([^()]*)\)", text)
    if matches:
        return matches[-1].strip()
    return ""


def _extract_meta_unit(key: str, description: str) -> str:
    unit = _extract_parenthesized_unit(description)
    if unit:
        return unit

    fallback_units = {
        "TITLE": "",
        "DATE": "",
        "TIME": "",
        "IPRESTEP": "A",
        "ISTEP1": "A",
        "ISTEP2": "A",
        "TPRESTEP": "s",
        "TSTEP1": "s",
        "TSTEP2": "s",
        "SAMPLETIME": "s",
        "AREA": "cm^2",
    }
    return fallback_units.get(key, "")


def _fmt_range_value(i_min: float | None, i_max: float | None) -> str:
    if i_min is None or i_max is None:
        return ""
    return f"{i_min:g} a {i_max:g}"


def _format_report_value(value: object, digits: int = 6) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if not isfinite(value):
            return ""
        if value == 0:
            return "0"
        abs_value = abs(value)
        if 1e-6 <= abs_value < 1e6:
            return f"{value:.{digits}f}".rstrip("0").rstrip(".")
        return f"{value:.{digits}g}"
    return str(value)


def _safe_filename_part(text: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]+', "_", text.strip())
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned.strip("._") or "curve"


def _column_index(parsed: ParsedDTA, column_name: str) -> int | None:
    try:
        return parsed.header.index(column_name)
    except ValueError:
        return None


def _required_float(row: list[str], idx: int | None) -> float:
    if idx is None or idx >= len(row):
        raise ValueError("Falta una columna requerida en la tabla CURVE.")
    num = to_float(row[idx])
    if num is None:
        raise ValueError(f"No se pudo convertir a número: {row[idx]!r}")
    return num


def _step_delta_from_file(item: PolarizationFile) -> float | None:
    parsed = parse_gamry_dta(item.path)
    i1 = to_float(parsed.meta_values.get("ISTEP1", ""))
    i2 = to_float(parsed.meta_values.get("ISTEP2", ""))
    if i1 is None or i2 is None:
        return None
    return abs(i2 - i1)


def infer_current_tolerance(files: list[PolarizationFile]) -> float:
    """Infer a tolerance to identify current plateaus from measured current."""
    for item in files:
        step_delta = _step_delta_from_file(item)
        if step_delta is not None and step_delta > 0:
            return max(min(step_delta * 0.1, 1e-3), 1e-5)
    return 1e-5
def _format_limit_value(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:.6g}"

def _round_down_dec(value: float, decimals: int = 1) -> float:
    scale = 10 ** decimals
    return floor(value * scale) / scale

def _round_up_dec(value: float, decimals: int = 1) -> float:
    scale = 10 ** decimals
    return ceil(value * scale) / scale

def _format_limit_value(value: float | None, decimals: int = 1) -> str:
    if value is None:
        return ""
    return f"{value:.{decimals}f}"

def _padded_limits(
    values: list[float],
    rel_pad: float = 0.05,
    decimals: int = 1,
) -> tuple[float | None, float | None]:
    if not values:
        return None, None

    vmin = min(values)
    vmax = max(values)

    if vmin == vmax:
        pad = max(abs(vmin) * rel_pad, 1e-6)
    else:
        pad = (vmax - vmin) * rel_pad

    lo = _round_down_dec(vmin - pad, decimals)
    hi = _round_up_dec(vmax + pad, decimals)
    return lo, hi


def _bundle_area_cm2(bundle: CurveBundle) -> float | None:
    reference_files = bundle.asc_files or bundle.dsc_files
    if not reference_files:
        return None

    parsed = parse_gamry_dta(reference_files[0].path)
    area = to_float(parsed.meta_values.get("AREA", ""))
    if area is None or area <= 0:
        return None
    return area


def _scaled_current(value: float, use_current_density: bool, area_cm2: float | None) -> float:
    if not use_current_density:
        return value
    if area_cm2 is None or area_cm2 <= 0:
        raise ValueError("No se pudo leer un AREA valida de la metadata para convertir a A/cm^2.")
    return value / area_cm2


def _scaled_dvdi(value: float, use_current_density: bool, area_cm2: float | None) -> float:
    if not use_current_density:
        return value
    if area_cm2 is None or area_cm2 <= 0:
        raise ValueError("No se pudo leer un AREA valida de la metadata para convertir dV/dI.")
    return value * area_cm2


def _dvdi_axis_unit(use_current_density: bool) -> str:
    return "V/(A/cm^2)" if use_current_density else "V/A"


def compute_default_v_vs_i_limits(
    bundle: CurveBundle,
    use_current_density: bool = False,
) -> dict[str, str]:
    curve_data = build_curve_bundle_data(bundle)
    area_cm2 = _bundle_area_cm2(bundle) if use_current_density else None
    area_cm2 = _bundle_area_cm2(bundle) if use_current_density else None

    asc_rows = (
        select_fractional_point_per_step(curve_data["asc_rows"], curve_data["asc_tol"], 1.0)
        if curve_data["asc_rows"] else []
    )
    dsc_rows = (
        select_fractional_point_per_step(curve_data["dsc_rows"], curve_data["dsc_tol"], 1.0)
        if curve_data["dsc_rows"] else []
    )

    rows = asc_rows + dsc_rows
    if not rows:
        return {
            "x_min": "",
            "x_max": "",
            "v_min": "",
            "v_max": "",
        }

    x_min, x_max = _padded_limits(
        [_scaled_current(r["Corriente"], use_current_density, area_cm2) for r in rows]
    )
    x_min = 0.0
    v_min, v_max = _padded_limits([r["Voltaje"] for r in rows])

    return {
        "x_min": _format_limit_value(x_min),
        "x_max": _format_limit_value(x_max),
        "v_min": _format_limit_value(v_min),
        "v_max": _format_limit_value(v_max),
    }


def compute_autofit_v_vs_i_limits(
    bundle: CurveBundle,
    show_asc: bool,
    show_dsc: bool,
    show_voltage: bool,
    show_temperature: bool,
    point_fraction: float,
    use_current_density: bool = False,
    decimals: int = 1,
) -> dict[str, str]:
    if not (show_asc or show_dsc):
        raise ValueError("Debe seleccionar Ascendente y/o Descendente para usar Autoescala.")

    if not (show_voltage or show_temperature):
        raise ValueError("Debe seleccionar al menos una magnitud para usar Autoescala.")

    curve_data = build_curve_bundle_data(bundle)
    area_cm2 = _bundle_area_cm2(bundle) if use_current_density else None

    asc_rows = (
        select_fractional_point_per_step(curve_data["asc_rows"], curve_data["asc_tol"], point_fraction)
        if show_asc and curve_data["asc_rows"] else []
    )
    dsc_rows = (
        select_fractional_point_per_step(curve_data["dsc_rows"], curve_data["dsc_tol"], point_fraction)
        if show_dsc and curve_data["dsc_rows"] else []
    )

    rows = asc_rows + dsc_rows
    if not rows:
        raise ValueError("No hay datos validos para ajustar los ejes.")

    out = {
        "x_min": "",
        "x_max": "",
        "v_min": "",
        "v_max": "",
        "t_min": "",
        "t_max": "",
    }

    i_values = [_scaled_current(r["Corriente"], use_current_density, area_cm2) for r in rows]
    if i_values:
        out["x_min"] = _format_limit_value(_round_down_dec(min(i_values), decimals), decimals)
        out["x_max"] = _format_limit_value(_round_up_dec(max(i_values), decimals), decimals)

    if show_voltage:
        v_values = [r["Voltaje"] for r in rows]
        if v_values:
            voltage_step = _autoscale_step_from_values(v_values)
            out["v_min"] = _format_step_aligned_value(_round_down_to_step(min(v_values), voltage_step), voltage_step)
            out["v_max"] = _format_step_aligned_value(_round_up_to_step(max(v_values), voltage_step), voltage_step)

    if show_temperature:
        t_values = [r["Temperatura"] for r in rows]
        if t_values:
            out["t_min"] = _format_limit_value(_round_down_dec(min(t_values), decimals), decimals)
            out["t_max"] = _format_limit_value(_round_up_dec(max(t_values), decimals), decimals)

    return out

def draw_v_vs_i_on_figure(
    fig: Figure,
    bundle: CurveBundle,
    show_asc: bool,
    show_dsc: bool,
    show_voltage: bool,
    show_temperature: bool,
    use_current_density: bool,
    point_fraction: float,
    asc_marker: str,
    dsc_marker: str,
    voltage_linestyle: str,
    temperature_linestyle: str,
    x_tick_count: int = 6,
    y_tick_count: int = 6,
    x_min: float | None = None,
    x_max: float | None = None,
    v_min: float | None = None,
    v_max: float | None = None,
    temp_min: float | None = None,
    temp_max: float | None = None,
    plot_title: str = "",
    show_title: bool = True,
    title_fontsize: float = 14,
    tick_fontsize: float = 10,
    label_fontsize: float = 11,
    legend_fontsize: float = 10,
    marker_size: float = 6,
    hollow_markers: bool = True,
    line_width: float = 1.5,
    show_slope_guides: bool = False,
    indicator_current: float | None = None,
    language: str | None = None,
) -> bool:
    fig.clear()
    language = _pc_language(language)

    if not (show_asc or show_dsc):
        return False
    if not (show_voltage or show_temperature):
        return False

    x_tick_count = max(2, int(x_tick_count))
    y_tick_count = max(2, int(y_tick_count))

    curve_data = build_curve_bundle_data(bundle)
    area_cm2 = _bundle_area_cm2(bundle) if use_current_density else None

    asc_rows = (
        select_fractional_point_per_step(
            curve_data["asc_rows"],
            curve_data["asc_tol"],
            point_fraction,
        )
        if show_asc and curve_data["asc_rows"]
        else []
    )

    dsc_rows = (
        select_fractional_point_per_step(
            curve_data["dsc_rows"],
            curve_data["dsc_tol"],
            point_fraction,
        )
        if show_dsc and curve_data["dsc_rows"]
        else []
    )

    if not asc_rows and not dsc_rows:
        return False

    ax_main = fig.add_subplot(111)
    ax_temp = None

    if show_voltage and show_temperature:
        ax_temp = ax_main.twinx()

    asc_marker_mpl = _mpl_marker(asc_marker)
    dsc_marker_mpl = _mpl_marker(dsc_marker)
    voltage_ls_mpl = _mpl_linestyle(voltage_linestyle)
    temp_ls_mpl = _mpl_linestyle(temperature_linestyle)

    def _series_visible(marker_value: str, line_value: str) -> bool:
        return not (marker_value == "none" and line_value == "none")

    def _line_kwargs(color: str, mpl_marker: str, mpl_linestyle: str) -> dict:
        kwargs = {
            "color": color,
            "marker": mpl_marker,
            "linestyle": mpl_linestyle,
            "linewidth": line_width,
            "markersize": marker_size,
        }
        

        if mpl_marker != "None":
            kwargs["markeredgecolor"] = color
            kwargs["markeredgewidth"] = 1.2

            # x and + are already line-only markers, so facecolor does not matter much
            if hollow_markers and mpl_marker not in {"x", "+"}:
                kwargs["markerfacecolor"] = "none"
            else:
                kwargs["markerfacecolor"] = color

        return kwargs

    def _draw_slope_guide(points: list[tuple[float, float]], color: str) -> None:
        if not show_slope_guides or indicator_current is None:
            return

        state = _voltage_state_at_current_from_points(points, indicator_current)
        if state is None:
            return

        x_values = [point[0] for point in points]
        if len(x_values) < 2:
            return

        guide_x_min = x_min if x_min is not None else min(x_values)
        guide_x_max = x_max if x_max is not None else max(x_values)
        if guide_x_max <= guide_x_min:
            guide_x_min = min(x_values)
            guide_x_max = max(x_values)
        if guide_x_max <= guide_x_min:
            return

        guide_x = [guide_x_min, guide_x_max]
        guide_y = [
            state["voltage"] + state["slope"] * (x_value - state["current"])
            for x_value in guide_x
        ]

        ax_main.plot(
            guide_x,
            guide_y,
            color=color,
            linestyle=":",
            linewidth=max(1.0, line_width),
            alpha=0.85,
            label="_nolegend_",
        )
        ax_main.plot(
            [state["current"]],
            [state["voltage"]],
            color=color,
            marker="o",
            linestyle="none",
            markersize=max(marker_size + 1.0, 4.0),
            markerfacecolor="white",
            markeredgecolor=color,
            markeredgewidth=1.2,
            label="_nolegend_",
        )

    # Voltage on main axis
    if show_voltage:
        asc_points = _v_vs_i_voltage_points(asc_rows, use_current_density, area_cm2)
        dsc_points = _v_vs_i_voltage_points(dsc_rows, use_current_density, area_cm2)

        if asc_rows and _series_visible(asc_marker, voltage_linestyle):
            ax_main.plot(
                [point[0] for point in asc_points],
                [point[1] for point in asc_points],
                label=f"{translate('ascending', language)} {translate('voltage_short', language)}",
                **_line_kwargs(
                    PC_PLOT_COLORS["asc_voltage"],
                    asc_marker_mpl,
                    voltage_ls_mpl,
                ),
            )
        if dsc_rows and _series_visible(dsc_marker, voltage_linestyle):
            ax_main.plot(
                [point[0] for point in dsc_points],
                [point[1] for point in dsc_points],
                label=f"{translate('descending', language)} {translate('voltage_short', language)}",
                **_line_kwargs(
                    PC_PLOT_COLORS["dsc_voltage"],
                    dsc_marker_mpl,
                    voltage_ls_mpl,
                ),
            )
        _draw_slope_guide(asc_points, PC_PLOT_COLORS["asc_voltage"])
        _draw_slope_guide(dsc_points, PC_PLOT_COLORS["dsc_voltage"])
        ax_main.set_ylabel(f"{translate('voltage', language)} (V)", fontsize=label_fontsize)

    # Temperature on second axis if needed
    if show_temperature:
        target_ax = ax_temp if ax_temp is not None else ax_main

        if asc_rows and _series_visible(asc_marker, temperature_linestyle):
            target_ax.plot(
                [_scaled_current(r["Corriente"], use_current_density, area_cm2) for r in asc_rows],
                [r["Temperatura"] for r in asc_rows],
                label=f"{translate('ascending', language)} T",
                **_line_kwargs(
                    PC_PLOT_COLORS["asc_temperature"],
                    asc_marker_mpl,
                    temp_ls_mpl,
                ),
            )
        if dsc_rows and _series_visible(dsc_marker, temperature_linestyle):
            target_ax.plot(
                [_scaled_current(r["Corriente"], use_current_density, area_cm2) for r in dsc_rows],
                [r["Temperatura"] for r in dsc_rows],
                label=f"{translate('descending', language)} T",
                **_line_kwargs(
                    PC_PLOT_COLORS["dsc_temperature"],
                    dsc_marker_mpl,
                    temp_ls_mpl,
                ),
            )
        target_ax.set_ylabel(f"{translate('temperature', language)} (°C)", fontsize=label_fontsize)

    handles, labels = ax_main.get_legend_handles_labels()
    if ax_temp is not None:
        h2, l2 = ax_temp.get_legend_handles_labels()
        handles += h2
        labels += l2

    if not handles:
        fig.clear()
        return False

    default_title = (
        f"{translate('v_vs_i', language)} - {bundle.description} #{bundle.curve_id} "
        f"(punto step = {point_fraction:.2f})"
    )
    final_title = plot_title.strip() if plot_title.strip() else default_title

    ax_main.set_xlabel(
        f"{translate('current_density', language)} (A/cm^2)"
        if use_current_density
        else f"{translate('current', language)} (A)",
        fontsize=label_fontsize,
    )
    ax_main.set_title(final_title if show_title else "", fontsize=title_fontsize)
    ax_main.grid(True)

    # Tick font size
    ax_main.tick_params(axis="both", labelsize=tick_fontsize)
    apply_x_tick_label_padding(ax_main, tick_fontsize)
    ax_main.xaxis.set_major_locator(MaxNLocator(nbins=x_tick_count))
    ax_main.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))

    if x_min is not None or x_max is not None:
        ax_main.set_xlim(left=x_min, right=x_max)
        apply_x_edge_ticks(ax_main, x_min, x_max, x_tick_count)

    # Voltage axis
    if show_voltage:
        if v_min is not None or v_max is not None:
            ax_main.set_ylim(bottom=v_min, top=v_max)
            ax_main.yaxis.set_major_locator(LinearLocator(y_tick_count))
        else:
            ax_main.yaxis.set_major_locator(MaxNLocator(nbins=y_tick_count))
        ax_main.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))

    

    # Temperature axis
    if ax_temp is not None:
        ax_temp.tick_params(axis="y", labelsize=tick_fontsize)

        temp_lines = ax_temp.get_lines()
        temp_values = []
        for line in temp_lines:
            temp_values.extend(line.get_ydata())

        apply_temperature_axis_scaling(
            ax_temp=ax_temp,
            temp_values=temp_values,
            tick_count=y_tick_count,
            t_min=temp_min,
            t_max=temp_max,
        )

    make_legend_draggable(ax_main.legend(handles, labels, fontsize=legend_fontsize))
    fig.tight_layout()
    return True


def build_dv_di_rows(
    rows: list[dict[str, float]],
    current_tolerance: float,
    point_fraction: float,
    smoothing_algorithm: str = "Median filter",
    smoothing_window: int = 1,
) -> list[dict[str, float]]:
    selected_rows = select_fractional_point_per_step(rows, current_tolerance, point_fraction)
    derivative_rows: list[dict[str, float]] = []

    for idx in range(len(selected_rows) - 1):
        row_a = selected_rows[idx]
        row_b = selected_rows[idx + 1]

        delta_i = row_b["Corriente"] - row_a["Corriente"]
        if abs(delta_i) <= 1e-12:
            continue

        delta_v = row_b["Voltaje"] - row_a["Voltaje"]
        derivative_rows.append(
            {
                "Corriente": (row_a["Corriente"] + row_b["Corriente"]) / 2.0,
                "dVdI": delta_v / delta_i,
                "Step": float(idx + 1),
            }
        )

    smoothed_values = apply_smoothing(
        [row["dVdI"] for row in derivative_rows],
        smoothing_algorithm,
        smoothing_window,
    )
    for row, value in zip(derivative_rows, smoothed_values):
        row["dVdI"] = value

    return derivative_rows


def _positive_dv_di_rows(rows: list[dict[str, float]]) -> list[dict[str, float]]:
    return [row for row in rows if row["dVdI"] > 0]


def _scale_compatible_dv_di_rows(
    rows: list[dict[str, float]],
    logarithmic_y: bool,
) -> list[dict[str, float]]:
    return _positive_dv_di_rows(rows) if logarithmic_y else rows


def _format_log_limit_value(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:.6g}"


def _format_adaptive_limit_value(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:.6g}"


def _log_axis_limits(values: list[float]) -> tuple[float | None, float | None]:
    positive_values = [value for value in values if value > 0]
    if not positive_values:
        return None, None

    vmin = min(positive_values)
    vmax = max(positive_values)

    lower = 10 ** floor(log10(vmin))
    upper = 10 ** ceil(log10(vmax))

    if lower == upper:
        upper *= 10.0

    return lower, upper


def _adaptive_linear_limits(values: list[float]) -> tuple[float | None, float | None]:
    if not values:
        return None, None

    vmin = min(values)
    vmax = max(values)

    if vmin == vmax:
        pad = max(abs(vmin) * 0.05, 1e-9)
    else:
        pad = (vmax - vmin) * 0.05

    lo = vmin - pad
    hi = vmax + pad

    if lo == hi:
        hi = lo + max(abs(lo) * 0.05, 1e-9)

    return lo, hi


def _autoscale_step_from_values(values: list[float]) -> float:
    if not values:
        return 0.1
    max_abs = max(abs(value) for value in values)
    if max_abs <= 0:
        return 0.1
    return 10 ** (floor(log10(max_abs)) - 1)


def _round_down_to_step(value: float, step: float) -> float:
    if step <= 0:
        return value
    return floor(value / step) * step


def _round_up_to_step(value: float, step: float) -> float:
    if step <= 0:
        return value
    return ceil(value / step) * step


def _format_step_aligned_value(value: float | None, step: float) -> str:
    if value is None:
        return ""
    if step <= 0:
        return f"{value:.6g}"

    decimals = max(0, -floor(log10(step))) if step < 1 else 0
    formatted = f"{value:.{decimals}f}"
    if decimals > 0:
        formatted = formatted.rstrip("0").rstrip(".")
    return formatted


def _normalize_median_window(window_size: int) -> int:
    size = max(1, int(window_size))
    if size % 2 == 0:
        size += 1
    return size


def _median_filter(values: list[float], window_size: int) -> list[float]:
    if not values:
        return []

    size = _normalize_median_window(window_size)
    if size == 1:
        return list(values)

    radius = size // 2
    filtered: list[float] = []
    for idx in range(len(values)):
        start = max(0, idx - radius)
        end = min(len(values), idx + radius + 1)
        window = sorted(values[start:end])
        filtered.append(window[len(window) // 2])
    return filtered


def _normalize_average_window(window_size: int) -> int:
    return max(1, int(window_size))


def _rolling_average_filter(values: list[float], window_size: int) -> list[float]:
    if not values:
        return []

    size = _normalize_average_window(window_size)
    if size == 1:
        return list(values)

    radius_left = (size - 1) // 2
    radius_right = size // 2
    filtered: list[float] = []
    for idx in range(len(values)):
        start = max(0, idx - radius_left)
        end = min(len(values), idx + radius_right + 1)
        window = values[start:end]
        filtered.append(sum(window) / len(window))
    return filtered


def apply_smoothing(values: list[float], algorithm: str, window_size: int) -> list[float]:
    if algorithm == "Median filter":
        return _median_filter(values, window_size)
    if algorithm == "Rolling average":
        return _rolling_average_filter(values, window_size)
    raise ValueError(f"Algoritmo de suavizado no soportado: {algorithm}")


def compute_default_dv_di_limits(
    bundle: CurveBundle,
    smoothing_algorithm: str = "Median filter",
    smoothing_window: int = 1,
    logarithmic_y: bool = True,
    use_current_density: bool = False,
) -> dict[str, str]:
    curve_data = build_curve_bundle_data(bundle)
    area_cm2 = _bundle_area_cm2(bundle) if use_current_density else None

    asc_rows = (
        build_dv_di_rows(
            curve_data["asc_rows"],
            curve_data["asc_tol"],
            1.0,
            smoothing_algorithm=smoothing_algorithm,
            smoothing_window=smoothing_window,
        )
        if curve_data["asc_rows"] else []
    )
    dsc_rows = (
        build_dv_di_rows(
            curve_data["dsc_rows"],
            curve_data["dsc_tol"],
            1.0,
            smoothing_algorithm=smoothing_algorithm,
            smoothing_window=smoothing_window,
        )
        if curve_data["dsc_rows"] else []
    )

    rows = _scale_compatible_dv_di_rows(asc_rows + dsc_rows, logarithmic_y=logarithmic_y)
    if not rows:
        return {
            "x_min": "",
            "x_max": "",
            "dvdi_min": "",
            "dvdi_max": "",
        }

    _x_min, x_max = _padded_limits([
        _scaled_current(r["Corriente"], use_current_density, area_cm2)
        for r in rows
    ])
    if logarithmic_y:
        dvdi_min, dvdi_max = _log_axis_limits([
            _scaled_dvdi(r["dVdI"], use_current_density, area_cm2)
            for r in rows
        ])
    else:
        dvdi_min, dvdi_max = _adaptive_linear_limits([
            _scaled_dvdi(r["dVdI"], use_current_density, area_cm2)
            for r in rows
        ])

    return {
        "x_min": _format_limit_value(0.0),
        "x_max": _format_limit_value(x_max),
        "dvdi_min": _format_log_limit_value(dvdi_min) if logarithmic_y else _format_adaptive_limit_value(dvdi_min),
        "dvdi_max": _format_log_limit_value(dvdi_max) if logarithmic_y else _format_adaptive_limit_value(dvdi_max),
    }

def compute_autofit_dv_di_limits(
    bundle: CurveBundle,
    show_asc: bool,
    show_dsc: bool,
    point_fraction: float,
    smoothing_algorithm: str = "Median filter",
    smoothing_window: int = 1,
    logarithmic_y: bool = True,
    use_current_density: bool = False,
    locked_x_min: float | None = None,
    locked_x_max: float | None = None,
    locked_dvdi_min: float | None = None,
    locked_dvdi_max: float | None = None,
    decimals: int = 1,
) -> dict[str, str]:
    if not (show_asc or show_dsc):
        raise ValueError("Debe seleccionar Ascendente y/o Descendente para usar Autoescala.")

    curve_data = build_curve_bundle_data(bundle)
    area_cm2 = _bundle_area_cm2(bundle) if use_current_density else None

    asc_rows = (
        build_dv_di_rows(
            curve_data["asc_rows"],
            curve_data["asc_tol"],
            point_fraction,
            smoothing_algorithm=smoothing_algorithm,
            smoothing_window=smoothing_window,
        )
        if show_asc and curve_data["asc_rows"] else []
    )
    dsc_rows = (
        build_dv_di_rows(
            curve_data["dsc_rows"],
            curve_data["dsc_tol"],
            point_fraction,
            smoothing_algorithm=smoothing_algorithm,
            smoothing_window=smoothing_window,
        )
        if show_dsc and curve_data["dsc_rows"] else []
    )

    rows = _scale_compatible_dv_di_rows(asc_rows + dsc_rows, logarithmic_y=logarithmic_y)
    display_rows = []
    for row in rows:
        display_row = dict(row)
        display_row["Corriente"] = _scaled_current(row["Corriente"], use_current_density, area_cm2)
        display_row["dVdI"] = _scaled_dvdi(row["dVdI"], use_current_density, area_cm2)
        display_rows.append(display_row)
    filtered_rows: list[dict[str, float]] = []
    for row in display_rows:
        if locked_x_min is not None and row["Corriente"] < locked_x_min:
            continue
        if locked_x_max is not None and row["Corriente"] > locked_x_max:
            continue
        if locked_dvdi_min is not None and row["dVdI"] < locked_dvdi_min:
            continue
        if locked_dvdi_max is not None and row["dVdI"] > locked_dvdi_max:
            continue
        filtered_rows.append(row)

    if not filtered_rows:
        if logarithmic_y:
            raise ValueError("No hay valores positivos de dV/dI dentro de los limites bloqueados para ajustar los ejes.")
        raise ValueError("No hay datos validos de dV/dI dentro de los limites bloqueados para ajustar los ejes.")

    out = {
        "x_min": "",
        "x_max": "",
        "dvdi_min": "",
        "dvdi_max": "",
    }

    i_values = [r["Corriente"] for r in filtered_rows]
    if i_values:
        out["x_min"] = _format_limit_value(0.0, decimals)
        out["x_max"] = _format_limit_value(_round_up_dec(max(i_values), decimals), decimals)

    dvdi_values = [r["dVdI"] for r in filtered_rows]
    if dvdi_values:
        if logarithmic_y:
            dvdi_min, dvdi_max = _log_axis_limits(dvdi_values)
            out["dvdi_min"] = _format_log_limit_value(dvdi_min)
            out["dvdi_max"] = _format_log_limit_value(dvdi_max)
        else:
            dvdi_min, dvdi_max = _padded_limits(dvdi_values, decimals=decimals)
            out["dvdi_min"] = _format_adaptive_limit_value(dvdi_min)
            out["dvdi_max"] = _format_adaptive_limit_value(dvdi_max)

    return out

def draw_dv_di_on_figure(
    fig: Figure,
    bundle: CurveBundle,
    show_asc: bool,
    show_dsc: bool,
    point_fraction: float,
    asc_marker: str,
    dsc_marker: str,
    dvdi_linestyle: str,
    smoothing_algorithm: str = "Median filter",
    smoothing_window: int = 1,
    logarithmic_y: bool = True,
    use_current_density: bool = False,
    x_tick_count: int = 6,
    y_tick_count: int = 6,
    x_min: float | None = None,
    x_max: float | None = None,
    dvdi_min: float | None = None,
    dvdi_max: float | None = None,
    plot_title: str = "",
    show_title: bool = True,
    title_fontsize: float = 14,
    tick_fontsize: float = 10,
    label_fontsize: float = 11,
    legend_fontsize: float = 10,
    marker_size: float = 6,
    hollow_markers: bool = True,
    line_width: float = 1.5,
    language: str | None = None,
) -> bool:
    fig.clear()
    language = _pc_language(language)

    if not (show_asc or show_dsc):
        return False

    x_tick_count = max(2, int(x_tick_count))
    y_tick_count = max(2, int(y_tick_count))

    curve_data = build_curve_bundle_data(bundle)
    area_cm2 = _bundle_area_cm2(bundle) if use_current_density else None

    asc_rows = (
        build_dv_di_rows(
            curve_data["asc_rows"],
            curve_data["asc_tol"],
            point_fraction,
            smoothing_algorithm=smoothing_algorithm,
            smoothing_window=smoothing_window,
        )
        if show_asc and curve_data["asc_rows"]
        else []
    )
    dsc_rows = (
        build_dv_di_rows(
            curve_data["dsc_rows"],
            curve_data["dsc_tol"],
            point_fraction,
            smoothing_algorithm=smoothing_algorithm,
            smoothing_window=smoothing_window,
        )
        if show_dsc and curve_data["dsc_rows"]
        else []
    )

    asc_rows = _scale_compatible_dv_di_rows(asc_rows, logarithmic_y=logarithmic_y)
    dsc_rows = _scale_compatible_dv_di_rows(dsc_rows, logarithmic_y=logarithmic_y)

    if not asc_rows and not dsc_rows:
        return False

    ax_main = fig.add_subplot(111)

    asc_marker_mpl = _mpl_marker(asc_marker)
    dsc_marker_mpl = _mpl_marker(dsc_marker)
    dvdi_ls_mpl = _mpl_linestyle(dvdi_linestyle)

    def _series_visible(marker_value: str, line_value: str) -> bool:
        return not (marker_value == "none" and line_value == "none")

    def _line_kwargs(color: str, mpl_marker: str, mpl_linestyle: str) -> dict:
        kwargs = {
            "color": color,
            "marker": mpl_marker,
            "linestyle": mpl_linestyle,
            "linewidth": line_width,
            "markersize": marker_size,
        }

        if mpl_marker != "None":
            kwargs["markeredgecolor"] = color
            kwargs["markeredgewidth"] = 1.2
            if hollow_markers and mpl_marker not in {"x", "+"}:
                kwargs["markerfacecolor"] = "none"
            else:
                kwargs["markerfacecolor"] = color

        return kwargs

    if asc_rows and _series_visible(asc_marker, dvdi_linestyle):
        ax_main.plot(
            [
                _scaled_current(r["Corriente"], use_current_density, area_cm2)
                for r in asc_rows
            ],
            [
                _scaled_dvdi(r["dVdI"], use_current_density, area_cm2)
                for r in asc_rows
            ],
            label=f"dV/dI {translate('ascending', language).lower()}",
            **_line_kwargs(
                PC_PLOT_COLORS["asc_voltage"],
                asc_marker_mpl,
                dvdi_ls_mpl,
            ),
        )

    if dsc_rows and _series_visible(dsc_marker, dvdi_linestyle):
        ax_main.plot(
            [
                _scaled_current(r["Corriente"], use_current_density, area_cm2)
                for r in dsc_rows
            ],
            [
                _scaled_dvdi(r["dVdI"], use_current_density, area_cm2)
                for r in dsc_rows
            ],
            label=f"dV/dI {translate('descending', language).lower()}",
            **_line_kwargs(
                PC_PLOT_COLORS["dsc_voltage"],
                dsc_marker_mpl,
                dvdi_ls_mpl,
            ),
        )

    handles, labels = ax_main.get_legend_handles_labels()
    if not handles:
        fig.clear()
        return False

    default_title = (
        f"{translate('dv_di', language)} vs "
        f"{translate('current_density', language) if use_current_density else translate('current', language)} "
        f"- {bundle.description} #{bundle.curve_id} "
        f"(punto step = {point_fraction:.2f})"
    )
    final_title = plot_title.strip() if plot_title.strip() else default_title

    ax_main.set_xlabel(
        f"{translate('current_density', language)} (A/cm^2)"
        if use_current_density
        else f"{translate('current', language)} (A)",
        fontsize=label_fontsize,
    )
    ax_main.set_ylabel(f"{translate('dv_di', language)} ({_dvdi_axis_unit(use_current_density)})", fontsize=label_fontsize)
    ax_main.set_title(final_title if show_title else "", fontsize=title_fontsize)
    ax_main.grid(True)
    ax_main.tick_params(axis="both", labelsize=tick_fontsize)
    apply_x_tick_label_padding(ax_main, tick_fontsize)
    ax_main.xaxis.set_major_locator(MaxNLocator(nbins=x_tick_count))
    ax_main.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))

    if x_min is not None or x_max is not None:
        ax_main.set_xlim(left=x_min, right=x_max)
        apply_x_edge_ticks(ax_main, x_min, x_max, x_tick_count)

    if logarithmic_y and dvdi_min is not None and dvdi_min <= 0:
        raise ValueError("dV/dI min debe ser mayor que 0 para usar escala logarítmica.")
    if logarithmic_y and dvdi_max is not None and dvdi_max <= 0:
        raise ValueError("dV/dI max debe ser mayor que 0 para usar escala logarítmica.")

    if logarithmic_y:
        ax_main.set_yscale("log")
        if dvdi_min is not None or dvdi_max is not None:
            ax_main.set_ylim(bottom=dvdi_min, top=dvdi_max)

        ax_main.yaxis.set_major_locator(LogLocator(base=10.0, numticks=max(2, y_tick_count)))
        ax_main.yaxis.set_minor_locator(LogLocator(base=10.0, subs=tuple(range(2, 10))))
        ax_main.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
        ax_main.yaxis.set_minor_formatter(LogFormatterSciNotation(base=10.0, labelOnlyBase=True))
        ax_main.grid(True, which="minor", axis="y", alpha=0.25)
    else:
        if dvdi_min is not None or dvdi_max is not None:
            ax_main.set_ylim(bottom=dvdi_min, top=dvdi_max)
            ax_main.yaxis.set_major_locator(LinearLocator(y_tick_count))
        else:
            ax_main.yaxis.set_major_locator(MaxNLocator(nbins=y_tick_count))
        ax_main.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))

    make_legend_draggable(ax_main.legend(handles, labels, fontsize=legend_fontsize))
    fig.tight_layout()
    return True

def _step_stability_direction_rows(
    bundle: CurveBundle,
    use_current_density: bool = True,
) -> dict[str, list[dict[str, float]]]:
    curve_data = build_curve_bundle_data(bundle)
    area_cm2 = _bundle_area_cm2(bundle) if use_current_density else None
    return {
        "asc_rows": (
            build_step_stability_rows(
                curve_data["asc_rows"],
                curve_data["asc_tol"],
                use_current_density,
                area_cm2,
            )
            if curve_data["asc_rows"] else []
        ),
        "dsc_rows": (
            build_step_stability_rows(
                curve_data["dsc_rows"],
                curve_data["dsc_tol"],
                use_current_density,
                area_cm2,
            )
            if curve_data["dsc_rows"] else []
        ),
    }


def compute_default_step_stability_limits(
    bundle: CurveBundle,
    use_current_density: bool = True,
) -> dict[str, str]:
    direction_rows = _step_stability_direction_rows(bundle, use_current_density=use_current_density)
    dv_values = [row["DeltaV"] for rows in direction_rows.values() for row in rows]
    rows = direction_rows["asc_rows"] + direction_rows["dsc_rows"]
    if not rows:
        return {
            "x_min": "0",
            "x_max": "",
            "dv_min": "0",
            "dv_max": "",
            "dt_min": "0",
            "dt_max": "",
        }

    _x_min, x_max = _padded_limits([row["Corriente"] for row in rows])
    _dv_min, dv_max = _padded_limits([row["DeltaV"] for row in rows])
    _dt_min, dt_max = _padded_limits([row["DeltaT"] for row in rows])

    return {
        "x_min": _format_limit_value(0.0),
        "x_max": _format_limit_value(x_max),
        "dv_min": _format_limit_value(_positive_min(dv_values) or 1e-6),
        "dv_max": _format_limit_value(dv_max),
        "dt_min": _format_limit_value(0.0),
        "dt_max": _format_limit_value(dt_max),
    }


def compute_autofit_step_stability_limits(
    bundle: CurveBundle,
    show_asc: bool,
    show_dsc: bool,
    show_voltage_delta: bool,
    show_temperature_delta: bool,
    use_current_density: bool = True,
    decimals: int = 1,
) -> dict[str, str]:
    if not (show_asc or show_dsc):
        raise ValueError("Debe seleccionar Ascendente y/o Descendente para usar Autoescala.")
    if not (show_voltage_delta or show_temperature_delta):
        raise ValueError("Debe seleccionar al menos una magnitud para usar Autoescala.")

    direction_rows = _step_stability_direction_rows(bundle, use_current_density=use_current_density)
    rows: list[dict[str, float]] = []
    if show_asc:
        rows.extend(direction_rows["asc_rows"])
    if show_dsc:
        rows.extend(direction_rows["dsc_rows"])
    if not rows:
        raise ValueError("No hay datos validos para ajustar los ejes.")

    out = {
        "x_min": _format_limit_value(0.0, decimals),
        "x_max": "",
        "dv_min": "",
        "dv_max": "",
        "dt_min": "",
        "dt_max": "",
    }

    current_values = [row["Corriente"] for row in rows]
    if current_values:
        out["x_max"] = _format_limit_value(_round_up_dec(max(current_values), decimals), decimals)

    if show_voltage_delta:
        delta_v_values = [row["DeltaV"] for row in rows]
        if delta_v_values:
            out["dv_min"] = _format_limit_value(0.0, decimals)
            out["dv_max"] = _format_limit_value(_round_up_dec(max(delta_v_values), decimals), decimals)

    if show_temperature_delta:
        delta_t_values = [row["DeltaT"] for row in rows]
        if delta_t_values:
            out["dt_min"] = _format_limit_value(0.0, decimals)
            out["dt_max"] = _format_limit_value(_round_up_dec(max(delta_t_values), decimals), decimals)

    return out


def draw_step_stability_on_figure(
    fig: Figure,
    bundle: CurveBundle,
    show_asc: bool,
    show_dsc: bool,
    show_voltage_delta: bool,
    show_temperature_delta: bool,
    use_current_density: bool,
    asc_marker: str,
    dsc_marker: str,
    voltage_linestyle: str,
    temperature_linestyle: str,
    x_tick_count: int = 6,
    y_tick_count: int = 6,
    x_min: float | None = None,
    x_max: float | None = None,
    dv_min: float | None = None,
    dv_max: float | None = None,
    dt_min: float | None = None,
    dt_max: float | None = None,
    plot_title: str = "",
    show_title: bool = True,
    title_fontsize: float = 14,
    tick_fontsize: float = 10,
    label_fontsize: float = 11,
    legend_fontsize: float = 10,
    marker_size: float = 6,
    hollow_markers: bool = False,
    line_width: float = 1.5,
    language: str | None = None,
) -> bool:
    fig.clear()
    language = _pc_language(language)

    if not (show_asc or show_dsc):
        return False
    if not (show_voltage_delta or show_temperature_delta):
        return False

    direction_rows = _step_stability_direction_rows(bundle, use_current_density=use_current_density)
    asc_rows = direction_rows["asc_rows"] if show_asc else []
    dsc_rows = direction_rows["dsc_rows"] if show_dsc else []
    if not asc_rows and not dsc_rows:
        return False

    ax_main = fig.add_subplot(111)
    ax_temp = None
    if show_voltage_delta and show_temperature_delta:
        ax_temp = ax_main.twinx()

    asc_marker_mpl = _mpl_marker(asc_marker)
    dsc_marker_mpl = _mpl_marker(dsc_marker)
    voltage_ls_mpl = _mpl_linestyle(voltage_linestyle)
    temp_ls_mpl = _mpl_linestyle(temperature_linestyle)

    def _series_visible(marker_value: str, line_value: str) -> bool:
        return not (marker_value == "none" and line_value == "none")

    def _line_kwargs(color: str, mpl_marker: str, mpl_linestyle: str) -> dict:
        kwargs = {
            "color": color,
            "marker": mpl_marker,
            "linestyle": mpl_linestyle,
            "linewidth": line_width,
            "markersize": marker_size,
        }
        if mpl_marker != "None":
            kwargs["markeredgecolor"] = color
            kwargs["markeredgewidth"] = 1.2
            kwargs["markerfacecolor"] = "none" if hollow_markers and mpl_marker not in {"x", "+"} else color
        return kwargs

    def _plot_direction(rows: list[dict[str, float]], prefix_key: str, marker_text: str, mpl_marker: str) -> None:
        if not rows:
            return
        prefix = translate(prefix_key, language)
        x_values = [row["Corriente"] for row in rows]

        if show_voltage_delta and _series_visible(marker_text, voltage_linestyle):
            ax_main.plot(
                x_values,
                [row["DeltaV"] for row in rows],
                label=f"{prefix} {translate('delta_voltage', language)}",
                **_line_kwargs(
                    PC_PLOT_COLORS["asc_voltage"] if prefix_key == "ascending" else PC_PLOT_COLORS["dsc_voltage"],
                    mpl_marker,
                    voltage_ls_mpl,
                ),
            )

        if show_temperature_delta and _series_visible(marker_text, temperature_linestyle):
            target_ax = ax_temp if ax_temp is not None else ax_main
            target_ax.plot(
                x_values,
                [row["DeltaT"] for row in rows],
                label=f"{prefix} {translate('delta_temperature', language)}",
                **_line_kwargs(
                    PC_PLOT_COLORS["asc_temperature"] if prefix_key == "ascending" else PC_PLOT_COLORS["dsc_temperature"],
                    mpl_marker,
                    temp_ls_mpl,
                ),
            )

    _plot_direction(asc_rows, "ascending", asc_marker, asc_marker_mpl)
    _plot_direction(dsc_rows, "descending", dsc_marker, dsc_marker_mpl)

    handles, labels = ax_main.get_legend_handles_labels()
    if ax_temp is not None:
        h2, l2 = ax_temp.get_legend_handles_labels()
        handles += h2
        labels += l2
    if not handles:
        fig.clear()
        return False

    default_title = f"{translate('step_stability', language)} - {bundle.description} #{bundle.curve_id}"
    final_title = plot_title.strip() if plot_title.strip() else default_title

    ax_main.set_xlabel(
        f"{translate('current_density', language)} (A/cm^2)"
        if use_current_density
        else f"{translate('current', language)} (A)",
        fontsize=label_fontsize,
    )
    ax_main.set_title(final_title if show_title else "", fontsize=title_fontsize)
    ax_main.grid(True)
    ax_main.tick_params(axis="both", labelsize=tick_fontsize)
    apply_x_tick_label_padding(ax_main, tick_fontsize)
    ax_main.xaxis.set_major_locator(MaxNLocator(nbins=max(2, int(x_tick_count))))
    ax_main.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))

    if x_min is not None or x_max is not None:
        ax_main.set_xlim(left=x_min, right=x_max)
        apply_x_edge_ticks(ax_main, x_min, x_max, x_tick_count)

    if show_voltage_delta:
        ax_main.set_ylabel(f"{translate('delta_voltage', language)} (V)", fontsize=label_fontsize)
        if dv_min is not None or dv_max is not None:
            ax_main.set_ylim(bottom=dv_min, top=dv_max)
            ax_main.yaxis.set_major_locator(LinearLocator(max(2, int(y_tick_count))))
        else:
            ax_main.yaxis.set_major_locator(MaxNLocator(nbins=max(2, int(y_tick_count))))
        ax_main.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
    elif show_temperature_delta:
        ax_main.set_ylabel(f"{translate('delta_temperature', language)} (C)", fontsize=label_fontsize)
        if dt_min is not None or dt_max is not None:
            ax_main.set_ylim(bottom=dt_min, top=dt_max)
            ax_main.yaxis.set_major_locator(LinearLocator(max(2, int(y_tick_count))))
        else:
            ax_main.yaxis.set_major_locator(MaxNLocator(nbins=max(2, int(y_tick_count))))
        ax_main.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))

    if ax_temp is not None:
        ax_temp.tick_params(axis="y", labelsize=tick_fontsize)
        ax_temp.set_ylabel(f"{translate('delta_temperature', language)} (C)", fontsize=label_fontsize)
        if dt_min is not None or dt_max is not None:
            ax_temp.set_ylim(bottom=dt_min, top=dt_max)
            ax_temp.yaxis.set_major_locator(LinearLocator(max(2, int(y_tick_count))))
            ax_temp.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
        else:
            temp_values = [row["DeltaT"] for row in asc_rows + dsc_rows]
            apply_secondary_axis_scaling(ax_temp, temp_values, max(2, int(y_tick_count)))

    make_legend_draggable(ax_main.legend(handles, labels, fontsize=legend_fontsize, loc="best"))
    fig.tight_layout()
    return True


def draw_series_by_time_on_figure(
    fig: Figure,
    bundle: CurveBundle,
    show_asc: bool,
    show_dsc: bool,
    show_voltage: bool,
    show_current: bool,
    show_temperature: bool,
    asc_marker: str,
    dsc_marker: str,
    voltage_linestyle: str,
    current_linestyle: str,
    temperature_linestyle: str,
    use_current_density: bool = False,
    time_unit: str = "s",
    x_tick_count: int = 6,
    y_tick_count: int = 6,
    t_min: float | None = None,
    t_max: float | None = None,
    v_min: float | None = None,
    v_max: float | None = None,
    i_min: float | None = None,
    i_max: float | None = None,
    temp_min: float | None = None,
    temp_max: float | None = None,
    plot_title: str = "",
    show_title: bool = True,
    title_fontsize: float = 14,
    tick_fontsize: float = 10,
    label_fontsize: float = 11,
    legend_fontsize: float = 10,
    marker_size: float = 6,
    line_width: float = 1.5,
    hollow_markers: bool = False,
    language: str = "es",
) -> bool:
    fig.clear()
    language = normalize_language(language)

    if not (show_asc or show_dsc):
        return False
    if not (show_voltage or show_current or show_temperature):
        return False

    x_tick_count = max(2, int(x_tick_count))
    y_tick_count = max(2, int(y_tick_count))

    plot_data = build_series_by_time_plot_data(bundle, time_unit=time_unit)
    area_cm2 = _bundle_area_cm2(bundle) if use_current_density else None
    asc_rows = plot_data["asc_rows"] if show_asc else []
    dsc_rows = plot_data["dsc_rows"] if show_dsc else []

    if not asc_rows and not dsc_rows:
        return False

    ax_main = fig.add_subplot(111)
    ax_current = None
    ax_temp = None

    # Axis layout
    if show_voltage:
        if show_current:
            ax_current = ax_main.twinx()
            ax_current.spines["left"].set_visible(False)
            ax_current.yaxis.tick_right()
            ax_current.yaxis.set_label_position("right")

        if show_temperature:
            ax_temp = ax_main.twinx()
            ax_temp.spines["left"].set_visible(False)
            ax_temp.yaxis.tick_right()
            ax_temp.yaxis.set_label_position("right")

            if ax_current is not None:
                ax_temp.spines["right"].set_position(("outward", 60))
    else:
        # No voltage selected: use main axis for current if present, otherwise temperature
        if show_current:
            ax_current = ax_main
        if show_temperature:
            if show_current:
                ax_temp = ax_main.twinx()
                ax_temp.spines["left"].set_visible(False)
                ax_temp.yaxis.tick_right()
                ax_temp.yaxis.set_label_position("right")
                ax_temp.spines["right"].set_position(("outward", 60))
            else:
                ax_temp = ax_main

    voltage_ls_mpl = _mpl_linestyle(voltage_linestyle)
    current_ls_mpl = _mpl_linestyle(current_linestyle)
    temp_ls_mpl = _mpl_linestyle(temperature_linestyle)

    def _series_visible(line_value: str) -> bool:
        return line_value != "none"

    def _line_kwargs(color: str, mpl_marker: str, mpl_linestyle: str) -> dict:
        kwargs = {
            "color": color,
            "marker": mpl_marker,
            "linestyle": mpl_linestyle,
            "linewidth": line_width,
            "markersize": marker_size,
        }

        if mpl_marker != "None":
            kwargs["markeredgecolor"] = color
            kwargs["markeredgewidth"] = 1.2
            kwargs["markerfacecolor"] = "none" if hollow_markers and mpl_marker not in {"x", "+"} else color

        return kwargs

    def _plot_rows(rows):
        if not rows:
            return

        x_vals = [r["plot_time"] for r in rows]

        if show_voltage and _series_visible(voltage_linestyle):
            ax_main.plot(
                x_vals,
                [r["Voltaje"] for r in rows],
                label=translate("voltage", language),
                **_line_kwargs(
                    PC_PLOT_COLORS["asc_voltage"],
                    "None",
                    voltage_ls_mpl,
                ),
            )

        if show_current and ax_current is not None and _series_visible(current_linestyle):
            ax_current.plot(
                x_vals,
                [_scaled_current(r["Corriente"], use_current_density, area_cm2) for r in rows],
                label=translate("current", language),
                **_line_kwargs(
                    PC_PLOT_COLORS["asc_current"],
                    "None",
                    current_ls_mpl,
                ),
            )

        if show_temperature and ax_temp is not None and _series_visible(temperature_linestyle):
            ax_temp.plot(
                x_vals,
                [r["Temperatura"] for r in rows],
                label=translate("temperature", language),
                **_line_kwargs(
                    PC_PLOT_COLORS["asc_temperature"],
                    "None",
                    temp_ls_mpl,
                ),
            )

    _plot_rows(asc_rows + dsc_rows)

    if asc_rows and dsc_rows:
        transition_time = plot_data["t_asc_end"] / _time_unit_scale(time_unit)
        ax_main.axvline(
            transition_time,
            color="#5b677a",
            linestyle=":",
            linewidth=max(1.0, line_width),
            alpha=0.9,
            label=translate("ascending_finish_marker", language),
        )

    handles, labels = ax_main.get_legend_handles_labels()
    for extra_ax in (ax_current, ax_temp):
        if extra_ax is not None and extra_ax is not ax_main:
            h2, l2 = extra_ax.get_legend_handles_labels()
            handles += h2
            labels += l2

    if ax_current is ax_main:
        current_label = (
            f"{translate('current_density', language)} (A/cm^2)"
            if use_current_density
            else f"{translate('current', language)} (A)"
        )
        ax_main.set_ylabel(current_label, fontsize=label_fontsize)
        current_values = [
            _scaled_current(r["Corriente"], use_current_density, area_cm2)
            for r in asc_rows + dsc_rows
        ]
        apply_current_axis_scaling(ax_main, current_values, y_tick_count, i_min, i_max)

    if not handles:
        fig.clear()
        return False

    default_title = f"{translate('series_by_time', language)} - {bundle.description} #{bundle.curve_id}"
    final_title = plot_title.strip() if plot_title.strip() else default_title

    ax_main.set_xlabel(f"{translate('time', language)} ({time_unit})", fontsize=label_fontsize)
    ax_main.set_title(final_title if show_title else "", fontsize=title_fontsize)
    ax_main.grid(True)
    ax_main.xaxis.set_major_locator(MaxNLocator(nbins=x_tick_count))
    ax_main.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
    ax_main.tick_params(axis="both", labelsize=tick_fontsize)
    apply_x_tick_label_padding(ax_main, tick_fontsize)

    if t_min is not None or t_max is not None:
        ax_main.set_xlim(left=t_min, right=t_max)
        apply_x_edge_ticks(ax_main, t_min, t_max, x_tick_count)
    ax_main.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))

    # Voltage axis
    if show_voltage:
        ax_main.set_ylabel(f"{translate('voltage', language)} (V)", fontsize=label_fontsize)
        if v_min is not None or v_max is not None:
            ax_main.set_ylim(bottom=v_min, top=v_max)
            ax_main.yaxis.set_major_locator(LinearLocator(y_tick_count))
        else:
            ax_main.yaxis.set_major_locator(MaxNLocator(nbins=y_tick_count))
        ax_main.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
    else:
        # Main axis is being used by current or temperature
        if ax_current is ax_main:
            current_label = (
                f"{translate('current_density', language)} (A/cm^2)"
                if use_current_density
                else f"{translate('current', language)} (A)"
            )
            ax_main.set_ylabel(current_label, fontsize=label_fontsize)
            current_values = [
                _scaled_current(r["Corriente"], use_current_density, area_cm2)
                for r in asc_rows + dsc_rows
            ]
            apply_current_axis_scaling(ax_main, current_values, y_tick_count, i_min, i_max)
        elif ax_temp is ax_main:
            ax_main.set_ylabel(f"{translate('temperature', language)} (°C)", fontsize=label_fontsize)
            temp_values = [r["Temperatura"] for r in asc_rows] + [r["Temperatura"] for r in dsc_rows]
            if temp_min is not None or temp_max is not None:
                current_lo, current_hi = ax_main.get_ylim()
                ax_main.set_ylim(
                    current_lo if temp_min is None else temp_min,
                    current_hi if temp_max is None else temp_max,
                )
                ax_main.yaxis.set_major_locator(LinearLocator(y_tick_count))
                ax_main.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
            else:
                apply_secondary_axis_scaling(ax_main, temp_values, y_tick_count)

    if ax_current is not None and ax_current is not ax_main:
        ax_current.tick_params(axis="y", labelsize=tick_fontsize)
        current_label = (
            f"{translate('current_density', language)} (A/cm^2)"
            if use_current_density
            else f"{translate('current', language)} (A)"
        )
        ax_current.set_ylabel(current_label, fontsize=label_fontsize)
        current_values = [
            _scaled_current(r["Corriente"], use_current_density, area_cm2)
            for r in asc_rows + dsc_rows
        ]
        apply_current_axis_scaling(ax_current, current_values, y_tick_count, i_min, i_max)

    if ax_temp is not None and ax_temp is not ax_main:
        ax_temp.tick_params(axis="y", labelsize=tick_fontsize)
        ax_temp.set_ylabel(f"{translate('temperature', language)} (°C)", fontsize=label_fontsize)
        temp_values = [r["Temperatura"] for r in asc_rows] + [r["Temperatura"] for r in dsc_rows]
        if temp_min is not None or temp_max is not None:
            current_lo, current_hi = ax_temp.get_ylim()
            ax_temp.set_ylim(
                current_lo if temp_min is None else temp_min,
                current_hi if temp_max is None else temp_max,
            )
            ax_temp.yaxis.set_major_locator(LinearLocator(y_tick_count))
            ax_temp.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
        else:
            apply_secondary_axis_scaling(ax_temp, temp_values, y_tick_count)

    # Extra room when 3 y-axes are visible
    if show_voltage and show_current and show_temperature:
        fig.subplots_adjust(right=0.80)

    def _best_legend_loc() -> str:
        candidate_boxes = {
            "upper right": (0.70, 0.72, 0.98, 0.96),
            "upper left": (0.02, 0.72, 0.30, 0.96),
            "lower right": (0.70, 0.04, 0.98, 0.28),
            "lower left": (0.02, 0.04, 0.30, 0.28),
            "center right": (0.70, 0.38, 0.98, 0.62),
            "center left": (0.02, 0.38, 0.30, 0.62),
        }
        axes_to_scan = [
            axis
            for axis in (ax_main, ax_current, ax_temp)
            if axis is not None
        ]
        best_loc = "best"
        best_score: tuple[int, float] | None = None

        for loc, (x0, y0, x1, y1) in candidate_boxes.items():
            overlap_count = 0
            nearest_distance = 1.0
            box_cx = (x0 + x1) / 2.0
            box_cy = (y0 + y1) / 2.0

            for axis in axes_to_scan:
                for line in axis.get_lines():
                    label = str(line.get_label() or "")
                    if label.startswith("_"):
                        continue
                    for x_value, y_value in zip(
                        line.get_xdata(orig=False),
                        line.get_ydata(orig=False),
                    ):
                        try:
                            display_xy = axis.transData.transform((float(x_value), float(y_value)))
                            axes_xy = ax_main.transAxes.inverted().transform(display_xy)
                        except (TypeError, ValueError):
                            continue

                        point_x = float(axes_xy[0])
                        point_y = float(axes_xy[1])
                        if not (0.0 <= point_x <= 1.0 and 0.0 <= point_y <= 1.0):
                            continue

                        if x0 <= point_x <= x1 and y0 <= point_y <= y1:
                            overlap_count += 1
                        nearest_distance = min(
                            nearest_distance,
                            ((point_x - box_cx) ** 2 + (point_y - box_cy) ** 2) ** 0.5,
                        )

            score = (overlap_count, -nearest_distance)
            if best_score is None or score < best_score:
                best_score = score
                best_loc = loc

        return best_loc

    make_legend_draggable(ax_main.legend(handles, labels, fontsize=legend_fontsize, loc=_best_legend_loc()))
    fig.tight_layout()
    return True

def _mpl_marker(value: str) -> str:
    return "None" if value == "none" else value


def _mpl_linestyle(value: str) -> str:
    return "None" if value == "none" else value


def _build_scrollable_controls(parent) -> ttk.Frame:
    _outer, controls_frame = create_scrollable_controls(parent)
    return controls_frame

def compute_autofit_series_by_time_limits(
    bundle: CurveBundle,
    show_asc: bool,
    show_dsc: bool,
    show_voltage: bool,
    show_current: bool,
    show_temperature: bool,
    decimals: int = 1,
) -> dict[str, str]:
    if not (show_asc or show_dsc):
        raise ValueError("Debe seleccionar Ascendente y/o Descendente para usar Autoescala.")

    if not (show_voltage or show_current or show_temperature):
        raise ValueError("Debe seleccionar al menos una magnitud para usar Autoescala.")

    plot_data = build_series_by_time_plot_data(bundle)

    rows = []
    if show_asc:
        rows.extend(plot_data["asc_rows"])
    if show_dsc:
        rows.extend(plot_data["dsc_rows"])

    if not rows:
        raise ValueError("No hay datos válidos para ajustar los ejes.")

    out = {
        "t_min": "",
        "t_max": "",
        "v_min": "",
        "v_max": "",
        "i_min": "0",
        "i_max": "",
    }

    # Time axis: natural visible range of selected rows
    t_values = [r["plot_time"] for r in rows]
    if t_values:
        out["t_min"] = _format_limit_value(_round_down_dec(min(t_values), decimals), decimals)
        out["t_max"] = _format_limit_value(_round_up_dec(max(t_values), decimals), decimals)

    # Voltage axis: same logic as V vs I autofit
    if show_voltage:
        v_values = [r["Voltaje"] for r in rows]
        if v_values:
            voltage_step = _autoscale_step_from_values(v_values)
            out["v_min"] = _format_step_aligned_value(_round_down_to_step(min(v_values), voltage_step), voltage_step)
            out["v_max"] = _format_step_aligned_value(_round_up_to_step(max(v_values), voltage_step), voltage_step)

    # Current axis: always start at 0
    if show_current:
        i_values = [r["Corriente"] for r in rows]
        if i_values:
            out["i_max"] = _format_limit_value(_round_up_dec(max(i_values), decimals), decimals)

    return out

# ---------------------------------------------------------------------------
# File discovery and grouping
# ---------------------------------------------------------------------------
def _parse_filename(path: Path) -> PolarizationFile | None:
    normalized_name = _normalize_filename_for_parser(path)
    match = FILE_RE.match(normalized_name)

    if not match:
        return None

    return PolarizationFile(
        path=path,
        direction=match.group("direction").title(),
        description=match.group("description"),
        curve_id=int(match.group("curve_id")),
        file_index=int(match.group("file_index")),
    )


def discover_curve_bundles(input_dir: Path) -> list[CurveBundle]:
    grouped: dict[tuple[str, int], dict[str, list[PolarizationFile]]] = defaultdict(
        lambda: {"Asc": [], "Dsc": []}
    )

    for path in sorted(input_dir.glob("*.DTA")):
        info = _parse_filename(path)
        if info is None:
            continue
        grouped[(info.description, info.curve_id)][info.direction].append(info)

    bundles: list[CurveBundle] = []
    for (description, curve_id), by_dir in sorted(grouped.items(), key=lambda x: (x[0][0], x[0][1])):
        asc_files = sorted(by_dir["Asc"], key=lambda item: item.file_index)
        dsc_files = sorted(by_dir["Dsc"], key=lambda item: item.file_index)
        bundles.append(
            CurveBundle(
                description=description,
                curve_id=curve_id,
                asc_files=asc_files,
                dsc_files=dsc_files,
            )
        )
    return bundles


# ---------------------------------------------------------------------------
# Parsing one Gamry CHRONOP file
# ---------------------------------------------------------------------------
def parse_gamry_dta(path: Path) -> ParsedDTA:
    text = path.read_text(encoding="latin-1", errors="replace")
    lines = text.splitlines()

    meta_values: dict[str, str] = {}
    meta_units: dict[str, str] = {}
    header: list[str] | None = None
    units: list[str] = []
    rows: list[list[str]] = []

    table_started = False

    for line in lines:
        if not table_started:
            if line.startswith("CURVE") and "TABLE" in line:
                table_started = True
                continue

            if not line.strip():
                continue

            parts = line.split("\t")
            if len(parts) >= 3 and parts[0].strip():
                key = parts[0].strip()
                value = parts[2].strip()
                description = " ".join(p.strip() for p in parts[3:] if p.strip())
                meta_values[key] = value
                meta_units[key] = _extract_meta_unit(key, description)
            continue

        if not line.strip():
            continue

        parts = _drop_leading_blank([p.strip() for p in line.rstrip("\r\n").split("\t")])
        if not parts:
            continue

        first = parts[0]
        if header is None:
            if first == "Pt":
                header = parts
            continue

        if not units:
            if first == "#":
                units = parts
            continue

        if re.fullmatch(r"-?\d+", first):
            rows.append(parts)

    if header is None:
        raise ValueError(f"No data header found in {path.name} (expected 'Pt ...')")

    return ParsedDTA(
        meta_values=meta_values,
        meta_units=meta_units,
        header=header,
        units=units,
        rows=rows,
    )


# ---------------------------------------------------------------------------
# Build metadata
# ---------------------------------------------------------------------------
def _collect_current_extremes(files: list[PolarizationFile]) -> tuple[float | None, float | None]:
    currents: list[float] = []
    for item in files:
        parsed = parse_gamry_dta(item.path)
        for key in ("ISTEP1", "ISTEP2"):
            val = to_float(parsed.meta_values.get(key, ""))
            if val is not None:
                currents.append(val)
    if not currents:
        return None, None
    return min(currents), max(currents)


def build_metadata(bundle: CurveBundle) -> list[tuple[str, object, str]]:
    reference_files = bundle.asc_files or bundle.dsc_files
    if not reference_files:
        raise ValueError("No se encontraron archivos Asc ni Dsc para exportar.")

    first_parsed = parse_gamry_dta(reference_files[0].path)

    step_duration = to_float(first_parsed.meta_values.get("TSTEP1", ""))
    sample_time = to_float(first_parsed.meta_values.get("SAMPLETIME", ""))
    area = to_float(first_parsed.meta_values.get("AREA", ""))

    i1 = to_float(first_parsed.meta_values.get("ISTEP1", ""))
    i2 = to_float(first_parsed.meta_values.get("ISTEP2", ""))
    delta_i = abs(i2 - i1) if i1 is not None and i2 is not None else None

    i_min, i_max = _collect_current_extremes(reference_files)

    metadata_map: dict[str, tuple[object, str]] = {
        "Técnica": (first_parsed.meta_values.get("TITLE", ""), ""),
        "Fecha": (first_parsed.meta_values.get("DATE", ""), ""),
        "Hora": (first_parsed.meta_values.get("TIME", ""), ""),
        "Duración del paso": (step_duration if step_duration is not None else "", "s"),
        "Rango I": (_fmt_range_value(i_min, i_max), "A"),
        "Paso I": (delta_i if delta_i is not None else "", "A"),
        "Tiempo de muestreo": (sample_time if sample_time is not None else "", "s"),
        "Área": (area if area is not None else "", "cm^2"),
    }

    return [(field, *metadata_map[field]) for field in META_ROWS_ORDER]


# ---------------------------------------------------------------------------
# Concatenate Asc / Dsc data
# ---------------------------------------------------------------------------
def _extract_local_rows(parsed: ParsedDTA) -> list[dict[str, float]]:
    idx_map = {source: _column_index(parsed, source) for source, _, _ in DATA_EXPORT}

    missing = [source for source, idx in idx_map.items() if idx is None]
    if missing:
        raise ValueError(
            "Faltan columnas requeridas en la tabla CURVE: " + ", ".join(missing)
        )

    out: list[dict[str, float]] = []
    for raw_row in parsed.rows:
        record: dict[str, float] = {}
        for source, export_name, _unit in DATA_EXPORT:
            record[export_name] = _required_float(raw_row, idx_map[source])
        out.append(record)
    return out


def concatenate_curve_data(files: list[PolarizationFile]) -> list[dict[str, float]]:
    all_rows: list[dict[str, float]] = []
    time_offset = 0.0
    global_pt = 0

    for item in files:
        parsed = parse_gamry_dta(item.path)
        local_rows = _extract_local_rows(parsed)
        if not local_rows:
            continue

        first_local_time = local_rows[0]["time"]

        for local in local_rows:
            row = {
                "Pt": float(global_pt),
                "time": local["time"] - first_local_time + time_offset,
                "Voltaje": local["Voltaje"],
                "Corriente": local["Corriente"],
                "Sig": local["Sig"],
                "Ach": local["Ach"],
                "Temperatura": local["Temperatura"],
            }
            all_rows.append(row)
            global_pt += 1

        sample_time = to_float(parsed.meta_values.get("SAMPLETIME", ""))
        if sample_time is None:
            # Fallback: infer it from the first two time values if possible.
            if len(local_rows) >= 2:
                sample_time = local_rows[1]["time"] - local_rows[0]["time"]
            else:
                sample_time = 0.0

        time_offset = all_rows[-1]["time"] + sample_time

    return all_rows

def _optional_float(value: str | None) -> float | None:
    if value is None:
        return None
    text = value.strip()
    if not text:
        return None
    return float(text.replace(",", "."))


def _time_unit_scale(time_unit: str) -> float:
    if time_unit == "min":
        return SECONDS_PER_MINUTE
    if time_unit == "h":
        return SECONDS_PER_HOUR
    return 1.0


def _time_unit_decimals(time_unit: str) -> int:
    if time_unit == "h":
        return 2
    if time_unit == "min":
        return 1
    return 1


def build_curve_bundle_data(bundle: CurveBundle) -> dict[str, object]:
    asc_rows = concatenate_curve_data(bundle.asc_files)
    dsc_rows = concatenate_curve_data(bundle.dsc_files)

    asc_tol = infer_current_tolerance(bundle.asc_files) if bundle.asc_files else 1e-5
    dsc_tol = infer_current_tolerance(bundle.dsc_files) if bundle.dsc_files else 1e-5

    return {
        "asc_rows": asc_rows,
        "dsc_rows": dsc_rows,
        "asc_tol": asc_tol,
        "dsc_tol": dsc_tol,
    }

def build_series_by_time_plot_data(bundle: CurveBundle, time_unit: str = "s") -> dict[str, object]:
    curve_data = build_curve_bundle_data(bundle)
    time_scale = _time_unit_scale(time_unit)

    raw_asc_rows = curve_data["asc_rows"]
    raw_dsc_rows = curve_data["dsc_rows"]

    t_asc_end = max((r["time"] for r in raw_asc_rows), default=0.0)
    dsc_time_offset = t_asc_end if raw_asc_rows else 0.0

    asc_rows = []
    for row in raw_asc_rows:
        new_row = dict(row)
        new_row["plot_time"] = row["time"] / time_scale
        asc_rows.append(new_row)

    dsc_rows = []
    for row in raw_dsc_rows:
        new_row = dict(row)
        new_row["plot_time"] = (dsc_time_offset + row["time"]) / time_scale
        dsc_rows.append(new_row)

    return {
        "asc_rows": asc_rows,
        "dsc_rows": dsc_rows,
        "t_asc_end": t_asc_end,
        "t_total": max((r["plot_time"] for r in asc_rows + dsc_rows), default=0.0),
    }


def compute_default_series_by_time_limits(
    bundle: CurveBundle,
    time_unit: str = "s",
    use_current_density: bool = False,
) -> dict[str, str]:
    plot_data = build_series_by_time_plot_data(bundle, time_unit=time_unit)
    time_decimals = _time_unit_decimals(time_unit)
    area_cm2 = _bundle_area_cm2(bundle) if use_current_density else None

    rows = plot_data["asc_rows"] + plot_data["dsc_rows"]
    if not rows:
        return {
            "t_min": "",
            "t_max": "",
            "v_min": "",
            "v_max": "",
            "i_min": "",
            "i_max": "",
            "temp_min": "",
            "temp_max": "",
        }

    t_values = [r["plot_time"] for r in rows]
    t_max = _round_up_dec(max(t_values), time_decimals) if t_values else None
    v_min, v_max = _padded_limits([r["Voltaje"] for r in rows])
    i_min, i_max = _padded_limits([
        _scaled_current(r["Corriente"], use_current_density, area_cm2)
        for r in rows
    ])
    temp_min, temp_max = _padded_limits([r["Temperatura"] for r in rows])

    return {
        "t_min": _format_limit_value(0.0, time_decimals),
        "t_max": _format_limit_value(t_max, time_decimals),
        "v_min": _format_limit_value(v_min),
        "v_max": _format_limit_value(v_max),
        "i_min": _format_limit_value(i_min),
        "i_max": _format_limit_value(i_max),
        "temp_min": _format_limit_value(temp_min),
        "temp_max": _format_limit_value(temp_max),
    }


def compute_autofit_series_by_time_limits(
    bundle: CurveBundle,
    show_asc: bool,
    show_dsc: bool,
    show_voltage: bool,
    show_current: bool,
    show_temperature: bool,
    time_unit: str = "s",
    use_current_density: bool = False,
) -> dict[str, str]:
    if not (show_asc or show_dsc):
        raise ValueError("Debe seleccionar Ascendente y/o Descendente para usar Autoescala.")

    if not (show_voltage or show_current or show_temperature):
        raise ValueError("Debe seleccionar al menos una magnitud para usar Autoescala.")

    plot_data = build_series_by_time_plot_data(bundle, time_unit=time_unit)
    area_cm2 = _bundle_area_cm2(bundle) if use_current_density else None

    rows = []
    if show_asc:
        rows.extend(plot_data["asc_rows"])
    if show_dsc:
        rows.extend(plot_data["dsc_rows"])

    if not rows:
        raise ValueError("No hay datos válidos para ajustar los ejes.")

    out = {
        "t_min": "",
        "t_max": "",
        "v_min": "",
        "v_max": "",
        "i_min": "",
        "i_max": "",
        "temp_min": "",
        "temp_max": "",
    }

    t_values = [r["plot_time"] for r in rows]
    if t_values:
        decimals = _time_unit_decimals(time_unit)
        out["t_min"] = _format_limit_value(0.0, decimals)
        out["t_max"] = _format_limit_value(_round_up_dec(max(t_values), decimals), decimals)

    if show_voltage:
        v_values = [r["Voltaje"] for r in rows]
        if v_values:
            voltage_step = _autoscale_step_from_values(v_values)
            out["v_min"] = _format_step_aligned_value(_round_down_to_step(min(v_values), voltage_step), voltage_step)
            out["v_max"] = _format_step_aligned_value(_round_up_to_step(max(v_values), voltage_step), voltage_step)

    if show_current:
        i_values = [_scaled_current(r["Corriente"], use_current_density, area_cm2) for r in rows]
        if i_values:
            current_step = _autoscale_step_from_values(i_values)
            out["i_min"] = _format_step_aligned_value(_round_down_to_step(min(i_values), current_step), current_step)
            out["i_max"] = _format_step_aligned_value(_round_up_to_step(max(i_values), current_step), current_step)

    if show_temperature:
        temp_values = [r["Temperatura"] for r in rows]
        if temp_values:
            out["temp_min"] = str(int(floor(min(temp_values))))
            out["temp_max"] = str(int(ceil(max(temp_values))))

    return out


def apply_secondary_axis_scaling(ax, values: list[float], tick_count: int) -> None:
    if not values:
        return

    y_min, y_max = _padded_limits(values)
    if y_min is None or y_max is None:
        return

    if y_min == y_max:
        y_max = y_min + 1.0

    tick_count = max(2, int(tick_count))

    ax.set_ylim(y_min, y_max)
    ax.yaxis.set_major_locator(LinearLocator(tick_count))
    ax.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))

def split_rows_into_steps(
    rows: list[dict[str, float]],
    current_tolerance: float,
) -> list[list[dict[str, float]]]:
    """Split concatenated rows into current plateaus (steps)."""
    if not rows:
        return []

    steps: list[list[dict[str, float]]] = []
    start = 0
    plateau_current = rows[0]["Corriente"]

    for idx in range(1, len(rows)):
        current = rows[idx]["Corriente"]

        if abs(current - plateau_current) > current_tolerance:
            steps.append(rows[start:idx])
            start = idx
            plateau_current = rows[start]["Corriente"]
        else:
            span = idx - start + 1
            plateau_current = ((plateau_current * (span - 1)) + current) / span

    steps.append(rows[start:])
    return steps


def pick_fractional_point_from_step(
    step_rows: list[dict[str, float]],
    fraction: float,
) -> dict[str, float]:
    """Pick one point inside a step:
    0.0 = first point, 1.0 = last point, values in between = proportional index.
    """
    if not step_rows:
        raise ValueError("El step está vacío.")

    f = max(0.0, min(1.0, float(fraction)))

    if len(step_rows) == 1:
        return dict(step_rows[0])

    idx = int(round(f * (len(step_rows) - 1)))
    idx = max(0, min(len(step_rows) - 1, idx))
    return dict(step_rows[idx])


def select_fractional_point_per_step(
    rows: list[dict[str, float]],
    current_tolerance: float,
    fraction: float,
) -> list[dict[str, float]]:
    """Return one representative point per step according to a fractional position."""
    steps = split_rows_into_steps(rows, current_tolerance)
    selected_rows: list[dict[str, float]] = []

    for step_number, step_rows in enumerate(steps, start=1):
        selected = pick_fractional_point_from_step(step_rows, fraction)
        selected["Step"] = float(step_number)
        selected_rows.append(selected)

    return selected_rows


def build_step_stability_rows(
    rows: list[dict[str, float]],
    current_tolerance: float,
    use_current_density: bool = False,
    area_cm2: float | None = None,
) -> list[dict[str, float]]:
    stability_rows: list[dict[str, float]] = []
    steps = split_rows_into_steps(rows, current_tolerance)

    for step_number, step_rows in enumerate(steps, start=1):
        if not step_rows:
            continue

        currents = [row["Corriente"] for row in step_rows]
        voltages = [row["Voltaje"] for row in step_rows]
        temperatures = [row["Temperatura"] for row in step_rows]

        avg_current = sum(currents) / len(currents)
        stability_rows.append(
            {
                "Step": float(step_number),
                "Corriente": _scaled_current(avg_current, use_current_density, area_cm2),
                "DeltaV": max(voltages) - min(voltages),
                "DeltaT": max(temperatures) - min(temperatures),
            }
        )

    return stability_rows


def _positive_min(values: list[float]) -> float | None:
    positives = [value for value in values if value > 0]
    if not positives:
        return None
    return min(positives)


def _selected_v_vs_i_rows(
    bundle: CurveBundle,
    show_asc: bool,
    show_dsc: bool,
    point_fraction: float,
) -> dict[str, list[dict[str, float]]]:
    curve_data = build_curve_bundle_data(bundle)

    asc_rows = (
        select_fractional_point_per_step(curve_data["asc_rows"], curve_data["asc_tol"], point_fraction)
        if show_asc and curve_data["asc_rows"]
        else []
    )
    dsc_rows = (
        select_fractional_point_per_step(curve_data["dsc_rows"], curve_data["dsc_tol"], point_fraction)
        if show_dsc and curve_data["dsc_rows"]
        else []
    )

    return {
        "asc_rows": asc_rows,
        "dsc_rows": dsc_rows,
    }


def _v_vs_i_voltage_points(
    rows: list[dict[str, float]],
    use_current_density: bool,
    area_cm2: float | None,
) -> list[tuple[float, float]]:
    points = [
        (
            _scaled_current(row["Corriente"], use_current_density, area_cm2),
            row["Voltaje"],
        )
        for row in rows
    ]
    points.sort(key=lambda item: item[0])
    return points


def _slope_at_current_from_points(
    points: list[tuple[float, float]],
    target_current: float,
    fit_point_count: int = PC_HIGH_CURRENT_FIT_POINTS,
) -> float | None:
    state = _voltage_state_at_current_from_points(points, target_current, fit_point_count=fit_point_count)
    return state["slope"] if state is not None else None


def _voltage_state_at_current_from_points(
    points: list[tuple[float, float]],
    target_current: float,
    fit_point_count: int = PC_HIGH_CURRENT_FIT_POINTS,
) -> dict[str, float] | None:
    if len(points) < 2:
        return None

    sorted_points = sorted(points, key=lambda item: item[0])
    fit_count = max(2, int(fit_point_count))
    if fit_count % 2 == 0:
        fit_count += 1
    fit_count = min(fit_count, len(sorted_points))
    nearest_index = min(
        range(len(sorted_points)),
        key=lambda index: abs(sorted_points[index][0] - target_current),
    )
    half_window = fit_count // 2
    start = nearest_index - half_window
    end = start + fit_count
    if start < 0:
        start = 0
        end = fit_count
    if end > len(sorted_points):
        end = len(sorted_points)
        start = max(0, end - fit_count)

    fit_points = sorted_points[start:end]
    fit = _linear_fit(fit_points)
    if fit is None:
        return None

    voltage = fit["slope"] * target_current + fit["intercept"]
    return {
        "current": target_current,
        "voltage": voltage,
        "slope": fit["slope"],
        "fit_center_current": sorted_points[nearest_index][0],
        "fit_points": float(len(fit_points)),
    }


def _linear_fit(points: list[tuple[float, float]]) -> dict[str, float] | None:
    if len(points) < 2:
        return None

    n = float(len(points))
    sum_x = sum(point[0] for point in points)
    sum_y = sum(point[1] for point in points)
    sum_xx = sum(point[0] * point[0] for point in points)
    sum_xy = sum(point[0] * point[1] for point in points)
    denominator = (n * sum_xx) - (sum_x * sum_x)
    if abs(denominator) <= 1e-18:
        return None

    slope = ((n * sum_xy) - (sum_x * sum_y)) / denominator
    intercept = (sum_y - slope * sum_x) / n
    mean_y = sum_y / n
    ss_tot = sum((point[1] - mean_y) ** 2 for point in points)
    ss_res = sum((point[1] - (slope * point[0] + intercept)) ** 2 for point in points)
    r2 = 1.0 if ss_tot <= 1e-18 else 1.0 - (ss_res / ss_tot)
    return {"slope": slope, "intercept": intercept, "r2": r2}


def _pc_report_direction_points(
    rows: list[dict[str, float]],
    current_tolerance: float,
    point_fraction: float,
    use_current_density: bool,
    area_cm2: float | None,
) -> list[tuple[float, float]]:
    selected_rows = select_fractional_point_per_step(rows, current_tolerance, point_fraction)
    return _v_vs_i_voltage_points(selected_rows, use_current_density, area_cm2)


def _pc_report_linear_points(
    points: list[tuple[float, float]],
    threshold_current: float,
) -> list[tuple[float, float]]:
    linear_points = [point for point in points if point[0] >= threshold_current]
    if len(linear_points) >= 2:
        return linear_points
    return points[-2:] if len(points) >= 2 else []


def build_pc_report_metadata(
    bundle: CurveBundle,
    use_current_density: bool = False,
    language: str = "es",
) -> list[tuple[str, str, str]]:
    language = normalize_language(language)
    reference_files = bundle.asc_files + bundle.dsc_files
    if not reference_files:
        raise ValueError("No se encontraron archivos Asc ni Dsc para generar el reporte.")

    first_parsed = parse_gamry_dta(reference_files[0].path)
    step_duration = to_float(first_parsed.meta_values.get("TSTEP1", ""))
    sample_time = to_float(first_parsed.meta_values.get("SAMPLETIME", ""))
    area = to_float(first_parsed.meta_values.get("AREA", ""))
    i1 = to_float(first_parsed.meta_values.get("ISTEP1", ""))
    i2 = to_float(first_parsed.meta_values.get("ISTEP2", ""))
    current_step = abs(i2 - i1) if i1 is not None and i2 is not None else None
    sweep_rate = (
        current_step / step_duration
        if current_step is not None and step_duration is not None and step_duration > 0
        else None
    )
    i_min, i_max = _collect_current_extremes(reference_files)
    current_range_text = _fmt_range_value(i_min, i_max)
    current_unit = "A"
    current_step_unit = "A"
    sweep_rate_unit = "A/s"

    if use_current_density and area is not None and area > 0:
        density_min = i_min / area if i_min is not None else None
        density_max = i_max / area if i_max is not None else None
        current_range_text = _fmt_range_value(density_min, density_max)
        current_step = current_step / area if current_step is not None else None
        sweep_rate = sweep_rate / area if sweep_rate is not None else None
        current_unit = "A/cm^2"
        current_step_unit = "A/cm^2"
        sweep_rate_unit = "A/cm^2/s"

    return [
        (translate("curve", language), f"{bundle.description} #{bundle.curve_id}", ""),
        (translate("technique", language), first_parsed.meta_values.get("TITLE", ""), ""),
        (translate("date", language), first_parsed.meta_values.get("DATE", ""), ""),
        (translate("start_time", language), first_parsed.meta_values.get("TIME", ""), ""),
        (translate("current_range", language), current_range_text, current_unit),
        (translate("step_duration", language), _format_report_value(step_duration), "s"),
        (translate("current_step", language), _format_report_value(current_step), current_step_unit),
        (translate("sweeping_rate", language), _format_report_value(sweep_rate), sweep_rate_unit),
        (translate("sampling_time", language), _format_report_value(sample_time), "s"),
        (translate("area", language), _format_report_value(area), "cm^2"),
        (translate("asc_files", language), str(len(bundle.asc_files)), ""),
        (translate("dsc_files", language), str(len(bundle.dsc_files)), ""),
    ]


def build_pc_report_indicators(
    bundle: CurveBundle,
    point_fraction: float = 1.0,
    use_current_density: bool = False,
    high_current_fraction: float = 0.90,
    linear_start_fraction: float = 0.20,
    language: str = "es",
    time_unit: str = "s",
) -> list[tuple[str, str, str]]:
    language = normalize_language(language)
    curve_data = build_curve_bundle_data(bundle)
    rows: list[tuple[str, str, str]] = []

    raw_rows = curve_data["asc_rows"] + curve_data["dsc_rows"]
    series_data = build_series_by_time_plot_data(bundle, time_unit=time_unit)
    time_values = [
        row["plot_time"]
        for row in series_data["asc_rows"] + series_data["dsc_rows"]
    ]
    transition_time = (
        series_data["t_asc_end"] / _time_unit_scale(time_unit)
        if series_data["asc_rows"] and series_data["dsc_rows"]
        else None
    )
    voltage_values = [row["Voltaje"] for row in raw_rows]
    temp_values = [row["Temperatura"] for row in raw_rows]

    if time_values:
        total_time = max(time_values) - min(time_values)
    else:
        total_time = None

    if voltage_values:
        voltage_range = max(voltage_values) - min(voltage_values)
    else:
        voltage_range = None

    if temp_values:
        avg_temp = sum(temp_values) / len(temp_values)
        max_delta_t = max(temp_values) - min(temp_values)
    else:
        avg_temp = None
        max_delta_t = None

    rows.append((translate("ascending_finish_time", language), _format_report_value(transition_time), time_unit))
    rows.append((translate("total_time", language), _format_report_value(total_time), time_unit))
    rows.append((translate("voltage_range", language), _format_report_value(voltage_range), "V"))
    rows.append((translate("average_temperature", language), _format_report_value(avg_temp), "C"))
    rows.append((translate("maximum_delta_t", language), _format_report_value(max_delta_t), "C"))
    return rows


def build_pc_dv_di_report_indicators(
    bundle: CurveBundle,
    point_fraction: float = 1.0,
    smoothing_algorithm: str = "Median filter",
    smoothing_window: int = 1,
    use_current_density: bool = True,
    high_current_fraction: float = 0.90,
    high_current_target_a: float | None = PC_HIGH_CURRENT_TARGET_A,
    high_current_fit_points: int = PC_HIGH_CURRENT_FIT_POINTS,
    language: str | None = None,
) -> list[tuple[str, str, str]]:
    language = _pc_language(language)
    curve_data = build_curve_bundle_data(bundle)
    area_cm2 = _bundle_area_cm2(bundle) if use_current_density else None

    def _direction_dvdi_points(rows: list[dict[str, float]], tolerance: float) -> list[tuple[float, float]]:
        derivative_rows = build_dv_di_rows(
            rows,
            tolerance,
            point_fraction,
            smoothing_algorithm=smoothing_algorithm,
            smoothing_window=smoothing_window,
        )
        return [
            (
                _scaled_current(row["Corriente"], use_current_density, area_cm2),
                _scaled_dvdi(row["dVdI"], use_current_density, area_cm2),
            )
            for row in derivative_rows
        ]

    def _direction_voltage_points(rows: list[dict[str, float]], tolerance: float) -> list[tuple[float, float]]:
        return _pc_report_direction_points(rows, tolerance, point_fraction, use_current_density, area_cm2)

    asc_dvdi_points = (
        _direction_dvdi_points(curve_data["asc_rows"], curve_data["asc_tol"])
        if curve_data["asc_rows"] else []
    )
    dsc_dvdi_points = (
        _direction_dvdi_points(curve_data["dsc_rows"], curve_data["dsc_tol"])
        if curve_data["dsc_rows"] else []
    )
    asc_voltage_points = (
        _direction_voltage_points(curve_data["asc_rows"], curve_data["asc_tol"])
        if curve_data["asc_rows"] else []
    )
    dsc_voltage_points = (
        _direction_voltage_points(curve_data["dsc_rows"], curve_data["dsc_tol"])
        if curve_data["dsc_rows"] else []
    )
    all_voltage_points = asc_voltage_points + dsc_voltage_points

    current_unit = "A/cm^2" if use_current_density else "A"
    dvdi_unit = _dvdi_axis_unit(use_current_density)
    rows: list[tuple[str, str, str]] = []

    if high_current_target_a is not None:
        high_current = _scaled_current(high_current_target_a, use_current_density, area_cm2)
    elif all_voltage_points:
        currents = [point[0] for point in all_voltage_points]
        current_min = min(currents)
        current_max = max(currents)
        high_current = current_min + high_current_fraction * (current_max - current_min)
    else:
        high_current = None

    if high_current is not None:
        rows.append((translate("high_current_target", language), _format_report_value(high_current), current_unit))

    for direction_key, voltage_points, dvdi_points in (
        ("ascending", asc_voltage_points, asc_dvdi_points),
        ("descending", dsc_voltage_points, dsc_dvdi_points),
    ):
        direction_label = translate(direction_key, language)
        high_value = (
            _slope_at_current_from_points(
                voltage_points,
                high_current,
                fit_point_count=high_current_fit_points,
            )
            if high_current is not None else None
        )
        rows.append((
            translate("high_current_dvdi", language, direction=direction_label),
            _format_report_value(high_value),
            dvdi_unit,
        ))

        if dvdi_points:
            max_current, max_dvdi = max(dvdi_points, key=lambda item: item[1])
        else:
            max_current = None
            max_dvdi = None
        rows.append((
            translate("maximum_dvdi", language, direction=direction_label),
            _format_report_value(max_dvdi),
            dvdi_unit,
        ))
        rows.append((
            translate("maximum_dvdi_current", language, direction=direction_label),
            _format_report_value(max_current),
            current_unit,
        ))

    return rows


def build_pc_step_stability_report_indicators(
    bundle: CurveBundle,
    use_current_density: bool = True,
    language: str | None = None,
) -> list[tuple[str, str, str]]:
    language = _pc_language(language)
    direction_rows = _step_stability_direction_rows(bundle, use_current_density=use_current_density)
    current_unit = "A/cm^2" if use_current_density else "A"
    rows: list[tuple[str, str, str]] = []

    for direction_key, data_key in (
        ("ascending", "asc_rows"),
        ("descending", "dsc_rows"),
    ):
        direction_label = translate(direction_key, language)
        step_rows = direction_rows.get(data_key, [])

        if step_rows:
            max_delta_v_row = max(step_rows, key=lambda row: row["DeltaV"])
            max_delta_t_row = max(step_rows, key=lambda row: row["DeltaT"])
        else:
            max_delta_v_row = None
            max_delta_t_row = None

        rows.append((
            translate("maximum_step_delta_voltage", language, direction=direction_label),
            _format_report_value(max_delta_v_row["DeltaV"] if max_delta_v_row else None),
            "V",
        ))
        rows.append((
            translate("maximum_step_delta_voltage_current", language, direction=direction_label),
            _format_report_value(max_delta_v_row["Corriente"] if max_delta_v_row else None),
            current_unit,
        ))
        rows.append((
            translate("maximum_step_delta_temperature", language, direction=direction_label),
            _format_report_value(max_delta_t_row["DeltaT"] if max_delta_t_row else None),
            "C",
        ))
        rows.append((
            translate("maximum_step_delta_temperature_current", language, direction=direction_label),
            _format_report_value(max_delta_t_row["Corriente"] if max_delta_t_row else None),
            current_unit,
        ))

    return rows


def draw_temperature_vs_current_on_figure(
    fig: Figure,
    bundle: CurveBundle,
    *,
    point_fraction: float = 1.0,
    use_current_density: bool = True,
    x_tick_count: int = 6,
    y_tick_count: int = 6,
    title_fontsize: float = 14,
    tick_fontsize: float = 10,
    label_fontsize: float = 11,
    legend_fontsize: float = 10,
    marker_size: float = 6,
    line_width: float = 1.5,
    hollow_markers: bool = False,
    language: str = "es",
) -> bool:
    fig.clear()
    language = normalize_language(language)
    curve_data = build_curve_bundle_data(bundle)
    area_cm2 = _bundle_area_cm2(bundle) if use_current_density else None
    x_values: list[float] = []
    temp_values: list[float] = []

    asc_rows = (
        select_fractional_point_per_step(curve_data["asc_rows"], curve_data["asc_tol"], point_fraction)
        if curve_data["asc_rows"] else []
    )
    dsc_rows = (
        select_fractional_point_per_step(curve_data["dsc_rows"], curve_data["dsc_tol"], point_fraction)
        if curve_data["dsc_rows"] else []
    )
    if not asc_rows and not dsc_rows:
        return False

    ax = fig.add_subplot(111)

    def _plot_rows(rows: list[dict[str, float]], label: str, color: str, marker: str) -> None:
        if not rows:
            return
        row_x_values = [
            _scaled_current(row["Corriente"], use_current_density, area_cm2)
            for row in rows
        ]
        row_temp_values = [row["Temperatura"] for row in rows]
        x_values.extend(row_x_values)
        temp_values.extend(row_temp_values)
        ax.plot(
            row_x_values,
            row_temp_values,
            color=color,
            marker=marker,
            linestyle="-",
            linewidth=line_width,
            markersize=marker_size,
            markerfacecolor="none" if hollow_markers else color,
            markeredgecolor=color,
            markeredgewidth=1.2,
            label=label,
        )

    language = _pc_language()
    _plot_rows(asc_rows, f"{translate('ascending', language)} T", PC_PLOT_COLORS["asc_temperature"], "^")
    _plot_rows(dsc_rows, f"{translate('descending', language)} T", PC_PLOT_COLORS["dsc_temperature"], "v")

    handles, labels = ax.get_legend_handles_labels()
    if not handles:
        fig.clear()
        return False

    ax.set_xlabel(
        f"{translate('current_density', language)} (A/cm^2)"
        if use_current_density
        else f"{translate('current', language)} (A)",
        fontsize=label_fontsize,
    )
    ax.set_ylabel(f"{translate('temperature', language)} (C)", fontsize=label_fontsize)
    ax.set_title(
        f"{translate('temperature', language)} vs {translate('current', language).lower()} - "
        f"{bundle.description} #{bundle.curve_id}",
        fontsize=title_fontsize,
    )
    ax.grid(True)
    ax.tick_params(axis="both", labelsize=tick_fontsize)
    apply_x_tick_label_padding(ax, tick_fontsize)
    x_tick_count = max(2, int(x_tick_count))
    y_tick_count = max(2, int(y_tick_count))

    if x_values:
        x_hi = _round_up_dec(max(x_values), 1)
        if x_hi <= 0.0:
            x_hi = 1.0
        ax.set_xlim(0.0, x_hi)
        ax.xaxis.set_major_locator(LinearLocator(x_tick_count))
    else:
        ax.xaxis.set_major_locator(MaxNLocator(nbins=x_tick_count))

    if temp_values:
        y_lo = floor(min(temp_values))
        y_hi = ceil(max(temp_values))
        if y_hi <= y_lo:
            y_hi = y_lo + 1
        ax.set_ylim(y_lo, y_hi)
        ax.yaxis.set_major_locator(LinearLocator(y_tick_count))
    else:
        ax.yaxis.set_major_locator(MaxNLocator(nbins=y_tick_count))

    ax.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
    ax.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
    make_legend_draggable(ax.legend(handles, labels, fontsize=legend_fontsize))
    fig.tight_layout()
    return True


def _add_report_table(
    ax,
    title: str,
    rows: list[tuple[str, str, str]],
    bbox: list[float],
    language: str = "es",
) -> None:
    table_width = min(float(bbox[2]), 0.82)
    bbox = [(1.0 - table_width) / 2.0, bbox[1], table_width, bbox[3]]
    ax.text(
        bbox[0],
        bbox[1] + bbox[3] + 0.025,
        title,
        fontsize=13,
        fontweight="bold",
        ha="left",
        va="bottom",
        transform=ax.transAxes,
    )
    table = ax.table(
        cellText=[[_format_report_value(name), _format_report_value(value), _format_report_value(unit)] for name, value, unit in rows],
        colLabels=[
            translate("name", language),
            translate("value", language),
            translate("unit", language),
        ],
        cellLoc="left",
        colLoc="left",
        bbox=bbox,
    )
    table.auto_set_font_size(False)
    table.set_fontsize(8.6)
    for (row_idx, _col_idx), cell in table.get_celld().items():
        cell.set_edgecolor("#b8c0c8")
        cell.set_linewidth(0.5)
        if row_idx == 0:
            cell.set_facecolor("#edf2f7")
            cell.set_text_props(weight="bold")
        else:
            cell.set_facecolor("white")


def export_pc_curve_report_pdf(
    bundle: CurveBundle,
    output_path: Path,
    *,
    plot_kwargs: dict[str, object],
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    plot_kwargs = dict(plot_kwargs)
    point_fraction = float(plot_kwargs.get("point_fraction", 1.0))
    use_current_density = bool(plot_kwargs.get("use_current_density", False))
    language = _pc_language(str(plot_kwargs.pop("language", "")))

    metadata_rows = build_pc_report_metadata(
        bundle,
        use_current_density=use_current_density,
        language=language,
    )
    indicator_rows = build_pc_report_indicators(
        bundle,
        point_fraction=point_fraction,
        use_current_density=use_current_density,
        language=language,
        time_unit=str(plot_kwargs.get("time_unit", "s")),
    )

    with PdfPages(output_path) as pdf:
        plot_fig = Figure(figsize=(9.5, 6.2), dpi=150)
        has_plot = draw_v_vs_i_on_figure(plot_fig, bundle=bundle, **plot_kwargs)
        if not has_plot:
            raise ValueError("No hay datos suficientes para generar el grafico del reporte.")
        pdf.savefig(plot_fig, bbox_inches="tight")

        temp_fig = Figure(figsize=(9.5, 6.2), dpi=150)
        has_temp_plot = draw_temperature_vs_current_on_figure(
            temp_fig,
            bundle,
            point_fraction=point_fraction,
            use_current_density=use_current_density,
            x_tick_count=int(plot_kwargs.get("x_tick_count", 6)),
            y_tick_count=int(plot_kwargs.get("y_tick_count", 6)),
            title_fontsize=float(plot_kwargs.get("title_fontsize", 14)),
            tick_fontsize=float(plot_kwargs.get("tick_fontsize", 10)),
            label_fontsize=float(plot_kwargs.get("label_fontsize", 11)),
            legend_fontsize=float(plot_kwargs.get("legend_fontsize", 10)),
            marker_size=float(plot_kwargs.get("marker_size", 6)),
            line_width=float(plot_kwargs.get("line_width", 1.5)),
            hollow_markers=bool(plot_kwargs.get("hollow_markers", False)),
        )
        if has_temp_plot:
            pdf.savefig(temp_fig, bbox_inches="tight")

        table_fig = Figure(figsize=(8.5, 11.0), dpi=150)
        ax = table_fig.add_subplot(111)
        ax.axis("off")
        table_fig.text(
            0.05,
            0.965,
            translate(
                "pc_report_title",
                language,
                curve=f"{bundle.description} #{bundle.curve_id}",
            ),
            fontsize=18,
            fontweight="bold",
            ha="left",
            va="top",
        )
        table_fig.text(
            0.05,
            0.935,
            translate("pc_report_subtitle", language),
            fontsize=10,
            ha="left",
            va="top",
            color="#4a5568",
        )
        _add_report_table(ax, translate("metadata", language), metadata_rows, [0.05, 0.53, 0.90, 0.34], language=language)
        _add_report_table(ax, translate("indicators", language), indicator_rows, [0.05, 0.08, 0.90, 0.36], language=language)
        pdf.savefig(table_fig, bbox_inches="tight")

    return output_path


def export_pc_series_by_time_report_pdf(
    bundle: CurveBundle,
    output_path: Path,
    *,
    plot_kwargs: dict[str, object],
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    plot_kwargs = dict(plot_kwargs)
    plot_kwargs.update(
        {
            "show_voltage": True,
            "show_current": True,
            "show_temperature": True,
            "use_current_density": bool(plot_kwargs.get("use_current_density", True)),
        }
    )
    point_fraction = float(plot_kwargs.pop("point_fraction", 1.0))
    use_current_density = bool(plot_kwargs.get("use_current_density", True))
    language = normalize_language(str(plot_kwargs.get("language", "es")))

    metadata_rows = build_pc_report_metadata(
        bundle,
        use_current_density=use_current_density,
        language=language,
    )
    indicator_rows = build_pc_report_indicators(
        bundle,
        point_fraction=point_fraction,
        use_current_density=use_current_density,
        language=language,
        time_unit=str(plot_kwargs.get("time_unit", "s")),
    )

    with PdfPages(output_path) as pdf:
        plot_fig = Figure(figsize=(10.0, 6.4), dpi=150)
        has_plot = draw_series_by_time_on_figure(plot_fig, bundle=bundle, **plot_kwargs)
        if not has_plot:
            raise ValueError("No hay datos suficientes para generar el grafico del reporte.")
        pdf.savefig(plot_fig, bbox_inches="tight")

        table_fig = Figure(figsize=(8.5, 11.0), dpi=150)
        ax = table_fig.add_subplot(111)
        ax.axis("off")
        table_fig.text(
            0.05,
            0.965,
            translate(
                "pc_series_report_title",
                language,
                curve=f"{bundle.description} #{bundle.curve_id}",
            ),
            fontsize=18,
            fontweight="bold",
            ha="left",
            va="top",
        )
        table_fig.text(
            0.05,
            0.935,
            translate("pc_series_report_subtitle", language),
            fontsize=10,
            ha="left",
            va="top",
            color="#4a5568",
        )
        _add_report_table(ax, translate("metadata", language), metadata_rows, [0.05, 0.53, 0.90, 0.34], language=language)
        _add_report_table(ax, translate("indicators", language), indicator_rows, [0.05, 0.08, 0.90, 0.36], language=language)
        pdf.savefig(table_fig, bbox_inches="tight")

    return output_path


def export_pc_dv_di_report_pdf(
    bundle: CurveBundle,
    output_path: Path,
    *,
    plot_kwargs: dict[str, object],
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    plot_kwargs = dict(plot_kwargs)
    language = _pc_language(str(plot_kwargs.get("language", "")))
    point_fraction = float(plot_kwargs.get("point_fraction", 1.0))
    use_current_density = bool(plot_kwargs.get("use_current_density", True))
    smoothing_algorithm = str(plot_kwargs.get("smoothing_algorithm", "Median filter"))
    smoothing_window = int(plot_kwargs.get("smoothing_window", 1))

    metadata_rows = build_pc_report_metadata(
        bundle,
        use_current_density=use_current_density,
        language=language,
    )
    indicator_rows = build_pc_dv_di_report_indicators(
        bundle,
        point_fraction=point_fraction,
        smoothing_algorithm=smoothing_algorithm,
        smoothing_window=smoothing_window,
        use_current_density=use_current_density,
        language=language,
    )

    with PdfPages(output_path) as pdf:
        plot_fig = Figure(figsize=(9.5, 6.2), dpi=150)
        has_plot = draw_dv_di_on_figure(plot_fig, bundle=bundle, **plot_kwargs)
        if not has_plot:
            raise ValueError("No hay datos suficientes para generar el grafico del reporte.")
        pdf.savefig(plot_fig, bbox_inches="tight")

        table_fig = Figure(figsize=(8.5, 11.0), dpi=150)
        ax = table_fig.add_subplot(111)
        ax.axis("off")
        table_fig.text(
            0.05,
            0.965,
            translate(
                "pc_dvdi_report_title",
                language,
                curve=f"{bundle.description} #{bundle.curve_id}",
            ),
            fontsize=18,
            fontweight="bold",
            ha="left",
            va="top",
        )
        table_fig.text(
            0.05,
            0.935,
            translate("pc_dvdi_report_subtitle", language),
            fontsize=10,
            ha="left",
            va="top",
            color="#4a5568",
        )
        _add_report_table(ax, translate("metadata", language), metadata_rows, [0.05, 0.53, 0.90, 0.34], language=language)
        _add_report_table(ax, translate("indicators", language), indicator_rows, [0.05, 0.08, 0.90, 0.36], language=language)
        pdf.savefig(table_fig, bbox_inches="tight")

    return output_path


def export_pc_step_stability_report_pdf(
    bundle: CurveBundle,
    output_path: Path,
    *,
    plot_kwargs: dict[str, object],
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    plot_kwargs = dict(plot_kwargs)
    language = _pc_language(str(plot_kwargs.get("language", "")))
    use_current_density = bool(plot_kwargs.get("use_current_density", True))

    metadata_rows = build_pc_report_metadata(
        bundle,
        use_current_density=use_current_density,
        language=language,
    )
    indicator_rows = build_pc_step_stability_report_indicators(
        bundle,
        use_current_density=use_current_density,
        language=language,
    )

    with PdfPages(output_path) as pdf:
        voltage_kwargs = dict(plot_kwargs)
        voltage_kwargs.update(
            {
                "show_voltage_delta": True,
                "show_temperature_delta": False,
                "plot_title": translate("voltage_step_range", language),
            }
        )
        voltage_fig = Figure(figsize=(9.5, 6.2), dpi=150)
        has_voltage_plot = draw_step_stability_on_figure(voltage_fig, bundle=bundle, **voltage_kwargs)
        if not has_voltage_plot:
            raise ValueError("No hay datos suficientes para generar el grafico de voltaje del reporte.")
        pdf.savefig(voltage_fig, bbox_inches="tight")

        temperature_kwargs = dict(plot_kwargs)
        temperature_kwargs.update(
            {
                "show_voltage_delta": False,
                "show_temperature_delta": True,
                "plot_title": translate("temperature_step_range", language),
                "dv_min": None,
                "dv_max": None,
            }
        )
        temperature_fig = Figure(figsize=(9.5, 6.2), dpi=150)
        has_temperature_plot = draw_step_stability_on_figure(
            temperature_fig,
            bundle=bundle,
            **temperature_kwargs,
        )
        if has_temperature_plot:
            pdf.savefig(temperature_fig, bbox_inches="tight")

        table_fig = Figure(figsize=(8.5, 11.0), dpi=150)
        ax = table_fig.add_subplot(111)
        ax.axis("off")
        table_fig.text(
            0.05,
            0.965,
            translate(
                "pc_step_stability_report_title",
                language,
                curve=f"{bundle.description} #{bundle.curve_id}",
            ),
            fontsize=18,
            fontweight="bold",
            ha="left",
            va="top",
        )
        table_fig.text(
            0.05,
            0.935,
            translate("pc_step_stability_report_subtitle", language),
            fontsize=10,
            ha="left",
            va="top",
            color="#4a5568",
        )
        _add_report_table(ax, translate("metadata", language), metadata_rows, [0.05, 0.53, 0.90, 0.34], language=language)
        _add_report_table(ax, translate("indicators", language), indicator_rows, [0.05, 0.08, 0.90, 0.36], language=language)
        pdf.savefig(table_fig, bbox_inches="tight")

    return output_path


def _open_v_vs_i_window_single(input_dir: Path, font_defaults: PlotFontDefaults | None = None) -> None:
    bundles = discover_curve_bundles(Path(input_dir))
    if not bundles:
        raise ValueError("No se encontraron curvas de polarización válidas.")

    font_defaults = resolve_plot_font_defaults(font_defaults)
    font_default_values = font_defaults.as_strings()
    bundle = bundles[0]

    default_limits = compute_default_v_vs_i_limits(bundle, use_current_density=True)

    win = tk.Toplevel()
    win.title(f"PC - V vs I - {bundle.description} #{bundle.curve_id}")
    win.geometry("1200x700")

    controls_host, plot_outer = create_resizable_plot_layout(win, sidebar_width=320)
    controls_frame = _build_scrollable_controls(controls_host)

    toolbar_frame = ttk.Frame(plot_outer)
    toolbar_frame.pack(side="top", fill="x")

    canvas_frame = ttk.Frame(plot_outer)
    canvas_frame.pack(side="top", fill="both", expand=True)

    fig = Figure(figsize=(9, 5.5), dpi=100)

    canvas = FigureCanvasTkAgg(fig, master=canvas_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    toolbar = NavigationToolbar2Tk(canvas, toolbar_frame, pack_toolbar=False)
    toolbar.update()
    toolbar.pack(side="left", fill="x")

    status_var = tk.StringVar(value="Listo.")

    asc_marker_var = tk.StringVar(value="^")
    dsc_marker_var = tk.StringVar(value="v")
    voltage_line_var = tk.StringVar(value="-")
    temperature_line_var = tk.StringVar(value="--")

    asc_var = tk.BooleanVar(value=True)
    dsc_var = tk.BooleanVar(value=True)
    voltage_var = tk.BooleanVar(value=True)
    temperature_var = tk.BooleanVar(value=False)
    current_density_var = tk.BooleanVar(value=True)
    point_fraction_var = tk.DoubleVar(value=1.0)
    indicator_current_var = tk.DoubleVar(value=0.0)
    show_slope_guides_var = tk.BooleanVar(value=False)

    x_min_var = tk.StringVar(value=default_limits["x_min"])
    x_max_var = tk.StringVar(value=default_limits["x_max"])
    v_min_var = tk.StringVar(value=default_limits["v_min"])
    v_max_var = tk.StringVar(value=default_limits["v_max"])
    temp_min_var = tk.StringVar(value="")
    temp_max_var = tk.StringVar(value="")

    x_tick_count_var = tk.IntVar(value=6)
    y_tick_count_var = tk.IntVar(value=6)

    plot_title_var = tk.StringVar(value="")
    show_title_var = tk.BooleanVar(value=True)
    title_fontsize_var = tk.StringVar(value=font_default_values["title"])
    tick_fontsize_var = tk.StringVar(value=font_default_values["tick"])
    label_fontsize_var = tk.StringVar(value=font_default_values["label"])
    legend_fontsize_var = tk.StringVar(value=font_default_values["legend"])
    marker_size_var = tk.StringVar(value="6")
    hollow_markers_var = tk.BooleanVar(value=False)
    line_width_var = tk.StringVar(value="1.5")
    indicator_current_value_var = tk.StringVar(value="-")
    indicator_asc_slope_var = tk.StringVar(value="-")
    indicator_dsc_slope_var = tk.StringVar(value="-")

    initial_state = {
        "asc": True,
        "dsc": True,
        "voltage": True,
        "temperature": False,
        "current_density": True,
        "fraction": 1.0,
        "x_min": default_limits["x_min"],
        "x_max": default_limits["x_max"],
        "v_min": default_limits["v_min"],
        "v_max": default_limits["v_max"],
        "temp_min": "",
        "temp_max": "",
        "asc_marker": "^",
        "dsc_marker": "v",
        "voltage_line": "-",
        "temperature_line": "--",
        "x_tick_count": 6,
        "y_tick_count": 6,
        "plot_title": "",
        "show_title": True,
        "title_fontsize": font_default_values["title"],
        "tick_fontsize": font_default_values["tick"],
        "label_fontsize": font_default_values["label"],
        "legend_fontsize": font_default_values["legend"],
        "marker_size": "6",
        "hollow_markers": False,
        "line_width": "1.5",
        "show_slope_guides": False,
    }
    current_density_state = {"value": True}
    indicator_scale_state = {"min": 0.0, "max": 1.0}

    def _schedule_plot(*_args):
        if suspend_events["value"]:
            return
        if plot_job["id"] is not None:
            win.after_cancel(plot_job["id"])
        plot_job["id"] = win.after(20, _plot)

    ttk.Label(
        controls_frame,
        text=f"Curva detectada:\n{bundle.description} #{bundle.curve_id}",
        justify="left",
    ).pack(anchor="w", pady=(0, 10))

    series_box = ttk.LabelFrame(controls_frame, text="Series")
    series_box.pack(fill="x", pady=5)

    indicators_box = ttk.LabelFrame(controls_frame, text="Indicadores")
    indicators_box.pack(fill="x", pady=5)

    style_box = ttk.LabelFrame(controls_frame, text="Estilo")
    style_box.pack(fill="x", pady=5)

    ttk.Label(style_box, text="Marcador ascendente").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    asc_marker_combo = ttk.Combobox(
        style_box,
        textvariable=asc_marker_var,
        values=MARKER_OPTIONS,
        state="readonly",
        width=10,
    )
    asc_marker_combo.grid(row=0, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Marcador descendente").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    dsc_marker_combo = ttk.Combobox(
        style_box,
        textvariable=dsc_marker_var,
        values=MARKER_OPTIONS,
        state="readonly",
        width=10,
    )
    dsc_marker_combo.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Línea de voltaje").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    voltage_line_combo = ttk.Combobox(
        style_box,
        textvariable=voltage_line_var,
        values=LINESTYLE_OPTIONS,
        state="readonly",
        width=10,
    )
    voltage_line_combo.grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Línea de temperatura").grid(row=3, column=0, sticky="w", padx=8, pady=3)
    temperature_line_combo = ttk.Combobox(
        style_box,
        textvariable=temperature_line_var,
        values=LINESTYLE_OPTIONS,
        state="readonly",
        width=10,
    )

    temperature_line_combo.grid(row=3, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Tamaño de marcador").grid(row=4, column=0, sticky="w", padx=8, pady=3)
    marker_size_entry = ttk.Spinbox(
        style_box,
        from_=0.0,
        to=20.0,
        increment=0.5,
        textvariable=marker_size_var,
        width=10,
    )
    marker_size_entry.grid(row=4, column=1, sticky="w", padx=8, pady=3)

    ttk.Checkbutton(
        style_box,
        text="Marcadores vacíos",
        variable=hollow_markers_var,
        command=_schedule_plot,
    ).grid(row=5, column=0, columnspan=2, sticky="w", padx=8, pady=4)

    ttk.Label(style_box, text="Grosor de línea").grid(row=6, column=0, sticky="w", padx=8, pady=3)
    line_width_entry = ttk.Spinbox(
        style_box,
        from_=0.0,
        to=10.0,
        increment=0.1,
        textvariable=line_width_var,
        width=10,
    )
    line_width_entry.grid(row=6, column=1, sticky="w", padx=8, pady=3)
    point_box = ttk.LabelFrame(controls_frame, text="Punto dentro de cada paso")
    point_box.pack(fill="x", pady=5)

    limits_box = ttk.LabelFrame(controls_frame, text="Límites de ejes")
    limits_box.pack(fill="x", pady=5)

    plot_job = {"id": None}
    suspend_events = {"value": False}

    text_box = ttk.LabelFrame(controls_frame, text="Texto / tamaños")
    text_box.pack(fill="x", pady=5)

    ttk.Label(text_box, text="Título").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    title_entry = ttk.Entry(text_box, textvariable=plot_title_var, width=28)
    title_entry.grid(row=0, column=1, sticky="we", padx=8, pady=3)
    ttk.Checkbutton(text_box, text="Mostrar titulo", variable=show_title_var, command=_schedule_plot).grid(
        row=5, column=0, columnspan=2, sticky="w", padx=8, pady=3
    )

    ttk.Label(text_box, text="Tamaño del título").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    title_size_entry = ttk.Spinbox(text_box, from_=6.0, to=50.0, increment=0.5, textvariable=title_fontsize_var, width=10)
    title_size_entry.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de ticks").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    tick_size_entry = ttk.Spinbox(text_box, from_=6.0, to=40.0, increment=0.5, textvariable=tick_fontsize_var, width=10)
    tick_size_entry.grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de etiquetas").grid(row=3, column=0, sticky="w", padx=8, pady=3)
    label_size_entry = ttk.Spinbox(text_box, from_=6.0, to=40.0, increment=0.5, textvariable=label_fontsize_var, width=10)
    label_size_entry.grid(row=3, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de leyenda").grid(row=4, column=0, sticky="w", padx=8, pady=3)
    legend_size_entry = ttk.Spinbox(text_box, from_=6.0, to=40.0, increment=0.5, textvariable=legend_fontsize_var, width=10)
    legend_size_entry.grid(row=4, column=1, sticky="w", padx=8, pady=3)

    def _positive_float(text: str, name: str) -> float:
        value = text.strip().replace(",", ".")
        if not value:
            raise ValueError(f"{name} no puede estar vacío.")
        num = float(value)
        if num <= 0:
            raise ValueError(f"{name} debe ser mayor que 0.")
        return num

    def _collect_limits():
        return dict(
            x_min=_optional_float(x_min_var.get()),
            x_max=_optional_float(x_max_var.get()),
            v_min=_optional_float(v_min_var.get()),
            v_max=_optional_float(v_max_var.get()),
            temp_min=_optional_float(temp_min_var.get()),
            temp_max=_optional_float(temp_max_var.get()),
        )

    def _current_axis_unit_text() -> str:
        return "A/cm^2" if current_density_var.get() else "A"

    def _slope_unit_text() -> str:
        return "V/(A/cm^2)" if current_density_var.get() else "V/A"

    def _set_indicator_state(current_text: str, asc_text: str = "-", dsc_text: str = "-") -> None:
        indicator_current_value_var.set(current_text)
        indicator_asc_slope_var.set(asc_text)
        indicator_dsc_slope_var.set(dsc_text)

    def _update_slope_indicator(_event=None, redraw_guides: bool = False) -> None:
        if not voltage_var.get():
            _set_indicator_state("Voltaje oculto")
            return

        try:
            area_cm2 = _bundle_area_cm2(bundle) if current_density_var.get() else None
            visible = _selected_v_vs_i_rows(
                bundle=bundle,
                show_asc=asc_var.get(),
                show_dsc=dsc_var.get(),
                point_fraction=point_fraction_var.get(),
            )
            asc_points = _v_vs_i_voltage_points(
                visible["asc_rows"],
                current_density_var.get(),
                area_cm2,
            )
            dsc_points = _v_vs_i_voltage_points(
                visible["dsc_rows"],
                current_density_var.get(),
                area_cm2,
            )
        except ValueError as exc:
            _set_indicator_state(f"Error: {exc}")
            return

        all_currents = [point[0] for point in asc_points] + [point[0] for point in dsc_points]
        if len(all_currents) < 2:
            _set_indicator_state("Sin puntos suficientes")
            return

        current_min = min(all_currents)
        current_max = max(all_currents)
        indicator_scale_state["min"] = current_min
        indicator_scale_state["max"] = current_max

        if current_max <= current_min:
            selected_current = current_min
        else:
            selected_current = min(max(indicator_current_var.get(), current_min), current_max)

        indicator_scale.configure(from_=current_min, to=current_max)
        indicator_current_var.set(selected_current)
        indicator_current_value_var.set(f"{selected_current:.6g} {_current_axis_unit_text()}")

        slope_unit = _slope_unit_text()
        asc_slope = _slope_at_current_from_points(asc_points, selected_current) if asc_points else None
        dsc_slope = _slope_at_current_from_points(dsc_points, selected_current) if dsc_points else None

        indicator_asc_slope_var.set(f"{asc_slope:.6g} {slope_unit}" if asc_slope is not None else "-")
        indicator_dsc_slope_var.set(f"{dsc_slope:.6g} {slope_unit}" if dsc_slope is not None else "-")

        if redraw_guides and show_slope_guides_var.get():
            _schedule_plot()

    def _on_indicator_scale_move(_value=None):
        _update_slope_indicator(redraw_guides=True)

    def _on_current_density_toggle():
        new_state = current_density_var.get()
        old_state = current_density_state["value"]
        if new_state == old_state:
            return

        try:
            area_cm2 = _bundle_area_cm2(bundle)
            if area_cm2 is None or area_cm2 <= 0:
                raise ValueError("No se pudo leer un AREA valida de la metadata para convertir la corriente.")
        except ValueError as exc:
            current_density_var.set(old_state)
            status_var.set(f"Error: {exc}")
            return

        factor = 1.0 / area_cm2 if new_state else area_cm2

        suspend_events["value"] = True
        try:
            for var in (x_min_var, x_max_var):
                value = _optional_float(var.get())
                if value is not None:
                    var.set(f"{value * factor:.6g}")
            indicator_current_var.set(indicator_current_var.get() * factor)
        finally:
            suspend_events["value"] = False

        current_density_state["value"] = new_state
        _plot()
        status_var.set("Unidad del eje x actualizada.")
    
    def _plot():
        plot_job["id"] = None

        try:
            has_plot = draw_v_vs_i_on_figure(
                fig=fig,
                bundle=bundle,
                show_asc=asc_var.get(),
                show_dsc=dsc_var.get(),
                asc_marker=asc_marker_var.get(),
                dsc_marker=dsc_marker_var.get(),
                voltage_linestyle=voltage_line_var.get(),
                temperature_linestyle=temperature_line_var.get(),
                show_voltage=voltage_var.get(),
                show_temperature=temperature_var.get(),
                use_current_density=current_density_var.get(),
                point_fraction=point_fraction_var.get(),
                x_tick_count=x_tick_count_var.get(),
                y_tick_count=y_tick_count_var.get(),
                plot_title=plot_title_var.get(),
                show_title=show_title_var.get(),
                title_fontsize=_positive_float(title_fontsize_var.get(), "Tamaño del título"),
                tick_fontsize=_positive_float(tick_fontsize_var.get(), "Tamaño de ticks"),
                label_fontsize=_positive_float(label_fontsize_var.get(), "Tamaño de etiquetas"),
                legend_fontsize=_positive_float(legend_fontsize_var.get(), "Tamaño de leyenda"),
                marker_size=_positive_float(marker_size_var.get(), "Tamaño de marcador"),
                hollow_markers=hollow_markers_var.get(),
                line_width=_positive_float(line_width_var.get(), "Grosor de línea"),
                show_slope_guides=show_slope_guides_var.get(),
                indicator_current=indicator_current_var.get(),
                language=language,
                **_collect_limits(),
            )

        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            return

        if not has_plot:
            fig.clear()
            canvas.draw_idle()
            status_var.set("No se muestra gráfico: seleccione al menos una dirección y una magnitud.")
            _update_slope_indicator()
            return

        canvas.draw_idle()
        _update_slope_indicator()
        status_var.set("Gráfico actualizado.")

    def _update_point_label(_event=None):
        point_value_label.config(text=f"{point_fraction_var.get():.2f}")

    def _on_scale_move(_value=None):
        _update_point_label()
        _plot()

    def _on_scale_release(_event=None):
        _schedule_plot()

    def _autofit():
        try:
            fitted = compute_autofit_v_vs_i_limits(
                bundle=bundle,
                show_asc=asc_var.get(),
                show_dsc=dsc_var.get(),
                show_voltage=voltage_var.get(),
                show_temperature=temperature_var.get(),
                point_fraction=point_fraction_var.get(),
                use_current_density=current_density_var.get(),
            )
        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            _update_slope_indicator()
            return

        suspend_events["value"] = True
        try:
            x_min_var.set(fitted["x_min"])
            x_max_var.set(fitted["x_max"])

            if fitted["v_min"] != "" or fitted["v_max"] != "":
                v_min_var.set(fitted["v_min"])
                v_max_var.set(fitted["v_max"])
            if fitted["t_min"] != "" or fitted["t_max"] != "":
                temp_min_var.set(fitted["t_min"])
                temp_max_var.set(fitted["t_max"])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Autoescala aplicada.")

    def _reset():
        suspend_events["value"] = True
        try:
            asc_var.set(initial_state["asc"])
            dsc_var.set(initial_state["dsc"])
            asc_marker_var.set(initial_state["asc_marker"])
            dsc_marker_var.set(initial_state["dsc_marker"])
            voltage_line_var.set(initial_state["voltage_line"])
            temperature_line_var.set(initial_state["temperature_line"])
            voltage_var.set(initial_state["voltage"])
            temperature_var.set(initial_state["temperature"])
            current_density_var.set(initial_state["current_density"])
            current_density_state["value"] = initial_state["current_density"]
            point_fraction_var.set(initial_state["fraction"])
            x_tick_count_var.set(initial_state["x_tick_count"])
            y_tick_count_var.set(initial_state["y_tick_count"])
            plot_title_var.set(initial_state["plot_title"])
            show_title_var.set(initial_state["show_title"])
            title_fontsize_var.set(initial_state["title_fontsize"])
            tick_fontsize_var.set(initial_state["tick_fontsize"])
            label_fontsize_var.set(initial_state["label_fontsize"])
            legend_fontsize_var.set(initial_state["legend_fontsize"])
            marker_size_var.set(initial_state["marker_size"])
            hollow_markers_var.set(initial_state["hollow_markers"])
            line_width_var.set(initial_state["line_width"])
            show_slope_guides_var.set(initial_state["show_slope_guides"])

            x_min_var.set(initial_state["x_min"])
            x_max_var.set(initial_state["x_max"])
            v_min_var.set(initial_state["v_min"])
            v_max_var.set(initial_state["v_max"])
            temp_min_var.set(initial_state["temp_min"])
            temp_max_var.set(initial_state["temp_max"])
            _update_point_label()
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Valores restaurados.")

    def _collect_report_plot_kwargs() -> dict[str, object]:
        return {
            "show_asc": bool(asc_var.get()),
            "show_dsc": bool(dsc_var.get()),
            "point_fraction": float(point_fraction_var.get()),
            **_smoothing_config(),
            **_axis_mode_config(),
            "use_current_density": bool(current_density_var.get()),
            "asc_marker": asc_marker_var.get(),
            "dsc_marker": dsc_marker_var.get(),
            "dvdi_linestyle": dvdi_line_var.get(),
            "x_tick_count": int(x_tick_count_var.get()),
            "y_tick_count": int(y_tick_count_var.get()),
            "plot_title": plot_title_var.get(),
            "show_title": bool(show_title_var.get()),
            "title_fontsize": _positive_float(title_fontsize_var.get(), "Tamaño del título"),
            "tick_fontsize": _positive_float(tick_fontsize_var.get(), "Tamaño de ticks"),
            "label_fontsize": _positive_float(label_fontsize_var.get(), "Tamaño de etiquetas"),
            "legend_fontsize": _positive_float(legend_fontsize_var.get(), "Tamaño de leyenda"),
            "marker_size": _positive_float(marker_size_var.get(), "Tamaño de marcador"),
            "hollow_markers": bool(hollow_markers_var.get()),
            "line_width": _positive_float(line_width_var.get(), "Grosor de línea"),
            "language": language,
            **_collect_limits(),
        }

    def _export_report() -> None:
        try:
            initial_dir = Path(output_dir) if output_dir is not None else Path.cwd()
            initial_dir.mkdir(parents=True, exist_ok=True)
            default_name = f"PC_dVdI_Report_{_safe_filename_part(bundle.description)}_#{bundle.curve_id}.pdf"
            path_text = filedialog.asksaveasfilename(
                title="Guardar reporte PC dV/dI",
                initialdir=str(initial_dir),
                initialfile=default_name,
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            )
            if not path_text:
                status_var.set("Exportacion de reporte cancelada.")
                return

            exported_path = export_pc_dv_di_report_pdf(
                bundle=bundle,
                output_path=Path(path_text),
                plot_kwargs=_collect_report_plot_kwargs(),
            )
        except Exception as exc:
            status_var.set(f"Error al exportar reporte: {type(exc).__name__}: {exc}")
            messagebox.showerror("PC dV/dI Report", str(exc))
            return

        status_var.set(f"Reporte exportado: {exported_path}")

    asc_marker_combo.bind("<<ComboboxSelected>>", _schedule_plot)
    dsc_marker_combo.bind("<<ComboboxSelected>>", _schedule_plot)
    voltage_line_combo.bind("<<ComboboxSelected>>", _schedule_plot)
    temperature_line_combo.bind("<<ComboboxSelected>>", _schedule_plot)
    for widget in (
        title_entry,
        title_size_entry,
        tick_size_entry,
        label_size_entry,
        legend_size_entry,
        marker_size_entry,
        line_width_entry,
    ):
        widget.bind("<Return>", _schedule_plot)
        widget.bind("<KP_Enter>", _schedule_plot)
        widget.bind("<FocusOut>", _schedule_plot)
    for spin in (
        title_size_entry,
        tick_size_entry,
        label_size_entry,
        legend_size_entry,
        marker_size_entry,
        line_width_entry,
    ):
        spin.configure(command=_schedule_plot)

    ttk.Checkbutton(series_box, text="Ascendente", variable=asc_var, command=_schedule_plot).pack(
        anchor="w", padx=8, pady=2
    )
    ttk.Checkbutton(series_box, text="Descendente", variable=dsc_var, command=_schedule_plot).pack(
        anchor="w", padx=8, pady=2
    )
    ttk.Checkbutton(series_box, text="Voltaje", variable=voltage_var, command=_schedule_plot).pack(
        anchor="w", padx=8, pady=2
    )
    ttk.Checkbutton(series_box, text="Temperatura", variable=temperature_var, command=_schedule_plot).pack(
        anchor="w", padx=8, pady=2
    )
    ttk.Checkbutton(
        series_box,
        text="x-axis A/cm^2",
        variable=current_density_var,
        command=_on_current_density_toggle,
    ).pack(anchor="w", padx=8, pady=2)

    indicators_box.columnconfigure(1, weight=1)
    ttk.Label(indicators_box, text="Pendiente dV/dI a corriente:").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    ttk.Label(indicators_box, textvariable=indicator_current_value_var).grid(row=0, column=1, sticky="w", padx=8, pady=3)

    indicator_scale = ttk.Scale(
        indicators_box,
        from_=0.0,
        to=1.0,
        orient="horizontal",
        variable=indicator_current_var,
        command=_on_indicator_scale_move,
    )
    indicator_scale.grid(row=1, column=0, columnspan=2, sticky="we", padx=8, pady=4)

    ttk.Label(indicators_box, text="dV/dI ascendente").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    ttk.Label(indicators_box, textvariable=indicator_asc_slope_var).grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(indicators_box, text="dV/dI descendente").grid(row=3, column=0, sticky="w", padx=8, pady=3)
    ttk.Label(indicators_box, textvariable=indicator_dsc_slope_var).grid(row=3, column=1, sticky="w", padx=8, pady=3)

    ttk.Checkbutton(
        indicators_box,
        text="Mostrar guía",
        variable=show_slope_guides_var,
        command=_schedule_plot,
    ).grid(row=4, column=0, columnspan=2, sticky="w", padx=8, pady=4)

    point_value_label = ttk.Label(point_box, text="1.00")
    point_value_label.pack(anchor="e", padx=8, pady=(4, 0))
    

    point_scale = ttk.Scale(
        point_box,
        from_=0.0,
        to=1.0,
        orient="horizontal",
        variable=point_fraction_var,
        command=_on_scale_move,
    )
    point_scale.pack(fill="x", padx=8, pady=6)

    ttk.Label(point_box, text="0 = primer punto, 1 = último punto").pack(
        anchor="w", padx=8, pady=(0, 6)
    )

    limit_specs = [
        ("I min", x_min_var),
        ("I max", x_max_var),
        ("V min", v_min_var),
        ("V max", v_max_var),
        ("T min", temp_min_var),
        ("T max", temp_max_var),
    ]

    entry_widgets = []
    for row_idx, (label, var) in enumerate(limit_specs):
        ttk.Label(limits_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        entry = ttk.Entry(limits_box, textvariable=var, width=12)
        entry.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        entry.bind("<KP_Enter>", _schedule_plot)
        entry.bind("<FocusOut>", _schedule_plot)
        entry.bind("<Return>", _schedule_plot)
        entry_widgets.append(entry)

    ttk.Label(limits_box, text="x-Ticks").grid(row=len(limit_specs), column=0, sticky="w", padx=8, pady=3)
    x_tick_spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=x_tick_count_var, width=8)
    x_tick_spin.grid(row=len(limit_specs), column=1, sticky="w", padx=8, pady=3)

    ttk.Label(limits_box, text="y-Ticks").grid(row=len(limit_specs) + 1, column=0, sticky="w", padx=8, pady=3)
    y_tick_spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=y_tick_count_var, width=8)
    y_tick_spin.grid(row=len(limit_specs) + 1, column=1, sticky="w", padx=8, pady=3)

    for spin in (x_tick_spin, y_tick_spin):
        spin.bind("<Return>", _schedule_plot)
        spin.bind("<FocusOut>", _schedule_plot)
        spin.config(command=_schedule_plot)

    ttk.Label(
        controls_frame,
        textvariable=status_var,
        wraplength=260,
        justify="left",
    ).pack(anchor="w", fill="x", pady=(10, 10))

    buttons_frame = ttk.Frame(controls_frame)
    buttons_frame.pack(fill="x", pady=(5, 0))

    ttk.Button(buttons_frame, text="Restablecer", command=_reset).pack(side="left", padx=(0, 6))
    ttk.Button(buttons_frame, text="Autoescala", command=_autofit).pack(side="left")

    _update_point_label()
    _plot()


def _copy_v_vs_i_line_style(src_line, dst_line) -> None:
    dst_line.set_color(src_line.get_color())
    dst_line.set_linestyle(src_line.get_linestyle())
    dst_line.set_marker(src_line.get_marker())
    dst_line.set_linewidth(src_line.get_linewidth())
    dst_line.set_markersize(src_line.get_markersize())
    try:
        dst_line.set_markerfacecolor(src_line.get_markerfacecolor())
    except Exception:
        pass
    try:
        dst_line.set_markeredgecolor(src_line.get_markeredgecolor())
    except Exception:
        pass
    try:
        dst_line.set_markeredgewidth(src_line.get_markeredgewidth())
    except Exception:
        pass
    try:
        dst_line.set_alpha(src_line.get_alpha())
    except Exception:
        pass


def _build_v_vs_i_tab(
    notebook: ttk.Notebook,
    bundle: CurveBundle,
    font_default_values: dict[str, str],
    composer_parent,
    source_contexts: dict[str, dict[str, object]],
    font_defaults: PlotFontDefaults,
    output_dir: Path | None = None,
    language: str = "es",
) -> dict[str, object]:
    language = normalize_language(language)
    tab_title = f"Etapa {bundle.curve_id} - {bundle.description}"
    default_limits = compute_default_v_vs_i_limits(bundle, use_current_density=True)

    tab = ttk.Frame(notebook)
    notebook.add(tab, text=tab_title[:28] + ("..." if len(tab_title) > 28 else ""))

    controls_host, plot_outer = create_resizable_plot_layout(tab, sidebar_width=320)
    controls_frame = _build_scrollable_controls(controls_host)
    ttk.Button(
        controls_frame,
        text="Componer",
        command=lambda: _open_v_vs_i_composer(composer_parent, source_contexts, font_defaults),
    ).pack(fill="x", pady=(0, 8))

    toolbar_frame = ttk.Frame(plot_outer)
    toolbar_frame.pack(side="top", fill="x")

    canvas_frame = ttk.Frame(plot_outer)
    canvas_frame.pack(side="top", fill="both", expand=True)

    fig = Figure(figsize=(9, 5.5), dpi=100)
    canvas = FigureCanvasTkAgg(fig, master=canvas_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    toolbar = NavigationToolbar2Tk(canvas, toolbar_frame, pack_toolbar=False)
    toolbar.update()
    toolbar.pack(side="left", fill="x")

    status_var = tk.StringVar(value="Listo.")

    asc_var = tk.BooleanVar(value=True)
    dsc_var = tk.BooleanVar(value=True)
    voltage_var = tk.BooleanVar(value=True)
    temperature_var = tk.BooleanVar(value=False)
    current_density_var = tk.BooleanVar(value=True)
    show_slope_guides_var = tk.BooleanVar(value=False)

    asc_marker_var = tk.StringVar(value="^")
    dsc_marker_var = tk.StringVar(value="v")
    voltage_line_var = tk.StringVar(value="-")
    temperature_line_var = tk.StringVar(value="--")
    marker_size_var = tk.StringVar(value="6")
    line_width_var = tk.StringVar(value="1.5")
    hollow_markers_var = tk.BooleanVar(value=False)

    point_fraction_var = tk.DoubleVar(value=1.0)
    indicator_current_var = tk.DoubleVar(value=0.0)
    indicator_current_value_var = tk.StringVar(value="-")
    indicator_asc_slope_var = tk.StringVar(value="-")
    indicator_dsc_slope_var = tk.StringVar(value="-")

    x_min_var = tk.StringVar(value=default_limits["x_min"])
    x_max_var = tk.StringVar(value=default_limits["x_max"])
    v_min_var = tk.StringVar(value=default_limits["v_min"])
    v_max_var = tk.StringVar(value=default_limits["v_max"])
    temp_min_var = tk.StringVar(value="")
    temp_max_var = tk.StringVar(value="")
    x_tick_count_var = tk.IntVar(value=6)
    y_tick_count_var = tk.IntVar(value=6)

    plot_title_var = tk.StringVar(value="")
    show_title_var = tk.BooleanVar(value=True)
    title_fontsize_var = tk.StringVar(value=font_default_values["title"])
    tick_fontsize_var = tk.StringVar(value=font_default_values["tick"])
    label_fontsize_var = tk.StringVar(value=font_default_values["label"])
    legend_fontsize_var = tk.StringVar(value=font_default_values["legend"])

    initial_state = {
        "asc": True,
        "dsc": True,
        "voltage": True,
        "temperature": False,
        "current_density": True,
        "show_slope_guides": False,
        "asc_marker": "^",
        "dsc_marker": "v",
        "voltage_line": "-",
        "temperature_line": "--",
        "marker_size": "6",
        "line_width": "1.5",
        "hollow_markers": False,
        "point_fraction": 1.0,
        "x_min": default_limits["x_min"],
        "x_max": default_limits["x_max"],
        "v_min": default_limits["v_min"],
        "v_max": default_limits["v_max"],
        "temp_min": "",
        "temp_max": "",
        "x_ticks": 6,
        "y_ticks": 6,
        "plot_title": "",
        "show_title": True,
        "title_size": font_default_values["title"],
        "tick_size": font_default_values["tick"],
        "label_size": font_default_values["label"],
        "legend_size": font_default_values["legend"],
    }

    plot_job = {"id": None}
    suspend_events = {"value": False}
    current_density_state = {"value": True}
    indicator_scale_state = {"min": 0.0, "max": 1.0}

    def _positive_float(text: str, name: str) -> float:
        value = text.strip().replace(",", ".")
        if not value:
            raise ValueError(f"{name} no puede estar vacÃ­o.")
        number = float(value)
        if number <= 0:
            raise ValueError(f"{name} debe ser mayor que 0.")
        return number

    def _collect_limits() -> dict[str, float | None]:
        return {
            "x_min": _optional_float(x_min_var.get()),
            "x_max": _optional_float(x_max_var.get()),
            "v_min": _optional_float(v_min_var.get()),
            "v_max": _optional_float(v_max_var.get()),
            "temp_min": _optional_float(temp_min_var.get()),
            "temp_max": _optional_float(temp_max_var.get()),
        }

    def _current_axis_unit_text() -> str:
        return "A/cm^2" if current_density_var.get() else "A"

    def _slope_unit_text() -> str:
        return "V/(A/cm^2)" if current_density_var.get() else "V/A"

    def _set_indicator_state(current_text: str, asc_text: str = "-", dsc_text: str = "-") -> None:
        indicator_current_value_var.set(current_text)
        indicator_asc_slope_var.set(asc_text)
        indicator_dsc_slope_var.set(dsc_text)

    def _schedule_plot(*_args) -> None:
        if suspend_events["value"]:
            return
        if plot_job["id"] is not None:
            tab.after_cancel(plot_job["id"])
        plot_job["id"] = tab.after(20, _plot)

    def _update_slope_indicator(_event=None, redraw_guides: bool = False) -> None:
        if not voltage_var.get():
            _set_indicator_state("Voltaje oculto")
            return

        try:
            area_cm2 = _bundle_area_cm2(bundle) if current_density_var.get() else None
            visible = _selected_v_vs_i_rows(
                bundle=bundle,
                show_asc=asc_var.get(),
                show_dsc=dsc_var.get(),
                point_fraction=point_fraction_var.get(),
            )
            asc_points = _v_vs_i_voltage_points(visible["asc_rows"], current_density_var.get(), area_cm2)
            dsc_points = _v_vs_i_voltage_points(visible["dsc_rows"], current_density_var.get(), area_cm2)
        except ValueError as exc:
            _set_indicator_state(f"Error: {exc}")
            return

        all_currents = [point[0] for point in asc_points] + [point[0] for point in dsc_points]
        if len(all_currents) < 2:
            _set_indicator_state("Sin puntos suficientes")
            return

        current_min = min(all_currents)
        current_max = max(all_currents)
        indicator_scale_state["min"] = current_min
        indicator_scale_state["max"] = current_max
        selected_current = current_min if current_max <= current_min else min(
            max(indicator_current_var.get(), current_min),
            current_max,
        )

        indicator_scale.configure(from_=current_min, to=current_max)
        indicator_current_var.set(selected_current)
        indicator_current_value_var.set(f"{selected_current:.6g} {_current_axis_unit_text()}")

        slope_unit = _slope_unit_text()
        asc_slope = _slope_at_current_from_points(asc_points, selected_current) if asc_points else None
        dsc_slope = _slope_at_current_from_points(dsc_points, selected_current) if dsc_points else None
        indicator_asc_slope_var.set(f"{asc_slope:.6g} {slope_unit}" if asc_slope is not None else "-")
        indicator_dsc_slope_var.set(f"{dsc_slope:.6g} {slope_unit}" if dsc_slope is not None else "-")

        if redraw_guides and show_slope_guides_var.get():
            _schedule_plot()

    def _plot() -> None:
        plot_job["id"] = None
        try:
            has_plot = draw_v_vs_i_on_figure(
                fig=fig,
                bundle=bundle,
                show_asc=asc_var.get(),
                show_dsc=dsc_var.get(),
                show_voltage=voltage_var.get(),
                show_temperature=temperature_var.get(),
                use_current_density=current_density_var.get(),
                point_fraction=point_fraction_var.get(),
                asc_marker=asc_marker_var.get(),
                dsc_marker=dsc_marker_var.get(),
                voltage_linestyle=voltage_line_var.get(),
                temperature_linestyle=temperature_line_var.get(),
                x_tick_count=x_tick_count_var.get(),
                y_tick_count=y_tick_count_var.get(),
                plot_title=plot_title_var.get(),
                show_title=show_title_var.get(),
                title_fontsize=_positive_float(title_fontsize_var.get(), "Tamaño del título"),
                tick_fontsize=_positive_float(tick_fontsize_var.get(), "Tamaño de ticks"),
                label_fontsize=_positive_float(label_fontsize_var.get(), "Tamaño de etiquetas"),
                legend_fontsize=_positive_float(legend_fontsize_var.get(), "Tamaño de leyenda"),
                marker_size=_positive_float(marker_size_var.get(), "Tamaño de marcador"),
                hollow_markers=hollow_markers_var.get(),
                line_width=_positive_float(line_width_var.get(), "Grosor de línea"),
                show_slope_guides=show_slope_guides_var.get(),
                indicator_current=indicator_current_var.get(),
                **_collect_limits(),
            )
        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            return

        if not has_plot:
            fig.clear()
            canvas.draw_idle()
            status_var.set("No se muestra grÃ¡fico: seleccione al menos una direcciÃ³n y una magnitud.")
            _update_slope_indicator()
            return

        canvas.draw_idle()
        _update_slope_indicator()
        status_var.set("GrÃ¡fico actualizado.")

    def _on_current_density_toggle() -> None:
        new_state = current_density_var.get()
        old_state = current_density_state["value"]
        if new_state == old_state:
            return

        area_cm2 = _bundle_area_cm2(bundle)
        if area_cm2 is None or area_cm2 <= 0:
            current_density_var.set(old_state)
            status_var.set("Error: no se pudo leer un AREA valida para convertir la corriente.")
            return

        factor = 1.0 / area_cm2 if new_state else area_cm2
        suspend_events["value"] = True
        try:
            for variable in (x_min_var, x_max_var):
                value = _optional_float(variable.get())
                if value is not None:
                    variable.set(f"{value * factor:.6g}")
            indicator_current_var.set(indicator_current_var.get() * factor)
        finally:
            suspend_events["value"] = False

        current_density_state["value"] = new_state
        _plot()
        status_var.set("Unidad del eje x actualizada.")

    def _autofit() -> None:
        try:
            fitted = compute_autofit_v_vs_i_limits(
                bundle=bundle,
                show_asc=asc_var.get(),
                show_dsc=dsc_var.get(),
                show_voltage=voltage_var.get(),
                show_temperature=temperature_var.get(),
                point_fraction=point_fraction_var.get(),
                use_current_density=current_density_var.get(),
            )
        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            _update_slope_indicator()
            return

        suspend_events["value"] = True
        try:
            x_min_var.set(fitted["x_min"])
            x_max_var.set(fitted["x_max"])
            if fitted["v_min"] != "" or fitted["v_max"] != "":
                v_min_var.set(fitted["v_min"])
                v_max_var.set(fitted["v_max"])
            if fitted["t_min"] != "" or fitted["t_max"] != "":
                temp_min_var.set(fitted["t_min"])
                temp_max_var.set(fitted["t_max"])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Autoescala aplicada.")

    def _reset() -> None:
        suspend_events["value"] = True
        try:
            asc_var.set(initial_state["asc"])
            dsc_var.set(initial_state["dsc"])
            voltage_var.set(initial_state["voltage"])
            temperature_var.set(initial_state["temperature"])
            current_density_var.set(initial_state["current_density"])
            show_slope_guides_var.set(initial_state["show_slope_guides"])
            asc_marker_var.set(initial_state["asc_marker"])
            dsc_marker_var.set(initial_state["dsc_marker"])
            voltage_line_var.set(initial_state["voltage_line"])
            temperature_line_var.set(initial_state["temperature_line"])
            marker_size_var.set(initial_state["marker_size"])
            line_width_var.set(initial_state["line_width"])
            hollow_markers_var.set(initial_state["hollow_markers"])
            point_fraction_var.set(initial_state["point_fraction"])
            x_min_var.set(initial_state["x_min"])
            x_max_var.set(initial_state["x_max"])
            v_min_var.set(initial_state["v_min"])
            v_max_var.set(initial_state["v_max"])
            temp_min_var.set(initial_state["temp_min"])
            temp_max_var.set(initial_state["temp_max"])
            x_tick_count_var.set(initial_state["x_ticks"])
            y_tick_count_var.set(initial_state["y_ticks"])
            plot_title_var.set(initial_state["plot_title"])
            show_title_var.set(initial_state["show_title"])
            title_fontsize_var.set(initial_state["title_size"])
            tick_fontsize_var.set(initial_state["tick_size"])
            label_fontsize_var.set(initial_state["label_size"])
            legend_fontsize_var.set(initial_state["legend_size"])
            current_density_state["value"] = initial_state["current_density"]
            _update_point_label()
        finally:
            suspend_events["value"] = False
        _plot()
        status_var.set("Valores restaurados.")

    def _collect_report_plot_kwargs() -> dict[str, object]:
        return {
            "show_asc": bool(asc_var.get()),
            "show_dsc": bool(dsc_var.get()),
            "show_voltage": bool(voltage_var.get()),
            "show_temperature": bool(temperature_var.get()),
            "use_current_density": bool(current_density_var.get()),
            "point_fraction": float(point_fraction_var.get()),
            "asc_marker": asc_marker_var.get(),
            "dsc_marker": dsc_marker_var.get(),
            "voltage_linestyle": voltage_line_var.get(),
            "temperature_linestyle": temperature_line_var.get(),
            "x_tick_count": int(x_tick_count_var.get()),
            "y_tick_count": int(y_tick_count_var.get()),
            "plot_title": plot_title_var.get(),
            "show_title": bool(show_title_var.get()),
            "title_fontsize": _positive_float(title_fontsize_var.get(), "TamaÃ±o del tÃ­tulo"),
            "tick_fontsize": _positive_float(tick_fontsize_var.get(), "TamaÃ±o de ticks"),
            "label_fontsize": _positive_float(label_fontsize_var.get(), "TamaÃ±o de etiquetas"),
            "legend_fontsize": _positive_float(legend_fontsize_var.get(), "TamaÃ±o de leyenda"),
            "marker_size": _positive_float(marker_size_var.get(), "TamaÃ±o de marcador"),
            "hollow_markers": bool(hollow_markers_var.get()),
            "line_width": _positive_float(line_width_var.get(), "Grosor de lÃ­nea"),
            "show_slope_guides": bool(show_slope_guides_var.get()),
            "indicator_current": float(indicator_current_var.get()),
            "language": language,
            **_collect_limits(),
        }

    def _export_report() -> None:
        try:
            initial_dir = Path(output_dir) if output_dir is not None else Path.cwd()
            initial_dir.mkdir(parents=True, exist_ok=True)
            default_name = f"PC_Report_{_safe_filename_part(bundle.description)}_#{bundle.curve_id}.pdf"
            path_text = filedialog.asksaveasfilename(
                title="Guardar reporte PC",
                initialdir=str(initial_dir),
                initialfile=default_name,
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            )
            if not path_text:
                status_var.set("Exportacion de reporte cancelada.")
                return

            exported_path = export_pc_curve_report_pdf(
                bundle=bundle,
                output_path=Path(path_text),
                plot_kwargs=_collect_report_plot_kwargs(),
            )
        except Exception as exc:
            status_var.set(f"Error al exportar reporte: {type(exc).__name__}: {exc}")
            messagebox.showerror("PC Report", str(exc))
            return

        status_var.set(f"Reporte exportado: {exported_path}")

    def _update_point_label(_event=None) -> None:
        point_value_label.config(text=f"{point_fraction_var.get():.2f}")

    def _bind_plot_widget(widget) -> None:
        widget.bind("<Return>", _schedule_plot)
        widget.bind("<KP_Enter>", _schedule_plot)
        widget.bind("<FocusOut>", _schedule_plot)
        try:
            widget.configure(command=_schedule_plot)
        except Exception:
            pass

    ttk.Label(
        controls_frame,
        text=f"Curva detectada:\nEtapa {bundle.curve_id} - {bundle.description}",
        justify="left",
    ).pack(anchor="w", pady=(0, 10))

    series_box = ttk.LabelFrame(controls_frame, text="Series")
    series_box.pack(fill="x", pady=5)
    for text, variable, command in (
        ("Ascendente", asc_var, _schedule_plot),
        ("Descendente", dsc_var, _schedule_plot),
        ("Voltaje", voltage_var, _schedule_plot),
        ("Temperatura", temperature_var, _schedule_plot),
        ("x-axis A/cm^2", current_density_var, _on_current_density_toggle),
    ):
        ttk.Checkbutton(series_box, text=text, variable=variable, command=command).pack(anchor="w", padx=8, pady=2)

    indicators_box = ttk.LabelFrame(controls_frame, text="Indicadores")
    indicators_box.pack(fill="x", pady=5)
    indicators_box.columnconfigure(1, weight=1)
    ttk.Label(indicators_box, text="Pendiente dV/dI a corriente:").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    ttk.Label(indicators_box, textvariable=indicator_current_value_var).grid(row=0, column=1, sticky="w", padx=8, pady=3)
    indicator_scale = ttk.Scale(
        indicators_box,
        from_=0.0,
        to=1.0,
        orient="horizontal",
        variable=indicator_current_var,
        command=lambda _value=None: _update_slope_indicator(redraw_guides=True),
    )
    indicator_scale.grid(row=1, column=0, columnspan=2, sticky="we", padx=8, pady=4)
    ttk.Label(indicators_box, text="dV/dI ascendente").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    ttk.Label(indicators_box, textvariable=indicator_asc_slope_var).grid(row=2, column=1, sticky="w", padx=8, pady=3)
    ttk.Label(indicators_box, text="dV/dI descendente").grid(row=3, column=0, sticky="w", padx=8, pady=3)
    ttk.Label(indicators_box, textvariable=indicator_dsc_slope_var).grid(row=3, column=1, sticky="w", padx=8, pady=3)
    ttk.Checkbutton(
        indicators_box,
        text="Mostrar guía",
        variable=show_slope_guides_var,
        command=_schedule_plot,
    ).grid(row=4, column=0, columnspan=2, sticky="w", padx=8, pady=4)

    style_box = ttk.LabelFrame(controls_frame, text="Estilo")
    style_box.pack(fill="x", pady=5)
    style_specs = [
        ("Marcador ascendente", asc_marker_var, MARKER_OPTIONS),
        ("Marcador descendente", dsc_marker_var, MARKER_OPTIONS),
        ("Línea de voltaje", voltage_line_var, LINESTYLE_OPTIONS),
        ("Línea de temperatura", temperature_line_var, LINESTYLE_OPTIONS),
    ]
    for row_idx, (label, variable, values) in enumerate(style_specs):
        ttk.Label(style_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        combo = ttk.Combobox(style_box, textvariable=variable, values=values, state="readonly", width=10)
        combo.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        combo.bind("<<ComboboxSelected>>", _schedule_plot)
    for row_idx, (label, variable, start, stop, step) in enumerate(
        (
            ("Tamaño de marcador", marker_size_var, 0.0, 20.0, 0.5),
            ("Grosor de línea", line_width_var, 0.0, 10.0, 0.1),
        ),
        start=len(style_specs),
    ):
        ttk.Label(style_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        spin = ttk.Spinbox(style_box, from_=start, to=stop, increment=step, textvariable=variable, width=10)
        spin.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        _bind_plot_widget(spin)
    ttk.Checkbutton(
        style_box,
        text="Marcadores vacíos",
        variable=hollow_markers_var,
        command=_schedule_plot,
    ).grid(row=len(style_specs) + 2, column=0, columnspan=2, sticky="w", padx=8, pady=4)

    point_box = ttk.LabelFrame(controls_frame, text="Punto dentro de cada paso")
    point_box.pack(fill="x", pady=5)
    point_value_label = ttk.Label(point_box, text="1.00")
    point_value_label.pack(anchor="e", padx=8, pady=(4, 0))
    point_scale = ttk.Scale(
        point_box,
        from_=0.0,
        to=1.0,
        orient="horizontal",
        variable=point_fraction_var,
        command=lambda _value=None: (_update_point_label(), _plot()),
    )
    point_scale.pack(fill="x", padx=8, pady=6)
    point_scale.bind("<ButtonRelease-1>", lambda _event: _schedule_plot())
    ttk.Label(point_box, text="0 = primer punto, 1 = último punto").pack(anchor="w", padx=8, pady=(0, 6))

    text_box = ttk.LabelFrame(controls_frame, text="Texto / tamaños")
    text_box.pack(fill="x", pady=5)
    text_specs = [
        ("Título", plot_title_var, None),
        ("Tamaño del título", title_fontsize_var, (6.0, 50.0, 0.5)),
        ("Tamaño de ticks", tick_fontsize_var, (6.0, 40.0, 0.5)),
        ("Tamaño de etiquetas", label_fontsize_var, (6.0, 40.0, 0.5)),
        ("Tamaño de leyenda", legend_fontsize_var, (6.0, 40.0, 0.5)),
    ]
    for row_idx, (label, variable, spin_cfg) in enumerate(text_specs):
        ttk.Label(text_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        if spin_cfg is None:
            widget = ttk.Entry(text_box, textvariable=variable, width=28)
        else:
            widget = ttk.Spinbox(
                text_box,
                from_=spin_cfg[0],
                to=spin_cfg[1],
                increment=spin_cfg[2],
                textvariable=variable,
                width=10,
            )
        widget.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        _bind_plot_widget(widget)
    ttk.Checkbutton(text_box, text="Mostrar titulo", variable=show_title_var, command=_schedule_plot).grid(
        row=len(text_specs), column=0, columnspan=2, sticky="w", padx=8, pady=3
    )

    limits_box = ttk.LabelFrame(controls_frame, text="Límites de ejes")
    limits_box.pack(fill="x", pady=5)
    for row_idx, (label, variable) in enumerate(
        (
            ("I min", x_min_var),
            ("I max", x_max_var),
            ("V min", v_min_var),
            ("V max", v_max_var),
            ("T min", temp_min_var),
            ("T max", temp_max_var),
        )
    ):
        ttk.Label(limits_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        entry = ttk.Entry(limits_box, textvariable=variable, width=12)
        entry.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        _bind_plot_widget(entry)
    for row_idx, (label, variable) in enumerate((("x-Ticks", x_tick_count_var), ("y-Ticks", y_tick_count_var)), start=6):
        ttk.Label(limits_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=variable, width=8)
        spin.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        spin.bind("<Return>", _schedule_plot)
        spin.bind("<FocusOut>", _schedule_plot)
        spin.config(command=_schedule_plot)

    ttk.Label(controls_frame, textvariable=status_var, wraplength=280, justify="left").pack(
        anchor="w",
        fill="x",
        pady=(10, 10),
    )

    buttons = ttk.Frame(controls_frame)
    buttons.pack(fill="x", pady=(5, 0))
    ttk.Button(buttons, text="Restablecer", command=_reset).pack(side="left", padx=(0, 6))
    ttk.Button(buttons, text="Autoescala", command=_autofit).pack(side="left")
    ttk.Button(buttons, text="Reporte PDF", command=_export_report).pack(side="left", padx=(6, 0))

    _update_point_label()
    _plot()

    return {
        "tab_title": tab_title,
        "figure": fig,
        "current_density_getter": lambda: bool(current_density_var.get()),
    }


def _open_v_vs_i_composer(
    parent,
    source_contexts: dict[str, dict[str, object]],
    font_defaults: PlotFontDefaults,
) -> None:
    if not source_contexts:
        return

    existing = getattr(parent, "_composer_win_v_vs_i", None)
    if existing is not None and existing.winfo_exists():
        existing.lift()
        existing.focus_force()
        return

    comp = tk.Toplevel(parent)
    parent._composer_win_v_vs_i = comp  # type: ignore[attr-defined]
    comp.title("Composite (PC V vs I)")
    comp.geometry("1240x760")
    comp.configure(
        bg=ttk.Style(comp).lookup("App.TFrame", "background")
        or ttk.Style(comp).lookup("TFrame", "background")
    )

    ctrl_host, plot_side = create_resizable_plot_layout(
        comp,
        sidebar_width=320,
        sidebar_side="right",
        plot_padding=0,
    )
    ctrl = _build_scrollable_controls(ctrl_host)

    figc = Figure(figsize=(8.5, 6.0), dpi=100)
    ax_main = figc.add_subplot(111)
    ax_temp = ax_main.twinx()

    canvas = FigureCanvasTkAgg(figc, master=plot_side)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    toolbar = NavigationToolbar2Tk(canvas, plot_side, pack_toolbar=False)
    toolbar.update()
    toolbar.pack(side="top", fill="x")

    title_var = tk.StringVar(value="Composición - V vs I")
    show_title_var = tk.BooleanVar(value=True)
    legend_var = tk.BooleanVar(value=True)
    status_var = tk.StringVar(value="Listo.")
    comp_lines: dict[str, list[object]] = {}
    density_mode = {"value": None}
    idx_to_key: list[str] = []
    title_fs_var = tk.DoubleVar(value=float(font_defaults.title))
    label_fs_var = tk.DoubleVar(value=float(font_defaults.label))
    legend_fs_var = tk.DoubleVar(value=float(font_defaults.legend))
    tick_fs_var = tk.DoubleVar(value=float(font_defaults.tick))
    x_tick_count_var = tk.IntVar(value=6)
    y_tick_count_var = tk.IntVar(value=6)
    xmin_var = tk.StringVar()
    xmax_var = tk.StringVar()
    vmin_var = tk.StringVar()
    vmax_var = tk.StringVar()
    tmin_var = tk.StringVar()
    tmax_var = tk.StringVar()

    def _source_title(key: str) -> str:
        fig = source_contexts[key]["figure"]
        if isinstance(fig, Figure) and fig.axes:
            title = (fig.axes[0].get_title() or "").strip()
            if title:
                return title
        return str(source_contexts[key]["tab_title"])

    def _source_density(key: str) -> bool:
        getter = source_contexts[key]["current_density_getter"]
        return bool(getter())

    def _source_lines(key: str) -> list[tuple[str, object]]:
        fig = source_contexts[key]["figure"]
        out: list[tuple[str, object]] = []
        if not isinstance(fig, Figure):
            return out
        for ax in fig.axes:
            for line in ax.get_lines():
                label = (line.get_label() or "").strip()
                if label and not label.startswith("_"):
                    out.append((label, line))
        return out

    def _refresh_listbox() -> None:
        lb.delete(0, "end")
        idx_to_key.clear()
        for key in sorted(source_contexts.keys(), key=lambda item: _source_title(item).lower()):
            lb.insert("end", f"{_source_title(key)}   [{key}]")
            idx_to_key.append(key)

    def _selected_keys() -> list[str]:
        return [idx_to_key[index] for index in lb.curselection()]

    def _apply_axes() -> None:
        ax_main.set_title(title_var.get() if show_title_var.get() else "")
        ax_main.set_xlabel("Densidad de corriente (A/cm^2)" if density_mode["value"] else "Corriente (A)")
        ax_main.set_ylabel("Voltaje (V)")
        ax_temp.set_ylabel("Temperatura (C)")
        ax_main.grid(True)
        ax_temp.grid(False)

        ax_main.yaxis.tick_left()
        ax_main.yaxis.set_label_position("left")
        ax_main.spines["left"].set_visible(True)
        ax_main.spines["right"].set_visible(False)

        ax_temp.yaxis.tick_right()
        ax_temp.yaxis.set_label_position("right")
        ax_temp.spines["left"].set_visible(False)
        ax_temp.spines["right"].set_position(("axes", 1.0))
        figc.subplots_adjust(right=0.88)

        handles, labels = ax_main.get_legend_handles_labels()
        h2, l2 = ax_temp.get_legend_handles_labels()
        leg = ax_main.get_legend()
        if leg is not None:
            leg.remove()
        if legend_var.get() and (handles or h2):
            make_legend_draggable(ax_main.legend(handles + h2, labels + l2, fontsize=float(font_defaults.legend)))

        has_temp = any(" T" in label for lines in comp_lines.values() for label in [line.get_label() for line in lines])
        ax_temp.yaxis.set_visible(has_temp)
        ax_temp.spines["right"].set_visible(has_temp)
        ax_temp.yaxis.label.set_visible(has_temp)
        canvas.draw_idle()

    def _has_temp_lines() -> bool:
        return any(
            str(line.get_label()).endswith(" T")
            for lines in comp_lines.values()
            for line in lines
        )

    def _fmt(v: float) -> str:
        return f"{v:.6g}"

    def _parse_float(s: str) -> float | None:
        s = s.strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None

    def _is_incomplete_number(s: str) -> bool:
        s = s.strip()
        if s in {"", "-", "+", ".", "-.", "+."}:
            return True
        if re.fullmatch(r"[+-]?\d+\.", s):
            return True
        if re.fullmatch(r"[+-]?(?:\d+\.?\d*|\.\d+)[eE][+-]?", s):
            return True
        return False

    def _apply_tick_settings() -> None:
        try:
            x_tick_count = max(2, int(x_tick_count_var.get()))
        except (tk.TclError, ValueError):
            x_tick_count = 6
        try:
            y_tick_count = max(2, int(y_tick_count_var.get()))
        except (tk.TclError, ValueError):
            y_tick_count = 6

        ax_main.xaxis.set_major_locator(MaxNLocator(nbins=x_tick_count))
        ax_main.yaxis.set_major_locator(MaxNLocator(nbins=y_tick_count))
        ax_temp.yaxis.set_major_locator(MaxNLocator(nbins=y_tick_count))

        ax_main.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
        ax_main.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
        ax_temp.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))

    def _sync_limit_entries() -> None:
        x0, x1 = ax_main.get_xlim()
        v0, v1 = ax_main.get_ylim()
        xmin_var.set(_fmt(x0))
        xmax_var.set(_fmt(x1))
        vmin_var.set(_fmt(v0))
        vmax_var.set(_fmt(v1))

        if _has_temp_lines():
            t0, t1 = ax_temp.get_ylim()
            tmin_var.set(_fmt(t0))
            tmax_var.set(_fmt(t1))
        else:
            tmin_var.set("")
            tmax_var.set("")

    def _apply_limits() -> None:
        raw_values = [
            xmin_var.get(),
            xmax_var.get(),
            vmin_var.get(),
            vmax_var.get(),
            tmin_var.get(),
            tmax_var.get(),
        ]
        if any(value.strip() and _is_incomplete_number(value) for value in raw_values):
            return

        cur_x0, cur_x1 = ax_main.get_xlim()
        cur_v0, cur_v1 = ax_main.get_ylim()
        cur_t0, cur_t1 = ax_temp.get_ylim()

        nx0 = _parse_float(xmin_var.get())
        nx1 = _parse_float(xmax_var.get())
        nv0 = _parse_float(vmin_var.get())
        nv1 = _parse_float(vmax_var.get())
        nt0 = _parse_float(tmin_var.get())
        nt1 = _parse_float(tmax_var.get())

        ax_main.set_xlim(cur_x0 if nx0 is None else nx0, cur_x1 if nx1 is None else nx1)
        ax_main.set_ylim(cur_v0 if nv0 is None else nv0, cur_v1 if nv1 is None else nv1)
        ax_temp.set_ylim(cur_t0 if nt0 is None else nt0, cur_t1 if nt1 is None else nt1)

        _apply_axes(redraw=False)
        canvas.draw_idle()
        _sync_limit_entries()

    def _apply_axes(redraw: bool = True) -> None:
        try:
            title_fs = float(title_fs_var.get())
            label_fs = float(label_fs_var.get())
            legend_fs = float(legend_fs_var.get())
            tick_fs = float(tick_fs_var.get())
        except (tk.TclError, ValueError):
            return

        ax_main.set_title(title_var.get() if show_title_var.get() else "", fontsize=title_fs)
        ax_main.set_xlabel(
            "Densidad de corriente (A/cm^2)" if density_mode["value"] else "Corriente (A)",
            fontsize=label_fs,
        )
        ax_main.set_ylabel("Voltaje (V)", fontsize=label_fs)
        ax_temp.set_ylabel("Temperatura (C)", fontsize=label_fs)
        ax_main.grid(True)
        ax_temp.grid(False)

        ax_main.tick_params(axis="both", labelsize=tick_fs)
        apply_x_tick_label_padding(ax_main, tick_fs)
        ax_temp.tick_params(axis="y", labelsize=tick_fs)
        _apply_tick_settings()

        ax_main.yaxis.tick_left()
        ax_main.yaxis.set_label_position("left")
        ax_main.spines["left"].set_visible(True)
        ax_main.spines["right"].set_visible(False)

        ax_temp.yaxis.tick_right()
        ax_temp.yaxis.set_label_position("right")
        ax_temp.spines["left"].set_visible(False)
        ax_temp.spines["right"].set_position(("axes", 1.0))
        figc.subplots_adjust(right=0.88)

        leg = ax_main.get_legend()
        if leg is not None:
            leg.remove()

        if legend_var.get():
            handles, labels = ax_main.get_legend_handles_labels()
            h2, l2 = ax_temp.get_legend_handles_labels()
            pairs = [
                (handle, label)
                for handle, label in zip(handles + h2, labels + l2)
                if label and not str(label).startswith("_")
            ]
            if pairs:
                legend_handles, legend_labels = zip(*pairs)
                make_legend_draggable(ax_main.legend(legend_handles, legend_labels, fontsize=legend_fs))

        has_temp = _has_temp_lines()
        ax_temp.yaxis.set_visible(has_temp)
        ax_temp.spines["right"].set_visible(has_temp)
        ax_temp.yaxis.label.set_visible(has_temp)

        if redraw:
            canvas.draw_idle()

    def _fit_all() -> None:
        if not comp_lines:
            _apply_axes(redraw=False)
            canvas.draw_idle()
            _sync_limit_entries()
            return

        xs: list[float] = []
        ys_main: list[float] = []
        ys_temp: list[float] = []
        for lines in comp_lines.values():
            for line in lines:
                label = line.get_label()
                xs.extend(float(value) for value in line.get_xdata(orig=False))
                if " T" in label:
                    ys_temp.extend(float(value) for value in line.get_ydata(orig=False))
                else:
                    ys_main.extend(float(value) for value in line.get_ydata(orig=False))
        if xs:
            x0 = min(xs)
            x1 = max(xs)
            dx = (x1 - x0) if x1 != x0 else (abs(x0) * 0.1 + 1.0)
            ax_main.set_xlim(x0 - 0.05 * dx, x1 + 0.05 * dx)
        if ys_main:
            y0 = min(ys_main)
            y1 = max(ys_main)
            dy = (y1 - y0) if y1 != y0 else (abs(y0) * 0.1 + 1.0)
            ax_main.set_ylim(y0 - 0.05 * dy, y1 + 0.05 * dy)
        if ys_temp:
            t0 = min(ys_temp)
            t1 = max(ys_temp)
            dt = (t1 - t0) if t1 != t0 else (abs(t0) * 0.1 + 1.0)
            ax_temp.set_ylim(t0 - 0.05 * dt, t1 + 0.05 * dt)
        _apply_axes(redraw=False)
        canvas.draw_idle()
        _sync_limit_entries()

    def _remove_key(key: str) -> None:
        for line in comp_lines.pop(key, []):
            try:
                line.remove()
            except Exception:
                pass
        if not comp_lines:
            density_mode["value"] = None

    def _add_selected() -> None:
        errors: list[str] = []
        for key in _selected_keys():
            source_lines = _source_lines(key)
            if not source_lines:
                errors.append(f"{_source_title(key)} no tiene series visibles.")
                continue
            source_density = _source_density(key)
            if density_mode["value"] is None:
                density_mode["value"] = source_density
            elif density_mode["value"] != source_density:
                errors.append(f"{_source_title(key)} usa una unidad distinta en el eje x.")
                continue

            _remove_key(key)
            new_lines: list[object] = []
            for label, src_line in source_lines:
                target_ax = ax_temp if label.endswith(" T") else ax_main
                x = list(src_line.get_xdata(orig=False))
                y = list(src_line.get_ydata(orig=False))
                (new_line,) = target_ax.plot(x, y, label=f"{_source_title(key)} - {label}")
                _copy_v_vs_i_line_style(src_line, new_line)
                new_lines.append(new_line)
            comp_lines[key] = new_lines

        _fit_all()
        status_var.set(" / ".join(errors) if errors else "Curvas agregadas al composite.")

    def _remove_selected() -> None:
        changed = False
        for key in _selected_keys():
            if key in comp_lines:
                _remove_key(key)
                changed = True
        if changed:
            _fit_all()
            status_var.set("Curvas removidas del composite.")

    def _clear_all() -> None:
        for key in list(comp_lines):
            _remove_key(key)
        ax_main.cla()
        ax_temp.cla()
        _apply_axes(redraw=False)
        canvas.draw_idle()
        _sync_limit_entries()
        status_var.set("Composición limpiada.")

    def _refresh_formatting() -> None:
        active = list(comp_lines.keys())
        _clear_all()
        for key in active:
            lb.selection_set(idx_to_key.index(key))
        _add_selected()

    src_box = ttk.LabelFrame(ctrl, text="Fuentes de curvas", padding=8)
    src_box.pack(fill="x", pady=(0, 10))
    lb = tk.Listbox(src_box, selectmode="extended", height=12, exportselection=False)
    lb.pack(fill="x")
    _refresh_listbox()

    btns = ttk.Frame(src_box)
    btns.pack(fill="x", pady=(8, 0))
    ttk.Button(btns, text="Añadir", command=_add_selected).pack(side="left", expand=True, fill="x", padx=(0, 6))
    ttk.Button(btns, text="Remover", command=_remove_selected).pack(side="left", expand=True, fill="x", padx=(0, 6))
    ttk.Button(btns, text="Limpiar", command=_clear_all).pack(side="left", expand=True, fill="x")
    ttk.Button(src_box, text="Actualizar formato", command=_refresh_formatting).pack(fill="x", pady=(8, 0))

    plot_box = ttk.LabelFrame(ctrl, text="Gráfico", padding=8)
    plot_box.pack(fill="x", pady=(0, 10))
    plot_box.columnconfigure(1, weight=1)
    ttk.Label(plot_box, text="Título").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=2)
    title_entry = ttk.Entry(plot_box, textvariable=title_var, width=22)
    title_entry.grid(row=0, column=1, sticky="ew", pady=2)
    title_entry.bind("<Return>", lambda _event: _apply_axes())
    title_entry.bind("<FocusOut>", lambda _event: _apply_axes())
    title_entry.bind("<KeyRelease>", lambda _event: _apply_axes())
    ttk.Checkbutton(plot_box, text="Mostrar titulo", variable=show_title_var, command=_apply_axes).grid(
        row=1,
        column=0,
        columnspan=2,
        sticky="w",
        pady=(4, 0),
    )
    ttk.Checkbutton(plot_box, text="Leyenda", variable=legend_var, command=_apply_axes).grid(
        row=2,
        column=0,
        columnspan=2,
        sticky="w",
        pady=(4, 4),
    )
    ttk.Button(plot_box, text="Ajustar todo", command=_fit_all).grid(row=3, column=0, columnspan=2, sticky="ew", pady=(4, 0))

    style_box = ttk.LabelFrame(ctrl, text="Apariencia", padding=8)
    style_box.pack(fill="x", pady=(0, 10))
    style_box.columnconfigure(1, weight=1)

    ttk.Label(style_box, text="Titulo").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=2)
    title_fs_spin = ttk.Spinbox(style_box, from_=6.0, to=50.0, increment=0.5, textvariable=title_fs_var, width=10)
    title_fs_spin.grid(row=0, column=1, sticky="w", pady=2)

    ttk.Label(style_box, text="Etiquetas").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=2)
    label_fs_spin = ttk.Spinbox(style_box, from_=6.0, to=40.0, increment=0.5, textvariable=label_fs_var, width=10)
    label_fs_spin.grid(row=1, column=1, sticky="w", pady=2)

    ttk.Label(style_box, text="Leyenda").grid(row=2, column=0, sticky="w", padx=(0, 6), pady=2)
    legend_fs_spin = ttk.Spinbox(style_box, from_=6.0, to=40.0, increment=0.5, textvariable=legend_fs_var, width=10)
    legend_fs_spin.grid(row=2, column=1, sticky="w", pady=2)

    ttk.Label(style_box, text="Ticks").grid(row=3, column=0, sticky="w", padx=(0, 6), pady=2)
    tick_fs_spin = ttk.Spinbox(style_box, from_=6.0, to=40.0, increment=0.5, textvariable=tick_fs_var, width=10)
    tick_fs_spin.grid(row=3, column=1, sticky="w", pady=2)

    ttk.Label(style_box, text="X ticks").grid(row=4, column=0, sticky="w", padx=(0, 6), pady=2)
    x_tick_spin = ttk.Spinbox(style_box, from_=2, to=20, increment=1, textvariable=x_tick_count_var, width=10)
    x_tick_spin.grid(row=4, column=1, sticky="w", pady=2)

    ttk.Label(style_box, text="Y ticks").grid(row=5, column=0, sticky="w", padx=(0, 6), pady=2)
    y_tick_spin = ttk.Spinbox(style_box, from_=2, to=20, increment=1, textvariable=y_tick_count_var, width=10)
    y_tick_spin.grid(row=5, column=1, sticky="w", pady=2)

    for spin in (
        title_fs_spin,
        label_fs_spin,
        legend_fs_spin,
        tick_fs_spin,
        x_tick_spin,
        y_tick_spin,
    ):
        spin.configure(command=_apply_axes)
        spin.bind("<Return>", lambda _event: _apply_axes())
        spin.bind("<FocusOut>", lambda _event: _apply_axes())

    lim_box = ttk.LabelFrame(ctrl, text="Limites de ejes", padding=8)
    lim_box.pack(fill="x", pady=(0, 10))

    def _limit_row(parent_widget, row_idx: int, label: str, variable: tk.StringVar):
        ttk.Label(parent_widget, text=label, width=6).grid(
            row=row_idx,
            column=0,
            sticky="w",
            padx=(0, 6),
            pady=2,
        )
        entry = ttk.Entry(parent_widget, textvariable=variable, width=12)
        entry.grid(row=row_idx, column=1, sticky="w", pady=2)
        entry.bind("<Return>", lambda _event: _apply_limits())
        entry.bind("<FocusOut>", lambda _event: _apply_limits())
        return entry

    _limit_row(lim_box, 0, "Xmin", xmin_var)
    _limit_row(lim_box, 1, "Xmax", xmax_var)
    _limit_row(lim_box, 2, "Vmin", vmin_var)
    _limit_row(lim_box, 3, "Vmax", vmax_var)
    _limit_row(lim_box, 4, "Tmin", tmin_var)
    _limit_row(lim_box, 5, "Tmax", tmax_var)

    lim_buttons = ttk.Frame(lim_box)
    lim_buttons.grid(row=6, column=0, columnspan=2, sticky="ew", pady=(8, 0))
    ttk.Button(lim_buttons, text="Ajustar todo", command=_fit_all).pack(side="left", expand=True, fill="x", padx=(0, 6))
    ttk.Button(lim_buttons, text="Refresh fields", command=_sync_limit_entries).pack(side="left", expand=True, fill="x")

    ttk.Label(ctrl, textvariable=status_var, wraplength=260, justify="left").pack(anchor="w", fill="x")

    _apply_axes(redraw=False)
    _sync_limit_entries()
    canvas.draw_idle()

    def _on_close() -> None:
        comp.destroy()
        try:
            delattr(parent, "_composer_win_v_vs_i")
        except Exception:
            pass

    comp.protocol("WM_DELETE_WINDOW", _on_close)


def open_v_vs_i_window(
    input_dir: Path,
    font_defaults: PlotFontDefaults | None = None,
    output_dir: Path | None = None,
    language: str = "es",
) -> None:
    bundles = discover_curve_bundles(Path(input_dir))
    if not bundles:
        raise ValueError("No se encontraron curvas de polarizaciÃ³n vÃ¡lidas.")

    global PC_LANGUAGE
    PC_LANGUAGE = _pc_language(language)
    language = PC_LANGUAGE
    font_defaults = resolve_plot_font_defaults(font_defaults)
    font_default_values = font_defaults.as_strings()

    root = tk._default_root
    created_root = False
    if root is None:
        root = tk.Tk()
        root.withdraw()
        created_root = True

    win = tk.Toplevel(root)
    win.title("PC - V vs I")
    win.geometry("1320x820")
    win.configure(
        bg=ttk.Style(win).lookup("App.TFrame", "background")
        or ttk.Style(win).lookup("TFrame", "background")
    )

    notebook = ttk.Notebook(win)
    notebook.pack(fill="both", expand=True, padx=8, pady=8)

    source_contexts: dict[str, dict[str, object]] = {}
    for bundle in bundles:
        context = _build_v_vs_i_tab(
            notebook,
            bundle,
            font_default_values,
            win,
            source_contexts,
            font_defaults,
            output_dir=output_dir,
            language=language,
        )
        source_contexts[str(context["tab_title"])] = context

    def _on_close() -> None:
        existing = getattr(win, "_composer_win_v_vs_i", None)
        if existing is not None and existing.winfo_exists():
            existing.destroy()
        win.destroy()
        if created_root:
            root.destroy()

    win.protocol("WM_DELETE_WINDOW", _on_close)


def open_dv_di_window(
    input_dir: Path,
    font_defaults: PlotFontDefaults | None = None,
    output_dir: Path | None = None,
    language: str = "es",
) -> None:
    bundles = discover_curve_bundles(Path(input_dir))
    if not bundles:
        raise ValueError("No se encontraron curvas de polarización válidas.")

    language = _pc_language(language)
    font_defaults = resolve_plot_font_defaults(font_defaults)
    font_default_values = font_defaults.as_strings()
    bundle = bundles[0]
    default_limits = compute_default_dv_di_limits(
        bundle,
        smoothing_algorithm="Median filter",
        smoothing_window=1,
        logarithmic_y=True,
        use_current_density=True,
    )

    win = tk.Toplevel()
    win.title(f"PC - dV/dI - {bundle.description} #{bundle.curve_id}")
    win.geometry("1200x700")

    controls_host, plot_outer = create_resizable_plot_layout(win, sidebar_width=320)
    controls_frame = _build_scrollable_controls(controls_host)

    toolbar_frame = ttk.Frame(plot_outer)
    toolbar_frame.pack(side="top", fill="x")

    canvas_frame = ttk.Frame(plot_outer)
    canvas_frame.pack(side="top", fill="both", expand=True)

    fig = Figure(figsize=(9, 5.5), dpi=100)

    canvas = FigureCanvasTkAgg(fig, master=canvas_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    toolbar = NavigationToolbar2Tk(canvas, toolbar_frame, pack_toolbar=False)
    toolbar.update()
    toolbar.pack(side="left", fill="x")

    status_var = tk.StringVar(value="Listo.")

    asc_marker_var = tk.StringVar(value="^")
    dsc_marker_var = tk.StringVar(value="v")
    dvdi_line_var = tk.StringVar(value="-")
    logarithmic_y_var = tk.BooleanVar(value=True)
    smoothing_algorithm_var = tk.StringVar(value="Median filter")
    smoothing_window_var = tk.IntVar(value=1)

    asc_var = tk.BooleanVar(value=True)
    dsc_var = tk.BooleanVar(value=True)
    current_density_var = tk.BooleanVar(value=True)
    point_fraction_var = tk.DoubleVar(value=1.0)

    x_min_var = tk.StringVar(value=default_limits["x_min"])
    x_max_var = tk.StringVar(value=default_limits["x_max"])
    dvdi_min_var = tk.StringVar(value=default_limits["dvdi_min"])
    dvdi_max_var = tk.StringVar(value=default_limits["dvdi_max"])
    x_min_lock_var = tk.BooleanVar(value=False)
    x_max_lock_var = tk.BooleanVar(value=False)
    dvdi_min_lock_var = tk.BooleanVar(value=False)
    dvdi_max_lock_var = tk.BooleanVar(value=False)

    x_tick_count_var = tk.IntVar(value=6)
    y_tick_count_var = tk.IntVar(value=6)

    plot_title_var = tk.StringVar(value="")
    show_title_var = tk.BooleanVar(value=True)
    title_fontsize_var = tk.StringVar(value=font_default_values["title"])
    tick_fontsize_var = tk.StringVar(value=font_default_values["tick"])
    label_fontsize_var = tk.StringVar(value=font_default_values["label"])
    legend_fontsize_var = tk.StringVar(value=font_default_values["legend"])
    marker_size_var = tk.StringVar(value="6")
    hollow_markers_var = tk.BooleanVar(value=False)
    line_width_var = tk.StringVar(value="1.5")

    initial_state = {
        "asc": True,
        "dsc": True,
        "current_density": True,
        "fraction": 1.0,
        "x_min": default_limits["x_min"],
        "x_max": default_limits["x_max"],
        "dvdi_min": default_limits["dvdi_min"],
        "dvdi_max": default_limits["dvdi_max"],
        "logarithmic_y": True,
        "smoothing_algorithm": "Median filter",
        "smoothing_window": 1,
        "asc_marker": "^",
        "dsc_marker": "v",
        "dvdi_line": "-",
        "x_tick_count": 6,
        "y_tick_count": 6,
        "plot_title": "",
        "show_title": True,
        "title_fontsize": font_default_values["title"],
        "tick_fontsize": font_default_values["tick"],
        "label_fontsize": font_default_values["label"],
        "legend_fontsize": font_default_values["legend"],
        "marker_size": "6",
        "hollow_markers": False,
        "line_width": "1.5",
        "x_min_lock": False,
        "x_max_lock": False,
        "dvdi_min_lock": False,
        "dvdi_max_lock": False,
    }

    plot_job = {"id": None}
    suspend_events = {"value": False}
    current_density_state = {"value": True}
    limit_entries: dict[str, ttk.Entry] = {}

    def _schedule_plot(*_args):
        if suspend_events["value"]:
            return
        if plot_job["id"] is not None:
            win.after_cancel(plot_job["id"])
        plot_job["id"] = win.after(20, _plot)

    ttk.Label(
        controls_frame,
        text=f"Curva detectada:\n{bundle.description} #{bundle.curve_id}",
        justify="left",
    ).pack(anchor="w", pady=(0, 10))

    series_box = ttk.LabelFrame(controls_frame, text="Series")
    series_box.pack(fill="x", pady=5)

    style_box = ttk.LabelFrame(controls_frame, text="Estilo")
    style_box.pack(fill="x", pady=5)

    smoothing_box = ttk.LabelFrame(controls_frame, text="Suavizado")
    smoothing_box.pack(fill="x", pady=5)

    ttk.Label(style_box, text="Marcador ascendente").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    asc_marker_combo = ttk.Combobox(
        style_box,
        textvariable=asc_marker_var,
        values=MARKER_OPTIONS,
        state="readonly",
        width=10,
    )
    asc_marker_combo.grid(row=0, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Marcador descendente").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    dsc_marker_combo = ttk.Combobox(
        style_box,
        textvariable=dsc_marker_var,
        values=MARKER_OPTIONS,
        state="readonly",
        width=10,
    )
    dsc_marker_combo.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Línea de dV/dI").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    dvdi_line_combo = ttk.Combobox(
        style_box,
        textvariable=dvdi_line_var,
        values=LINESTYLE_OPTIONS,
        state="readonly",
        width=10,
    )
    dvdi_line_combo.grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(smoothing_box, text="Algoritmo").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    smoothing_algo_combo = ttk.Combobox(
        smoothing_box,
        textvariable=smoothing_algorithm_var,
        values=SMOOTHING_ALGORITHMS,
        state="readonly",
        width=16,
    )
    smoothing_algo_combo.grid(row=0, column=1, sticky="w", padx=8, pady=3)

    smoothing_window_label = ttk.Label(smoothing_box, text="Ventana de mediana")
    smoothing_window_label.grid(row=1, column=0, sticky="w", padx=8, pady=3)
    smoothing_window_spin = tk.Spinbox(
        smoothing_box,
        from_=1,
        to=500,
        increment=1,
        textvariable=smoothing_window_var,
        width=8,
    )
    smoothing_window_spin.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    def _update_smoothing_ui(*_args):
        algorithm = smoothing_algorithm_var.get()
        if algorithm == "Rolling average":
            smoothing_window_label.config(text="Ventana de promedio")
        else:
            smoothing_window_label.config(text="Ventana de mediana")
        _schedule_plot()

    point_box = ttk.LabelFrame(controls_frame, text="Punto dentro de cada paso")
    point_box.pack(fill="x", pady=5)

    limits_box = ttk.LabelFrame(controls_frame, text="Límites de ejes")
    limits_box.pack(fill="x", pady=5)

    text_box = ttk.LabelFrame(controls_frame, text="Texto / tamaños")
    text_box.pack(fill="x", pady=5)

    ttk.Label(text_box, text="Título").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    title_entry = ttk.Entry(text_box, textvariable=plot_title_var, width=28)
    title_entry.grid(row=0, column=1, sticky="we", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño del título").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    title_size_entry = ttk.Spinbox(text_box, from_=6.0, to=50.0, increment=0.5, textvariable=title_fontsize_var, width=10)
    title_size_entry.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de ticks").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    tick_size_entry = ttk.Spinbox(text_box, from_=6.0, to=40.0, increment=0.5, textvariable=tick_fontsize_var, width=10)
    tick_size_entry.grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de etiquetas").grid(row=3, column=0, sticky="w", padx=8, pady=3)
    label_size_entry = ttk.Spinbox(text_box, from_=6.0, to=40.0, increment=0.5, textvariable=label_fontsize_var, width=10)
    label_size_entry.grid(row=3, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de leyenda").grid(row=4, column=0, sticky="w", padx=8, pady=3)
    legend_size_entry = ttk.Spinbox(text_box, from_=6.0, to=40.0, increment=0.5, textvariable=legend_fontsize_var, width=10)
    legend_size_entry.grid(row=4, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de marcador").grid(row=5, column=0, sticky="w", padx=8, pady=3)
    marker_size_entry = ttk.Spinbox(text_box, from_=0.0, to=20.0, increment=0.5, textvariable=marker_size_var, width=10)
    marker_size_entry.grid(row=5, column=1, sticky="w", padx=8, pady=3)

    ttk.Checkbutton(
        text_box,
        text="Marcadores vacíos",
        variable=hollow_markers_var,
        command=_schedule_plot,
    ).grid(row=6, column=0, columnspan=2, sticky="w", padx=8, pady=4)

    ttk.Label(text_box, text="Grosor de línea").grid(row=7, column=0, sticky="w", padx=8, pady=3)
    line_width_entry = ttk.Spinbox(text_box, from_=0.0, to=10.0, increment=0.1, textvariable=line_width_var, width=10)
    line_width_entry.grid(row=7, column=1, sticky="w", padx=8, pady=3)
    ttk.Checkbutton(text_box, text="Mostrar titulo", variable=show_title_var, command=_schedule_plot).grid(
        row=8, column=0, columnspan=2, sticky="w", padx=8, pady=3
    )

    def _positive_float(text: str, name: str) -> float:
        value = text.strip().replace(",", ".")
        if not value:
            raise ValueError(f"{name} no puede estar vacío.")
        num = float(value)
        if num <= 0:
            raise ValueError(f"{name} debe ser mayor que 0.")
        return num

    def _collect_limits():
        return dict(
            x_min=_optional_float(x_min_var.get()),
            x_max=_optional_float(x_max_var.get()),
            dvdi_min=_optional_float(dvdi_min_var.get()),
            dvdi_max=_optional_float(dvdi_max_var.get()),
        )

    def _smoothing_config() -> dict[str, object]:
        return {
            "smoothing_algorithm": smoothing_algorithm_var.get(),
            "smoothing_window": smoothing_window_var.get(),
        }

    def _axis_mode_config() -> dict[str, object]:
        return {
            "logarithmic_y": logarithmic_y_var.get(),
        }

    def _apply_lock_states() -> None:
        states = {
            "x_min": x_min_lock_var.get(),
            "x_max": x_max_lock_var.get(),
            "dvdi_min": dvdi_min_lock_var.get(),
            "dvdi_max": dvdi_max_lock_var.get(),
        }
        for key, entry in limit_entries.items():
            entry.state(["disabled"] if states[key] else ["!disabled"])

    def _on_current_density_toggle() -> None:
        new_state = current_density_var.get()
        old_state = current_density_state["value"]
        if new_state == old_state:
            return

        area_cm2 = _bundle_area_cm2(bundle)
        if area_cm2 is None or area_cm2 <= 0:
            current_density_var.set(old_state)
            status_var.set("Error: no se pudo leer un AREA valida para convertir la corriente.")
            return

        current_factor = 1.0 / area_cm2 if new_state else area_cm2
        dvdi_factor = area_cm2 if new_state else 1.0 / area_cm2

        suspend_events["value"] = True
        try:
            for variable in (x_min_var, x_max_var):
                value = _optional_float(variable.get())
                if value is not None:
                    variable.set(f"{value * current_factor:.6g}")
            for variable in (dvdi_min_var, dvdi_max_var):
                value = _optional_float(variable.get())
                if value is not None:
                    variable.set(f"{value * dvdi_factor:.6g}")
        finally:
            suspend_events["value"] = False

        current_density_state["value"] = new_state
        _plot()
        status_var.set("Unidad del eje x actualizada.")

    def _plot():
        plot_job["id"] = None

        try:
            has_plot = draw_dv_di_on_figure(
                fig=fig,
                bundle=bundle,
                show_asc=asc_var.get(),
                show_dsc=dsc_var.get(),
                point_fraction=point_fraction_var.get(),
                **_smoothing_config(),
                **_axis_mode_config(),
                use_current_density=current_density_var.get(),
                asc_marker=asc_marker_var.get(),
                dsc_marker=dsc_marker_var.get(),
                dvdi_linestyle=dvdi_line_var.get(),
                x_tick_count=x_tick_count_var.get(),
                y_tick_count=y_tick_count_var.get(),
                plot_title=plot_title_var.get(),
                show_title=show_title_var.get(),
                title_fontsize=_positive_float(title_fontsize_var.get(), "Tamaño del título"),
                tick_fontsize=_positive_float(tick_fontsize_var.get(), "Tamaño de ticks"),
                label_fontsize=_positive_float(label_fontsize_var.get(), "Tamaño de etiquetas"),
                legend_fontsize=_positive_float(legend_fontsize_var.get(), "Tamaño de leyenda"),
                marker_size=_positive_float(marker_size_var.get(), "Tamaño de marcador"),
                hollow_markers=hollow_markers_var.get(),
                line_width=_positive_float(line_width_var.get(), "Grosor de línea"),
                language=language,
                **_collect_limits(),
            )
        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            return

        if not has_plot:
            fig.clear()
            canvas.draw_idle()
            status_var.set("No se muestra gráfico: seleccione al menos una dirección.")
            return

        canvas.draw_idle()
        status_var.set("Gráfico actualizado.")

    def _update_point_label(_event=None):
        point_value_label.config(text=f"{point_fraction_var.get():.2f}")

    def _on_scale_move(_value=None):
        _update_point_label()
        _plot()

    def _on_scale_release(_event=None):
        _schedule_plot()

    def _autofit():
        try:
            fitted = compute_autofit_dv_di_limits(
                bundle=bundle,
                show_asc=asc_var.get(),
                show_dsc=dsc_var.get(),
                point_fraction=point_fraction_var.get(),
                **_smoothing_config(),
                **_axis_mode_config(),
                use_current_density=current_density_var.get(),
                locked_x_min=_optional_float(x_min_var.get()) if x_min_lock_var.get() else None,
                locked_x_max=_optional_float(x_max_var.get()) if x_max_lock_var.get() else None,
                locked_dvdi_min=_optional_float(dvdi_min_var.get()) if dvdi_min_lock_var.get() else None,
                locked_dvdi_max=_optional_float(dvdi_max_var.get()) if dvdi_max_lock_var.get() else None,
            )
        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            return

        suspend_events["value"] = True
        try:
            if not x_min_lock_var.get():
                x_min_var.set(fitted["x_min"])
            if not x_max_lock_var.get():
                x_max_var.set(fitted["x_max"])
            if not dvdi_min_lock_var.get():
                dvdi_min_var.set(fitted["dvdi_min"])
            if not dvdi_max_lock_var.get():
                dvdi_max_var.set(fitted["dvdi_max"])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Autoescala aplicada.")

    def _reset():
        suspend_events["value"] = True
        try:
            asc_var.set(initial_state["asc"])
            dsc_var.set(initial_state["dsc"])
            current_density_var.set(initial_state["current_density"])
            logarithmic_y_var.set(initial_state["logarithmic_y"])
            smoothing_algorithm_var.set(initial_state["smoothing_algorithm"])
            smoothing_window_var.set(initial_state["smoothing_window"])
            asc_marker_var.set(initial_state["asc_marker"])
            dsc_marker_var.set(initial_state["dsc_marker"])
            dvdi_line_var.set(initial_state["dvdi_line"])
            point_fraction_var.set(initial_state["fraction"])
            x_tick_count_var.set(initial_state["x_tick_count"])
            y_tick_count_var.set(initial_state["y_tick_count"])
            plot_title_var.set(initial_state["plot_title"])
            show_title_var.set(initial_state["show_title"])
            title_fontsize_var.set(initial_state["title_fontsize"])
            tick_fontsize_var.set(initial_state["tick_fontsize"])
            label_fontsize_var.set(initial_state["label_fontsize"])
            legend_fontsize_var.set(initial_state["legend_fontsize"])
            marker_size_var.set(initial_state["marker_size"])
            hollow_markers_var.set(initial_state["hollow_markers"])
            line_width_var.set(initial_state["line_width"])
            x_min_var.set(initial_state["x_min"])
            x_max_var.set(initial_state["x_max"])
            dvdi_min_var.set(initial_state["dvdi_min"])
            dvdi_max_var.set(initial_state["dvdi_max"])
            x_min_lock_var.set(initial_state["x_min_lock"])
            x_max_lock_var.set(initial_state["x_max_lock"])
            dvdi_min_lock_var.set(initial_state["dvdi_min_lock"])
            dvdi_max_lock_var.set(initial_state["dvdi_max_lock"])
            current_density_state["value"] = initial_state["current_density"]
            _apply_lock_states()
            _update_point_label()
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Valores restaurados.")

    asc_marker_combo.bind("<<ComboboxSelected>>", _schedule_plot)
    dsc_marker_combo.bind("<<ComboboxSelected>>", _schedule_plot)
    dvdi_line_combo.bind("<<ComboboxSelected>>", _schedule_plot)
    def _collect_report_plot_kwargs() -> dict[str, object]:
        return {
            "show_asc": bool(asc_var.get()),
            "show_dsc": bool(dsc_var.get()),
            "point_fraction": float(point_fraction_var.get()),
            **_smoothing_config(),
            **_axis_mode_config(),
            "use_current_density": bool(current_density_var.get()),
            "asc_marker": asc_marker_var.get(),
            "dsc_marker": dsc_marker_var.get(),
            "dvdi_linestyle": dvdi_line_var.get(),
            "x_tick_count": int(x_tick_count_var.get()),
            "y_tick_count": int(y_tick_count_var.get()),
            "plot_title": plot_title_var.get(),
            "show_title": bool(show_title_var.get()),
            "title_fontsize": _positive_float(title_fontsize_var.get(), "Tamaño del título"),
            "tick_fontsize": _positive_float(tick_fontsize_var.get(), "Tamaño de ticks"),
            "label_fontsize": _positive_float(label_fontsize_var.get(), "Tamaño de etiquetas"),
            "legend_fontsize": _positive_float(legend_fontsize_var.get(), "Tamaño de leyenda"),
            "marker_size": _positive_float(marker_size_var.get(), "Tamaño de marcador"),
            "hollow_markers": bool(hollow_markers_var.get()),
            "line_width": _positive_float(line_width_var.get(), "Grosor de línea"),
            "language": language,
            **_collect_limits(),
        }

    def _export_report() -> None:
        try:
            initial_dir = Path(output_dir) if output_dir is not None else Path.cwd()
            initial_dir.mkdir(parents=True, exist_ok=True)
            default_name = f"PC_dVdI_Report_{_safe_filename_part(bundle.description)}_#{bundle.curve_id}.pdf"
            path_text = filedialog.asksaveasfilename(
                title="Guardar reporte PC dV/dI",
                initialdir=str(initial_dir),
                initialfile=default_name,
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            )
            if not path_text:
                status_var.set("Exportacion de reporte cancelada.")
                return

            exported_path = export_pc_dv_di_report_pdf(
                bundle=bundle,
                output_path=Path(path_text),
                plot_kwargs=_collect_report_plot_kwargs(),
            )
        except Exception as exc:
            status_var.set(f"Error al exportar reporte: {type(exc).__name__}: {exc}")
            messagebox.showerror("PC dV/dI Report", str(exc))
            return

        status_var.set(f"Reporte exportado: {exported_path}")

    smoothing_algo_combo.bind("<<ComboboxSelected>>", _update_smoothing_ui)
    for widget in (
        title_entry,
        title_size_entry,
        tick_size_entry,
        label_size_entry,
        legend_size_entry,
        marker_size_entry,
        line_width_entry,
    ):
        widget.bind("<Return>", _schedule_plot)
        widget.bind("<KP_Enter>", _schedule_plot)
        widget.bind("<FocusOut>", _schedule_plot)
    for spin in (
        title_size_entry,
        tick_size_entry,
        label_size_entry,
        legend_size_entry,
        marker_size_entry,
        line_width_entry,
    ):
        spin.configure(command=_schedule_plot)

    ttk.Checkbutton(series_box, text="Ascendente", variable=asc_var, command=_schedule_plot).pack(
        anchor="w", padx=8, pady=2
    )
    ttk.Checkbutton(series_box, text="Descendente", variable=dsc_var, command=_schedule_plot).pack(
        anchor="w", padx=8, pady=2
    )
    ttk.Checkbutton(
        series_box,
        text="x-axis A/cm^2",
        variable=current_density_var,
        command=_on_current_density_toggle,
    ).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="Eje y logarítmico", variable=logarithmic_y_var, command=_schedule_plot).pack(
        anchor="w", padx=8, pady=2
    )

    point_value_label = ttk.Label(point_box, text="1.00")
    point_value_label.pack(anchor="e", padx=8, pady=(4, 0))

    point_scale = ttk.Scale(
        point_box,
        from_=0.0,
        to=1.0,
        orient="horizontal",
        variable=point_fraction_var,
        command=_on_scale_move,
    )
    point_scale.pack(fill="x", padx=8, pady=6)

    ttk.Label(point_box, text="0 = primer punto, 1 = último punto").pack(
        anchor="w", padx=8, pady=(0, 6)
    )

    limit_specs = [
        ("I min", "x_min", x_min_var, x_min_lock_var),
        ("I max", "x_max", x_max_var, x_max_lock_var),
        ("dV/dI min", "dvdi_min", dvdi_min_var, dvdi_min_lock_var),
        ("dV/dI max", "dvdi_max", dvdi_max_var, dvdi_max_lock_var),
    ]

    for row_idx, (label, key, var, lock_var) in enumerate(limit_specs):
        ttk.Label(limits_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        entry = ttk.Entry(limits_box, textvariable=var, width=12)
        entry.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        entry.bind("<KP_Enter>", _schedule_plot)
        entry.bind("<FocusOut>", _schedule_plot)
        entry.bind("<Return>", _schedule_plot)
        limit_entries[key] = entry
        ttk.Checkbutton(
            limits_box,
            text="Lock",
            variable=lock_var,
            command=_apply_lock_states,
        ).grid(row=row_idx, column=2, sticky="w", padx=8, pady=3)

    ttk.Label(limits_box, text="x-Ticks").grid(row=len(limit_specs), column=0, sticky="w", padx=8, pady=3)
    x_tick_spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=x_tick_count_var, width=8)
    x_tick_spin.grid(row=len(limit_specs), column=1, sticky="w", padx=8, pady=3)

    ttk.Label(limits_box, text="y-Ticks").grid(row=len(limit_specs) + 1, column=0, sticky="w", padx=8, pady=3)
    y_tick_spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=y_tick_count_var, width=8)
    y_tick_spin.grid(row=len(limit_specs) + 1, column=1, sticky="w", padx=8, pady=3)

    for spin in (smoothing_window_spin, x_tick_spin, y_tick_spin):
        spin.bind("<Return>", _schedule_plot)
        spin.bind("<FocusOut>", _schedule_plot)
        spin.config(command=_schedule_plot)

    ttk.Label(
        controls_frame,
        textvariable=status_var,
        wraplength=260,
        justify="left",
    ).pack(anchor="w", fill="x", pady=(10, 10))

    buttons_frame = ttk.Frame(controls_frame)
    buttons_frame.pack(fill="x", pady=(5, 0))

    ttk.Button(buttons_frame, text="Restablecer", command=_reset).pack(side="left", padx=(0, 6))
    ttk.Button(buttons_frame, text="Autoescala", command=_autofit).pack(side="left")
    ttk.Button(buttons_frame, text="Reporte PDF", command=_export_report).pack(side="left", padx=(6, 0))

    point_scale.bind("<ButtonRelease-1>", _on_scale_release)
    point_scale.bind("<B1-Motion>", _on_scale_move)

    _apply_lock_states()
    _update_smoothing_ui()
    _update_point_label()
    _plot()
# ---------------------------------------------------------------------------
# Stable-point helper
# ---------------------------------------------------------------------------
def find_last_point_of_each_step(
    rows: list[dict[str, float]],
    current_tolerance: float,
) -> list[dict[str, float]]:
    """Return the last point of each current plateau."""
    if not rows:
        return []

    stable_rows: list[dict[str, float]] = []
    plateau_start = 0
    plateau_current = rows[0]["Corriente"]
    step_number = 1

    for idx in range(1, len(rows)):
        current = rows[idx]["Corriente"]
        if abs(current - plateau_current) > current_tolerance:
            last_row = dict(rows[idx - 1])
            last_row["Step"] = float(step_number)
            stable_rows.append(last_row)

            plateau_start = idx
            plateau_current = rows[plateau_start]["Corriente"]
            step_number += 1
        else:
            span = idx - plateau_start + 1
            plateau_current = ((plateau_current * (span - 1)) + current) / span

    last_row = dict(rows[-1])
    last_row["Step"] = float(step_number)
    stable_rows.append(last_row)
    return stable_rows


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
def _write_metadata_sheet(ws, metadata_rows: list[tuple[str, object, str]]) -> None:
    ws.title = "Metadata"
    ws["A1"] = "Campo"
    ws["B1"] = "Valor"
    ws["C1"] = "Unidad"
    for ref in ("A1", "B1", "C1"):
        ws[ref].font = Font(bold=True)

    for row_idx, (field, value, unit) in enumerate(metadata_rows, start=2):
        ws.cell(row=row_idx, column=1, value=field)
        ws.cell(row=row_idx, column=2, value=value)
        ws.cell(row=row_idx, column=3, value=unit)

    ws.freeze_panes = "A2"


def _write_data_sheet(ws, rows: list[dict[str, float]], include_step: bool = False) -> None:
    headers = [label for _src, label, _unit in DATA_EXPORT]
    units = [unit for _src, _label, unit in DATA_EXPORT]

    if include_step:
        headers = ["Step"] + headers
        units = [""] + units

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for col_num, unit in enumerate(units, start=1):
        ws.cell(row=2, column=col_num, value=unit)

    for row_num, record in enumerate(rows, start=3):
        for col_num, header in enumerate(headers, start=1):
            value = record.get(header, "")
            if header in {"Pt", "Step"} and value != "":
                value = int(value)
            ws.cell(row=row_num, column=col_num, value=value)

    ws.freeze_panes = "A3"


def _auto_format_sheet(ws) -> None:
    for col_num in range(1, ws.max_column + 1):
        max_len = 0
        for row_num in range(1, min(ws.max_row, 100) + 1):
            value = ws.cell(row=row_num, column=col_num).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_num)].width = min(max(10, max_len + 2), 45)

    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top")


def export_curve_bundle(bundle: CurveBundle, out_path: Path) -> None:
    metadata_rows = build_metadata(bundle)
    asc_rows = concatenate_curve_data(bundle.asc_files)
    dsc_rows = concatenate_curve_data(bundle.dsc_files)

    asc_last_rows = find_last_point_of_each_step(
        asc_rows,
        infer_current_tolerance(bundle.asc_files),
    )
    dsc_last_rows = find_last_point_of_each_step(
        dsc_rows,
        infer_current_tolerance(bundle.dsc_files),
    )

    wb = Workbook()
    wb.remove(wb.active)

    ws_meta = wb.create_sheet("Metadata")
    _write_metadata_sheet(ws_meta, metadata_rows)

    ws_asc = wb.create_sheet("Asc")
    _write_data_sheet(ws_asc, asc_rows)

    ws_dsc = wb.create_sheet("Dsc")
    _write_data_sheet(ws_dsc, dsc_rows)

    ws_asc_last = wb.create_sheet("Asc_last")
    _write_data_sheet(ws_asc_last, asc_last_rows, include_step=True)

    ws_dsc_last = wb.create_sheet("Dsc_last")
    _write_data_sheet(ws_dsc_last, dsc_last_rows, include_step=True)

    for ws in (ws_meta, ws_asc, ws_dsc, ws_asc_last, ws_dsc_last):
        _auto_format_sheet(ws)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ---------------------------------------------------------------------------
# Public entry point (GUI-compatible)
# ---------------------------------------------------------------------------
def export_folder(input_dir: Path, output_dir: Path) -> list[Path]:
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)

    bundles = discover_curve_bundles(input_dir)
    exported_files: list[Path] = []

    for bundle in bundles:
        out_name = f"Curva_Polarizacion_{bundle.description}_#{bundle.curve_id}.xlsx"
        out_path = output_dir / out_name
        export_curve_bundle(bundle, out_path)
        exported_files.append(out_path)

    return exported_files


_draw_step_stability_on_figure_linear = draw_step_stability_on_figure


def draw_step_stability_on_figure(*args, voltage_log_scale: bool = True, **kwargs) -> bool:
    has_plot = _draw_step_stability_on_figure_linear(*args, **kwargs)
    if not has_plot or not voltage_log_scale:
        return has_plot

    show_delta_voltage = bool(kwargs.get("show_voltage_delta", kwargs.get("show_delta_voltage", True)))
    if not show_delta_voltage:
        return has_plot

    fig = kwargs.get("fig")
    if fig is None and args:
        fig = args[0]
    bundle = kwargs.get("bundle")
    if bundle is None and len(args) > 1:
        bundle = args[1]
    if fig is None or bundle is None or not fig.axes:
        return has_plot

    use_current_density = bool(kwargs.get("use_current_density", True))
    step_rows = _step_stability_direction_rows(bundle, use_current_density=use_current_density)
    selected_values: list[float] = []
    if isinstance(step_rows, dict):
        row_groups = step_rows.items()
    else:
        row_groups = [("all", step_rows)]
    for direction, rows in row_groups:
        direction_name = str(direction).lower()
        is_ascending = direction_name.startswith("asc")
        is_descending = direction_name.startswith("dsc") or direction_name.startswith("desc")
        if is_ascending and not bool(kwargs.get("show_asc", True)):
            continue
        if is_descending and not bool(kwargs.get("show_dsc", True)):
            continue
        selected_values.extend(row["DeltaV"] for row in rows if "DeltaV" in row)

    positive_values = [value for value in selected_values if value > 0]
    if not positive_values:
        return has_plot

    voltage_axis = fig.axes[0]
    for axis in fig.axes:
        ylabel = axis.get_ylabel().lower()
        if "delta v" in ylabel or "voltage" in ylabel or "voltaje" in ylabel:
            voltage_axis = axis
            break
    data_min = min(positive_values)
    data_max = max(positive_values)
    requested_min = kwargs.get("dv_min")
    requested_max = kwargs.get("dv_max")

    bottom = requested_min if isinstance(requested_min, (int, float)) and requested_min > 0 else data_min / 1.25
    top = requested_max if isinstance(requested_max, (int, float)) and requested_max > bottom else data_max * 1.25
    if bottom <= 0:
        bottom = data_min
    if top <= bottom:
        bottom = data_min / 1.25
        top = data_max * 1.25
    if top <= bottom:
        top = bottom * 10.0

    voltage_axis.set_yscale("log")
    voltage_axis.set_ylim(bottom, top)
    fig.tight_layout()
    return has_plot


def open_step_stability_window(
    input_dir: Path,
    font_defaults: PlotFontDefaults | None = None,
    output_dir: Path | None = None,
    language: str = "es",
    *,
    parent: tk.Misc | None = None,
    bundle: CurveBundle | None = None,
) -> None:
    language = _pc_language(language)

    if bundle is None:
        bundles = discover_curve_bundles(Path(input_dir))
        if not bundles:
            raise ValueError("No se encontraron curvas de polarizacion validas.")
    else:
        bundles = [bundle]

    font_defaults = resolve_plot_font_defaults(font_defaults)
    font_default_values = font_defaults.as_strings()

    if parent is None and len(bundles) > 1:
        root = tk._default_root
        created_root = False
        if root is None:
            root = tk.Tk()
            root.withdraw()
            created_root = True

        win = tk.Toplevel(root)
        win.title("PC - Step Stability")
        win.geometry("1320x820")
        win.configure(
            bg=ttk.Style(win).lookup("App.TFrame", "background")
            or ttk.Style(win).lookup("TFrame", "background")
        )

        notebook = ttk.Notebook(win)
        notebook.pack(fill="both", expand=True, padx=8, pady=8)

        for tab_bundle in bundles:
            tab_title = f"Etapa {tab_bundle.curve_id} - {tab_bundle.description}"
            tab = ttk.Frame(notebook)
            notebook.add(tab, text=tab_title[:28] + ("..." if len(tab_title) > 28 else ""))
            open_step_stability_window(
                input_dir,
                font_defaults=font_defaults,
                output_dir=output_dir,
                language=language,
                parent=tab,
                bundle=tab_bundle,
            )

        def _on_close() -> None:
            win.destroy()
            if created_root:
                root.destroy()

        win.protocol("WM_DELETE_WINDOW", _on_close)
        return

    bundle = bundles[0]
    default_limits = compute_default_step_stability_limits(bundle, use_current_density=True)

    if parent is None:
        win = tk.Toplevel()
        win.title(f"PC - Step Stability - {bundle.description} #{bundle.curve_id}")
        win.geometry("1200x720")
    else:
        win = parent

    controls_host, plot_outer = create_resizable_plot_layout(win, sidebar_width=320)
    controls_frame = _build_scrollable_controls(controls_host)

    toolbar_frame = ttk.Frame(plot_outer)
    toolbar_frame.pack(side="top", fill="x")

    canvas_frame = ttk.Frame(plot_outer)
    canvas_frame.pack(side="top", fill="both", expand=True)

    fig = Figure(figsize=(9, 5.5), dpi=100)
    canvas = FigureCanvasTkAgg(fig, master=canvas_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    toolbar = NavigationToolbar2Tk(canvas, toolbar_frame, pack_toolbar=False)
    toolbar.update()
    toolbar.pack(side="left", fill="x")

    status_var = tk.StringVar(value="Listo.")
    asc_var = tk.BooleanVar(value=True)
    dsc_var = tk.BooleanVar(value=True)
    voltage_delta_var = tk.BooleanVar(value=True)
    temperature_delta_var = tk.BooleanVar(value=True)
    current_density_var = tk.BooleanVar(value=True)
    current_density_state = {"value": True}

    asc_marker_var = tk.StringVar(value="^")
    dsc_marker_var = tk.StringVar(value="v")
    voltage_line_var = tk.StringVar(value="-")
    temperature_line_var = tk.StringVar(value="--")
    marker_size_var = tk.StringVar(value="6")
    line_width_var = tk.StringVar(value="1.5")
    hollow_markers_var = tk.BooleanVar(value=False)

    x_min_var = tk.StringVar(value=default_limits["x_min"])
    x_max_var = tk.StringVar(value=default_limits["x_max"])
    dv_min_var = tk.StringVar(value=default_limits["dv_min"])
    dv_max_var = tk.StringVar(value=default_limits["dv_max"])
    dt_min_var = tk.StringVar(value=default_limits["dt_min"])
    dt_max_var = tk.StringVar(value=default_limits["dt_max"])
    x_tick_count_var = tk.IntVar(value=6)
    y_tick_count_var = tk.IntVar(value=6)

    plot_title_var = tk.StringVar(value="")
    show_title_var = tk.BooleanVar(value=True)
    title_fontsize_var = tk.StringVar(value=font_default_values["title"])
    tick_fontsize_var = tk.StringVar(value=font_default_values["tick"])
    label_fontsize_var = tk.StringVar(value=font_default_values["label"])
    legend_fontsize_var = tk.StringVar(value=font_default_values["legend"])

    initial_state = {
        "asc": True,
        "dsc": True,
        "voltage_delta": True,
        "temperature_delta": True,
        "current_density": True,
        "asc_marker": "^",
        "dsc_marker": "v",
        "voltage_line": "-",
        "temperature_line": "--",
        "marker_size": "6",
        "line_width": "1.5",
        "hollow_markers": False,
        "x_min": default_limits["x_min"],
        "x_max": default_limits["x_max"],
        "dv_min": default_limits["dv_min"],
        "dv_max": default_limits["dv_max"],
        "dt_min": default_limits["dt_min"],
        "dt_max": default_limits["dt_max"],
        "x_tick_count": 6,
        "y_tick_count": 6,
        "plot_title": "",
        "show_title": True,
        "title_fontsize": font_default_values["title"],
        "tick_fontsize": font_default_values["tick"],
        "label_fontsize": font_default_values["label"],
        "legend_fontsize": font_default_values["legend"],
    }

    plot_job = {"id": None}
    suspend_events = {"value": False}

    def _positive_float(text: str, name: str) -> float:
        value = text.strip().replace(",", ".")
        if not value:
            raise ValueError(f"{name} no puede estar vacio.")
        number = float(value)
        if number <= 0:
            raise ValueError(f"{name} debe ser mayor que 0.")
        return number

    def _collect_limits() -> dict[str, float | None]:
        return {
            "x_min": _optional_float(x_min_var.get()),
            "x_max": _optional_float(x_max_var.get()),
            "dv_min": _optional_float(dv_min_var.get()),
            "dv_max": _optional_float(dv_max_var.get()),
            "dt_min": _optional_float(dt_min_var.get()),
            "dt_max": _optional_float(dt_max_var.get()),
        }

    def _schedule_plot(*_args) -> None:
        if suspend_events["value"]:
            return
        if plot_job["id"] is not None:
            win.after_cancel(plot_job["id"])
        plot_job["id"] = win.after(20, _plot)

    def _plot() -> None:
        plot_job["id"] = None
        try:
            has_plot = draw_step_stability_on_figure(
                fig=fig,
                bundle=bundle,
                show_asc=asc_var.get(),
                show_dsc=dsc_var.get(),
                show_voltage_delta=voltage_delta_var.get(),
                show_temperature_delta=temperature_delta_var.get(),
                use_current_density=current_density_var.get(),
                asc_marker=asc_marker_var.get(),
                dsc_marker=dsc_marker_var.get(),
                voltage_linestyle=voltage_line_var.get(),
                temperature_linestyle=temperature_line_var.get(),
                x_tick_count=x_tick_count_var.get(),
                y_tick_count=y_tick_count_var.get(),
                plot_title=plot_title_var.get(),
                show_title=show_title_var.get(),
                title_fontsize=_positive_float(title_fontsize_var.get(), "Tamaño del titulo"),
                tick_fontsize=_positive_float(tick_fontsize_var.get(), "Tamaño de ticks"),
                label_fontsize=_positive_float(label_fontsize_var.get(), "Tamaño de etiquetas"),
                legend_fontsize=_positive_float(legend_fontsize_var.get(), "Tamaño de leyenda"),
                marker_size=_positive_float(marker_size_var.get(), "Tamaño de marcador"),
                hollow_markers=hollow_markers_var.get(),
                line_width=_positive_float(line_width_var.get(), "Grosor de linea"),
                language=language,
                **_collect_limits(),
            )
        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            return

        if not has_plot:
            fig.clear()
            canvas.draw_idle()
            status_var.set("No se muestra grafico: seleccione al menos una direccion y una magnitud.")
            return

        canvas.draw_idle()
        status_var.set("Grafico actualizado.")

    def _autofit() -> None:
        try:
            fitted = compute_autofit_step_stability_limits(
                bundle=bundle,
                show_asc=asc_var.get(),
                show_dsc=dsc_var.get(),
                show_voltage_delta=voltage_delta_var.get(),
                show_temperature_delta=temperature_delta_var.get(),
                use_current_density=current_density_var.get(),
            )
        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            return

        suspend_events["value"] = True
        try:
            x_min_var.set(fitted["x_min"])
            x_max_var.set(fitted["x_max"])
            if fitted["dv_min"] != "" or fitted["dv_max"] != "":
                dv_min_var.set(fitted["dv_min"])
                dv_max_var.set(fitted["dv_max"])
            if fitted["dt_min"] != "" or fitted["dt_max"] != "":
                dt_min_var.set(fitted["dt_min"])
                dt_max_var.set(fitted["dt_max"])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Autoescala aplicada.")

    def _on_current_density_toggle() -> None:
        new_state = current_density_var.get()
        old_state = current_density_state["value"]
        if new_state == old_state:
            return

        area_cm2 = _bundle_area_cm2(bundle)
        if area_cm2 is None or area_cm2 <= 0:
            current_density_var.set(old_state)
            status_var.set("Error: no se pudo leer un AREA valida para convertir la corriente.")
            return

        factor = 1.0 / area_cm2 if new_state else area_cm2
        suspend_events["value"] = True
        try:
            for variable in (x_min_var, x_max_var):
                value = _optional_float(variable.get())
                if value is not None:
                    variable.set(f"{value * factor:.6g}")
        finally:
            suspend_events["value"] = False

        current_density_state["value"] = new_state
        _plot()
        status_var.set("Unidad del eje x actualizada.")

    def _reset() -> None:
        suspend_events["value"] = True
        try:
            asc_var.set(initial_state["asc"])
            dsc_var.set(initial_state["dsc"])
            voltage_delta_var.set(initial_state["voltage_delta"])
            temperature_delta_var.set(initial_state["temperature_delta"])
            current_density_var.set(initial_state["current_density"])
            current_density_state["value"] = initial_state["current_density"]
            asc_marker_var.set(initial_state["asc_marker"])
            dsc_marker_var.set(initial_state["dsc_marker"])
            voltage_line_var.set(initial_state["voltage_line"])
            temperature_line_var.set(initial_state["temperature_line"])
            marker_size_var.set(initial_state["marker_size"])
            line_width_var.set(initial_state["line_width"])
            hollow_markers_var.set(initial_state["hollow_markers"])
            x_min_var.set(initial_state["x_min"])
            x_max_var.set(initial_state["x_max"])
            dv_min_var.set(initial_state["dv_min"])
            dv_max_var.set(initial_state["dv_max"])
            dt_min_var.set(initial_state["dt_min"])
            dt_max_var.set(initial_state["dt_max"])
            x_tick_count_var.set(initial_state["x_tick_count"])
            y_tick_count_var.set(initial_state["y_tick_count"])
            plot_title_var.set(initial_state["plot_title"])
            show_title_var.set(initial_state["show_title"])
            title_fontsize_var.set(initial_state["title_fontsize"])
            tick_fontsize_var.set(initial_state["tick_fontsize"])
            label_fontsize_var.set(initial_state["label_fontsize"])
            legend_fontsize_var.set(initial_state["legend_fontsize"])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Valores restaurados.")

    def _collect_report_plot_kwargs() -> dict[str, object]:
        return {
            "show_asc": bool(asc_var.get()),
            "show_dsc": bool(dsc_var.get()),
            "show_voltage_delta": True,
            "show_temperature_delta": True,
            "use_current_density": bool(current_density_var.get()),
            "asc_marker": asc_marker_var.get(),
            "dsc_marker": dsc_marker_var.get(),
            "voltage_linestyle": voltage_line_var.get(),
            "temperature_linestyle": temperature_line_var.get(),
            "x_tick_count": int(x_tick_count_var.get()),
            "y_tick_count": int(y_tick_count_var.get()),
            "plot_title": plot_title_var.get(),
            "show_title": bool(show_title_var.get()),
            "title_fontsize": _positive_float(title_fontsize_var.get(), "Tamaño del titulo"),
            "tick_fontsize": _positive_float(tick_fontsize_var.get(), "Tamaño de ticks"),
            "label_fontsize": _positive_float(label_fontsize_var.get(), "Tamaño de etiquetas"),
            "legend_fontsize": _positive_float(legend_fontsize_var.get(), "Tamaño de leyenda"),
            "marker_size": _positive_float(marker_size_var.get(), "Tamaño de marcador"),
            "hollow_markers": bool(hollow_markers_var.get()),
            "line_width": _positive_float(line_width_var.get(), "Grosor de linea"),
            "language": language,
            **_collect_limits(),
        }

    def _export_report() -> None:
        try:
            initial_dir = Path(output_dir) if output_dir is not None else Path.cwd()
            initial_dir.mkdir(parents=True, exist_ok=True)
            default_name = f"PC_Step_Stability_Report_{_safe_filename_part(bundle.description)}_#{bundle.curve_id}.pdf"
            path_text = filedialog.asksaveasfilename(
                title="Guardar reporte PC Step Stability",
                initialdir=str(initial_dir),
                initialfile=default_name,
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            )
            if not path_text:
                status_var.set("Exportacion de reporte cancelada.")
                return

            exported_path = export_pc_step_stability_report_pdf(
                bundle=bundle,
                output_path=Path(path_text),
                plot_kwargs=_collect_report_plot_kwargs(),
            )
        except Exception as exc:
            status_var.set(f"Error al exportar reporte: {type(exc).__name__}: {exc}")
            messagebox.showerror("PC Step Stability Report", str(exc))
            return

        status_var.set(f"Reporte exportado: {exported_path}")

    ttk.Label(
        controls_frame,
        text=f"Curva detectada:\nEtapa {bundle.curve_id} - {bundle.description}",
        justify="left",
    ).pack(anchor="w", pady=(0, 10))

    series_box = ttk.LabelFrame(controls_frame, text="Series")
    series_box.pack(fill="x", pady=5)
    for text, variable, command in (
        ("Ascendente", asc_var, _schedule_plot),
        ("Descendente", dsc_var, _schedule_plot),
        ("Delta V", voltage_delta_var, _schedule_plot),
        ("Delta T", temperature_delta_var, _schedule_plot),
        ("x-axis A/cm^2", current_density_var, _on_current_density_toggle),
    ):
        ttk.Checkbutton(series_box, text=text, variable=variable, command=command).pack(anchor="w", padx=8, pady=2)

    style_box = ttk.LabelFrame(controls_frame, text="Estilo")
    style_box.pack(fill="x", pady=5)
    style_specs = [
        ("Marcador ascendente", asc_marker_var, MARKER_OPTIONS),
        ("Marcador descendente", dsc_marker_var, MARKER_OPTIONS),
        ("Linea de Delta V", voltage_line_var, LINESTYLE_OPTIONS),
        ("Linea de Delta T", temperature_line_var, LINESTYLE_OPTIONS),
    ]
    for row_idx, (label, variable, values) in enumerate(style_specs):
        ttk.Label(style_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        combo = ttk.Combobox(style_box, textvariable=variable, values=values, state="readonly", width=10)
        combo.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        combo.bind("<<ComboboxSelected>>", _schedule_plot)
    for row_idx, (label, variable, start, stop, step) in enumerate(
        (
            ("Tamaño de marcador", marker_size_var, 0.0, 20.0, 0.5),
            ("Grosor de linea", line_width_var, 0.0, 10.0, 0.1),
        ),
        start=len(style_specs),
    ):
        ttk.Label(style_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        spin = ttk.Spinbox(style_box, from_=start, to=stop, increment=step, textvariable=variable, width=10)
        spin.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        spin.bind("<Return>", _schedule_plot)
        spin.bind("<KP_Enter>", _schedule_plot)
        spin.bind("<FocusOut>", _schedule_plot)
        spin.configure(command=_schedule_plot)
    ttk.Checkbutton(
        style_box,
        text="Marcadores vacios",
        variable=hollow_markers_var,
        command=_schedule_plot,
    ).grid(row=len(style_specs) + 2, column=0, columnspan=2, sticky="w", padx=8, pady=4)

    text_box = ttk.LabelFrame(controls_frame, text="Texto / tamaños")
    text_box.pack(fill="x", pady=5)
    text_specs = [
        ("Titulo", plot_title_var, None),
        ("Tamaño del titulo", title_fontsize_var, (6.0, 50.0, 0.5)),
        ("Tamaño de ticks", tick_fontsize_var, (6.0, 40.0, 0.5)),
        ("Tamaño de etiquetas", label_fontsize_var, (6.0, 40.0, 0.5)),
        ("Tamaño de leyenda", legend_fontsize_var, (6.0, 40.0, 0.5)),
    ]
    for row_idx, (label, variable, spin_cfg) in enumerate(text_specs):
        ttk.Label(text_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        if spin_cfg is None:
            widget = ttk.Entry(text_box, textvariable=variable, width=28)
        else:
            widget = ttk.Spinbox(
                text_box,
                from_=spin_cfg[0],
                to=spin_cfg[1],
                increment=spin_cfg[2],
                textvariable=variable,
                width=10,
            )
        widget.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        widget.bind("<Return>", _schedule_plot)
        widget.bind("<KP_Enter>", _schedule_plot)
        widget.bind("<FocusOut>", _schedule_plot)
        try:
            widget.configure(command=_schedule_plot)
        except Exception:
            pass
    ttk.Checkbutton(text_box, text="Mostrar titulo", variable=show_title_var, command=_schedule_plot).grid(
        row=len(text_specs), column=0, columnspan=2, sticky="w", padx=8, pady=3
    )

    limits_box = ttk.LabelFrame(controls_frame, text="Limites de ejes")
    limits_box.pack(fill="x", pady=5)
    limit_specs = [
        ("I min", x_min_var),
        ("I max", x_max_var),
        ("Delta V min", dv_min_var),
        ("Delta V max", dv_max_var),
        ("Delta T min", dt_min_var),
        ("Delta T max", dt_max_var),
    ]
    for row_idx, (label, variable) in enumerate(limit_specs):
        ttk.Label(limits_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        entry = ttk.Entry(limits_box, textvariable=variable, width=12)
        entry.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        entry.bind("<Return>", _schedule_plot)
        entry.bind("<KP_Enter>", _schedule_plot)
        entry.bind("<FocusOut>", _schedule_plot)
    for row_idx, (label, variable) in enumerate((("x-Ticks", x_tick_count_var), ("y-Ticks", y_tick_count_var)), start=len(limit_specs)):
        ttk.Label(limits_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=variable, width=8)
        spin.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        spin.bind("<Return>", _schedule_plot)
        spin.bind("<FocusOut>", _schedule_plot)
        spin.config(command=_schedule_plot)

    ttk.Label(controls_frame, textvariable=status_var, wraplength=280, justify="left").pack(
        anchor="w",
        fill="x",
        pady=(10, 10),
    )

    buttons = ttk.Frame(controls_frame)
    buttons.pack(fill="x", pady=(5, 0))
    ttk.Button(buttons, text="Restablecer", command=_reset).pack(side="left", padx=(0, 6))
    ttk.Button(buttons, text="Autoescala", command=_autofit).pack(side="left")
    ttk.Button(buttons, text="Reporte PDF", command=_export_report).pack(side="left", padx=(6, 0))

    _plot()


def _show_pc_stub(title: str) -> None:
    messagebox.showinfo("PC", f"{title} aún no está implementado.")


def run_pipeline(
    input_dir: Path,
    output_dir: Path,
    selected_options: list[str] | None = None,
    font_defaults: PlotFontDefaults | None = None,
    language: str = "es",
) -> list[Path]:
    global PC_LANGUAGE
    PC_LANGUAGE = _pc_language(language)
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)

    exported_files = export_folder(input_dir, output_dir)
    if not exported_files:
        return []

    chosen = set(selected_options or [])

    if "V vs I" in chosen:
        open_v_vs_i_window(
            input_dir,
            font_defaults=font_defaults,
            output_dir=output_dir,
            language=language,
        )

    if "Series by time" in chosen:
        open_series_by_time_window(
            input_dir,
            font_defaults=font_defaults,
            output_dir=output_dir,
            language=language,
        )

    if "dV/dI" in chosen:
        open_dv_di_window(
            input_dir,
            font_defaults=font_defaults,
            output_dir=output_dir,
            language=language,
        )

    if "Step Stability" in chosen:
        open_step_stability_window(
            input_dir,
            font_defaults=font_defaults,
            output_dir=output_dir,
            language=language,
        )

    return exported_files



def open_series_by_time_window(
    input_dir: Path,
    font_defaults: PlotFontDefaults | None = None,
    output_dir: Path | None = None,
    language: str = "es",
    *,
    parent: tk.Misc | None = None,
    bundle: CurveBundle | None = None,
) -> None:
    language = normalize_language(language)
    if bundle is None:
        bundles = discover_curve_bundles(Path(input_dir))
        if not bundles:
            raise ValueError("No se encontraron curvas de polarización válidas.")
    else:
        bundles = [bundle]

    font_defaults = resolve_plot_font_defaults(font_defaults)
    font_default_values = font_defaults.as_strings()

    if parent is None and len(bundles) > 1:
        root = tk._default_root
        created_root = False
        if root is None:
            root = tk.Tk()
            root.withdraw()
            created_root = True

        win = tk.Toplevel(root)
        win.title("PC - Series by time")
        win.geometry("1320x820")
        win.configure(
            bg=ttk.Style(win).lookup("App.TFrame", "background")
            or ttk.Style(win).lookup("TFrame", "background")
        )

        notebook = ttk.Notebook(win)
        notebook.pack(fill="both", expand=True, padx=8, pady=8)

        for tab_bundle in bundles:
            tab_title = f"Etapa {tab_bundle.curve_id} - {tab_bundle.description}"
            tab = ttk.Frame(notebook)
            notebook.add(tab, text=tab_title[:28] + ("..." if len(tab_title) > 28 else ""))
            open_series_by_time_window(
                input_dir,
                font_defaults=font_defaults,
                output_dir=output_dir,
                language=language,
                parent=tab,
                bundle=tab_bundle,
            )

        def _on_close() -> None:
            win.destroy()
            if created_root:
                root.destroy()

        win.protocol("WM_DELETE_WINDOW", _on_close)
        return

    bundle = bundles[0]
    default_time_unit = "min"
    default_limits = compute_default_series_by_time_limits(
        bundle,
        time_unit=default_time_unit,
        use_current_density=True,
    )

    if parent is None:
        win = tk.Toplevel()
        win.title(f"PC - Series by time - {bundle.description} #{bundle.curve_id}")
        win.geometry("1200x720")
    else:
        win = parent

    controls_host, plot_outer = create_resizable_plot_layout(win, sidebar_width=320)
    controls_frame = _build_scrollable_controls(controls_host)

    toolbar_frame = ttk.Frame(plot_outer)
    toolbar_frame.pack(side="top", fill="x")

    canvas_frame = ttk.Frame(plot_outer)
    canvas_frame.pack(side="top", fill="both", expand=True)

    fig = Figure(figsize=(9, 5.5), dpi=100)

    canvas = FigureCanvasTkAgg(fig, master=canvas_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    toolbar = NavigationToolbar2Tk(canvas, toolbar_frame, pack_toolbar=False)
    toolbar.update()
    toolbar.pack(side="left", fill="x")

    status_var = tk.StringVar(value="Listo.")

    voltage_line_var = tk.StringVar(value="-")
    current_line_var = tk.StringVar(value="-.")
    temperature_line_var = tk.StringVar(value="--")

    asc_var = tk.BooleanVar(value=True)
    dsc_var = tk.BooleanVar(value=True)
    voltage_var = tk.BooleanVar(value=True)
    current_var = tk.BooleanVar(value=True)
    temperature_var = tk.BooleanVar(value=True)
    current_density_var = tk.BooleanVar(value=True)
    time_unit_var = tk.StringVar(value=default_time_unit)

    t_min_var = tk.StringVar(value=default_limits["t_min"])
    t_max_var = tk.StringVar(value=default_limits["t_max"])
    v_min_var = tk.StringVar(value=default_limits["v_min"])
    v_max_var = tk.StringVar(value=default_limits["v_max"])
    i_min_var = tk.StringVar(value=default_limits["i_min"])
    i_max_var = tk.StringVar(value=default_limits["i_max"])
    temp_min_var = tk.StringVar(value=default_limits["temp_min"])
    temp_max_var = tk.StringVar(value=default_limits["temp_max"])

    x_tick_count_var = tk.IntVar(value=6)
    y_tick_count_var = tk.IntVar(value=6)

    plot_title_var = tk.StringVar(value="")
    show_title_var = tk.BooleanVar(value=True)
    title_fontsize_var = tk.StringVar(value=font_default_values["title"])
    tick_fontsize_var = tk.StringVar(value=font_default_values["tick"])
    label_fontsize_var = tk.StringVar(value=font_default_values["label"])
    legend_fontsize_var = tk.StringVar(value=font_default_values["legend"])
    marker_size_var = tk.StringVar(value="6")
    line_width_var = tk.StringVar(value="1.5")
    hollow_markers_var = tk.BooleanVar(value=False)

    initial_state = {
        "asc": True,
        "dsc": True,
        "voltage": True,
        "current": True,
        "temperature": True,
        "current_density": True,
        "time_unit": default_time_unit,
        "t_min": default_limits["t_min"],
        "t_max": default_limits["t_max"],
        "v_min": default_limits["v_min"],
        "v_max": default_limits["v_max"],
        "i_min": default_limits["i_min"],
        "i_max": default_limits["i_max"],
        "temp_min": default_limits["temp_min"],
        "temp_max": default_limits["temp_max"],
        "voltage_line": "-",
        "current_line": "-.",
        "temperature_line": "--",
        "x_tick_count": 6,
        "y_tick_count": 6,
        "plot_title": "",
        "show_title": True,
        "title_fontsize": font_default_values["title"],
        "tick_fontsize": font_default_values["tick"],
        "label_fontsize": font_default_values["label"],
        "legend_fontsize": font_default_values["legend"],
        "marker_size": "6",
        "line_width": "1.5",
        "hollow_markers": False,
    }

    def _positive_float(text: str, name: str) -> float:
        value = text.strip().replace(",", ".")
        if not value:
            raise ValueError(f"{name} no puede estar vacío.")
        num = float(value)
        if num <= 0:
            raise ValueError(f"{name} debe ser mayor que 0.")
        return num

    def _collect_limits():
        return dict(
            t_min=_optional_float(t_min_var.get()),
            t_max=_optional_float(t_max_var.get()),
            v_min=_optional_float(v_min_var.get()),
            v_max=_optional_float(v_max_var.get()),
            i_min=_optional_float(i_min_var.get()),
            i_max=_optional_float(i_max_var.get()),
            temp_min=_optional_float(temp_min_var.get()),
            temp_max=_optional_float(temp_max_var.get()),
        )

    def _format_time_limit(value: float | None, time_unit: str) -> str:
        if value is None:
            return ""
        decimals = _time_unit_decimals(time_unit)
        return _format_limit_value(value, decimals)

    def _convert_time_limit_text(value_text: str, from_unit: str, to_unit: str) -> str:
        value = _optional_float(value_text)
        if value is None or from_unit == to_unit:
            return value_text

        seconds = value * _time_unit_scale(from_unit)
        converted = seconds / _time_unit_scale(to_unit)
        return _format_time_limit(converted, to_unit)

    plot_job = {"id": None}
    current_time_unit = {"value": time_unit_var.get()}
    current_density_state = {"value": True}
    suspend_events = {"value": False}

    def _plot():
        plot_job["id"] = None
        try:
            has_plot = draw_series_by_time_on_figure(
                fig=fig,
                bundle=bundle,
                show_asc=asc_var.get(),
                show_dsc=dsc_var.get(),
                show_voltage=voltage_var.get(),
                show_current=current_var.get(),
                show_temperature=temperature_var.get(),
                asc_marker="^",
                dsc_marker="v",
                voltage_linestyle=voltage_line_var.get(),
                current_linestyle=current_line_var.get(),
                temperature_linestyle=temperature_line_var.get(),
                use_current_density=current_density_var.get(),
                time_unit=time_unit_var.get(),
                x_tick_count=x_tick_count_var.get(),
                y_tick_count=y_tick_count_var.get(),
                plot_title=plot_title_var.get(),
                show_title=show_title_var.get(),
                title_fontsize=_positive_float(title_fontsize_var.get(), "Tamaño del título"),
                tick_fontsize=_positive_float(tick_fontsize_var.get(), "Tamaño de ticks"),
                label_fontsize=_positive_float(label_fontsize_var.get(), "Tamaño de etiquetas"),
                legend_fontsize=_positive_float(legend_fontsize_var.get(), "Tamaño de leyenda"),
                marker_size=_positive_float(marker_size_var.get(), "Tamaño de marcador"),
                line_width=_positive_float(line_width_var.get(), "Grosor de línea"),
                hollow_markers=hollow_markers_var.get(),
                language=language,
                **_collect_limits(),
            )
        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            return

        if not has_plot:
            fig.clear()
            canvas.draw_idle()
            status_var.set("No se muestra gráfico: seleccione al menos una dirección y una magnitud.")
            return

        canvas.draw_idle()
        status_var.set("Gráfico actualizado.")

    def _schedule_plot(*_args):
        if suspend_events["value"]:
            return
        if plot_job["id"] is not None:
            win.after_cancel(plot_job["id"])
        plot_job["id"] = win.after(20, _plot)

    def _autofit():
        try:
            fitted = compute_autofit_series_by_time_limits(
                bundle=bundle,
                show_asc=asc_var.get(),
                show_dsc=dsc_var.get(),
                show_voltage=voltage_var.get(),
                show_current=current_var.get(),
                show_temperature=temperature_var.get(),
                time_unit=time_unit_var.get(),
                use_current_density=current_density_var.get(),
            )
        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            return

        suspend_events["value"] = True
        try:
            t_min_var.set(fitted["t_min"])
            t_max_var.set(fitted["t_max"])
            if fitted["v_min"] != "" or fitted["v_max"] != "":
                v_min_var.set(fitted["v_min"])
                v_max_var.set(fitted["v_max"])
            if fitted["i_min"] != "" or fitted["i_max"] != "":
                i_min_var.set(fitted["i_min"])
                i_max_var.set(fitted["i_max"])
            if fitted["temp_min"] != "" or fitted["temp_max"] != "":
                temp_min_var.set(fitted["temp_min"])
                temp_max_var.set(fitted["temp_max"])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Autoescala aplicada.")

    def _on_time_unit_changed(*_args):
        if suspend_events["value"]:
            return

        old_unit = current_time_unit["value"]
        new_unit = time_unit_var.get()

        suspend_events["value"] = True
        try:
            if t_min_var.get().strip():
                t_min_var.set(_convert_time_limit_text(t_min_var.get(), old_unit, new_unit))
            else:
                defaults = compute_default_series_by_time_limits(
                    bundle,
                    time_unit=new_unit,
                    use_current_density=current_density_var.get(),
                )
                t_min_var.set(defaults["t_min"])

            if t_max_var.get().strip():
                t_max_var.set(_convert_time_limit_text(t_max_var.get(), old_unit, new_unit))
            else:
                defaults = compute_default_series_by_time_limits(
                    bundle,
                    time_unit=new_unit,
                    use_current_density=current_density_var.get(),
                )
                t_max_var.set(defaults["t_max"])

            current_time_unit["value"] = new_unit
        finally:
            suspend_events["value"] = False

        _schedule_plot()

    def _on_current_density_toggle():
        new_state = current_density_var.get()
        old_state = current_density_state["value"]
        if new_state == old_state:
            return

        try:
            area_cm2 = _bundle_area_cm2(bundle)
            if area_cm2 is None or area_cm2 <= 0:
                raise ValueError("No se pudo leer un AREA valida de la metadata para convertir la corriente.")
        except ValueError as exc:
            current_density_var.set(old_state)
            status_var.set(f"Error: {exc}")
            return

        factor = 1.0 / area_cm2 if new_state else area_cm2
        suspend_events["value"] = True
        try:
            for var in (i_min_var, i_max_var):
                value = _optional_float(var.get())
                if value is not None:
                    var.set(f"{value * factor:.6g}")
        finally:
            suspend_events["value"] = False

        current_density_state["value"] = new_state
        _plot()
        status_var.set("Unidad de corriente actualizada.")

    def _reset():
        suspend_events["value"] = True
        try:
            asc_var.set(initial_state["asc"])
            dsc_var.set(initial_state["dsc"])
            voltage_var.set(initial_state["voltage"])
            current_var.set(initial_state["current"])
            temperature_var.set(initial_state["temperature"])
            current_density_var.set(initial_state["current_density"])
            current_density_state["value"] = initial_state["current_density"]
            current_time_unit["value"] = initial_state["time_unit"]
            time_unit_var.set(initial_state["time_unit"])

            voltage_line_var.set(initial_state["voltage_line"])
            current_line_var.set(initial_state["current_line"])
            temperature_line_var.set(initial_state["temperature_line"])

            t_min_var.set(initial_state["t_min"])
            t_max_var.set(initial_state["t_max"])
            v_min_var.set(initial_state["v_min"])
            v_max_var.set(initial_state["v_max"])
            i_min_var.set(initial_state["i_min"])
            i_max_var.set(initial_state["i_max"])
            temp_min_var.set(initial_state["temp_min"])
            temp_max_var.set(initial_state["temp_max"])

            x_tick_count_var.set(initial_state["x_tick_count"])
            y_tick_count_var.set(initial_state["y_tick_count"])
            plot_title_var.set(initial_state["plot_title"])
            show_title_var.set(initial_state["show_title"])
            title_fontsize_var.set(initial_state["title_fontsize"])
            tick_fontsize_var.set(initial_state["tick_fontsize"])
            label_fontsize_var.set(initial_state["label_fontsize"])
            legend_fontsize_var.set(initial_state["legend_fontsize"])
            marker_size_var.set(initial_state["marker_size"])
            line_width_var.set(initial_state["line_width"])
            hollow_markers_var.set(initial_state["hollow_markers"])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Valores restaurados.")

    def _collect_report_plot_kwargs() -> dict[str, object]:
        return {
            "show_asc": bool(asc_var.get()),
            "show_dsc": bool(dsc_var.get()),
            "show_voltage": True,
            "show_current": True,
            "show_temperature": True,
            "asc_marker": "^",
            "dsc_marker": "v",
            "voltage_linestyle": voltage_line_var.get(),
            "current_linestyle": current_line_var.get(),
            "temperature_linestyle": temperature_line_var.get(),
            "use_current_density": bool(current_density_var.get()),
            "time_unit": time_unit_var.get(),
            "x_tick_count": int(x_tick_count_var.get()),
            "y_tick_count": int(y_tick_count_var.get()),
            "plot_title": plot_title_var.get(),
            "show_title": bool(show_title_var.get()),
            "title_fontsize": _positive_float(title_fontsize_var.get(), "TamaÃ±o del tÃ­tulo"),
            "tick_fontsize": _positive_float(tick_fontsize_var.get(), "TamaÃ±o de ticks"),
            "label_fontsize": _positive_float(label_fontsize_var.get(), "TamaÃ±o de etiquetas"),
            "legend_fontsize": _positive_float(legend_fontsize_var.get(), "TamaÃ±o de leyenda"),
            "marker_size": _positive_float(marker_size_var.get(), "TamaÃ±o de marcador"),
            "line_width": _positive_float(line_width_var.get(), "Grosor de lÃ­nea"),
            "hollow_markers": bool(hollow_markers_var.get()),
            "language": language,
            "point_fraction": 1.0,
            **_collect_limits(),
        }

    def _export_report() -> None:
        try:
            initial_dir = Path(output_dir) if output_dir is not None else Path.cwd()
            initial_dir.mkdir(parents=True, exist_ok=True)
            default_name = f"PC_Series_by_time_Report_{_safe_filename_part(bundle.description)}_#{bundle.curve_id}.pdf"
            path_text = filedialog.asksaveasfilename(
                title="Guardar reporte PC Series by time",
                initialdir=str(initial_dir),
                initialfile=default_name,
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            )
            if not path_text:
                status_var.set("Exportacion de reporte cancelada.")
                return

            exported_path = export_pc_series_by_time_report_pdf(
                bundle=bundle,
                output_path=Path(path_text),
                plot_kwargs=_collect_report_plot_kwargs(),
            )
        except Exception as exc:
            status_var.set(f"Error al exportar reporte: {type(exc).__name__}: {exc}")
            messagebox.showerror("PC Series by time Report", str(exc))
            return

        status_var.set(f"Reporte exportado: {exported_path}")

    ttk.Label(
        controls_frame,
        text=f"Curva detectada:\n{bundle.description} #{bundle.curve_id}",
        justify="left",
    ).pack(anchor="w", pady=(0, 10))

    series_box = ttk.LabelFrame(controls_frame, text="Series")
    series_box.pack(fill="x", pady=5)

    style_box = ttk.LabelFrame(controls_frame, text="Estilo")
    style_box.pack(fill="x", pady=5)

    text_box = ttk.LabelFrame(controls_frame, text="Texto / tamaños")
    text_box.pack(fill="x", pady=5)

    limits_box = ttk.LabelFrame(controls_frame, text="Límites de ejes")
    limits_box.pack(fill="x", pady=5)

    ttk.Checkbutton(series_box, text="Voltaje", variable=voltage_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="Corriente", variable=current_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="Temperatura", variable=temperature_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(
        series_box,
        text="Corriente en A/cm^2",
        variable=current_density_var,
        command=_on_current_density_toggle,
    ).pack(anchor="w", padx=8, pady=2)
    ttk.Label(series_box, text="Unidad de tiempo").pack(anchor="w", padx=8, pady=(8, 2))
    time_unit_combo = ttk.Combobox(
        series_box,
        textvariable=time_unit_var,
        values=TIME_UNIT_OPTIONS,
        state="readonly",
        width=8,
    )
    time_unit_combo.pack(anchor="w", padx=8, pady=(0, 4))

    ttk.Label(style_box, text="Línea de voltaje").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    voltage_line_combo = ttk.Combobox(style_box, textvariable=voltage_line_var, values=LINESTYLE_OPTIONS, state="readonly", width=10)
    voltage_line_combo.grid(row=0, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Línea de corriente").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    current_line_combo = ttk.Combobox(style_box, textvariable=current_line_var, values=LINESTYLE_OPTIONS, state="readonly", width=10)
    current_line_combo.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Línea de temperatura").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    temperature_line_combo = ttk.Combobox(style_box, textvariable=temperature_line_var, values=LINESTYLE_OPTIONS, state="readonly", width=10)
    temperature_line_combo.grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Título").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    title_entry = ttk.Entry(text_box, textvariable=plot_title_var, width=28)
    title_entry.grid(row=0, column=1, sticky="we", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño del título").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    title_size_entry = ttk.Spinbox(text_box, from_=6.0, to=50.0, increment=0.5, textvariable=title_fontsize_var, width=10)
    title_size_entry.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de ticks").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    tick_size_entry = ttk.Spinbox(text_box, from_=6.0, to=40.0, increment=0.5, textvariable=tick_fontsize_var, width=10)
    tick_size_entry.grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de etiquetas").grid(row=3, column=0, sticky="w", padx=8, pady=3)
    label_size_entry = ttk.Spinbox(text_box, from_=6.0, to=40.0, increment=0.5, textvariable=label_fontsize_var, width=10)
    label_size_entry.grid(row=3, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de leyenda").grid(row=4, column=0, sticky="w", padx=8, pady=3)
    legend_size_entry = ttk.Spinbox(text_box, from_=6.0, to=40.0, increment=0.5, textvariable=legend_fontsize_var, width=10)
    legend_size_entry.grid(row=4, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Grosor de línea").grid(row=5, column=0, sticky="w", padx=8, pady=3)
    line_width_entry = ttk.Spinbox(text_box, from_=0.0, to=10.0, increment=0.1, textvariable=line_width_var, width=10)
    line_width_entry.grid(row=5, column=1, sticky="w", padx=8, pady=3)

    ttk.Checkbutton(text_box, text="Mostrar titulo", variable=show_title_var, command=_schedule_plot).grid(
        row=6, column=0, columnspan=2, sticky="w", padx=8, pady=3
    )

    limit_specs = [
        ("t min", t_min_var),
        ("t max", t_max_var),
        ("V min", v_min_var),
        ("V max", v_max_var),
        ("I min", i_min_var),
        ("I max", i_max_var),
        ("T min", temp_min_var),
        ("T max", temp_max_var),
    ]

    for row_idx, (label, var) in enumerate(limit_specs):
        ttk.Label(limits_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        entry = ttk.Entry(limits_box, textvariable=var, width=12)
        entry.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        entry.bind("<Return>", _schedule_plot)
        entry.bind("<KP_Enter>", _schedule_plot)
        entry.bind("<FocusOut>", _schedule_plot)

    ttk.Label(limits_box, text="x-Ticks").grid(row=len(limit_specs), column=0, sticky="w", padx=8, pady=3)
    x_tick_spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=x_tick_count_var, width=8)
    x_tick_spin.grid(row=len(limit_specs), column=1, sticky="w", padx=8, pady=3)

    ttk.Label(limits_box, text="y-Ticks").grid(row=len(limit_specs) + 1, column=0, sticky="w", padx=8, pady=3)
    y_tick_spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=y_tick_count_var, width=8)
    y_tick_spin.grid(row=len(limit_specs) + 1, column=1, sticky="w", padx=8, pady=3)

    for combo in (voltage_line_combo, current_line_combo, temperature_line_combo):
        combo.bind("<<ComboboxSelected>>", _schedule_plot)

    time_unit_combo.bind("<<ComboboxSelected>>", _on_time_unit_changed)

    for spin in (x_tick_spin, y_tick_spin):
        spin.bind("<Return>", _schedule_plot)
        spin.bind("<FocusOut>", _schedule_plot)
        spin.config(command=_schedule_plot)

    for widget in (
        title_entry,
        title_size_entry,
        tick_size_entry,
        label_size_entry,
        legend_size_entry,
        line_width_entry,
    ):
        widget.bind("<Return>", _schedule_plot)
        widget.bind("<KP_Enter>", _schedule_plot)
        widget.bind("<FocusOut>", _schedule_plot)
    for spin in (
        title_size_entry,
        tick_size_entry,
        label_size_entry,
        legend_size_entry,
        line_width_entry,
    ):
        spin.configure(command=_schedule_plot)

    ttk.Label(
        controls_frame,
        textvariable=status_var,
        wraplength=260,
        justify="left",
    ).pack(anchor="w", fill="x", pady=(10, 10))

    buttons_frame = ttk.Frame(controls_frame)
    buttons_frame.pack(fill="x", pady=(5, 0))

    ttk.Button(buttons_frame, text="Restablecer", command=_reset).pack(side="left", padx=(0, 6))
    ttk.Button(buttons_frame, text="Autoescala", command=_autofit).pack(side="left")
    ttk.Button(buttons_frame, text="Reporte PDF", command=_export_report).pack(side="left", padx=(6, 0))

    _plot()


if __name__ == "__main__":
    repo_dir = Path(__file__).resolve().parents[1]
    input_dir = repo_dir / "data"
    output_dir = repo_dir / "outputs"

    exported = export_folder(input_dir, output_dir)
    if exported:
        print("Archivos exportados:")
        for path in exported:
            print(f" - {path}")
    else:
        print("No se encontraron archivos de polarización para exportar.")
