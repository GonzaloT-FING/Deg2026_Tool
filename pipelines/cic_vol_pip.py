"""Cyclic voltammetry (.DTA) -> Excel (.xlsx) exporter for Gamry CV files."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from statistics import median

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from matplotlib.backends.backend_agg import FigureCanvasAgg
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.ticker import MaxNLocator, StrMethodFormatter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from plot_defaults import (
    PlotFontDefaults,
    apply_x_tick_label_padding,
    ensure_axis_bottom_margin,
    resolve_plot_font_defaults,
)
from ui_layout import create_resizable_plot_layout
from i18n import normalize_language, translate

from pipelines.activ_pip import (
    ACTIV_CYCLE_GRADIENTS,
    ACTIV_PLOT_COLORS,
    LINESTYLE_OPTIONS,
    _axis_right_footprint_px,
    _build_scrollable_controls,
    _build_scrollable_cycle_selector,
    _cycle_gradient_color,
    _draw_cycle_scale_bars,
    _format_limit_value,
    _mpl_linestyle,
    _optional_float,
    _padded_limits,
    _positive_float,
    _round_down_dec,
    _round_up_dec,
    apply_current_axis_scaling,
    apply_temperature_axis_scaling,
    apply_x_edge_ticks,
)

CV_LANGUAGE = "es"


META_FIELDS = [
    ("TITLE", "Técnica"),
    ("DATE", "Fecha"),
    ("TIME", "Hora"),
    ("VLIMIT1", "Vlimit1"),
    ("VLIMIT2", "Vlimit2"),
    ("SCANRATE", "Scanrate"),
    ("STEPSIZE", "Stepsize"),
    ("CYCLES", "Cycles"),
    ("AREA", "Área"),
]

CV_META_LABEL_KEYS = {
    "TITLE": "technique",
    "DATE": "date",
    "TIME": "start_time",
    "VLIMIT1": "v_limit_1",
    "VLIMIT2": "v_limit_2",
    "SCANRATE": "scan_rate",
    "STEPSIZE": "step_size",
    "CYCLES": "cycles",
    "AREA": "area",
}

DATA_EXPORT = [
    ("Pt", "Pt", ""),
    ("T", "time", "s"),
    ("Vf", "Voltaje", "V"),
    ("Im", "Corriente", "A"),
    ("Vu", "Vu", "V"),
    ("Sig", "Sig", "V"),
    ("Temp", "Temperatura", "deg C"),
]

CV_DATA_LABEL_KEYS = {
    "T": "time",
    "Vf": "voltage",
    "Im": "current",
    "Temp": "temperature",
}

CV_FILE_RE = re.compile(r"^Voltametria_ciclica(?:_.+)?\.DTA$", re.IGNORECASE)
CV_STAGE_RE = re.compile(r"#(\d+)(?!.*#\d+)", re.IGNORECASE)


def _cv_language(language: str | None = None) -> str:
    return normalize_language(language or CV_LANGUAGE)


@dataclass
class ParsedDTA:
    meta_values: dict[str, str]
    meta_units: dict[str, str]
    header: list[str]
    units: list[str]
    rows: list[list[str]]


@dataclass
class CVCycleSegment:
    key: str
    cycle: int
    direction: str
    order: int
    rows: list[dict[str, float]]


@dataclass
class CVDataset:
    path: Path
    parsed: ParsedDTA
    cycle_rows: list[list[dict[str, float]]]
    segments: list[CVCycleSegment]
    stage_number: int | None = None

    @property
    def display_name(self) -> str:
        return self.path.stem


def _extract_stage_number(stem: str) -> int | None:
    match = CV_STAGE_RE.search(stem)
    if match is None:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def _dataset_stage_label(dataset: CVDataset) -> str:
    if dataset.stage_number is None:
        return dataset.display_name
    return f"Etapa {dataset.stage_number}"


def _dataset_summary_label(dataset: CVDataset) -> str:
    if dataset.stage_number is None:
        return dataset.display_name
    return f"{_dataset_stage_label(dataset)} - {dataset.display_name}"


def to_float(val: str) -> float | None:
    s = val.strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _metadata_area_cm2(parsed: ParsedDTA) -> float | None:
    area = to_float(parsed.meta_values.get("AREA", ""))
    if area is None or area <= 0:
        return None
    return area


def _current_density_values(rows: list[dict[str, float]], area_cm2: float | None) -> list[float]:
    if area_cm2 is None or area_cm2 <= 0:
        return [row["Corriente"] for row in rows]
    return [row["Corriente"] / area_cm2 for row in rows]


def _current_density_axis_label(language: str | None = None) -> str:
    return f"{translate('current_density', _cv_language(language))} (A/cm^2)"


def _current_magnitude_axis_label(parsed: ParsedDTA, language: str | None = None) -> str:
    if _metadata_area_cm2(parsed) is None:
        return f"{translate('current', _cv_language(language))} (A)"
    return _current_density_axis_label(language)


def _current_limit_decimals(values: list[float]) -> int:
    finite_values = [abs(value) for value in values if value is not None]
    if not finite_values:
        return 1
    max_abs = max(finite_values)
    if max_abs >= 1:
        return 1
    if max_abs >= 0.1:
        return 2
    if max_abs >= 0.01:
        return 3
    if max_abs >= 0.001:
        return 4
    return 6


def _format_current_limit_value(value: float | None, decimals: int) -> str:
    if value is None:
        return ""
    return f"{value:.{decimals}f}"


def _drop_leading_blank(parts: list[str]) -> list[str]:
    if parts and parts[0] == "":
        return parts[1:]
    return parts


def _extract_parenthesized_unit(text: str) -> str:
    matches = re.findall(r"\(([^()]*)\)", text)
    if matches:
        return matches[-1].strip()
    return ""


def _extract_meta_unit(key: str, description: str) -> str:
    unit = _extract_parenthesized_unit(description)
    if unit:
        return unit

    fallback_units = {
        "TITLE": "",
        "DATE": "",
        "TIME": "",
        "VLIMIT1": "V",
        "VLIMIT2": "V",
        "SCANRATE": "mV/s",
        "STEPSIZE": "mV",
        "CYCLES": "#",
        "AREA": "cm^2",
    }
    return fallback_units.get(key, "")


def parse_gamry_dta(path: Path) -> ParsedDTA:
    text = path.read_text(encoding="latin-1", errors="replace")
    lines = text.splitlines()

    meta_values: dict[str, str] = {}
    meta_units: dict[str, str] = {}
    header: list[str] | None = None
    units: list[str] = []
    rows: list[list[str]] = []
    table_started = False

    for line in lines:
        if not table_started:
            if line.startswith("CURVE") and "TABLE" in line:
                table_started = True
                continue

            if not line.strip():
                continue

            parts = line.split("\t")
            if len(parts) >= 3 and parts[0].strip():
                key = parts[0].strip()
                value = parts[2].strip()
                description = " ".join(part.strip() for part in parts[3:] if part.strip())
                meta_values[key] = value
                meta_units[key] = _extract_meta_unit(key, description)
            continue

        if not line.strip():
            continue

        parts = _drop_leading_blank([part.strip() for part in line.rstrip("\r\n").split("\t")])
        if not parts:
            continue

        first = parts[0]
        if header is None:
            if first == "Pt":
                header = parts
            continue

        if not units:
            if first == "#":
                units = parts
            continue

        if re.fullmatch(r"-?\d+", first):
            rows.append(parts)

    if header is None:
        raise ValueError(f"No data header found in {path.name} (expected 'Pt ...')")

    return ParsedDTA(
        meta_values=meta_values,
        meta_units=meta_units,
        header=header,
        units=units,
        rows=rows,
    )


def find_cv_files(input_dir: Path) -> list[Path]:
    return sorted(
        [
            path
            for path in Path(input_dir).iterdir()
            if path.is_file() and path.suffix.lower() == ".dta" and CV_FILE_RE.match(path.name)
        ]
    )


def _column_index(parsed: ParsedDTA, column_name: str) -> int | None:
    try:
        return parsed.header.index(column_name)
    except ValueError:
        return None


def _format_report_value(value: object, digits: int = 6) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if not (-float("inf") < value < float("inf")):
            return ""
        if value == 0:
            return "0"
        abs_value = abs(value)
        if 1e-6 <= abs_value < 1e6:
            return f"{value:.{digits}f}".rstrip("0").rstrip(".")
        return f"{value:.{digits}g}"
    return str(value)


def _build_metadata_rows(parsed: ParsedDTA, language: str | None = None) -> list[tuple[str, object, str]]:
    language = _cv_language(language)
    numeric_meta_keys = {"VLIMIT1", "VLIMIT2", "SCANRATE", "STEPSIZE", "CYCLES", "AREA"}
    metadata_rows: list[tuple[str, object, str]] = []

    for key, label in META_FIELDS:
        raw_value = parsed.meta_values.get(key, "")
        value: object = raw_value
        if key in numeric_meta_keys:
            num = to_float(raw_value)
            value = num if num is not None else raw_value
        label_key = CV_META_LABEL_KEYS.get(key)
        metadata_rows.append((translate(label_key, language) if label_key else label, value, parsed.meta_units.get(key, "")))

    return metadata_rows


def _required_float(row: list[str], idx: int | None, column_name: str) -> float:
    if idx is None or idx >= len(row):
        raise ValueError(f"Falta la columna requerida {column_name!r} en la tabla CURVE.")
    num = to_float(row[idx])
    if num is None:
        raise ValueError(f"No se pudo convertir a numero el valor {row[idx]!r} de la columna {column_name!r}.")
    return num


def _extract_data_rows(parsed: ParsedDTA) -> list[dict[str, float]]:
    idx_map = {source: _column_index(parsed, source) for source, _label, _unit in DATA_EXPORT}
    missing = [source for source, idx in idx_map.items() if idx is None]
    if missing:
        raise ValueError("Faltan columnas requeridas en la tabla CURVE: " + ", ".join(missing))

    rows: list[dict[str, float]] = []
    for raw_row in parsed.rows:
        record: dict[str, float] = {}
        for source, label, _unit in DATA_EXPORT:
            record[label] = _required_float(raw_row, idx_map[source], source)
        rows.append(record)
    return rows


def _extract_cycle_count(parsed: ParsedDTA) -> int:
    raw_cycles = to_float(parsed.meta_values.get("CYCLES", ""))
    if raw_cycles is None:
        return 1
    return max(1, int(round(raw_cycles)))


def _median_positive_step(values: list[float]) -> float | None:
    diffs = [curr - prev for prev, curr in zip(values, values[1:]) if curr > prev]
    if not diffs:
        return None
    return median(diffs)


def _estimate_cycle_point_count(parsed: ParsedDTA, rows: list[dict[str, float]]) -> int:
    cycle_count = _extract_cycle_count(parsed)
    if len(rows) <= 1:
        return 1

    vlimit1 = to_float(parsed.meta_values.get("VLIMIT1", ""))
    vlimit2 = to_float(parsed.meta_values.get("VLIMIT2", ""))
    voltage_span = None
    if vlimit1 is not None and vlimit2 is not None:
        voltage_span = abs(vlimit2 - vlimit1)

    step_size_mv = to_float(parsed.meta_values.get("STEPSIZE", ""))
    if voltage_span is not None and step_size_mv not in (None, 0):
        points = int(round((2.0 * voltage_span * 1000.0) / abs(step_size_mv)))
        if points >= 2:
            return points

    scan_rate_mv_s = to_float(parsed.meta_values.get("SCANRATE", ""))
    time_step = _median_positive_step([row["time"] for row in rows])
    if voltage_span is not None and scan_rate_mv_s not in (None, 0) and time_step not in (None, 0):
        cycle_seconds = (2.0 * voltage_span * 1000.0) / abs(scan_rate_mv_s)
        points = int(round(cycle_seconds / time_step))
        if points >= 2:
            return points

    return max(2, int(round((len(rows) - 1) / cycle_count)))


def _cycle_target_and_direction(parsed: ParsedDTA, rows: list[dict[str, float]]) -> tuple[float, bool]:
    sig_values = [row["Sig"] for row in rows]
    start_reference = to_float(parsed.meta_values.get("VINIT", ""))
    if start_reference is None:
        start_reference = sig_values[0]

    bounds = [value for value in (to_float(parsed.meta_values.get("VLIMIT1", "")), to_float(parsed.meta_values.get("VLIMIT2", ""))) if value is not None]
    if bounds:
        target = min(bounds, key=lambda value: abs(value - start_reference))
        low_bound = min(bounds)
        high_bound = max(bounds)
    else:
        target = start_reference
        low_bound = min(sig_values)
        high_bound = max(sig_values)

    use_minimum = abs(target - low_bound) <= abs(target - high_bound)
    return target, use_minimum


def _boundary_score(
    sig_values: list[float],
    idx: int,
    expected_idx: int,
    target: float,
    use_minimum: bool,
) -> tuple[int, float, int]:
    value = sig_values[idx]
    is_turning_point = False

    if 0 < idx < len(sig_values) - 1:
        prev_value = sig_values[idx - 1]
        next_value = sig_values[idx + 1]
        if use_minimum:
            is_turning_point = value <= prev_value and value <= next_value
        else:
            is_turning_point = value >= prev_value and value >= next_value

    return (
        0 if is_turning_point else 1,
        abs(value - target),
        abs(idx - expected_idx),
    )


def detect_cycle_boundaries(parsed: ParsedDTA, rows: list[dict[str, float]]) -> list[int]:
    cycle_count = _extract_cycle_count(parsed)
    if cycle_count <= 1 or len(rows) <= 1:
        return [0, len(rows) - 1]

    target, use_minimum = _cycle_target_and_direction(parsed, rows)
    cycle_point_count = _estimate_cycle_point_count(parsed, rows)
    sig_values = [row["Sig"] for row in rows]
    last_idx = len(rows) - 1

    search_radius = max(5, min(max(20, cycle_point_count // 5), max(20, len(rows) // max(2, cycle_count * 2))))
    boundaries = [0]

    for cycle_number in range(1, cycle_count + 1):
        expected_idx = min(last_idx, int(round(cycle_number * cycle_point_count)))
        lo = max(boundaries[-1] + 1, expected_idx - search_radius)
        hi = min(last_idx, expected_idx + search_radius)

        if lo > hi:
            lo = boundaries[-1] + 1
            hi = last_idx

        best_idx = min(
            range(lo, hi + 1),
            key=lambda idx: _boundary_score(sig_values, idx, expected_idx, target, use_minimum),
        )
        boundaries.append(best_idx)

    if boundaries[-1] != last_idx:
        last_value = sig_values[last_idx]
        detected_value = sig_values[boundaries[-1]]
        if abs(last_value - target) <= abs(detected_value - target):
            boundaries[-1] = last_idx

    return boundaries


def split_rows_by_cycle(parsed: ParsedDTA, rows: list[dict[str, float]]) -> list[list[dict[str, float]]]:
    boundaries = detect_cycle_boundaries(parsed, rows)
    cycles: list[list[dict[str, float]]] = []

    for cycle_number, (start_idx, end_idx) in enumerate(zip(boundaries, boundaries[1:]), start=1):
        raw_cycle_rows = rows[start_idx : end_idx + 1]
        if not raw_cycle_rows:
            continue

        time_origin = raw_cycle_rows[0]["time"]
        cycle_rows: list[dict[str, float]] = []
        for local_pt, row in enumerate(raw_cycle_rows):
            cycle_row = dict(row)
            cycle_row["Pt"] = float(local_pt)
            cycle_row["time"] = row["time"] - time_origin
            cycle_rows.append(cycle_row)

        cycles.append(cycle_rows)

    return cycles


def _cycle_voltage_targets(parsed: ParsedDTA, cycle_rows: list[dict[str, float]]) -> tuple[float, float]:
    sig_values = [row["Sig"] for row in cycle_rows]
    start_reference = to_float(parsed.meta_values.get("VINIT", ""))
    if start_reference is None:
        start_reference = sig_values[0]

    bounds = [
        value
        for value in (
            to_float(parsed.meta_values.get("VLIMIT1", "")),
            to_float(parsed.meta_values.get("VLIMIT2", "")),
        )
        if value is not None
    ]
    if len(bounds) >= 2:
        start_target = min(bounds, key=lambda value: abs(value - start_reference))
        opposite_target = max(bounds, key=lambda value: abs(value - start_target))
        return start_target, opposite_target

    low_sig = min(sig_values)
    high_sig = max(sig_values)
    start_target = low_sig if abs(start_reference - low_sig) <= abs(start_reference - high_sig) else high_sig
    opposite_target = high_sig if start_target == low_sig else low_sig
    return start_target, opposite_target


def _first_voltage_direction(cycle_rows: list[dict[str, float]]) -> str:
    for prev_row, curr_row in zip(cycle_rows, cycle_rows[1:]):
        delta = curr_row["Sig"] - prev_row["Sig"]
        if delta > 0:
            return "Asc"
        if delta < 0:
            return "Dsc"
    return "Asc"


def split_cycle_rows_by_direction(
    parsed: ParsedDTA,
    cycle_rows: list[dict[str, float]],
) -> list[tuple[str, list[dict[str, float]]]]:
    if not cycle_rows:
        return []
    if len(cycle_rows) == 1:
        return [("Asc", [dict(cycle_rows[0])])]

    _start_target, opposite_target = _cycle_voltage_targets(parsed, cycle_rows)
    first_direction = _first_voltage_direction(cycle_rows)
    midpoint = len(cycle_rows) / 2.0
    turning_idx = min(
        range(len(cycle_rows)),
        key=lambda idx: (abs(cycle_rows[idx]["Sig"] - opposite_target), abs(idx - midpoint)),
    )

    if turning_idx <= 0:
        turning_idx = min(len(cycle_rows) - 1, max(1, len(cycle_rows) // 2))
    elif turning_idx >= len(cycle_rows) - 1:
        turning_idx = max(0, len(cycle_rows) // 2)

    first_rows = [dict(row) for row in cycle_rows[: turning_idx + 1]]
    second_rows = [dict(row) for row in cycle_rows[turning_idx:]]

    if first_direction == "Asc":
        return [("Asc", first_rows), ("Dsc", second_rows)]
    return [("Dsc", first_rows), ("Asc", second_rows)]


def build_cycle_segments(parsed: ParsedDTA, cycle_rows: list[list[dict[str, float]]]) -> list[CVCycleSegment]:
    segments: list[CVCycleSegment] = []
    for cycle_number, rows_in_cycle in enumerate(cycle_rows, start=1):
        for order, (direction, segment_rows) in enumerate(split_cycle_rows_by_direction(parsed, rows_in_cycle)):
            if not segment_rows:
                continue
            segments.append(
                CVCycleSegment(
                    key=f"{direction}_{cycle_number}",
                    cycle=cycle_number,
                    direction=direction,
                    order=order,
                    rows=segment_rows,
                )
            )
    return segments


def load_cv_dataset(path: Path) -> CVDataset:
    parsed = parse_gamry_dta(path)
    data_rows = _extract_data_rows(parsed)
    cycle_rows = split_rows_by_cycle(parsed, data_rows)
    return CVDataset(
        path=path,
        parsed=parsed,
        cycle_rows=cycle_rows,
        segments=build_cycle_segments(parsed, cycle_rows),
        stage_number=_extract_stage_number(path.stem),
    )


def discover_cv_datasets(input_dir: Path) -> list[CVDataset]:
    datasets = [load_cv_dataset(path) for path in find_cv_files(input_dir)]
    datasets.sort(
        key=lambda dataset: (
            dataset.stage_number is None,
            dataset.stage_number if dataset.stage_number is not None else float("inf"),
            dataset.path.stem.lower(),
        )
    )
    return datasets


def build_visible_cycle_i_vs_v_data(
    dataset: CVDataset,
    visible_segment_keys: set[str],
) -> dict[str, object]:
    visible_segments = [segment for segment in dataset.segments if segment.key in visible_segment_keys]
    cycle_ids = sorted({segment.cycle for segment in visible_segments})

    cycles: list[dict[str, object]] = []
    for cycle in cycle_ids:
        cycle_segments = [segment for segment in visible_segments if segment.cycle == cycle]
        cycle_segments.sort(key=lambda item: item.order)

        cycle_rows: list[dict[str, float]] = []
        for segment in cycle_segments:
            cycle_rows.extend(dict(row) for row in segment.rows)

        if cycle_rows:
            cycles.append({"cycle": cycle, "rows": cycle_rows})

    return {"cycles": cycles}


def compute_default_i_vs_v_limits(dataset: CVDataset) -> dict[str, str]:
    plot_data = build_visible_cycle_i_vs_v_data(dataset, {segment.key for segment in dataset.segments})
    rows = [row for cycle in plot_data["cycles"] for row in cycle["rows"]]
    area_cm2 = _metadata_area_cm2(dataset.parsed)

    if not rows:
        return {
            "v_min": "",
            "v_max": "",
            "i_min": "",
            "i_max": "",
            "temp_min": "",
            "temp_max": "",
        }

    v_min, v_max = _padded_limits([row["Voltaje"] for row in rows])
    i_values = _current_density_values(rows, area_cm2)
    i_decimals = _current_limit_decimals(i_values)
    i_min, i_max = _padded_limits(i_values, decimals=i_decimals)
    temp_min, temp_max = _padded_limits([row["Temperatura"] for row in rows])

    return {
        "v_min": _format_limit_value(v_min),
        "v_max": _format_limit_value(v_max),
        "i_min": _format_current_limit_value(i_min, i_decimals),
        "i_max": _format_current_limit_value(i_max, i_decimals),
        "temp_min": _format_limit_value(temp_min),
        "temp_max": _format_limit_value(temp_max),
    }


def compute_autofit_i_vs_v_limits(
    dataset: CVDataset,
    visible_segment_keys: set[str],
    show_current: bool,
    show_temperature: bool,
) -> dict[str, str]:
    if not visible_segment_keys:
        raise ValueError("Debe seleccionar al menos una rampa para usar Autoescala.")
    if not (show_current or show_temperature):
        raise ValueError("Debe seleccionar al menos una magnitud para usar Autoescala.")

    plot_data = build_visible_cycle_i_vs_v_data(dataset, visible_segment_keys=visible_segment_keys)
    rows = [row for cycle in plot_data["cycles"] for row in cycle["rows"]]
    area_cm2 = _metadata_area_cm2(dataset.parsed)
    if not rows:
        raise ValueError("No hay datos validos para ajustar los ejes.")

    out = {
        "v_min": "",
        "v_max": "",
        "i_min": "",
        "i_max": "",
        "temp_min": "",
        "temp_max": "",
    }

    v_values = [row["Voltaje"] for row in rows]
    if v_values:
        out["v_min"] = _format_limit_value(_round_down_dec(min(v_values)), 1)
        out["v_max"] = _format_limit_value(_round_up_dec(max(v_values)), 1)

    if show_current:
        i_values = _current_density_values(rows, area_cm2)
        if i_values:
            i_decimals = _current_limit_decimals(i_values)
            out["i_min"] = _format_current_limit_value(_round_down_dec(min(i_values), i_decimals), i_decimals)
            out["i_max"] = _format_current_limit_value(_round_up_dec(max(i_values), i_decimals), i_decimals)

    if show_temperature:
        temp_values = [row["Temperatura"] for row in rows]
        if temp_values:
            out["temp_min"] = _format_limit_value(_round_down_dec(min(temp_values)), 1)
            out["temp_max"] = _format_limit_value(_round_up_dec(max(temp_values)), 1)

    return out


def _write_metadata_sheet(ws, metadata_rows: list[tuple[str, object, str]], language: str | None = None) -> None:
    language = _cv_language(language)
    ws.title = "Metadata"
    ws["A1"] = translate("name", language)
    ws["B1"] = translate("value", language)
    ws["C1"] = translate("unit", language)

    for ref in ("A1", "B1", "C1"):
        ws[ref].font = Font(bold=True)

    for row_idx, (field, value, unit) in enumerate(metadata_rows, start=2):
        ws.cell(row=row_idx, column=1, value=field)
        ws.cell(row=row_idx, column=2, value=value)
        ws.cell(row=row_idx, column=3, value=unit)

    ws.freeze_panes = "A2"


def _write_data_sheet(ws, rows: list[dict[str, float]], language: str | None = None) -> None:
    language = _cv_language(language)
    headers = [label for _source, label, _unit in DATA_EXPORT]
    display_headers = [
        translate(CV_DATA_LABEL_KEYS[source], language) if source in CV_DATA_LABEL_KEYS else label
        for source, label, _unit in DATA_EXPORT
    ]
    units = [unit for _source, _label, unit in DATA_EXPORT]

    for col_num, header in enumerate(display_headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for col_num, unit in enumerate(units, start=1):
        ws.cell(row=2, column=col_num, value=unit)

    for row_num, record in enumerate(rows, start=3):
        for col_num, header in enumerate(headers, start=1):
            value = record.get(header, "")
            if header == "Pt" and value != "":
                value = int(value)
            ws.cell(row=row_num, column=col_num, value=value)

    ws.freeze_panes = "A3"


def _auto_format_sheet(ws) -> None:
    for col_num in range(1, ws.max_column + 1):
        max_len = 0
        for row_num in range(1, min(ws.max_row, 100) + 1):
            value = ws.cell(row=row_num, column=col_num).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_num)].width = min(max(10, max_len + 2), 45)

    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top")


def draw_i_vs_v_on_figure(
    fig: Figure,
    dataset: CVDataset,
    visible_segment_keys: set[str],
    show_current: bool,
    show_temperature: bool,
    current_linestyle: str,
    temperature_linestyle: str,
    x_tick_count: int = 6,
    y_tick_count: int = 6,
    v_min: float | None = None,
    v_max: float | None = None,
    i_min: float | None = None,
    i_max: float | None = None,
    temp_min: float | None = None,
    temp_max: float | None = None,
    plot_title: str = "",
    show_title: bool = True,
    title_fontsize: float = 14,
    tick_fontsize: float = 10,
    label_fontsize: float = 11,
    legend_fontsize: float = 10,
    legend_scale: float = 1.0,
    color_axes_by_magnitude: bool = False,
    line_width: float = 1.5,
    language: str | None = None,
) -> bool:
    language = _cv_language(language)
    fig.clear()

    if not visible_segment_keys:
        return False
    if not (show_current or show_temperature):
        return False

    plot_data = build_visible_cycle_i_vs_v_data(dataset, visible_segment_keys=visible_segment_keys)
    visible_cycles = plot_data["cycles"]
    if not visible_cycles:
        return False

    cycle_ids = [cycle_data["cycle"] for cycle_data in visible_cycles]
    area_cm2 = _metadata_area_cm2(dataset.parsed)
    x_tick_count = max(2, int(x_tick_count))
    y_tick_count = max(2, int(y_tick_count))

    ax_main = fig.add_subplot(111)
    ax_temp = None

    if show_current and show_temperature:
        ax_temp = ax_main.twinx()
        ax_temp.spines["left"].set_visible(False)
        ax_temp.yaxis.tick_right()
        ax_temp.yaxis.set_label_position("right")

    current_ls = _mpl_linestyle(current_linestyle)
    temp_ls = _mpl_linestyle(temperature_linestyle)

    axis_colors = {
        "current": ACTIV_PLOT_COLORS["current"],
        "temperature": ACTIV_PLOT_COLORS["temperature"],
    }

    for color_idx, cycle_data in enumerate(visible_cycles):
        grouped_rows = cycle_data["rows"]
        x_vals = [row["Voltaje"] for row in grouped_rows]
        current_color = _cycle_gradient_color(ACTIV_CYCLE_GRADIENTS["current"], color_idx, len(cycle_ids))
        temp_color = _cycle_gradient_color(ACTIV_CYCLE_GRADIENTS["temperature"], color_idx, len(cycle_ids))

        if show_current and current_ls != "None":
            ax_main.plot(
                x_vals,
                _current_density_values(grouped_rows, area_cm2),
                color=current_color,
                linestyle=current_ls,
                linewidth=line_width,
            )

        if show_temperature:
            target_ax = ax_temp if ax_temp is not None else ax_main
            if temp_ls != "None":
                target_ax.plot(
                    x_vals,
                    [row["Temperatura"] for row in grouped_rows],
                    color=temp_color,
                    linestyle=temp_ls,
                    linewidth=line_width,
                )

    if not (show_current and current_ls != "None") and not (show_temperature and temp_ls != "None"):
        fig.clear()
        return False

    default_title = f"j vs V - CV {_dataset_stage_label(dataset)}"
    final_title = plot_title.strip() if plot_title.strip() else default_title

    ax_main.set_xlabel(f"{translate('voltage', language)} (V)", fontsize=label_fontsize)
    ax_main.set_title(final_title if show_title else "", fontsize=title_fontsize)
    ax_main.grid(True)
    ax_main.xaxis.set_major_locator(MaxNLocator(nbins=x_tick_count))
    ax_main.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
    ax_main.tick_params(axis="both", labelsize=tick_fontsize)
    apply_x_tick_label_padding(ax_main, tick_fontsize)

    if v_min is not None or v_max is not None:
        ax_main.set_xlim(left=v_min, right=v_max)
        apply_x_edge_ticks(ax_main, v_min, v_max, x_tick_count)

    if show_current:
        current_values = _current_density_values(
            [row for cycle_data in visible_cycles for row in cycle_data["rows"]],
            area_cm2,
        )
        ax_main.set_ylabel(_current_magnitude_axis_label(dataset.parsed, language), fontsize=label_fontsize)
        apply_current_axis_scaling(ax_main, current_values, y_tick_count, i_min, i_max)
        if color_axes_by_magnitude:
            ax_main.yaxis.label.set_color(axis_colors["current"])
            ax_main.tick_params(axis="y", colors=axis_colors["current"], labelsize=tick_fontsize)
            ax_main.spines["left"].set_color(axis_colors["current"])
    else:
        temp_values = [row["Temperatura"] for cycle_data in visible_cycles for row in cycle_data["rows"]]
        ax_main.set_ylabel("Temperatura (C)", fontsize=label_fontsize)
        apply_temperature_axis_scaling(ax_main, temp_values, y_tick_count, temp_min, temp_max)
        if color_axes_by_magnitude:
            ax_main.yaxis.label.set_color(axis_colors["temperature"])
            ax_main.tick_params(axis="y", colors=axis_colors["temperature"], labelsize=tick_fontsize)
            ax_main.spines["left"].set_color(axis_colors["temperature"])

    if ax_temp is not None:
        temp_values = [row["Temperatura"] for cycle_data in visible_cycles for row in cycle_data["rows"]]
        ax_temp.tick_params(axis="y", labelsize=tick_fontsize)
        ax_temp.set_ylabel("Temperatura (C)", fontsize=label_fontsize)
        apply_temperature_axis_scaling(ax_temp, temp_values, y_tick_count, temp_min, temp_max)
        if color_axes_by_magnitude:
            ax_temp.yaxis.label.set_color(axis_colors["temperature"])
            ax_temp.tick_params(axis="y", colors=axis_colors["temperature"], labelsize=tick_fontsize)
            ax_temp.spines["right"].set_color(axis_colors["temperature"])

    legend_rows = sum(
        [
            1 if show_current and current_ls != "None" else 0,
            1 if show_temperature and temp_ls != "None" else 0,
        ]
    )
    if legend_rows:
        bottom_margin = min(0.32, 0.12 + legend_rows * (0.04 * max(0.5, legend_scale)))
        fig.subplots_adjust(right=0.90, bottom=bottom_margin)
        bottom_margin = ensure_axis_bottom_margin(
            fig,
            ax_main,
            bottom_margin,
            tick_fontsize,
            max_bottom_margin=0.38,
        )
        fig.subplots_adjust(right=0.90, bottom=bottom_margin)

        if ax_temp is not None:
            fig.canvas.draw()
            total_right_px = _axis_right_footprint_px(fig, ax_temp)
            outer_pad_px = max(10.0, tick_fontsize)
            fig_width_px = fig.get_size_inches()[0] * fig.dpi
            dynamic_right_margin = 1.0 - ((total_right_px + outer_pad_px) / fig_width_px)
            fig.subplots_adjust(right=min(0.96, max(0.55, dynamic_right_margin)), bottom=bottom_margin)

        _draw_cycle_scale_bars(
            fig,
            cycle_ids,
            False,
            show_current and current_ls != "None",
            show_temperature and temp_ls != "None",
            legend_fontsize,
            legend_scale,
        )

    return True


def _add_cv_report_table(
    ax,
    title: str,
    rows: list[tuple[str, object, str]],
    bbox: list[float],
    language: str | None = None,
) -> None:
    language = _cv_language(language)
    table_width = min(float(bbox[2]), 0.82)
    bbox = [(1.0 - table_width) / 2.0, bbox[1], table_width, bbox[3]]
    ax.text(
        bbox[0],
        bbox[1] + bbox[3] + 0.025,
        title,
        fontsize=13,
        fontweight="bold",
        ha="left",
        va="bottom",
        transform=ax.transAxes,
    )
    table = ax.table(
        cellText=[[_format_report_value(name), _format_report_value(value), _format_report_value(unit)] for name, value, unit in rows],
        colLabels=[
            translate("name", language),
            translate("value", language),
            translate("unit", language),
        ],
        cellLoc="left",
        colLoc="left",
        bbox=bbox,
    )
    table.auto_set_font_size(False)
    table.set_fontsize(8.6)
    for (row_idx, _col_idx), cell in table.get_celld().items():
        cell.set_edgecolor("#b8c0c8")
        cell.set_linewidth(0.5)
        if row_idx == 0:
            cell.set_facecolor("#edf2f7")
            cell.set_text_props(weight="bold")
        else:
            cell.set_facecolor("white")


def _build_cv_report_indicator_rows(
    dataset: CVDataset,
    visible_segment_keys: set[str],
    language: str | None = None,
) -> list[tuple[str, object, str]]:
    language = _cv_language(language)
    plot_data = build_visible_cycle_i_vs_v_data(dataset, visible_segment_keys)
    report_cycles = [
        cycle_data for cycle_data in plot_data["cycles"]
        if int(cycle_data["cycle"]) >= 2
    ]
    rows = [row for cycle_data in report_cycles for row in cycle_data["rows"]]
    if not rows:
        return []

    temp_values = [row["Temperatura"] for row in rows]
    cycle_delta_t = []
    for cycle_data in report_cycles:
        cycle_temp_values = [row["Temperatura"] for row in cycle_data["rows"]]
        if cycle_temp_values:
            cycle_delta_t.append(max(cycle_temp_values) - min(cycle_temp_values))

    def _fmt_sig(value: float | None, sig: int = 6) -> str:
        return _format_report_value(value, digits=sig)

    return [
        (
            translate("average_temperature_report_cycles", language),
            _fmt_sig(sum(temp_values) / len(temp_values)),
            "deg C",
        ),
        (
            translate("maximum_cycle_delta_temperature", language),
            _fmt_sig(max(cycle_delta_t) if cycle_delta_t else None),
            "deg C",
        ),
    ]


def export_i_vs_v_report_pdf(
    dataset: CVDataset,
    output_path: Path,
    *,
    visible_segment_keys: set[str],
    x_tick_count: int = 6,
    y_tick_count: int = 6,
    v_min: float | None = None,
    v_max: float | None = None,
    i_min: float | None = None,
    i_max: float | None = None,
    plot_title: str = "",
    show_title: bool = True,
    title_fontsize: float = 14,
    tick_fontsize: float = 10,
    label_fontsize: float = 11,
    legend_fontsize: float = 10,
    legend_scale: float = 1.0,
    line_width: float = 1.5,
    language: str | None = None,
) -> Path:
    language = _cv_language(language)
    report_segment_keys = {
        segment.key
        for segment in dataset.segments
        if segment.key in visible_segment_keys and segment.cycle >= 2
    }
    if not report_segment_keys:
        raise ValueError("No hay ciclos validos para exportar el reporte CV.")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    source_title = plot_title.strip() if plot_title.strip() else f"j vs V - CV {_dataset_stage_label(dataset)}"
    metadata_rows = _build_metadata_rows(dataset.parsed, language=language)
    indicator_rows = _build_cv_report_indicator_rows(dataset, report_segment_keys, language=language)
    report_limits = compute_autofit_i_vs_v_limits(
        dataset=dataset,
        visible_segment_keys=report_segment_keys,
        show_current=True,
        show_temperature=False,
    )

    with PdfPages(output_path) as pdf:
        plot_fig = Figure(figsize=(9.5, 6.2), dpi=150)
        FigureCanvasAgg(plot_fig)
        has_plot = draw_i_vs_v_on_figure(
            fig=plot_fig,
            dataset=dataset,
            visible_segment_keys=report_segment_keys,
            show_current=True,
            show_temperature=False,
            current_linestyle="-",
            temperature_linestyle="none",
            color_axes_by_magnitude=False,
            x_tick_count=x_tick_count,
            y_tick_count=y_tick_count,
            v_min=_optional_float(report_limits["v_min"]),
            v_max=_optional_float(report_limits["v_max"]),
            i_min=_optional_float(report_limits["i_min"]),
            i_max=_optional_float(report_limits["i_max"]),
            temp_min=None,
            temp_max=None,
            plot_title=source_title,
            show_title=show_title,
            title_fontsize=title_fontsize,
            tick_fontsize=tick_fontsize,
            label_fontsize=label_fontsize,
            legend_fontsize=legend_fontsize,
            legend_scale=legend_scale,
            line_width=line_width,
            language=language,
        )
        if not has_plot:
            raise ValueError("No hay datos validos para exportar el reporte CV.")
        pdf.savefig(plot_fig, bbox_inches="tight")

        table_fig = Figure(figsize=(8.5, 11.0), dpi=150)
        table_ax = table_fig.add_subplot(111)
        table_ax.axis("off")
        table_fig.text(
            0.05,
            0.965,
            translate("cv_i_vs_v_report_title", language, curve=source_title),
            fontsize=18,
            fontweight="bold",
            ha="left",
            va="top",
        )
        _add_cv_report_table(table_ax, translate("metadata", language), metadata_rows, [0.05, 0.53, 0.90, 0.34], language=language)
        _add_cv_report_table(table_ax, translate("cv_indicators", language), indicator_rows, [0.05, 0.32, 0.90, 0.12], language=language)
        pdf.savefig(table_fig, bbox_inches="tight")

    return output_path


def _build_i_vs_v_tab(
    notebook: ttk.Notebook,
    dataset: CVDataset,
    font_default_values: dict[str, str],
    language: str | None = None,
) -> None:
    default_limits = compute_default_i_vs_v_limits(dataset)
    cycle_ids = [cycle_number for cycle_number in range(1, len(dataset.cycle_rows) + 1)]

    tab = ttk.Frame(notebook)
    tab_title = _dataset_stage_label(dataset)
    notebook.add(tab, text=tab_title[:28] + ("..." if len(tab_title) > 28 else ""))

    controls_host, plot_outer = create_resizable_plot_layout(tab, sidebar_width=320)
    controls_frame = _build_scrollable_controls(controls_host)

    toolbar_frame = ttk.Frame(plot_outer)
    toolbar_frame.pack(side="top", fill="x")

    canvas_frame = ttk.Frame(plot_outer)
    canvas_frame.pack(side="top", fill="both", expand=True)

    fig = Figure(figsize=(9, 5.5), dpi=100)
    canvas = FigureCanvasTkAgg(fig, master=canvas_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    toolbar = NavigationToolbar2Tk(canvas, toolbar_frame, pack_toolbar=False)
    toolbar.update()
    toolbar.pack(side="left", fill="x")

    status_var = tk.StringVar(value="Listo.")

    asc_var = tk.BooleanVar(value=True)
    dsc_var = tk.BooleanVar(value=True)
    current_var = tk.BooleanVar(value=True)
    temperature_var = tk.BooleanVar(value=False)
    color_axes_var = tk.BooleanVar(value=False)

    current_line_var = tk.StringVar(value="-")
    temperature_line_var = tk.StringVar(value="--")

    v_min_var = tk.StringVar(value=default_limits["v_min"])
    v_max_var = tk.StringVar(value=default_limits["v_max"])
    i_min_var = tk.StringVar(value=default_limits["i_min"])
    i_max_var = tk.StringVar(value=default_limits["i_max"])
    temp_min_var = tk.StringVar(value=default_limits["temp_min"])
    temp_max_var = tk.StringVar(value=default_limits["temp_max"])

    x_tick_count_var = tk.IntVar(value=6)
    y_tick_count_var = tk.IntVar(value=6)

    plot_title_var = tk.StringVar(value="")
    show_title_var = tk.BooleanVar(value=True)
    title_fontsize_var = tk.StringVar(value=font_default_values["title"])
    tick_fontsize_var = tk.StringVar(value=font_default_values["tick"])
    label_fontsize_var = tk.StringVar(value=font_default_values["label"])
    legend_fontsize_var = tk.StringVar(value=font_default_values["legend"])
    legend_scale_var = tk.StringVar(value="1.0")
    line_width_var = tk.StringVar(value="1.5")

    cycle_vars = {cycle: tk.BooleanVar(value=True) for cycle in cycle_ids}

    initial_state = {
        "asc": True,
        "dsc": True,
        "current": True,
        "temperature": False,
        "color_axes": False,
        "current_line": "-",
        "temperature_line": "--",
        "v_min": default_limits["v_min"],
        "v_max": default_limits["v_max"],
        "i_min": default_limits["i_min"],
        "i_max": default_limits["i_max"],
        "temp_min": default_limits["temp_min"],
        "temp_max": default_limits["temp_max"],
        "x_tick_count": 6,
        "y_tick_count": 6,
        "plot_title": "",
        "show_title": True,
        "title_fontsize": font_default_values["title"],
        "tick_fontsize": font_default_values["tick"],
        "label_fontsize": font_default_values["label"],
        "legend_fontsize": font_default_values["legend"],
        "legend_scale": "1.0",
        "line_width": "1.5",
        "visible_cycles": {cycle: True for cycle in cycle_ids},
    }

    def _visible_segment_keys() -> set[str]:
        visible_cycles = {cycle for cycle, var in cycle_vars.items() if var.get()}
        keys: set[str] = set()
        for segment in dataset.segments:
            if segment.cycle not in visible_cycles:
                continue
            if segment.direction == "Asc" and not asc_var.get():
                continue
            if segment.direction == "Dsc" and not dsc_var.get():
                continue
            keys.add(segment.key)
        return keys

    def _collect_limits():
        return dict(
            v_min=_optional_float(v_min_var.get()),
            v_max=_optional_float(v_max_var.get()),
            i_min=_optional_float(i_min_var.get()),
            i_max=_optional_float(i_max_var.get()),
            temp_min=_optional_float(temp_min_var.get()),
            temp_max=_optional_float(temp_max_var.get()),
        )

    plot_job = {"id": None}
    suspend_events = {"value": False}

    def _schedule_plot(*_args):
        if suspend_events["value"]:
            return
        if plot_job["id"] is not None:
            tab.after_cancel(plot_job["id"])
        plot_job["id"] = tab.after(20, _plot)

    def _plot():
        plot_job["id"] = None
        try:
            has_plot = draw_i_vs_v_on_figure(
                fig=fig,
                dataset=dataset,
                visible_segment_keys=_visible_segment_keys(),
                show_current=current_var.get(),
                show_temperature=temperature_var.get(),
                current_linestyle=current_line_var.get(),
                temperature_linestyle=temperature_line_var.get(),
                color_axes_by_magnitude=color_axes_var.get(),
                x_tick_count=x_tick_count_var.get(),
                y_tick_count=y_tick_count_var.get(),
                plot_title=plot_title_var.get(),
                show_title=show_title_var.get(),
                title_fontsize=_positive_float(title_fontsize_var.get(), "Tamaño del título"),
                tick_fontsize=_positive_float(tick_fontsize_var.get(), "Tamaño de ticks"),
                label_fontsize=_positive_float(label_fontsize_var.get(), "Tamaño de etiquetas"),
                legend_fontsize=_positive_float(legend_fontsize_var.get(), "Tamaño de leyenda"),
                legend_scale=_positive_float(legend_scale_var.get(), "Escala del gradiente"),
                line_width=_positive_float(line_width_var.get(), "Grosor de línea"),
                language=language,
                **_collect_limits(),
            )
        except ValueError as exc:
            fig.clear()
            canvas.draw_idle()
            status_var.set(f"Error: {exc}")
            return

        if not has_plot:
            fig.clear()
            canvas.draw_idle()
            status_var.set("No se muestra grafico: seleccione al menos una rampa y una magnitud.")
            return

        canvas.draw_idle()
        status_var.set("Grafico actualizado.")

    def _autofit():
        try:
            fitted = compute_autofit_i_vs_v_limits(
                dataset=dataset,
                visible_segment_keys=_visible_segment_keys(),
                show_current=current_var.get(),
                show_temperature=temperature_var.get(),
            )
        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            return

        suspend_events["value"] = True
        try:
            v_min_var.set(fitted["v_min"])
            v_max_var.set(fitted["v_max"])
            if fitted["i_min"] != "" or fitted["i_max"] != "":
                i_min_var.set(fitted["i_min"])
                i_max_var.set(fitted["i_max"])
            if fitted["temp_min"] != "" or fitted["temp_max"] != "":
                temp_min_var.set(fitted["temp_min"])
                temp_max_var.set(fitted["temp_max"])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Autoescala aplicada.")

    def _export_report() -> None:
        language_code = _cv_language(language)
        try:
            default_name = f"CV_j_vs_V_Report_{dataset.path.stem}.pdf"
            path_text = filedialog.asksaveasfilename(
                title=translate("save_cv_i_vs_v_report", language_code),
                initialfile=default_name,
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            )
            if not path_text:
                return
            exported_path = export_i_vs_v_report_pdf(
                dataset,
                Path(path_text),
                visible_segment_keys=_visible_segment_keys(),
                x_tick_count=x_tick_count_var.get(),
                y_tick_count=y_tick_count_var.get(),
                v_min=_optional_float(v_min_var.get()),
                v_max=_optional_float(v_max_var.get()),
                i_min=_optional_float(i_min_var.get()),
                i_max=_optional_float(i_max_var.get()),
                plot_title=plot_title_var.get(),
                show_title=show_title_var.get(),
                title_fontsize=_positive_float(title_fontsize_var.get(), "Tamaño del título"),
                tick_fontsize=_positive_float(tick_fontsize_var.get(), "Tamaño de ticks"),
                label_fontsize=_positive_float(label_fontsize_var.get(), "Tamaño de etiquetas"),
                legend_fontsize=_positive_float(legend_fontsize_var.get(), "Tamaño de leyenda"),
                legend_scale=_positive_float(legend_scale_var.get(), "Escala del gradiente"),
                line_width=_positive_float(line_width_var.get(), "Grosor de línea"),
                language=language_code,
            )
        except Exception as exc:
            messagebox.showerror("CV Report", str(exc))
            return
        messagebox.showinfo("CV Report", f"{translate('report_exported', language_code)}:\n{exported_path}")

    def _reset():
        suspend_events["value"] = True
        try:
            asc_var.set(initial_state["asc"])
            dsc_var.set(initial_state["dsc"])
            current_var.set(initial_state["current"])
            temperature_var.set(initial_state["temperature"])
            color_axes_var.set(initial_state["color_axes"])
            current_line_var.set(initial_state["current_line"])
            temperature_line_var.set(initial_state["temperature_line"])
            v_min_var.set(initial_state["v_min"])
            v_max_var.set(initial_state["v_max"])
            i_min_var.set(initial_state["i_min"])
            i_max_var.set(initial_state["i_max"])
            temp_min_var.set(initial_state["temp_min"])
            temp_max_var.set(initial_state["temp_max"])
            x_tick_count_var.set(initial_state["x_tick_count"])
            y_tick_count_var.set(initial_state["y_tick_count"])
            plot_title_var.set(initial_state["plot_title"])
            show_title_var.set(initial_state["show_title"])
            title_fontsize_var.set(initial_state["title_fontsize"])
            tick_fontsize_var.set(initial_state["tick_fontsize"])
            label_fontsize_var.set(initial_state["label_fontsize"])
            legend_fontsize_var.set(initial_state["legend_fontsize"])
            legend_scale_var.set(initial_state["legend_scale"])
            line_width_var.set(initial_state["line_width"])
            for cycle, var in cycle_vars.items():
                var.set(initial_state["visible_cycles"][cycle])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Valores restaurados.")

    ttk.Label(
        controls_frame,
        text=f"Serie detectada:\nCV {_dataset_summary_label(dataset)}",
        justify="left",
        wraplength=260,
    ).pack(anchor="w", pady=(0, 10))

    cycles_box = ttk.LabelFrame(controls_frame, text="Ciclos")
    cycles_box.pack(fill="x", pady=5)

    series_box = ttk.LabelFrame(controls_frame, text="Series")
    series_box.pack(fill="x", pady=5)

    style_box = ttk.LabelFrame(controls_frame, text="Estilo")
    style_box.pack(fill="x", pady=5)

    text_box = ttk.LabelFrame(controls_frame, text="Texto / tamaños")
    text_box.pack(fill="x", pady=5)

    limits_box = ttk.LabelFrame(controls_frame, text="Límites de ejes")
    limits_box.pack(fill="x", pady=5)

    cycles_inner, bind_cycle_scroll = _build_scrollable_cycle_selector(cycles_box, height=120)

    for cycle in cycle_ids:
        cycle_toggle = ttk.Checkbutton(
            cycles_inner,
            text=f"Ciclo #{cycle}",
            variable=cycle_vars[cycle],
            command=_schedule_plot,
        )
        cycle_toggle.pack(anchor="w", padx=4, pady=2)
        bind_cycle_scroll(cycle_toggle)

    ttk.Checkbutton(series_box, text="Ascendente", variable=asc_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="Descendente", variable=dsc_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="Corriente", variable=current_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="Temperatura", variable=temperature_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="y-axes a color", variable=color_axes_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)

    ttk.Label(style_box, text="Línea de corriente").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    current_line_combo = ttk.Combobox(style_box, textvariable=current_line_var, values=LINESTYLE_OPTIONS, state="readonly", width=10)
    current_line_combo.grid(row=0, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Línea de temperatura").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    temperature_line_combo = ttk.Combobox(style_box, textvariable=temperature_line_var, values=LINESTYLE_OPTIONS, state="readonly", width=10)
    temperature_line_combo.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Título").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    title_entry = ttk.Entry(text_box, textvariable=plot_title_var, width=28)
    title_entry.grid(row=0, column=1, sticky="we", padx=8, pady=3)
    ttk.Checkbutton(text_box, text="Mostrar titulo", variable=show_title_var, command=_schedule_plot).grid(
        row=7, column=0, columnspan=2, sticky="w", padx=8, pady=3
    )

    ttk.Label(text_box, text="Tamaño del título").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    title_size_spin = tk.Spinbox(text_box, from_=6, to=30, increment=0.5, textvariable=title_fontsize_var, width=8)
    title_size_spin.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de ticks").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    tick_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=tick_fontsize_var, width=8)
    tick_size_spin.grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de etiquetas").grid(row=3, column=0, sticky="w", padx=8, pady=3)
    label_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=label_fontsize_var, width=8)
    label_size_spin.grid(row=3, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de leyenda").grid(row=4, column=0, sticky="w", padx=8, pady=3)
    legend_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=legend_fontsize_var, width=8)
    legend_size_spin.grid(row=4, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Escala del gradiente").grid(row=5, column=0, sticky="w", padx=8, pady=3)
    legend_scale_spin = tk.Spinbox(text_box, from_=0.5, to=3.0, increment=0.1, textvariable=legend_scale_var, width=8)
    legend_scale_spin.grid(row=5, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Grosor de línea").grid(row=6, column=0, sticky="w", padx=8, pady=3)
    line_width_spin = tk.Spinbox(text_box, from_=0.0, to=10.0, increment=0.1, textvariable=line_width_var, width=8)
    line_width_spin.grid(row=6, column=1, sticky="w", padx=8, pady=3)

    limit_specs = [
        ("V min", v_min_var),
        ("V max", v_max_var),
        ("j min", i_min_var),
        ("j max", i_max_var),
        ("T min", temp_min_var),
        ("T max", temp_max_var),
    ]
    for row_idx, (label, var) in enumerate(limit_specs):
        ttk.Label(limits_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        entry = ttk.Entry(limits_box, textvariable=var, width=12)
        entry.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        entry.bind("<Return>", _schedule_plot)
        entry.bind("<KP_Enter>", _schedule_plot)
        entry.bind("<FocusOut>", _schedule_plot)

    ttk.Label(limits_box, text="x-Ticks").grid(row=len(limit_specs), column=0, sticky="w", padx=8, pady=3)
    x_tick_spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=x_tick_count_var, width=8)
    x_tick_spin.grid(row=len(limit_specs), column=1, sticky="w", padx=8, pady=3)

    ttk.Label(limits_box, text="y-Ticks").grid(row=len(limit_specs) + 1, column=0, sticky="w", padx=8, pady=3)
    y_tick_spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=y_tick_count_var, width=8)
    y_tick_spin.grid(row=len(limit_specs) + 1, column=1, sticky="w", padx=8, pady=3)

    for combo in (current_line_combo, temperature_line_combo):
        combo.bind("<<ComboboxSelected>>", _schedule_plot)

    for widget in (
        title_entry,
        title_size_spin,
        tick_size_spin,
        label_size_spin,
        legend_size_spin,
        legend_scale_spin,
        line_width_spin,
    ):
        widget.bind("<Return>", _schedule_plot)
        widget.bind("<KP_Enter>", _schedule_plot)
        widget.bind("<FocusOut>", _schedule_plot)

    for spin in (title_size_spin, tick_size_spin, label_size_spin, legend_size_spin, legend_scale_spin, line_width_spin, x_tick_spin, y_tick_spin):
        spin.config(command=_schedule_plot)
        spin.bind("<Return>", _schedule_plot)
        spin.bind("<FocusOut>", _schedule_plot)

    ttk.Label(
        controls_frame,
        textvariable=status_var,
        wraplength=240,
        justify="left",
    ).pack(anchor="w", fill="x", pady=(10, 10))

    buttons_frame = ttk.Frame(controls_frame)
    buttons_frame.pack(fill="x", pady=(5, 0))
    ttk.Button(buttons_frame, text="Restablecer", command=_reset).pack(side="left", padx=(0, 6))
    ttk.Button(buttons_frame, text="Autoescala", command=_autofit).pack(side="left")
    ttk.Button(buttons_frame, text=translate("pdf_report", _cv_language(language)), command=_export_report).pack(side="left", padx=(6, 0))

    _plot()


def open_i_vs_v_window(
    input_dir: Path,
    font_defaults: PlotFontDefaults | None = None,
    language: str | None = None,
) -> None:
    language = _cv_language(language)
    datasets = discover_cv_datasets(Path(input_dir))
    if not datasets:
        raise ValueError("No se encontraron archivos de voltametria ciclica validos.")

    font_defaults = resolve_plot_font_defaults(font_defaults)
    font_default_values = font_defaults.as_strings()

    root = tk._default_root
    created_root = False
    if root is None:
        root = tk.Tk()
        root.withdraw()
        created_root = True

    win = tk.Toplevel(root)
    win.title("CV - I vs V")
    win.geometry("1320x820")
    win.configure(
        bg=ttk.Style(win).lookup("App.TFrame", "background")
        or ttk.Style(win).lookup("TFrame", "background")
    )

    notebook = ttk.Notebook(win)
    notebook.pack(fill="both", expand=True, padx=8, pady=8)

    for dataset in datasets:
        _build_i_vs_v_tab(notebook, dataset, font_default_values, language=language)

    def _on_close() -> None:
        win.destroy()
        if created_root:
            root.destroy()

    win.protocol("WM_DELETE_WINDOW", _on_close)


def export_cv_file(source_path: Path, output_path: Path, language: str | None = None) -> None:
    parsed = parse_gamry_dta(source_path)
    metadata_rows = _build_metadata_rows(parsed, language=language)
    data_rows = _extract_data_rows(parsed)
    cycle_rows = split_rows_by_cycle(parsed, data_rows)

    wb = Workbook()
    ws_meta = wb.active
    _write_metadata_sheet(ws_meta, metadata_rows, language=language)

    cycle_sheets = []
    for cycle_number, rows_in_cycle in enumerate(cycle_rows, start=1):
        ws_cycle = wb.create_sheet(f"Cycle_#{cycle_number}")
        _write_data_sheet(ws_cycle, rows_in_cycle, language=language)
        cycle_sheets.append(ws_cycle)

    for ws in (ws_meta, *cycle_sheets):
        _auto_format_sheet(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def export_folder(input_dir: Path, output_dir: Path, language: str | None = None) -> list[Path]:
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)

    exported_files: list[Path] = []
    for cv_file in find_cv_files(input_dir):
        out_path = output_dir / f"{cv_file.stem}.xlsx"
        export_cv_file(cv_file, out_path, language=language)
        exported_files.append(out_path)

    return exported_files


def _show_cv_stub(selected_options: set[str]) -> None:
    pending = [option for option in ("Peak current", "Onset potential", "I vs t") if option in selected_options]
    if not pending:
        return
    messagebox.showinfo("CV", ", ".join(pending) + " aún no está implementado.")


def run_pipeline(
    input_dir: Path,
    output_dir: Path,
    selected_options: list[str] | None = None,
    font_defaults: PlotFontDefaults | None = None,
    language: str = "es",
) -> list[Path]:
    global CV_LANGUAGE
    CV_LANGUAGE = _cv_language(language)
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)

    exported_files = export_folder(input_dir, output_dir, language=CV_LANGUAGE)
    if not exported_files:
        return []

    chosen = set(selected_options or [])

    if "I vs V" in chosen:
        open_i_vs_v_window(input_dir, font_defaults=font_defaults, language=CV_LANGUAGE)

    _show_cv_stub(chosen - {"I vs V"})
    return exported_files
