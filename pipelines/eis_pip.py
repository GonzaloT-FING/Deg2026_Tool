
"""EIS (.DTA) -> Excel (.xlsx) exporter for Gamry Potentiostatic EIS.

What this version does:
  - Finds all .DTA files whose filename contains 'EISPOT'
  - Parses selected metadata fields
  - Parses the ZCURVE table
  - Exports ONE .xlsx per input file with two sheets:
        1) Metadata  -> Campo / Valor / Unidad
        2) Data      -> headers row, units row, then numeric data
  - Optionally creates plots depending on the GUI selection:
        * Nyquist plot        -> Zimag vs Zreal
        * Bode plot           -> Zmod vs Freq, and Zphz vs Freq
        * I vs pt             -> Idc vs Pt
        * T vs pt / T vs t    -> Temp vs Pt

This version is written to match the real structure of the uploaded Gamry files.
"""

from __future__ import annotations

import colorsys
from decimal import Decimal, InvalidOperation
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
import re
import math

from matplotlib.figure import Figure
from matplotlib.ticker import LinearLocator, LogLocator, MaxNLocator, StrMethodFormatter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import tkinter.messagebox as mb
import tkinter as tk
from tkinter import ttk, colorchooser
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

from plot_defaults import (
    PlotFontDefaults,
    apply_plot_font_defaults,
    apply_x_tick_label_padding,
    make_legend_draggable,
    resolve_plot_font_defaults,
)
from ui_layout import create_resizable_plot_layout


# ---------------------------------------------------------------------------
# Labels to export (Spanish-friendly names)
# ---------------------------------------------------------------------------

META_FIELDS = [
    ("TITLE", "Técnica"),
    ("DATE", "Fecha"),
    ("TIME", "Hora"),
    ("VDC", "Vdc"),
    ("FREQINIT", "Frecuencia inicial"),
    ("FREQFINAL", "Frecuencia final"),
    ("PTSPERDEC", "Puntos por década"),
    ("VAC", "Amplitud"),
    ("AREA", "Área"),
]

DATA_MAP = {
    "Pt": "Pt",
    "Freq": "Frecuencia",
    "Zreal": "Zreal",
    "Zimag": "Zimag",
    "Zsig": "Zsig",
    "Zmod": "Zmod",
    "Zphz": "Zphz",
    "Idc": "Idc",
    "Vdc": "Vdc",
    "Temp": "Temperatura",
}

PRE_STAB_META_FIELDS = [
    ("TITLE", "Tecnica"),
    ("DATE", "Fecha"),
    ("TIME", "Hora"),
    ("ISTEP1", "Corriente del paso"),
    ("TSTEP1", "Duracion del paso"),
    ("SAMPLETIME", "Tiempo de muestreo"),
    ("AREA", "Area"),
]

PRE_STAB_DATA_MAP = {
    "Pt": "Pt",
    "T": "time",
    "Vf": "Voltaje",
    "Im": "Corriente",
    "Sig": "Sig",
    "Ach": "Ach",
    "Temp": "Temperatura",
}

EIS_STAGE_MARKER_SEQUENCE = ["o", "s", "^", "D", "v", "P", "X", "<", ">"]
EIS_MARKER_OPTIONS = EIS_STAGE_MARKER_SEQUENCE + [".", "x", "+", "*", "None"]


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def to_float(val: str) -> float | None:
    """Convert Gamry-style numbers (decimal comma) to float."""
    s = val.strip()
    if not s:
        return None

    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _drop_leading_blank(parts: list[str]) -> list[str]:
    """Gamry ZCURVE rows usually start with a leading tab."""
    if parts and parts[0] == "":
        return parts[1:]
    return parts


def _extract_parenthesized_unit(text: str) -> str:
    """Extract the last '(...)' group from a description line."""
    matches = re.findall(r"\(([^()]*)\)", text)
    if matches:
        return matches[-1].strip()
    return ""


def _extract_meta_unit(key: str, description: str) -> str:
    """Extract a clean metadata unit from the descriptive text."""
    unit = _extract_parenthesized_unit(description)
    if unit:
        return unit

    if key == "PTSPERDEC":
        return "puntos/década"

    if key in {"TITLE", "DATE", "TIME"}:
        return ""

    return ""

def _is_incomplete_number(s: str) -> bool:
    s = s.strip()
    if s in {"", "-", "+", ".", "-.", "+."}:
        return True
    if re.fullmatch(r"[+-]?\d+\.", s):
        return True
    if re.fullmatch(r"[+-]?(?:\d+\.?\d*|\.\d+)[eE][+-]?", s):
        return True
    return False


def _build_scrollable_controls(parent) -> tuple[ttk.Frame, ttk.Frame]:
    outer = ttk.Frame(parent, padding=(8, 6))
    outer.pack(fill="both", expand=True)

    canvas = tk.Canvas(outer, highlightthickness=0, borderwidth=0)
    scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
    controls_frame = ttk.Frame(canvas, padding=(0, 0, 6, 0))

    controls_frame.bind(
        "<Configure>",
        lambda _event: canvas.configure(scrollregion=canvas.bbox("all")),
    )

    canvas_window = canvas.create_window((0, 0), window=controls_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    def _resize_controls(event):
        canvas.itemconfigure(canvas_window, width=event.width)

    def _bind_mousewheel(_event):
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

    def _unbind_mousewheel(_event):
        canvas.unbind_all("<MouseWheel>")

    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    canvas.bind("<Configure>", _resize_controls)
    canvas.bind("<Enter>", _bind_mousewheel)
    canvas.bind("<Leave>", _unbind_mousewheel)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    top_frame = ttk.Frame(controls_frame)
    top_frame.pack(fill="x")

    host_frame = ttk.Frame(controls_frame)
    host_frame.pack(fill="both", expand=True, pady=(10, 0))

    return top_frame, host_frame


# ---------------------------------------------------------------------------
# Parsed container
# ---------------------------------------------------------------------------

@dataclass
class ParsedDTA:
    meta_values: dict[str, str]
    meta_units: dict[str, str]
    header: list[str]
    units: list[str]
    rows: list[list[str]]


@dataclass
class EISPlotEntry:
    path: Path
    parsed: ParsedDTA
    display_name: str
    stage_number: int | None = None
    current_label: str | None = None
    voltage_label: str | None = None
    current_value: float | None = None
    nyquist_color: str | None = None
    default_marker: str | None = None


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def parse_gamry_dta(path: Path) -> ParsedDTA:
    """Parse one Gamry .DTA file containing a ZCURVE table."""
    text = path.read_text(encoding="latin-1", errors="replace")
    lines = text.splitlines()

    meta_values: dict[str, str] = {}
    meta_units: dict[str, str] = {}
    header: list[str] | None = None
    units: list[str] = []
    rows: list[list[str]] = []

    table_started = False

    for line in lines:
        if not table_started:
            if line.startswith("ZCURVE") and "TABLE" in line:
                table_started = True
                continue

            if not line.strip():
                continue

            parts = line.split("\t")
            if len(parts) >= 3 and parts[0].strip():
                key = parts[0].strip()
                value = parts[2].strip()
                description = parts[-1].strip() if len(parts) >= 4 else ""

                meta_values[key] = value
                meta_units[key] = _extract_meta_unit(key, description)

            continue

        if not line.strip():
            continue

        parts = _drop_leading_blank(line.rstrip("\r\n").split("\t"))
        if not parts:
            continue

        if header is None:
            if parts[0] == "Pt":
                header = parts
            continue

        if parts[0] == "#":
            units = parts
            continue

        if re.fullmatch(r"-?\d+", parts[0]):
            rows.append(parts)

    if header is None:
        raise ValueError(f"No data header found in {path.name} (expected 'Pt ...')")

    return ParsedDTA(
        meta_values=meta_values,
        meta_units=meta_units,
        header=header,
        units=units,
        rows=rows,
    )


def parse_gamry_curve_dta(path: Path) -> ParsedDTA:
    """Parse one Gamry .DTA file containing a CURVE table."""
    text = path.read_text(encoding="latin-1", errors="replace")
    lines = text.splitlines()

    meta_values: dict[str, str] = {}
    meta_units: dict[str, str] = {}
    header: list[str] | None = None
    units: list[str] = []
    rows: list[list[str]] = []

    table_started = False

    for line in lines:
        if not table_started:
            if line.startswith("CURVE") and "TABLE" in line:
                table_started = True
                continue

            if not line.strip():
                continue

            parts = line.split("\t")
            if len(parts) >= 3 and parts[0].strip():
                key = parts[0].strip()
                value = parts[2].strip()
                description = " ".join(part.strip() for part in parts[3:] if part.strip())
                meta_values[key] = value
                meta_units[key] = _extract_meta_unit(key, description)

            continue

        if not line.strip():
            continue

        parts = _drop_leading_blank([part.strip() for part in line.rstrip("\r\n").split("\t")])
        if not parts:
            continue

        first = parts[0]
        if header is None:
            if first == "Pt":
                header = parts
            continue

        if not units:
            if first == "#":
                units = parts
            continue

        if re.fullmatch(r"-?\d+", first):
            rows.append(parts)

    if header is None:
        raise ValueError(f"No data header found in {path.name} (expected 'Pt ...')")

    return ParsedDTA(
        meta_values=meta_values,
        meta_units=meta_units,
        header=header,
        units=units,
        rows=rows,
    )


# ---------------------------------------------------------------------------
# Data extraction for plotting
# ---------------------------------------------------------------------------

def _column_index(parsed: ParsedDTA, column_name: str) -> int | None:
    try:
        return parsed.header.index(column_name)
    except ValueError:
        return None


def _column_unit(parsed: ParsedDTA, column_name: str) -> str:
    idx = _column_index(parsed, column_name)
    if idx is None or not parsed.units or idx >= len(parsed.units):
        return ""
    unit = parsed.units[idx]
    if unit == "#":
        return ""
    return unit

def _impedance_unit(parsed: ParsedDTA, column_name: str) -> str:
    unit = _column_unit(parsed, column_name).strip()
    if not unit:
        return "ohm"
    if unit.lower() in {"ohm", "ohms"}:
        return "ohm"
    return unit

def _metadata_area_cm2(parsed: ParsedDTA) -> float | None:
    area = to_float(parsed.meta_values.get("AREA", ""))
    if area is None or area <= 0:
        return None
    return area

def _paired_series(
    parsed: ParsedDTA,
    x_name: str,
    y_name: str,
    *,
    require_positive_x: bool = False,
) -> tuple[list[float], list[float]]:
    x_idx = _column_index(parsed, x_name)
    y_idx = _column_index(parsed, y_name)
    if x_idx is None or y_idx is None:
        return [], []

    xs: list[float] = []
    ys: list[float] = []

    for row in parsed.rows:
        if x_idx >= len(row) or y_idx >= len(row):
            continue

        x_val = to_float(row[x_idx])
        y_val = to_float(row[y_idx])

        if x_val is None or y_val is None:
            continue
        if require_positive_x and x_val <= 0:
            continue

        xs.append(x_val)
        ys.append(y_val)

    return xs, ys

def _triplet_series(
    parsed: ParsedDTA,
    x_name: str,
    y_name: str,
    f_name: str,
    *,
    require_positive_f: bool = True,
) -> tuple[list[float], list[float], list[float]]:
    x_idx = _column_index(parsed, x_name)
    y_idx = _column_index(parsed, y_name)
    f_idx = _column_index(parsed, f_name)

    if x_idx is None or y_idx is None or f_idx is None:
        return [], [], []

    xs: list[float] = []
    ys: list[float] = []
    fs: list[float] = []

    for row in parsed.rows:
        if x_idx >= len(row) or y_idx >= len(row) or f_idx >= len(row):
            continue

        x_val = to_float(row[x_idx])
        y_val = to_float(row[y_idx])
        f_val = to_float(row[f_idx])

        if x_val is None or y_val is None or f_val is None:
            continue
        if require_positive_f and f_val <= 0:
            continue

        xs.append(x_val)
        ys.append(y_val)
        fs.append(f_val)

    return xs, ys, fs

def _technique_name(parsed: ParsedDTA) -> str:
    return parsed.meta_values.get("TITLE", "").strip() or "Potentiostatic EIS"


def _extract_stage_number(stem: str) -> int | None:
    match = re.search(r"#(\d+)", stem)
    if match is None:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def _extract_current_label(stem: str) -> tuple[str | None, float | None]:
    for token in re.split(r"[_\s]+", stem):
        clean = token.strip()
        if not clean:
            continue
        if re.fullmatch(r"[+-]?\d+(?:[.,]\d+)?A", clean, flags=re.IGNORECASE):
            return clean, to_float(clean[:-1])
    return None, None


def _format_decimal_for_plot_label(raw_value: str) -> str | None:
    raw_number = raw_value.strip()
    if not raw_number:
        return None

    try:
        value = Decimal(raw_number.replace(",", "."))
    except InvalidOperation:
        return None

    normalized = value.normalize()
    if normalized == 0:
        return "0"

    abs_value = normalized.copy_abs()
    if Decimal("0.001") <= abs_value < Decimal("10000"):
        text = format(value, "f").rstrip("0").rstrip(".")
    else:
        mantissa, exponent = format(normalized, "E").split("E", maxsplit=1)
        mantissa = mantissa.rstrip("0").rstrip(".")
        text = f"{mantissa}E{int(exponent):+d}"

    return text.replace(".", ",")


def _format_voltage_label(parsed: ParsedDTA) -> str | None:
    raw_value = parsed.meta_values.get("VDC", "").strip()
    if not raw_value:
        return None

    if raw_value.upper().endswith("V"):
        raw_value = raw_value[:-1].strip()

    formatted = _format_decimal_for_plot_label(raw_value)
    if formatted is None:
        compact = raw_value.replace(" ", "").replace(".", ",")
        return f"{compact}V" if compact else None
    return f"{formatted}V"


def _extract_characterization_label(
    stem: str,
    *,
    current_label: str | None,
    voltage_label: str | None,
) -> str | None:
    if current_label is not None or voltage_label != "0V":
        return None

    match = re.fullmatch(r"EISPOT_[^_]+_([23])_#\d+", stem, flags=re.IGNORECASE)
    if match is None:
        return None

    return {
        "2": "Pre-char",
        "3": "Post-char",
    }.get(match.group(1))


def _build_eis_display_name(path: Path, parsed: ParsedDTA) -> tuple[str, int | None, str | None, str | None, float | None]:
    stem = path.stem
    stage_number = _extract_stage_number(stem)
    current_label, current_value = _extract_current_label(stem)
    voltage_label = _format_voltage_label(parsed)
    characterization_label = _extract_characterization_label(
        stem,
        current_label=current_label,
        voltage_label=voltage_label,
    )

    parts: list[str] = []
    if current_label:
        if stage_number is not None:
            parts.append(f"Etapa {stage_number}")
        parts.append(current_label)
        if voltage_label:
            parts.append(voltage_label)
    elif characterization_label:
        if stage_number is not None:
            parts.append(f"Etapa {stage_number}")
        if voltage_label:
            parts.append(voltage_label)
        parts.append(characterization_label)

    # Only prettify filenames when they match the expected naming schemes.
    # Non-standard 0 V measurements should keep their stem so the composer
    # treats them as independent series instead of merging them under a
    # generic label such as "Etapa N / 0V".
    display_name = " / ".join(parts) if parts else stem
    return display_name, stage_number, current_label, voltage_label, current_value


def _hls_to_hex(hue: float, lightness: float, saturation: float) -> str:
    r, g, b = colorsys.hls_to_rgb(hue % 1.0, max(0.0, min(1.0, lightness)), max(0.0, min(1.0, saturation)))
    return f"#{int(round(r * 255)):02x}{int(round(g * 255)):02x}{int(round(b * 255)):02x}"


def _assign_stage_visual_defaults(entries: list[EISPlotEntry]) -> None:
    if not entries:
        return

    def _stage_entry_sort_key(entry: EISPlotEntry) -> tuple[bool, float, str, str]:
        return (
            entry.current_value is None,
            entry.current_value if entry.current_value is not None else math.inf,
            entry.current_label or "",
            entry.path.stem.lower(),
        )

    grouped: dict[int | None, list[EISPlotEntry]] = {}
    for entry in entries:
        grouped.setdefault(entry.stage_number, []).append(entry)

    ordered_stage_keys = sorted(
        grouped,
        key=lambda stage: (stage is None, stage if stage is not None else math.inf),
    )

    styled_stage_keys = [stage_key for stage_key in ordered_stage_keys if stage_key is not None]
    stage_count = max(1, len(styled_stage_keys))

    for stage_entries in grouped.values():
        stage_entries.sort(key=_stage_entry_sort_key)

    for stage_index, stage_key in enumerate(styled_stage_keys):
        stage_entries = grouped[stage_key]
        base_hue = (0.58 + (stage_index / stage_count)) % 1.0
        stage_color = _hls_to_hex(base_hue, 0.46, 0.80)
        for idx, entry in enumerate(stage_entries):
            entry.nyquist_color = stage_color
            entry.default_marker = EIS_STAGE_MARKER_SEQUENCE[idx] if idx < len(EIS_STAGE_MARKER_SEQUENCE) else None

    for entry in grouped.get(None, []):
        entry.nyquist_color = None
        entry.default_marker = None


def _collect_eis_plot_entries(dta_files: list[Path]) -> list[EISPlotEntry]:
    entries: list[EISPlotEntry] = []
    for dta_file in dta_files:
        parsed = parse_gamry_dta(dta_file)
        display_name, stage_number, current_label, voltage_label, current_value = _build_eis_display_name(dta_file, parsed)
        entries.append(
            EISPlotEntry(
                path=dta_file,
                parsed=parsed,
                display_name=display_name,
                stage_number=stage_number,
                current_label=current_label,
                voltage_label=voltage_label,
                current_value=current_value,
            )
        )

    _assign_stage_visual_defaults(entries)
    entries.sort(
        key=lambda entry: (
            entry.stage_number is None,
            entry.stage_number if entry.stage_number is not None else math.inf,
            entry.current_value is None,
            entry.current_value if entry.current_value is not None else math.inf,
            entry.current_label or "",
            entry.path.stem.lower(),
        )
    )
    return entries


def _build_pre_stabilization_display_name(path: Path) -> tuple[str, int | None, str | None, float | None]:
    stage_number = _extract_stage_number(path.stem)
    current_label, current_value = _extract_current_label(path.stem)

    parts: list[str] = []
    if stage_number is not None:
        parts.append(f"Etapa {stage_number}")
    if current_label:
        parts.append(current_label)

    return (" - ".join(parts) if parts else path.stem, stage_number, current_label, current_value)


def _collect_pre_stabilization_entries(dta_files: list[Path]) -> list[EISPlotEntry]:
    entries: list[EISPlotEntry] = []
    for dta_file in dta_files:
        parsed = parse_gamry_curve_dta(dta_file)
        display_name, stage_number, current_label, current_value = _build_pre_stabilization_display_name(dta_file)
        entries.append(
            EISPlotEntry(
                path=dta_file,
                parsed=parsed,
                display_name=display_name,
                stage_number=stage_number,
                current_label=current_label,
                current_value=current_value,
            )
        )

    entries.sort(
        key=lambda entry: (
            entry.stage_number is None,
            entry.stage_number if entry.stage_number is not None else math.inf,
            entry.current_value is None,
            entry.current_value if entry.current_value is not None else math.inf,
            entry.current_label or "",
            entry.path.stem.lower(),
        )
    )
    return entries


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_to_xlsx(
    parsed: ParsedDTA,
    out_path: Path,
    *,
    meta_fields: list[tuple[str, str]] | None = None,
    data_map: dict[str, str] | None = None,
    numeric_meta_keys: set[str] | None = None,
) -> None:
    """Create one .xlsx with Metadata + Data sheets."""
    meta_fields = META_FIELDS if meta_fields is None else meta_fields
    data_map = DATA_MAP if data_map is None else data_map
    if numeric_meta_keys is None:
        numeric_meta_keys = {"VDC", "FREQINIT", "FREQFINAL", "PTSPERDEC", "VAC", "AREA"}

    wb = Workbook()
    wb.remove(wb.active)

    ws_meta = wb.create_sheet("Metadata")
    ws_data = wb.create_sheet("Data")

    # ---------------- Metadata sheet ---------------------------------------
    ws_meta["A1"] = "Campo"
    ws_meta["B1"] = "Valor"
    ws_meta["C1"] = "Unidad"
    for ref in ("A1", "B1", "C1"):
        ws_meta[ref].font = Font(bold=True)
    ws_meta.freeze_panes = "A2"

    for row_idx, (key, label) in enumerate(meta_fields, start=2):
        raw_value = parsed.meta_values.get(key, "")
        raw_unit = parsed.meta_units.get(key, "")

        ws_meta.cell(row=row_idx, column=1, value=label)

        if key in numeric_meta_keys:
            num = to_float(raw_value)
            ws_meta.cell(row=row_idx, column=2, value=num if num is not None else raw_value)
        else:
            ws_meta.cell(row=row_idx, column=2, value=raw_value)

        ws_meta.cell(row=row_idx, column=3, value=raw_unit)

    # ---------------- Data sheet -------------------------------------------
    col_idx = {name: idx for idx, name in enumerate(parsed.header)}
    selected_source_cols = [name for name in data_map if name in col_idx]
    selected_output_headers = [data_map[name] for name in selected_source_cols]

    # Row 1 = names
    for col_num, header_name in enumerate(selected_output_headers, start=1):
        cell = ws_data.cell(row=1, column=col_num, value=header_name)
        cell.font = Font(bold=True)

    # Row 2 = units
    for col_num, source_name in enumerate(selected_source_cols, start=1):
        unit_value = ""
        source_index = col_idx[source_name]

        if parsed.units and source_index < len(parsed.units):
            unit_value = parsed.units[source_index]
            if unit_value == "#":
                unit_value = ""

        ws_data.cell(row=2, column=col_num, value=unit_value)

    ws_data.freeze_panes = "A3"

    # Row 3 onward = numeric data
    for row_num, raw_parts in enumerate(parsed.rows, start=3):
        for col_num, source_name in enumerate(selected_source_cols, start=1):
            source_index = col_idx[source_name]
            raw_value = raw_parts[source_index] if source_index < len(raw_parts) else ""

            num = to_float(raw_value)
            ws_data.cell(row=row_num, column=col_num, value=num if num is not None else raw_value)

    # ---------------- Light formatting -------------------------------------
    for ws in (ws_meta, ws_data):
        for col_num in range(1, ws.max_column + 1):
            max_len = 0
            for row_num in range(1, min(ws.max_row, 50) + 1):
                value = ws.cell(row=row_num, column=col_num).value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))

            ws.column_dimensions[get_column_letter(col_num)].width = min(max(10, max_len + 2), 45)

        for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ---------------------------------------------------------------------------
# Plotting (interactive: build Figures and show them in a Tk window)
# ---------------------------------------------------------------------------

def _new_figure() -> Figure:
    # You can tweak size if you want bigger/smaller default tabs
    return Figure(figsize=(6.8, 4.8), dpi=100)

def fig_nyquist(
    parsed: ParsedDTA,
    *,
    plot_title: str | None = None,
    line_color: str | None = None,
    marker_style: str | None = None,
    font_defaults: PlotFontDefaults | None = None,
) -> Figure | None:
    x, y, f = _triplet_series(parsed, "Zreal", "Zimag", "Freq")
    if not x or not y or not f:
        return None

    y_plot = [-v for v in y]

    fig = _new_figure()
    ax = fig.add_subplot(111)

    line_kwargs = {
        "marker": marker_style or "o",
        "linestyle": "-",
        "markerfacecolor": "none",
    }
    if line_color:
        line_kwargs["color"] = line_color
        line_kwargs["markeredgecolor"] = line_color

    (line,) = ax.plot(x, y_plot, **line_kwargs)
    line._eis_freq = f  # attach frequency array (same index as points)

    ax.set_aspect("equal", adjustable="box")
    ax.margins(0.05)

    x_unit = _impedance_unit(parsed, "Zreal")
    y_unit = _impedance_unit(parsed, "Zimag")
    ax.set_title(plot_title or _technique_name(parsed))
    ax.set_xlabel(f"Zreal ({x_unit})" if x_unit else "Zreal")
    ax.set_ylabel(f"-Zimag ({y_unit})" if y_unit else "-Zimag")
    ax.grid(True)

    apply_plot_font_defaults(fig, font_defaults)
    fig.tight_layout()
    return fig

def fig_bode(
    parsed: ParsedDTA,
    *,
    plot_title: str | None = None,
    marker_style: str | None = None,
    font_defaults: PlotFontDefaults | None = None,
) -> Figure | None:
    freq_unit = _column_unit(parsed, "Freq")
    zmod_unit = _column_unit(parsed, "Zmod")
    zphz_unit = _column_unit(parsed, "Zphz")

    x1, y1 = _paired_series(parsed, "Freq", "Zmod", require_positive_x=True)
    x2, y2 = _paired_series(parsed, "Freq", "Zphz", require_positive_x=True)
    if not ((x1 and y1) or (x2 and y2)):
        return None

    fig = _new_figure()
    ax_mod = fig.add_subplot(111)
    ax_phz = ax_mod.twinx()
    ax_phz.spines["left"].set_visible(False)
    ax_phz.yaxis.tick_right()
    ax_phz.yaxis.set_label_position("right")

    bode_lines: dict[str, object] = {}
    bode_axes: dict[str, object] = {"mod": ax_mod, "phz": ax_phz}

    if x1 and y1:
        (ln_mod,) = ax_mod.semilogx(
            x1, y1,
            marker=marker_style or "o",
            linestyle="-",
            markerfacecolor="none",
            color="#1f77b4",
            label="Zmod",
        )
        bode_lines["mod"] = ln_mod

    if x2 and y2:
        (ln_phz,) = ax_phz.semilogx(
            x2, y2,
            marker=marker_style or "o",
            linestyle="--",
            markerfacecolor="none",
            color="#c40000",
            label="Zphz",
        )
        bode_lines["phz"] = ln_phz

    if "mod" in bode_lines:
        ax_mod.set_ylabel(f"Zmod ({zmod_unit})" if zmod_unit else "Zmod")
    if "phz" in bode_lines:
        ax_phz.set_ylabel(f"Zphz ({zphz_unit})" if zphz_unit else "Zphz")

    ax_mod.set_title(plot_title or f"{_technique_name(parsed)} - Bode")
    ax_mod.set_xlabel(f"Frecuencia ({freq_unit})" if freq_unit else "Frecuencia")
    ax_mod.grid(True, which="both")

    fig._bode_plot = True
    fig._bode_lines = bode_lines
    fig._bode_axes = bode_axes
    apply_plot_font_defaults(fig, font_defaults)
    fig.tight_layout()
    return fig

def _update_right_axis_spacing(fig, canvas, axes, pad_px: float = 12.0, min_outward_pt: float = 30.0):
    axV = axes.get("V")
    axT = axes.get("T")
    if axV is None or axT is None:
        return

    # Force a draw so text extents are up-to-date
    canvas.draw()
    renderer = canvas.get_renderer()
    px_per_pt = float(renderer.points_to_pixels(1.0))

    # Right edge (in display px) of the *base* right spine (x=1 in axes coords)
    base_right_x = axV.transAxes.transform((1.0, 0.0))[0]

    def _rightmost_x1_of_v_axis(ax_):
        x1 = base_right_x

        # tick labels
        for t in ax_.get_yticklabels():
            if t.get_visible() and t.get_text():
                bb = t.get_window_extent(renderer=renderer)
                x1 = max(x1, bb.x1)

        # y-axis label
        lab = ax_.yaxis.label
        if lab.get_visible() and lab.get_text():
            bb = lab.get_window_extent(renderer=renderer)
            x1 = max(x1, bb.x1)

        return x1

    # How far to the right the V-axis content reaches (ticks + ylabel)
    v_rightmost_x1 = _rightmost_x1_of_v_axis(axV)

    # Required outward shift in pixels so the Temp spine is beyond V content
    required_px = (v_rightmost_x1 - base_right_x) + pad_px
    outward_pt = max(min_outward_pt, required_px / px_per_pt)

    axT.spines["right"].set_position(("outward", outward_pt))

    # Optional: ensure figure margin doesn't clip the outer axis label/ticks
    # Store a baseline right margin once so it doesn't "ratchet" smaller over time.
    if not hasattr(fig, "_default_right_margin"):
        fig._default_right_margin = fig.subplotpars.right
    fig.subplots_adjust(right=fig._default_right_margin)

    canvas.draw()
    renderer = canvas.get_renderer()

    # Check if Temp axis content clips beyond figure width; if so, reduce right margin.
    fig_px = fig.get_size_inches()[0] * fig.dpi
    t_rightmost = base_right_x

    for t in axT.get_yticklabels():
        if t.get_visible() and t.get_text():
            bb = t.get_window_extent(renderer=renderer)
            t_rightmost = max(t_rightmost, bb.x1)

    labT = axT.yaxis.label
    if labT.get_visible() and labT.get_text():
        bb = labT.get_window_extent(renderer=renderer)
        t_rightmost = max(t_rightmost, bb.x1)

    overflow = t_rightmost - (fig_px - 8.0)  # 8px padding
    if overflow > 0:
        extra = overflow / fig_px
        fig.subplots_adjust(right=max(0.50, fig._default_right_margin - extra))

def fig_series_vs_pt(
    parsed: ParsedDTA,
    *,
    plot_title: str | None = None,
    font_defaults: PlotFontDefaults | None = None,
) -> Figure | None:
    fig = _new_figure()
    axI = fig.add_subplot(111)

    pt_default_colors = {
        "I": "#2f9e44",
        "V": "#1f77b4",
        "T": "#c40000",
    }

    lines: dict[str, object] = {}
    axes: dict[str, object] = {"I": axI}
    ylabels: dict[str, str] = {}

    # Create extra axes (always created; lines may or may not exist)
    axV = axI.twinx()
    axT = axI.twinx()
    axT.spines["right"].set_position(("outward", 60))

    # After:
    # axV = axI.twinx()
    # axT = axI.twinx()
    # axT.spines["right"].set_position(("outward", 60))

    axV.spines["left"].set_visible(False)
    axT.spines["left"].set_visible(False)
    # make sure their ticks are on the right (usually automatic, but explicit is fine)
    axV.yaxis.tick_right()
    axV.yaxis.set_label_position("right")
    axT.yaxis.tick_right()
    axT.yaxis.set_label_position("right")

    axes["V"] = axV
    axes["T"] = axT

    # X label
    x_unit = _column_unit(parsed, "Pt")
    axI.set_xlabel(f"Pt ({x_unit})" if x_unit else "Pt")

    # Grid only on base axis to avoid clutter
    axI.grid(True)
    axV.grid(False)
    axT.grid(False)

    def _add_series(key: str, col: str, label: str, ax):
        x, y, f = _triplet_series(parsed, "Pt", col, "Freq")
        if not x or not y or not f:
            return

        unit = _column_unit(parsed, col)
        lab = f"{label} ({unit})" if unit else label

        (ln,) = ax.plot(
            x,
            y,
            marker="o",
            linestyle="-",
            markerfacecolor="none",
            color=pt_default_colors.get(key),
            label=lab,
        )
        ln._eis_freq = f
        if key == "I":
            ln._eis_abs_ydata = tuple(y)
            ln._eis_abs_label = lab
            ln._eis_density_label = "Densidad de corriente (A/cm^2)"
        lines[key] = ln
        ylabels[key] = lab

    _add_series("I", "Idc", "Idc", axI)
    _add_series("V", "Vdc", "Vdc", axV)
    _add_series("T", "Temp", "Temp", axT)

    if not lines:
        return None

    # Default visibility: show only Idc if present, else first available
    for k, ln in lines.items():
        ln.set_visible(False)

    base_key = "I" if "I" in lines else next(iter(lines.keys()))
    lines[base_key].set_visible(True)

    # Set ylabels for each axis (they can be colored later in UI)
    if "I" in lines: axI.set_ylabel(ylabels["I"])
    if "V" in lines: axV.set_ylabel(ylabels["V"])
    if "T" in lines: axT.set_ylabel(ylabels["T"])

    # Title
    axI.set_title(plot_title or f"{_technique_name(parsed)} - Series vs Pt")

    # Save metadata for the UI
    fig._pt_series = True
    fig._pt_lines = lines
    fig._pt_axes = axes
    fig._pt_ylabels = ylabels
    fig._pt_base_key = base_key  # which axis drives the grid/tick alignment
    fig._pt_current_area_cm2 = _metadata_area_cm2(parsed)

    apply_plot_font_defaults(fig, font_defaults)
    fig.tight_layout()
    return fig


def fig_pre_stabilization(
    entry: EISPlotEntry,
    *,
    font_defaults: PlotFontDefaults | None = None,
) -> Figure | None:
    parsed = entry.parsed
    fig = _new_figure()
    axV = fig.add_subplot(111)
    axI = axV.twinx()
    axT = axV.twinx()
    axT.spines["right"].set_position(("outward", 60))

    axI.spines["left"].set_visible(False)
    axT.spines["left"].set_visible(False)
    axI.yaxis.tick_right()
    axI.yaxis.set_label_position("right")
    axT.yaxis.tick_right()
    axT.yaxis.set_label_position("right")

    axes = {"I": axI, "V": axV, "T": axT}
    lines: dict[str, object] = {}

    time_unit = _column_unit(parsed, "T")
    axV.set_xlabel(f"Tiempo ({time_unit})" if time_unit else "Tiempo")
    axV.grid(True)
    axI.grid(False)
    axT.grid(False)

    def _add_series(key: str, col: str, label: str, ax, color: str):
        x, y = _paired_series(parsed, "T", col)
        if not x or not y:
            return

        unit = _column_unit(parsed, col)
        series_label = f"{label} ({unit})" if unit else label
        (ln,) = ax.plot(
            x,
            y,
            marker="o",
            linestyle="-",
            markerfacecolor="none",
            color=color,
            label=series_label,
        )
        lines[key] = ln
        ax.set_ylabel(series_label)

    _add_series("V", "Vf", "Voltaje", axV, "#1f77b4")
    _add_series("I", "Im", "Corriente", axI, "#2f9e44")
    _add_series("T", "Temp", "Temperatura", axT, "#c40000")

    if not lines:
        return None

    axV.set_title(f"{entry.display_name} - Pre-estabilizacion")

    fig._pre_stab_series = True
    fig._pre_stab_lines = lines
    fig._pre_stab_axes = axes

    apply_plot_font_defaults(fig, font_defaults)
    fig.tight_layout()
    return fig


def build_figures(
    entry: EISPlotEntry,
    selected_options: Iterable[str] | None,
    font_defaults: PlotFontDefaults | None = None,
) -> list[tuple[str, Figure]]:
    """Create only the figures requested by the GUI."""
    if not selected_options:
        return []

    chosen = set(selected_options)
    figs: list[tuple[str, Figure]] = []
    parsed = entry.parsed
    base_name = entry.display_name

    if "Nyquist plot" in chosen:
        f = fig_nyquist(
            parsed,
            plot_title=base_name,
            line_color=entry.nyquist_color,
            marker_style=entry.default_marker,
            font_defaults=font_defaults,
        )
        if f is not None:
            figs.append((f"{base_name} — Nyquist", f))

    if "Bode plot" in chosen:
        f = fig_bode(
            parsed,
            plot_title=f"{base_name} - Bode",
            marker_style=entry.default_marker,
            font_defaults=font_defaults,
        )
        if f is not None:
            figs.append((f"{base_name} — Bode", f))

    if "Series by Pt" in chosen:
        f = fig_series_vs_pt(parsed, plot_title=f"{base_name} - Series vs Pt", font_defaults=font_defaults)
        if f is not None:
            figs.append((f"{base_name} — Series vs Pt", f))

    # "Equivalent circuit fit" is listed in GUI but not implemented here yet.
    return figs


def _open_composer_pt(
    win,
    pt_sources: dict[str, dict],
    font_defaults: PlotFontDefaults,
    *,
    window_title: str = "Composite (Series vs Pt)",
    default_title: str = "Composite - Series vs Pt",
    sources_title: str = "Series vs Pt sources",
) -> None:
    source_keys = sorted(pt_sources.keys(), key=lambda k: k.lower())
    if not source_keys:
        return

    existing = getattr(win, "_composer_win_pt", None)
    if existing is not None and existing.winfo_exists():
        existing.lift()
        existing.focus_force()
        return

    series_order = ("I", "V", "T")
    series_names = {"I": "Idc", "V": "Vdc", "T": "Temp"}

    def _fmt(v: float) -> str:
        return f"{v:.6g}"

    def _source_line(key: str, kind: str):
        src = pt_sources.get(key, {})
        return src.get("lines", {}).get(kind)

    def _source_axis(key: str, kind: str):
        src = pt_sources.get(key, {})
        return src.get("axes", {}).get(kind)

    def _src_label(key: str) -> str:
        for kind in series_order:
            ax_src = _source_axis(key, kind)
            if ax_src is not None:
                title = (ax_src.get_title() or "").strip()
                if title:
                    return title
        return key

    def _source_axis_label(key: str, kind: str) -> str:
        ln = _source_line(key, kind)
        if ln is not None:
            label = (ln.get_label() or "").strip()
            if label and not label.startswith("_"):
                return label
        ax_src = _source_axis(key, kind)
        if ax_src is not None:
            label = (ax_src.get_ylabel() or "").strip()
            if label:
                return label
        return series_names[kind]

    def _legend_series_name(key: str, kind: str) -> str:
        axis_label = _source_axis_label(key, kind)
        if kind == "I" and ("A/cm" in axis_label or "densidad" in axis_label.lower()):
            return "Densidad de corriente"
        return series_names[kind]

    def _default_xlabel() -> str:
        for key in source_keys:
            for kind in series_order:
                ax_src = _source_axis(key, kind)
                if ax_src is not None:
                    label = (ax_src.get_xlabel() or "").strip()
                    if label:
                        return label
        return "Pt"

    def _default_ylabel(kind: str) -> str:
        for key in source_keys:
            if _source_line(key, kind) is not None:
                return _source_axis_label(key, kind)
        return series_names[kind]

    default_xlabel = _default_xlabel()
    default_ylabels = {kind: _default_ylabel(kind) for kind in series_order}

    comp = tk.Toplevel(win)
    win._composer_win_pt = comp  # type: ignore[attr-defined]
    comp.title(window_title)
    comp.geometry("1280x800")

    ctrl_host, plot_frame = create_resizable_plot_layout(
        comp,
        sidebar_width=320,
        sidebar_side="right",
        plot_padding=0,
    )

    figc = _new_figure()
    axc_I = figc.add_subplot(111)
    axc_V = axc_I.twinx()
    axc_T = axc_I.twinx()
    axes = {"I": axc_I, "V": axc_V, "T": axc_T}
    comp._pt_legend = None  # type: ignore[attr-defined]
    comp._pt_legend_dragger = None  # type: ignore[attr-defined]

    def _reset_axes():
        for ax_reset in axes.values():
            ax_reset.cla()
        axc_I.grid(True)
        axc_I.set_title(default_title)
        axc_I.set_xlabel(default_xlabel)
        axc_I.set_ylabel(default_ylabels["I"])
        axc_V.set_ylabel(default_ylabels["V"])
        axc_T.set_ylabel(default_ylabels["T"])

        axc_V.spines["left"].set_visible(False)
        axc_T.spines["left"].set_visible(False)
        axc_V.yaxis.tick_right()
        axc_V.yaxis.set_label_position("right")
        axc_T.yaxis.tick_right()
        axc_T.yaxis.set_label_position("right")
        axc_T.spines["right"].set_position(("outward", 60))

    _reset_axes()

    canvas = FigureCanvasTkAgg(figc, master=plot_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    toolbar = NavigationToolbar2Tk(canvas, plot_frame)
    toolbar.update()

    _topbar, ctrl_scroll = _build_scrollable_controls(ctrl_host)
    ctrl = ttk.Frame(ctrl_scroll, padding=10)
    ctrl.pack(fill="both", expand=True)

    src_box = ttk.LabelFrame(ctrl, text=sources_title, padding=8)
    src_box.pack(fill="x", pady=(0, 10))

    lb = tk.Listbox(src_box, selectmode="extended", height=12, exportselection=False)
    lb.pack(fill="x", expand=False)

    idx_to_key: list[str] = []

    def _rebuild_listbox():
        nonlocal idx_to_key
        lb.delete(0, "end")
        idx_to_key = []
        keys = sorted(source_keys, key=lambda k: _src_label(k).lower())
        for key in keys:
            lb.insert("end", f"{_src_label(key)}   [{key}]")
            idx_to_key.append(key)

    _rebuild_listbox()

    def _selected_keys() -> list[str]:
        return [idx_to_key[i] for i in lb.curselection()]

    comp_lines: dict[str, dict[str, object]] = {"I": {}, "V": {}, "T": {}}
    legend_var = tk.BooleanVar(value=True)
    title_text_var = tk.StringVar(value=default_title)
    show_title_var = tk.BooleanVar(value=True)
    title_fs_var = tk.DoubleVar(value=float(font_defaults.title))
    label_fs_var = tk.DoubleVar(value=float(font_defaults.label))
    legend_fs_var = tk.DoubleVar(value=float(font_defaults.legend))
    tick_fs_var = tk.DoubleVar(value=float(font_defaults.tick))
    x_tick_count_var = tk.IntVar(value=6)
    y_tick_count_var = tk.IntVar(value=6)

    def _copy_style(src_line, dst_line):
        dst_line.set_color(src_line.get_color())
        dst_line.set_linestyle(src_line.get_linestyle())
        dst_line.set_marker(src_line.get_marker())
        dst_line.set_linewidth(src_line.get_linewidth())
        dst_line.set_markersize(src_line.get_markersize())
        try:
            dst_line.set_markerfacecolor(src_line.get_markerfacecolor())
        except Exception:
            pass
        try:
            dst_line.set_markeredgecolor(src_line.get_markeredgecolor())
        except Exception:
            pass
        try:
            dst_line.set_alpha(src_line.get_alpha())
        except Exception:
            pass

    def _copy_data_and_style(key: str, kind: str, dst_line) -> bool:
        src_line = _source_line(key, kind)
        if src_line is None:
            return False
        dst_line.set_xdata(list(src_line.get_xdata(orig=False)))
        dst_line.set_ydata(list(src_line.get_ydata(orig=False)))
        _copy_style(src_line, dst_line)
        dst_line.set_label(f"{_src_label(key)} - {_legend_series_name(key, kind)}")
        dst_line._pt_axis_label = _source_axis_label(key, kind)  # type: ignore[attr-defined]
        return True

    def _source_kind_visible(key: str, kind: str) -> bool:
        src_line = _source_line(key, kind)
        return bool(src_line is not None and src_line.get_visible())

    def _visible_source_kinds(key: str) -> list[str]:
        return [kind for kind in series_order if _source_kind_visible(key, kind)]

    def _active_lines(kind: str) -> list[object]:
        return [ln for ln in comp_lines[kind].values() if ln.get_visible()]

    def _axis_label_for(kind: str) -> str:
        labels = []
        for ln in comp_lines[kind].values():
            label = getattr(ln, "_pt_axis_label", "")
            if label:
                labels.append(str(label))
        unique = sorted(set(labels))
        if len(unique) == 1:
            return unique[0]
        if unique and kind == "I":
            return "Idc / densidad de corriente"
        return series_names[kind] if unique else default_ylabels[kind]

    def _apply_legend(redraw: bool = True):
        for axis in axes.values():
            leg = axis.get_legend()
            if leg is not None:
                leg.remove()
        comp._pt_legend = None  # type: ignore[attr-defined]
        comp._pt_legend_dragger = None  # type: ignore[attr-defined]

        if not legend_var.get():
            if redraw:
                canvas.draw_idle()
            return

        handles = []
        labels = []
        for kind in series_order:
            for key in sorted(comp_lines[kind].keys(), key=lambda k: _src_label(k).lower()):
                ln = comp_lines[kind][key]
                if ln.get_visible():
                    handles.append(ln)
                    labels.append(ln.get_label())

        if handles:
            try:
                legend_fs = float(legend_fs_var.get())
            except (tk.TclError, ValueError):
                legend_fs = float(font_defaults.legend)
            # The Pt composer uses twinx axes; attach the legend to the topmost
            # axis so draggable legend events are not intercepted by a twin axis.
            legend = axc_T.legend(handles, labels, loc="best", fontsize=legend_fs)
            make_legend_draggable(legend)
            comp._pt_legend = legend  # type: ignore[attr-defined]
            comp._pt_legend_dragger = getattr(legend, "_draggable", None)  # type: ignore[attr-defined]
        if redraw:
            canvas.draw_idle()

    def _apply_tick_settings():
        try:
            x_tick_count = max(2, int(x_tick_count_var.get()))
        except (tk.TclError, ValueError):
            x_tick_count = 6
        try:
            y_tick_count = max(2, int(y_tick_count_var.get()))
        except (tk.TclError, ValueError):
            y_tick_count = 6
        axc_I.xaxis.set_major_locator(MaxNLocator(nbins=x_tick_count))
        axc_I.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
        for axis in axes.values():
            axis.yaxis.set_major_locator(LinearLocator(y_tick_count))
            axis.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
            axis.yaxis.get_offset_text().set_visible(False)

    def _sync_axis_visibility(redraw: bool = True):
        for kind in series_order:
            ax_kind = axes[kind]
            for ln in comp_lines[kind].values():
                ln.set_visible(True)

            axis_visible = bool(comp_lines[kind])
            ax_kind.yaxis.set_visible(axis_visible)
            ax_kind.yaxis.label.set_visible(axis_visible)
            ax_kind.set_ylabel(_axis_label_for(kind))
            if kind == "I":
                ax_kind.tick_params(
                    axis="y",
                    labelleft=axis_visible,
                    left=axis_visible,
                    labelright=False,
                    right=False,
                )
                ax_kind.spines["left"].set_visible(axis_visible)
            else:
                ax_kind.tick_params(
                    axis="y",
                    labelleft=False,
                    left=False,
                    labelright=axis_visible,
                    right=axis_visible,
                )
                ax_kind.spines["right"].set_visible(axis_visible)

        axc_I.grid(True)
        _apply_tick_settings()
        _apply_legend(redraw=False)
        _update_right_axis_spacing(figc, canvas, axes)
        if redraw:
            canvas.draw_idle()

    def _apply_plot_settings(redraw: bool = True):
        try:
            title_fs = float(title_fs_var.get())
            label_fs = float(label_fs_var.get())
            tick_fs = float(tick_fs_var.get())
        except (tk.TclError, ValueError):
            return

        axc_I.set_title(title_text_var.get() if show_title_var.get() else "", fontsize=title_fs)
        axc_I.xaxis.label.set_fontsize(label_fs)
        for axis in axes.values():
            axis.yaxis.label.set_fontsize(label_fs)
            axis.tick_params(axis="both", labelsize=tick_fs)
        apply_x_tick_label_padding(axc_I, tick_fs)
        _apply_tick_settings()
        _apply_legend(redraw=False)
        _update_right_axis_spacing(figc, canvas, axes)
        if redraw:
            canvas.draw_idle()

    pending_plot_settings = {"id": None}

    def _schedule_plot_settings(_evt=None):
        if pending_plot_settings["id"] is not None:
            comp.after_cancel(pending_plot_settings["id"])
        pending_plot_settings["id"] = comp.after(180, _apply_plot_settings)

    def _sync_limit_entries():
        x0, x1 = axc_I.get_xlim()
        xmin_var.set(_fmt(x0))
        xmax_var.set(_fmt(x1))
        for kind in series_order:
            y0, y1 = axes[kind].get_ylim()
            ymin_vars[kind].set(_fmt(y0))
            ymax_vars[kind].set(_fmt(y1))

    def _fit_all():
        active_by_kind = {kind: _active_lines(kind) for kind in series_order}
        active_all = [ln for lineset in active_by_kind.values() for ln in lineset]
        if not active_all:
            return

        xs: list[float] = []
        for ln in active_all:
            xs.extend([float(v) for v in ln.get_xdata(orig=False)])
        if not xs:
            return

        x0, x1 = min(xs), max(xs)
        dx = (x1 - x0) if x1 != x0 else (abs(x0) * 0.1 + 1.0)
        for axis in axes.values():
            axis.set_xlim(x0 - 0.05 * dx, x1 + 0.05 * dx)

        for kind, lineset in active_by_kind.items():
            if not lineset:
                continue
            ys: list[float] = []
            for ln in lineset:
                ys.extend([float(v) for v in ln.get_ydata(orig=False)])
            if not ys:
                continue
            y0, y1 = min(ys), max(ys)
            dy = (y1 - y0) if y1 != y0 else (abs(y0) * 0.1 + 1.0)
            axes[kind].set_ylim(y0 - 0.05 * dy, y1 + 0.05 * dy)

        _sync_axis_visibility(redraw=False)
        _apply_plot_settings(redraw=False)
        canvas.draw_idle()
        _sync_limit_entries()

    def add_selected():
        added = False
        for key in _selected_keys():
            for kind in _visible_source_kinds(key):
                if key in comp_lines[kind]:
                    continue
                src_line = _source_line(key, kind)
                if src_line is None:
                    continue
                axis = axes[kind]
                x = list(src_line.get_xdata(orig=False))
                y = list(src_line.get_ydata(orig=False))
                (ln,) = axis.plot(x, y, label=f"{_src_label(key)} - {_legend_series_name(key, kind)}")
                _copy_style(src_line, ln)
                ln._pt_axis_label = _source_axis_label(key, kind)  # type: ignore[attr-defined]
                comp_lines[kind][key] = ln
                added = True

        if added:
            _sync_axis_visibility(redraw=False)
            _fit_all()

    def remove_selected():
        removed = False
        for key in _selected_keys():
            for kind in series_order:
                ln = comp_lines[kind].pop(key, None)
                if ln is not None:
                    try:
                        ln.remove()
                    except Exception:
                        pass
                    removed = True
        if removed:
            _sync_axis_visibility(redraw=False)
            _fit_all()

    def clear_all():
        for kind in series_order:
            comp_lines[kind].clear()
        _reset_axes()
        _sync_axis_visibility(redraw=False)
        _apply_plot_settings(redraw=False)
        canvas.draw_idle()
        _sync_limit_entries()
        _apply_legend()
        try:
            lb.selection_clear(0, "end")
        except Exception:
            pass

    def refresh_formatting():
        source_keys_in_composite = sorted(
            {key for kind in series_order for key in comp_lines[kind].keys()},
            key=lambda item: _src_label(item).lower(),
        )
        changed = False
        for key in source_keys_in_composite:
            for kind in series_order:
                ln = comp_lines[kind].get(key)
                if _source_kind_visible(key, kind):
                    if ln is None:
                        src_line = _source_line(key, kind)
                        if src_line is None:
                            continue
                        (ln,) = axes[kind].plot([], [])
                        comp_lines[kind][key] = ln
                    _copy_data_and_style(key, kind, ln)
                    changed = True
                    continue

                if ln is not None:
                    comp_lines[kind].pop(key, None)
                    try:
                        ln.remove()
                    except Exception:
                        pass
                    changed = True
        _rebuild_listbox()
        if changed:
            _sync_axis_visibility(redraw=False)
            _fit_all()
        else:
            _apply_legend()
            canvas.draw_idle()

    btns = ttk.Frame(src_box)
    btns.pack(fill="x", pady=(8, 0))
    ttk.Button(btns, text="Anadir", command=add_selected).pack(side="left", expand=True, fill="x", padx=(0, 6))
    ttk.Button(btns, text="Remover", command=remove_selected).pack(side="left", expand=True, fill="x", padx=(0, 6))
    ttk.Button(btns, text="Limpiar", command=clear_all).pack(side="left", expand=True, fill="x")

    ttk.Button(src_box, text="Actualizar formato", command=refresh_formatting).pack(fill="x", pady=(8, 0))
    ttk.Checkbutton(ctrl, text="Leyenda", variable=legend_var, command=_apply_legend).pack(anchor="w", pady=(0, 10))

    plot_box = ttk.LabelFrame(ctrl, text="Grafico", padding=8)
    plot_box.pack(fill="x", pady=(0, 10))
    plot_box.columnconfigure(1, weight=1)

    ttk.Label(plot_box, text="Titulo").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=2)
    title_entry = ttk.Entry(plot_box, textvariable=title_text_var, width=20)
    title_entry.grid(row=0, column=1, sticky="ew", pady=2)

    ttk.Label(plot_box, text="Tamano del titulo").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=2)
    title_spin = ttk.Spinbox(plot_box, from_=6.0, to=50.0, increment=0.5, textvariable=title_fs_var, width=10)
    title_spin.grid(row=1, column=1, sticky="w", pady=2)

    ttk.Label(plot_box, text="Tamano de etiquetas").grid(row=2, column=0, sticky="w", padx=(0, 6), pady=2)
    label_spin = ttk.Spinbox(plot_box, from_=6.0, to=40.0, increment=0.5, textvariable=label_fs_var, width=10)
    label_spin.grid(row=2, column=1, sticky="w", pady=2)

    ttk.Label(plot_box, text="Tamano de leyenda").grid(row=3, column=0, sticky="w", padx=(0, 6), pady=2)
    legend_spin = ttk.Spinbox(plot_box, from_=6.0, to=40.0, increment=0.5, textvariable=legend_fs_var, width=10)
    legend_spin.grid(row=3, column=1, sticky="w", pady=2)

    ttk.Label(plot_box, text="Tamano de ticks").grid(row=4, column=0, sticky="w", padx=(0, 6), pady=2)
    tick_spin = ttk.Spinbox(plot_box, from_=6.0, to=40.0, increment=0.5, textvariable=tick_fs_var, width=10)
    tick_spin.grid(row=4, column=1, sticky="w", pady=2)

    ttk.Label(plot_box, text="X ticks").grid(row=5, column=0, sticky="w", padx=(0, 6), pady=2)
    x_tick_spin = ttk.Spinbox(plot_box, from_=2, to=20, increment=1, textvariable=x_tick_count_var, width=10)
    x_tick_spin.grid(row=5, column=1, sticky="w", pady=2)

    ttk.Label(plot_box, text="Y ticks").grid(row=6, column=0, sticky="w", padx=(0, 6), pady=2)
    y_tick_spin = ttk.Spinbox(plot_box, from_=2, to=20, increment=1, textvariable=y_tick_count_var, width=10)
    y_tick_spin.grid(row=6, column=1, sticky="w", pady=2)

    ttk.Checkbutton(plot_box, text="Mostrar titulo", variable=show_title_var, command=_apply_plot_settings).grid(
        row=7, column=0, columnspan=2, sticky="w", pady=(6, 0)
    )

    lim_box = ttk.LabelFrame(ctrl, text="Limites de ejes", padding=8)
    lim_box.pack(fill="x", pady=(0, 10))

    xmin_var = tk.StringVar()
    xmax_var = tk.StringVar()
    ymin_vars = {kind: tk.StringVar() for kind in series_order}
    ymax_vars = {kind: tk.StringVar() for kind in series_order}

    def _parse_float(s: str) -> float | None:
        s = s.strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None

    def apply_limits():
        raw_values = [xmin_var.get(), xmax_var.get()]
        for kind in series_order:
            raw_values.extend([ymin_vars[kind].get(), ymax_vars[kind].get()])
        if any(_is_incomplete_number(raw) for raw in raw_values):
            return

        cur_x0, cur_x1 = axc_I.get_xlim()
        nx0 = _parse_float(xmin_var.get())
        nx1 = _parse_float(xmax_var.get())
        new_x0 = cur_x0 if nx0 is None else nx0
        new_x1 = cur_x1 if nx1 is None else nx1
        for axis in axes.values():
            axis.set_xlim(new_x0, new_x1)

        for kind in series_order:
            cur_y0, cur_y1 = axes[kind].get_ylim()
            ny0 = _parse_float(ymin_vars[kind].get())
            ny1 = _parse_float(ymax_vars[kind].get())
            axes[kind].set_ylim(cur_y0 if ny0 is None else ny0, cur_y1 if ny1 is None else ny1)

        _apply_plot_settings(redraw=False)
        canvas.draw_idle()
        _sync_limit_entries()

    def _row(parent, row: int, label: str, var):
        ttk.Label(parent, text=label, width=6).grid(row=row, column=0, sticky="w", padx=(0, 6), pady=2)
        entry = ttk.Entry(parent, textvariable=var, width=12)
        entry.grid(row=row, column=1, sticky="w", pady=2)
        return entry

    entries = []
    next_row = 0
    entries.append(_row(lim_box, next_row, "Xmin", xmin_var))
    next_row += 1
    entries.append(_row(lim_box, next_row, "Xmax", xmax_var))
    next_row += 1
    for kind, label in (("I", "I"), ("V", "V"), ("T", "T")):
        entries.append(_row(lim_box, next_row, f"{label}min", ymin_vars[kind]))
        next_row += 1
        entries.append(_row(lim_box, next_row, f"{label}max", ymax_vars[kind]))
        next_row += 1

    for entry in entries:
        entry.bind("<Return>", lambda ev: apply_limits())
        entry.bind("<FocusOut>", lambda ev: apply_limits())

    title_entry.bind("<Return>", lambda ev: _apply_plot_settings())
    title_entry.bind("<FocusOut>", lambda ev: _apply_plot_settings())
    title_entry.bind("<KeyRelease>", _schedule_plot_settings)

    for spin in (title_spin, label_spin, legend_spin, tick_spin, x_tick_spin, y_tick_spin):
        spin.configure(command=_apply_plot_settings)
        spin.bind("<Return>", lambda ev: _apply_plot_settings())
        spin.bind("<FocusOut>", lambda ev: _apply_plot_settings())

    limit_buttons = ttk.Frame(lim_box)
    limit_buttons.grid(row=next_row, column=0, columnspan=2, sticky="ew", pady=(8, 0))
    ttk.Button(limit_buttons, text="Ajustar todo", command=_fit_all).pack(side="left", expand=True, fill="x", padx=(0, 6))
    ttk.Button(limit_buttons, text="Refresh fields", command=_sync_limit_entries).pack(side="left", expand=True, fill="x")

    _sync_limit_entries()
    _sync_axis_visibility(redraw=False)
    _apply_plot_settings(redraw=False)

    def _on_close():
        comp.destroy()
        try:
            delattr(win, "_composer_win_pt")
        except Exception:
            pass

    comp.protocol("WM_DELETE_WINDOW", _on_close)


def show_figures_tk(
    figures: list[tuple[str, Figure]],
    window_title: str = "EIS plots",
    font_defaults: PlotFontDefaults | None = None,
) -> None:
    font_defaults = resolve_plot_font_defaults(font_defaults)

    tab_id_by_title: dict[str, str] = {}
    title_by_tab_id: dict[str, str] = {}

    if not figures:
        return

    for _tab_title, fig in figures:
        apply_plot_font_defaults(fig, font_defaults)

    root = tk._default_root
    created_root = False
    if root is None:
        root = tk.Tk()
        root.withdraw()
        created_root = True

    win = tk.Toplevel(root)
    win.title(window_title)
    win.geometry("1200x780")

    controls_host, plot_frame = create_resizable_plot_layout(win, sidebar_width=340, plot_padding=0)
    topbar, ctrl_host = _build_scrollable_controls(controls_host)

    ttk.Label(topbar, text="Seleccionar:").pack(side="top", anchor="w")

    nyquist_sources: dict[str, dict] = {}
    bode_sources = {"zmod": {}, "zphz": {}}
    pt_sources: dict[str, dict] = {}
    pre_stabilization_sources: dict[str, dict] = {}

    plot_names_var = tk.StringVar()
    plot_select = ttk.Combobox(topbar, textvariable=plot_names_var, state="readonly", width=32)
    plot_select.pack(side="top", pady=(6, 0), fill="x")

    compose_btn = ttk.Button(topbar, text="Componer")
    compose_btn.pack(side="top", pady=(8, 0), fill="x")

    composer_enabled = True

    nb = ttk.Notebook(plot_frame)
    nb.pack(fill="both", expand=True)

    ctrl_frame_by_tab_id: dict[str, ttk.Frame] = {}

    def _goto_selected(_evt=None):
        wanted = plot_names_var.get()
        tab_id = tab_id_by_title.get(wanted)
        if tab_id is not None:
            nb.select(tab_id)

    plot_select.bind("<<ComboboxSelected>>", _goto_selected)

    def _sync_combo(_evt=None):
        tab_id = nb.select()
        title = title_by_tab_id.get(tab_id)
        if title:
            plot_names_var.set(title)
        if composer_enabled:
            if compose_btn.winfo_manager() != "pack":
                compose_btn.pack(side="top", pady=(8, 0), fill="x")
        else:
            if compose_btn.winfo_manager():
                compose_btn.pack_forget()
        for other_tab_id, frame in ctrl_frame_by_tab_id.items():
            if other_tab_id == tab_id:
                frame.pack(fill="both", expand=True)
            else:
                frame.pack_forget()

    nb.bind("<<NotebookTabChanged>>", _sync_combo)

    win._mpl_refs = []  # type: ignore[attr-defined]

    def _on_close():
        for _, fig in figures:
            fig.clear()
        win.destroy()
        if created_root:
            root.destroy()

    win.protocol("WM_DELETE_WINDOW", _on_close)

    marker_opts = EIS_MARKER_OPTIONS
    linestyle_opts = ["-", "--", "-.", ":", "None"]

    def _fmt(v: float) -> str:
        return f"{v:.6g}"

    def _nice_step(raw_step: float) -> float:
        if raw_step <= 0:
            return 1.0
        exponent = math.floor(math.log10(raw_step))
        fraction = raw_step / (10 ** exponent)
        if fraction <= 1:
            nice_fraction = 1
        elif fraction <= 2:
            nice_fraction = 2
        elif fraction <= 5:
            nice_fraction = 5
        else:
            nice_fraction = 10
        return nice_fraction * (10 ** exponent)

    def _snap_linear_limits(
        vmin: float,
        vmax: float,
        nbins: int = 6,
        *,
        force_zero_floor: bool = False,
    ) -> tuple[float, float]:
        if vmax < vmin:
            vmin, vmax = vmax, vmin

        if vmin == vmax:
            pad = 1.0 if vmin == 0 else abs(vmin) * 0.1
            vmin -= pad
            vmax += pad

        span = vmax - vmin
        step = _nice_step(span / max(1, nbins - 1))

        lo = math.floor(vmin / step) * step
        hi = math.ceil(vmax / step) * step

        if force_zero_floor and lo > 0:
            lo = 0.0
        if force_zero_floor and vmin >= 0:
            lo = 0.0
        if hi <= lo:
            hi = lo + step

        return float(lo), float(hi)

    def _add_tab(tab_title: str, fig: Figure) -> None:
        is_nyquist = ("nyquist" in tab_title.lower())
        tab = ttk.Frame(nb)
        nb.add(tab, text=tab_title[:28] + ("…" if len(tab_title) > 28 else ""))

        tab_id = str(tab)  # notebook tab identifier
        tab_id_by_title[tab_title] = tab_id
        title_by_tab_id[tab_id] = tab_title

        is_pt_series = bool(getattr(fig, "_pt_series", False))
        is_pre_stabilization = bool(getattr(fig, "_pre_stab_series", False))
        is_multi_series = is_pt_series or is_pre_stabilization
        is_bode_plot = bool(getattr(fig, "_bode_plot", False))

        tlow = tab_title.lower()

        def _refresh_plot_dropdown():
            visible_titles = []
            for tab_id in nb.tabs():
                if nb.tab(tab_id, "state") != "hidden":
                    title = title_by_tab_id.get(tab_id)
                    if title:
                        visible_titles.append(title)

            plot_select["values"] = visible_titles

            # if current selection got hidden, jump to first visible
            if plot_names_var.get() not in visible_titles and visible_titles:
                plot_names_var.set(visible_titles[0])
                nb.select(tab_id_by_title[visible_titles[0]])

        # right after nb.add(tab, text=...)
        current = list(plot_select["values"])
        current.append(tab_title)
        plot_select["values"] = current

        # also set initial value once
        if not plot_names_var.get():
            plot_names_var.set(tab_title)

        outer = ttk.Frame(tab)
        outer.pack(fill="both", expand=True)

        plot_frame = ttk.Frame(outer)
        plot_frame.pack(fill="both", expand=True)

        ctrl_frame = ttk.Frame(ctrl_host, padding=10)
        ctrl_frame_by_tab_id[tab_id] = ctrl_frame
        if len(ctrl_frame_by_tab_id) == 1:
            ctrl_frame.pack(fill="both", expand=True)

        canvas = FigureCanvasTkAgg(fig, master=plot_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

        toolbar = NavigationToolbar2Tk(canvas, plot_frame)
        toolbar.update()

        ax = fig.axes[0] if fig.axes else fig.add_subplot(111)
        line = ax.lines[0] if ax.lines else None

        # --- Multi-axis series controls (Series vs Pt / Pre-estabilizacion) ---
        if is_multi_series:

            def _apply_axis_color(ax, color: str, enabled: bool, side: str):
                # side: "left" or "right"
                c = color if enabled else "black"

                if side == "left":
                    ax.tick_params(axis="y", colors=c, labelleft=True, labelright=False)
                    ax.yaxis.label.set_color(c)
                    if "left" in ax.spines:
                        ax.spines["left"].set_color(c)
                    # do NOT touch right spine here
                else:
                    ax.tick_params(axis="y", colors=c, labelright=True, labelleft=False)
                    ax.yaxis.label.set_color(c)
                    if "right" in ax.spines:
                        ax.spines["right"].set_color(c)
                    # do NOT touch left spine here

            color_axes_var = tk.BooleanVar(value=True)

            def _refresh_axis_colors():
                for key, side in series_sides:
                    if key not in lines or key not in axes:
                        continue
                    _apply_axis_color(
                        axes[key],
                        lines[key].get_color(),
                        enabled=(color_axes_var.get() and lines[key].get_visible()),
                        side=side,
                    )
            
            def _pick_series_color(k: str):
                if k not in lines:
                    return
                chosen = colorchooser.askcolor(title=f"Choose color for {k}")
                if chosen and chosen[1]:
                    ln = lines[k]
                    ln.set_color(chosen[1])
                    # keep the UI var in sync (prevents "reset to default" on next style change)
                    try:
                        style_vars[k]["color"].set(chosen[1])
                    except Exception:
                        pass
                    # keep hollow marker consistent
                    try:
                        ln.set_markerfacecolor("none")
                        ln.set_markeredgecolor(chosen[1])
                    except Exception:
                        pass
                    _refresh_axis_colors()
                    _update_legend()
                    canvas.draw_idle()

            multi_title = "Series vs Pt" if is_pt_series else "Pre-estabilizacion"
            pt_box = ttk.LabelFrame(ctrl_frame, text=multi_title, padding=8)
            pt_box.pack(fill="x", pady=(0, 10))

            lines = getattr(fig, "_pt_lines", {}) if is_pt_series else getattr(fig, "_pre_stab_lines", {})
            ylabels = getattr(fig, "_pt_ylabels", {}) if is_pt_series else {
                key: line_obj.get_label() for key, line_obj in lines.items()
            }

            style_nb = ttk.Notebook(pt_box)
            style_nb.pack(fill="x", pady=(8, 0))

            style_vars = {}

            ttk.Checkbutton(pt_box, text="y-axes a color", variable=color_axes_var,
                command=lambda: (_refresh_axis_colors(), canvas.draw_idle())).pack(anchor="w", pady=(6,0))

            current_density_var = tk.BooleanVar(value=False)
            current_density_state = {"value": False}
            show_I = tk.BooleanVar(value=("I" in lines and lines["I"].get_visible()))
            show_V = tk.BooleanVar(value=("V" in lines and lines["V"].get_visible()))
            show_T = tk.BooleanVar(value=("T" in lines and lines["T"].get_visible()))

            axes = getattr(fig, "_pt_axes", {}) if is_pt_series else getattr(fig, "_pre_stab_axes", {})
            base_key = getattr(fig, "_pt_base_key", "I") if is_pt_series else "V"
            base_ax = axes.get(base_key, ax)
            series_sides = (
                (("I", "left"), ("V", "right"), ("T", "right"))
                if is_pt_series
                else (("V", "left"), ("I", "right"), ("T", "right"))
            )

            def _right_spacing_axes():
                if is_pt_series:
                    return axes
                return {"V": axes.get("I"), "T": axes.get("T")}

            def _update_legend():
                handles = []
                labels = []
                for k, ln in lines.items():
                    if ln.get_visible():
                        handles.append(ln)
                        labels.append(ln.get_label())

                try:
                    legend_fs = float(legend_fs_var.get())
                except Exception:
                    legend_fs = float(font_defaults.legend)

                leg = base_ax.get_legend()
                if leg is not None:
                    leg.remove()

                if handles:
                    make_legend_draggable(base_ax.legend(handles, labels, loc="best", fontsize=legend_fs))

            def _autoscale_visible_axes():
                for k, ln in lines.items():
                    axk = axes.get(k)
                    if axk is None:
                        continue
                    if ln.get_visible():
                        axk.relim()
                        axk.autoscale_view()

            def _current_tick_count() -> int:
                try:
                    return max(2, int(tick_count_var.get()))
                except Exception:
                    return 6

            def _current_x_tick_count() -> int:
                try:
                    return max(2, int(x_tick_count_var.get()))
                except Exception:
                    return 6

            def _apply_pt_x_axis_scaling(nbins: int | None = None):
                tick_count = _current_x_tick_count() if nbins is None else nbins
                for axx in axes.values():
                    axx.xaxis.set_major_locator(LinearLocator(tick_count))
                    axx.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
                    axx.xaxis.get_offset_text().set_visible(False)

            def _apply_pt_axis_scaling(axk, nbins: int | None = None):
                axk.yaxis.set_major_locator(LinearLocator(_current_tick_count() if nbins is None else nbins))
                axk.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
                axk.yaxis.get_offset_text().set_visible(False)

            def _autoscale_pt_axis(axk, ln, *, force_zero_floor: bool = False):
                if ln is None or not ln.get_visible():
                    return

                xdata = [float(x) for x in ln.get_xdata(orig=False) if x is not None]
                ydata = [float(y) for y in ln.get_ydata(orig=False) if y is not None]
                if not xdata or not ydata:
                    return

                x_tick_count = _current_x_tick_count()
                y_tick_count = _current_tick_count()
                x0, x1 = _snap_linear_limits(min(xdata), max(xdata), nbins=x_tick_count)
                y0, y1 = _snap_linear_limits(min(ydata), max(ydata), nbins=y_tick_count, force_zero_floor=force_zero_floor)

                axk.set_xlim(x0, x1)
                axk.set_ylim(y0, y1)
                _apply_pt_x_axis_scaling(nbins=x_tick_count)
                _apply_pt_axis_scaling(axk, nbins=y_tick_count)

            def _apply_current_density():
                if "I" not in lines or "I" not in axes:
                    return

                new_state = bool(current_density_var.get())
                old_state = bool(current_density_state["value"])
                area_cm2 = getattr(fig, "_pt_current_area_cm2", None)

                if new_state and (area_cm2 is None or area_cm2 <= 0):
                    mb.showerror(
                        "Series vs Pt",
                        "No se pudo leer un AREA valida de la metadata para convertir la corriente.",
                    )
                    current_density_var.set(old_state)
                    return

                ln = lines["I"]
                absolute_values = getattr(ln, "_eis_abs_ydata", None)
                if absolute_values is None:
                    absolute_values = tuple(float(y) for y in ln.get_ydata(orig=False) if y is not None)
                    ln._eis_abs_ydata = absolute_values

                if new_state:
                    y_values = [float(y) / area_cm2 for y in absolute_values]
                    label = getattr(ln, "_eis_density_label", "Densidad de corriente (A/cm^2)")
                else:
                    y_values = [float(y) for y in absolute_values]
                    label = getattr(ln, "_eis_abs_label", "Idc")

                ln.set_ydata(y_values)
                ln.set_label(label)
                ylabels["I"] = label
                axes["I"].set_ylabel(label)
                current_density_state["value"] = new_state

                if ln.get_visible():
                    _autoscale_pt_axis(axes["I"], ln, force_zero_floor=False)
                    try:
                        pt_init_limits["I"] = axes["I"].get_ylim()
                    except NameError:
                        pass

                _update_legend()
                _refresh_axis_colors()
                _update_right_axis_spacing(fig, canvas, _right_spacing_axes())
                try:
                    _update_limit_entries()
                except NameError:
                    pass
                canvas.draw_idle()

            def _apply_visibility():
                if "I" in lines:
                    lines["I"].set_visible(True if is_pt_series else bool(show_I.get()))
                if "V" in lines: lines["V"].set_visible(bool(show_V.get()))
                if "T" in lines: lines["T"].set_visible(bool(show_T.get()))

                for key, side in series_sides:
                    axk = axes.get(key)
                    ln = lines.get(key)
                    if axk is None:
                        continue
                    axis_visible = bool(ln is not None and ln.get_visible())
                    axk.yaxis.set_visible(axis_visible)
                    axk.yaxis.label.set_visible(axis_visible)
                    if side == "left":
                        axk.tick_params(axis="y", labelleft=axis_visible, left=axis_visible, labelright=False, right=False)
                    else:
                        axk.tick_params(axis="y", labelleft=False, left=False, labelright=axis_visible, right=axis_visible)
                    if side in axk.spines:
                        axk.spines[side].set_visible(axis_visible)

                # decide master (Idc preferred)
                visible_keys = [k for k, _side in series_sides if k in lines and lines[k].get_visible()]
                for axx in axes.values():
                    axx.grid(False)

                if visible_keys:
                    master_key = (
                        "I" if is_pt_series and "I" in visible_keys
                        else "V" if is_pre_stabilization and "V" in visible_keys
                        else visible_keys[0]
                    )
                    master_ax = axes[master_key]
                    master_ax.grid(True)

                    for key in visible_keys:
                        _autoscale_pt_axis(axes[key], lines[key], force_zero_floor=False)

                _update_right_axis_spacing(fig, canvas, _right_spacing_axes())

                _update_legend()
                _refresh_axis_colors()
                canvas.draw_idle()

            # ---- Per-series style notebook (Idc / Vdc / Temp) ----
            style_vars = {}

            def _apply_series_style(k: str):
                ln = lines[k]
                v = style_vars[k]

                c = style_vars[k]["color"].get().strip()
                if c:
                    try:
                        ln.set_color(c)
                    except Exception:
                        pass  # keep current color if user typed something invalid

                ls = v["ls"].get()
                mk = v["mk"].get()
                ln.set_linestyle("" if ls == "None" else ls)
                ln.set_marker("" if mk == "None" else mk)
                ln.set_linewidth(float(v["lw"].get()))
                ln.set_markersize(float(v["ms"].get()))

                # keep hollow markers
                if ln.get_marker() not in ("", None):
                    ln.set_markerfacecolor("none")
                    ln.set_markeredgecolor(ln.get_color())

                _update_legend()
                _refresh_axis_colors()
                canvas.draw_idle()

            def _safe_fg_for_bg(bg: str) -> str:
                if not (isinstance(bg, str) and bg.startswith("#") and len(bg) == 7):
                    return "black"
                try:
                    r = int(bg[1:3], 16); g = int(bg[3:5], 16); b = int(bg[5:7], 16)
                except ValueError:
                    return "black"
                lum = 0.2126*r + 0.7152*g + 0.0722*b
                return "black" if lum > 140 else "white"

            def _update_color_entry(entry, color: str):
                c = (color or "").strip()
                if c.startswith("#") and len(c) == 7:
                    entry.configure(bg=c, fg=_safe_fg_for_bg(c))
                else:
                    entry.configure(bg="white", fg="black")

            series_titles = (
                (("I", "Idc"), ("V", "Vdc"), ("T", "Temp"))
                if is_pt_series
                else (("I", "Corriente"), ("V", "Voltaje"), ("T", "Temperatura"))
            )
            for k, title in series_titles:
                if k not in lines:
                    continue

                f = ttk.Frame(style_nb, padding=6)
                style_nb.add(f, text=title)

                ln = lines[k]
                style_vars[k] = {
                    "color": tk.StringVar(value=str(ln.get_color())),
                    "ls": tk.StringVar(value=str(ln.get_linestyle() or "-") or "-"),
                    "mk": tk.StringVar(value=str(ln.get_marker() or "o") or "o"),
                    "lw": tk.DoubleVar(value=float(ln.get_linewidth())),
                    "ms": tk.DoubleVar(value=float(ln.get_markersize())),
                }

                ttk.Label(f, text="Color").grid(row=0, column=0, sticky="w")
                ce = tk.Entry(f, textvariable=style_vars[k]["color"], width=10)
                ce.grid(row=0, column=1, sticky="w", padx=(6, 0))

                # initialize + keep it updated as user types or picker sets value
                _update_color_entry(ce, style_vars[k]["color"].get())
                style_vars[k]["color"].trace_add("write", lambda *_ , kk=k, e=ce: _update_color_entry(e, style_vars[kk]["color"].get()))
                ttk.Button(f, text="Pick…", command=lambda kk=k: _pick_series_color(kk)).grid(row=0, column=2, sticky="w", padx=(6, 0))

                ttk.Label(f, text="Line").grid(row=1, column=0, sticky="w", pady=(6, 0))
                cb_ls = ttk.Combobox(f, textvariable=style_vars[k]["ls"], values=linestyle_opts, state="readonly", width=8)
                cb_ls.grid(row=1, column=1, sticky="w", padx=(6, 0), pady=(6, 0))

                ttk.Label(f, text="Marker").grid(row=2, column=0, sticky="w", pady=(6, 0))
                cb_mk = ttk.Combobox(f, textvariable=style_vars[k]["mk"], values=marker_opts, state="readonly", width=8)
                cb_mk.grid(row=2, column=1, sticky="w", padx=(6, 0), pady=(6, 0))

                ttk.Label(f, text="LW").grid(row=3, column=0, sticky="w", pady=(6, 0))
                sp_lw = ttk.Spinbox(f, from_=0.0, to=10.0, increment=0.1, textvariable=style_vars[k]["lw"], width=8)
                sp_lw.grid(row=3, column=1, sticky="w", padx=(6, 0), pady=(6, 0))

                ttk.Label(f, text="MS").grid(row=4, column=0, sticky="w", pady=(6, 0))
                sp_ms = ttk.Spinbox(f, from_=0.0, to=20.0, increment=0.5, textvariable=style_vars[k]["ms"], width=8)
                sp_ms.grid(row=4, column=1, sticky="w", padx=(6, 0), pady=(6, 0))

                # auto-apply
                ce.bind("<Return>", lambda e, kk=k: _apply_series_style(kk))
                ce.bind("<FocusOut>", lambda e, kk=k: _apply_series_style(kk))
                cb_ls.bind("<<ComboboxSelected>>", lambda e, kk=k: _apply_series_style(kk))
                cb_mk.bind("<<ComboboxSelected>>", lambda e, kk=k: _apply_series_style(kk))
                sp_lw.configure(command=lambda kk=k: _apply_series_style(kk))
                sp_ms.configure(command=lambda kk=k: _apply_series_style(kk))
                sp_lw.bind("<KeyRelease>", lambda e, kk=k: _apply_series_style(kk))
                sp_ms.bind("<KeyRelease>", lambda e, kk=k: _apply_series_style(kk))

                # auto-apply bindings
                ce.bind("<Return>", lambda e, kk=k: _apply_series_style(kk))
                ce.bind("<FocusOut>", lambda e, kk=k: _apply_series_style(kk))
                cb_ls.bind("<<ComboboxSelected>>", lambda e, kk=k: _apply_series_style(kk))
                cb_mk.bind("<<ComboboxSelected>>", lambda e, kk=k: _apply_series_style(kk))
                sp_lw.configure(command=lambda kk=k: _apply_series_style(kk))
                sp_ms.configure(command=lambda kk=k: _apply_series_style(kk))
                sp_lw.bind("<KeyRelease>", lambda e, kk=k: _apply_series_style(kk))
                sp_ms.bind("<KeyRelease>", lambda e, kk=k: _apply_series_style(kk))

            if is_pt_series and "I" in lines:
                ttk.Checkbutton(
                    pt_box,
                    text="Densidad de corriente",
                    variable=current_density_var,
                    command=_apply_current_density,
                ).pack(anchor="w", pady=(8, 0))
            if is_pre_stabilization and "I" in lines:
                ttk.Checkbutton(pt_box, text="Corriente", variable=show_I, command=_apply_visibility).pack(anchor="w")
            ttk.Checkbutton(
                pt_box,
                text="Vdc" if is_pt_series else "Voltaje",
                variable=show_V,
                command=_apply_visibility,
            ).pack(anchor="w")
            ttk.Checkbutton(
                pt_box,
                text="Temp" if is_pt_series else "Temperatura",
                variable=show_T,
                command=_apply_visibility,
            ).pack(anchor="w")
            if is_pt_series:
                pt_sources[tab_title] = {
                    "lines": lines,
                    "axes": axes,
                    "fig": fig,
                }
            else:
                pre_stabilization_sources[tab_title] = {
                    "lines": lines,
                    "axes": axes,
                    "fig": fig,
                }

        if is_pre_stabilization:
            pre_stabilization_sources[tab_title] = {
                "lines": getattr(fig, "_pre_stab_lines", {}),
                "axes": getattr(fig, "_pre_stab_axes", {}),
                "fig": fig,
            }

        if line is not None and not is_bode_plot:
            if "nyquist" in tlow:
                nyquist_sources[tab_title] = {"line": line, "ax": ax, "fig": fig}

        cat = None
        if "vs pt" in tlow:
            if "idc" in tlow:
                cat = "pt_i"
            elif "vdc" in tlow:
                cat = "pt_v"
            elif "temp" in tlow:
                cat = "pt_t"
        

        if is_nyquist and line is not None:
            nyquist_sources[tab_title] = {
                "line": line,   # live reference (so we can copy current formatting)
                "ax": ax,
                "fig": fig,
            }

        init_title_text = ax.get_title()
        title_text_var = tk.StringVar(value=init_title_text)
        show_title_var = tk.BooleanVar(value=True)

        def apply_title_text():
            ax.set_title(title_text_var.get() if show_title_var.get() else "", fontsize=float(title_fs_var.get()))
            try:
                fig.tight_layout()
            except Exception:
                pass
            canvas.draw_idle()

        if is_nyquist:
            ax.set_aspect("equal", adjustable="box")

        x0, x1 = ax.get_xlim()
        y0, y1 = ax.get_ylim()
        xmin_var = tk.StringVar(value=_fmt(x0))
        xmax_var = tk.StringVar(value=_fmt(x1))
        ymin_var = tk.StringVar(value=_fmt(y0))
        ymax_var = tk.StringVar(value=_fmt(y1))

        init_xlim = (x0, x1)
        init_ylim = (y0, y1)

        init_color = line.get_color() if line else ""
        init_marker = (line.get_marker() if line else "o") or "None"
        init_ls = (line.get_linestyle() if line else "-") or "None"
        init_lw = float(line.get_linewidth()) if line else 1.0
        init_ms = float(line.get_markersize()) if line else 4.0
        tick_count_var = tk.IntVar(value=6)
        x_tick_count_var = tk.IntVar(value=6)
        pt_axes = (
            getattr(fig, "_pt_axes", {})
            if is_pt_series
            else getattr(fig, "_pre_stab_axes", {}) if is_pre_stabilization else {}
        )
        pt_limit_vars: dict[str, dict[str, tk.StringVar]] = {}
        pt_init_limits: dict[str, tuple[float, float]] = {}
        bode_axes = getattr(fig, "_bode_axes", {}) if is_bode_plot else {}
        bode_lines = getattr(fig, "_bode_lines", {}) if is_bode_plot else {}
        bode_limit_vars: dict[str, dict[str, tk.StringVar]] = {}
        bode_init_limits: dict[str, tuple[float, float]] = {}
        bode_show_vars: dict[str, tk.BooleanVar] = {}
        bode_style_vars: dict[str, dict[str, object]] = {}
        bode_color_axes_var = tk.BooleanVar(value=True) if is_bode_plot else None

        if is_multi_series:
            for key in ("I", "V", "T"):
                axk = pt_axes.get(key)
                if axk is None:
                    continue
                y0k, y1k = axk.get_ylim()
                pt_limit_vars[key] = {
                    "min": tk.StringVar(value=_fmt(y0k)),
                    "max": tk.StringVar(value=_fmt(y1k)),
                }
                pt_init_limits[key] = (y0k, y1k)

        if is_bode_plot:
            for key in ("mod", "phz"):
                axk = bode_axes.get(key)
                ln = bode_lines.get(key)
                if axk is None or ln is None:
                    continue
                y0k, y1k = axk.get_ylim()
                bode_limit_vars[key] = {
                    "min": tk.StringVar(value=_fmt(y0k)),
                    "max": tk.StringVar(value=_fmt(y1k)),
                }
                bode_init_limits[key] = (y0k, y1k)
                bode_show_vars[key] = tk.BooleanVar(value=bool(ln.get_visible()))
                bode_style_vars[key] = {
                    "color": tk.StringVar(value=str(ln.get_color())),
                    "ls": tk.StringVar(value=str(ln.get_linestyle() or ("-" if key == "mod" else "--")) or ("-" if key == "mod" else "--")),
                    "mk": tk.StringVar(value=str(ln.get_marker() or "o") or "o"),
                    "lw": tk.DoubleVar(value=float(ln.get_linewidth())),
                    "ms": tk.DoubleVar(value=float(ln.get_markersize())),
                }
            if "mod" in bode_lines and "mod" in bode_axes:
                bode_sources["zmod"][tab_title] = {"line": bode_lines["mod"], "ax": bode_axes["mod"], "fig": fig}
            if "phz" in bode_lines and "phz" in bode_axes:
                bode_sources["zphz"][tab_title] = {"line": bode_lines["phz"], "ax": bode_axes["phz"], "fig": fig}

        if init_marker in (None, "", "None"):
            init_marker = "None"
        if init_ls in (None, "", "None"):
            init_ls = "None"

        color_var = tk.StringVar(value=str(init_color))
        marker_var = tk.StringVar(value=str(init_marker))
        linestyle_var = tk.StringVar(value=str(init_ls))
        lw_var = tk.DoubleVar(value=init_lw)
        ms_var = tk.DoubleVar(value=init_ms)

        def _update_limit_entries():
            a0, a1 = ax.get_xlim()
            b0, b1 = ax.get_ylim()
            xmin_var.set(_fmt(a0))
            xmax_var.set(_fmt(a1))
            ymin_var.set(_fmt(b0))
            ymax_var.set(_fmt(b1))
            if is_multi_series:
                for key, vars_map in pt_limit_vars.items():
                    axk = pt_axes.get(key)
                    if axk is None:
                        continue
                    lo, hi = axk.get_ylim()
                    vars_map["min"].set(_fmt(lo))
                    vars_map["max"].set(_fmt(hi))
            if is_bode_plot:
                for key, vars_map in bode_limit_vars.items():
                    axk = bode_axes.get(key)
                    if axk is None:
                        continue
                    lo, hi = axk.get_ylim()
                    vars_map["min"].set(_fmt(lo))
                    vars_map["max"].set(_fmt(hi))

        def _parse_float(s: str) -> float | None:
            s = s.strip()
            if not s:
                return None
            try:
                return float(s)
            except ValueError:
                return None

        def apply_axes(live: bool = False):
            if is_multi_series:
                raw_values = [xmin_var.get(), xmax_var.get()]
                for vars_map in pt_limit_vars.values():
                    raw_values.extend([vars_map["min"].get(), vars_map["max"].get()])
                if any(_is_incomplete_number(r) for r in raw_values):
                    return
                if live and any(not r.strip() for r in raw_values):
                    return

                cur_x0, cur_x1 = ax.get_xlim()
                nx0 = _parse_float(xmin_var.get())
                nx1 = _parse_float(xmax_var.get())
                new_x0 = cur_x0 if nx0 is None else nx0
                new_x1 = cur_x1 if nx1 is None else nx1

                for key, vars_map in pt_limit_vars.items():
                    axk = pt_axes.get(key)
                    if axk is None:
                        continue
                    cur_y0, cur_y1 = axk.get_ylim()
                    ny0 = _parse_float(vars_map["min"].get())
                    ny1 = _parse_float(vars_map["max"].get())
                    axk.set_xlim(new_x0, new_x1)
                    axk.set_ylim(cur_y0 if ny0 is None else ny0, cur_y1 if ny1 is None else ny1)
                    _apply_pt_axis_scaling(axk)
                _apply_pt_x_axis_scaling()

                canvas.draw_idle()
                if not live:
                    _update_limit_entries()
                return
            if is_bode_plot:
                raw_values = [xmin_var.get(), xmax_var.get()]
                for vars_map in bode_limit_vars.values():
                    raw_values.extend([vars_map["min"].get(), vars_map["max"].get()])
                if any(_is_incomplete_number(r) for r in raw_values):
                    return
                if live and any(not r.strip() for r in raw_values):
                    return

                cur_x0, cur_x1 = ax.get_xlim()
                nx0 = _parse_float(xmin_var.get())
                nx1 = _parse_float(xmax_var.get())
                new_x0 = cur_x0 if nx0 is None else nx0
                new_x1 = cur_x1 if nx1 is None else nx1
                if new_x0 <= 0 or new_x1 <= 0:
                    if not live:
                        _update_limit_entries()
                    return

                for key, vars_map in bode_limit_vars.items():
                    axk = bode_axes.get(key)
                    if axk is None:
                        continue
                    cur_y0, cur_y1 = axk.get_ylim()
                    ny0 = _parse_float(vars_map["min"].get())
                    ny1 = _parse_float(vars_map["max"].get())
                    axk.set_xlim(new_x0, new_x1)
                    axk.set_ylim(cur_y0 if ny0 is None else ny0, cur_y1 if ny1 is None else ny1)
                    axk.yaxis.set_major_locator(LinearLocator(max(2, int(tick_count_var.get()))))
                    axk.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
                    axk.yaxis.get_offset_text().set_visible(False)

                canvas.draw_idle()
                if not live:
                    _update_limit_entries()
                return
            raws = [xmin_var.get(), xmax_var.get(), ymin_var.get(), ymax_var.get()]
            if any(_is_incomplete_number(r) for r in raws):
                return  # don't apply AND don't refresh the fields
            if live and any(not r.strip() for r in raws):
                return
            cur_x0, cur_x1 = ax.get_xlim()
            cur_y0, cur_y1 = ax.get_ylim()

            nx0 = _parse_float(xmin_var.get())
            nx1 = _parse_float(xmax_var.get())
            ny0 = _parse_float(ymin_var.get())
            ny1 = _parse_float(ymax_var.get())

            new_x0 = cur_x0 if nx0 is None else nx0
            new_x1 = cur_x1 if nx1 is None else nx1
            new_y0 = cur_y0 if ny0 is None else ny0
            new_y1 = cur_y1 if ny1 is None else ny1

            if ax.get_xscale() == "log" and (new_x0 <= 0 or new_x1 <= 0):
                if not live:
                    _update_limit_entries()
                return
            if ax.get_yscale() == "log" and (new_y0 <= 0 or new_y1 <= 0):
                if not live:
                    _update_limit_entries()
                return

            ax.set_xlim(new_x0, new_x1)
            ax.set_ylim(new_y0, new_y1)
            ax.yaxis.set_major_locator(LinearLocator(max(2, int(tick_count_var.get()))))

            if is_nyquist:
                ax.set_aspect("equal", adjustable="box")

            canvas.draw_idle()
            if not live:
                _update_limit_entries()

        def reset_axes():
            if is_multi_series:
                ax.set_xlim(*init_xlim)
                for key, limits in pt_init_limits.items():
                    axk = pt_axes.get(key)
                    if axk is None:
                        continue
                    axk.set_xlim(*init_xlim)
                    axk.set_ylim(*limits)
                    _apply_pt_axis_scaling(axk)
                _apply_pt_x_axis_scaling()
                canvas.draw_idle()
                _update_limit_entries()
                return
            if is_bode_plot:
                ax.set_xlim(*init_xlim)
                for key, limits in bode_init_limits.items():
                    axk = bode_axes.get(key)
                    if axk is None:
                        continue
                    axk.set_xlim(*init_xlim)
                    axk.set_ylim(*limits)
                    axk.yaxis.set_major_locator(LinearLocator(max(2, int(tick_count_var.get()))))
                    axk.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
                    axk.yaxis.get_offset_text().set_visible(False)
                canvas.draw_idle()
                _update_limit_entries()
                return

            ax.set_xlim(*init_xlim)
            ax.set_ylim(*init_ylim)
            ax.yaxis.set_major_locator(LinearLocator(max(2, int(tick_count_var.get()))))
            if is_nyquist:
                ax.set_aspect("equal", adjustable="box")
            canvas.draw_idle()
            _update_limit_entries()

        def autoscale_axes():
            tick_count = max(2, int(tick_count_var.get()))
            if is_multi_series:
                _apply_visibility()
                _update_limit_entries()
                return
            if is_bode_plot:
                visible_keys = [k for k, var in bode_show_vars.items() if var.get() and k in bode_lines]
                if not visible_keys:
                    return

                x_values: list[float] = []
                for key in visible_keys:
                    ln = bode_lines[key]
                    x_values.extend([float(x) for x in ln.get_xdata(orig=False) if x is not None and float(x) > 0])

                if not x_values:
                    return

                x0, x1 = min(x_values), max(x_values)
                ax.set_xlim(x0, x1)

                for key in visible_keys:
                    ln = bode_lines[key]
                    axk = bode_axes[key]
                    y_values = [float(y) for y in ln.get_ydata(orig=False) if y is not None]
                    if not y_values:
                        continue
                    y0, y1 = _snap_linear_limits(min(y_values), max(y_values), nbins=tick_count)
                    axk.set_xlim(x0, x1)
                    axk.set_ylim(y0, y1)
                    axk.yaxis.set_major_locator(LinearLocator(tick_count))
                    axk.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
                    axk.yaxis.get_offset_text().set_visible(False)

                canvas.draw_idle()
                _update_limit_entries()
                return

            if line is None:
                ax.relim()
                ax.autoscale_view()
                x0, x1 = ax.get_xlim()
                y0, y1 = ax.get_ylim()
            else:
                xdata = [float(x) for x in line.get_xdata(orig=False) if x is not None]
                ydata = [float(y) for y in line.get_ydata(orig=False) if y is not None]
                if not xdata or not ydata:
                    return
                x0, x1 = min(xdata), max(xdata)
                y0, y1 = min(ydata), max(ydata)

            if ax.get_xscale() == "log":
                ax.set_xlim(x0, x1)
            else:
                x0, x1 = _snap_linear_limits(x0, x1, nbins=tick_count)
                ax.set_xlim(x0, x1)

            if ax.get_yscale() == "log":
                ax.set_ylim(y0, y1)
            else:
                y0, y1 = _snap_linear_limits(
                    y0,
                    y1,
                    nbins=tick_count,
                    force_zero_floor=is_nyquist,
                )
                ax.set_ylim(y0, y1)

            ax.yaxis.set_major_locator(LinearLocator(tick_count))

            if is_nyquist:
                ax.set_aspect("equal", adjustable="box")
            canvas.draw_idle()
            _update_limit_entries()

        def _safe_fg_for_bg(bg: str) -> str:
            # bg like "#RRGGBB"; returns black/white for readability
            if not (isinstance(bg, str) and bg.startswith("#") and len(bg) == 7):
                return "black"
            try:
                r = int(bg[1:3], 16)
                g = int(bg[3:5], 16)
                b = int(bg[5:7], 16)
            except ValueError:
                return "black"
            # perceived luminance
            lum = 0.2126*r + 0.7152*g + 0.0722*b
            return "black" if lum > 140 else "white"

        def _update_color_entry_bg():
            c = color_var.get().strip()
            if c.startswith("#") and len(c) == 7:
                try:
                    color_entry.configure(background=c, foreground=_safe_fg_for_bg(c))
                except Exception:
                    # ttk.Entry may ignore background on some themes; fallback below
                    pass

        def apply_style():
            if line is None:
                return

            c = color_var.get().strip()
            if c:
                line.set_color(c)

            ls = linestyle_var.get().strip()
            line.set_linestyle("" if ls == "None" else ls)

            mk = marker_var.get().strip()
            line.set_marker("" if mk == "None" else mk)

            # Keep markers hollow by default
            if line.get_marker() not in ("", None):
                line.set_markerfacecolor("none")
                line.set_markeredgecolor(line.get_color())

            try:
                line.set_linewidth(float(lw_var.get()))
            except Exception:
                pass
            try:
                line.set_markersize(float(ms_var.get()))
            except Exception:
                pass
            _update_color_entry_bg()
            canvas.draw_idle()

        def reset_style():
            if line is None:
                return
            color_var.set(str(init_color))
            marker_var.set(str(init_marker))
            linestyle_var.set(str(init_ls))
            lw_var.set(init_lw)
            ms_var.set(init_ms)
            apply_style()

        def _apply_bode_visibility():
            if not is_bode_plot:
                return

            visible_keys = []
            for key, var in bode_show_vars.items():
                ln = bode_lines.get(key)
                axk = bode_axes.get(key)
                if ln is None or axk is None:
                    continue
                visible = bool(var.get())
                ln.set_visible(visible)
                if visible:
                    visible_keys.append(key)
                    axk.tick_params(axis="y", labelleft=(key == "mod"), labelright=(key == "phz"))
                    axk.yaxis.label.set_visible(True)
                else:
                    axk.tick_params(axis="y", labelleft=False, labelright=False)
                    axk.yaxis.label.set_visible(False)

            handles = []
            labels = []
            for key in ("mod", "phz"):
                ln = bode_lines.get(key)
                if ln is not None and ln.get_visible():
                    handles.append(ln)
                    labels.append(ln.get_label())

            leg = ax.get_legend()
            if leg is not None:
                leg.remove()
            if handles:
                try:
                    legend_fs = float(legend_fs_var.get())
                except Exception:
                    legend_fs = float(font_defaults.legend)
                make_legend_draggable(ax.legend(handles, labels, loc="best", fontsize=legend_fs))

            for key in visible_keys:
                bode_axes[key].yaxis.set_major_locator(LinearLocator(max(2, int(tick_count_var.get()))))
                bode_axes[key].yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
                bode_axes[key].yaxis.get_offset_text().set_visible(False)

            use_axis_colors = bool(bode_color_axes_var.get()) if bode_color_axes_var is not None else True
            for key in ("mod", "phz"):
                axk = bode_axes.get(key)
                ln = bode_lines.get(key)
                if axk is None:
                    continue
                axis_color = "black"
                if use_axis_colors and ln is not None and ln.get_visible():
                    axis_color = ln.get_color()
                axk.yaxis.label.set_color(axis_color)
                axk.tick_params(axis="y", colors=axis_color)
                side = "left" if key == "mod" else "right"
                if side in axk.spines:
                    axk.spines[side].set_color(axis_color)

            canvas.draw_idle()

        def _apply_bode_style(key: str):
            if not is_bode_plot or key not in bode_lines or key not in bode_style_vars:
                return
            ln = bode_lines[key]
            vals = bode_style_vars[key]
            color = vals["color"].get().strip()
            if color:
                ln.set_color(color)
            ln.set_linestyle("" if vals["ls"].get() == "None" else vals["ls"].get())
            ln.set_marker("" if vals["mk"].get() == "None" else vals["mk"].get())
            ln.set_linewidth(float(vals["lw"].get()))
            ln.set_markersize(float(vals["ms"].get()))
            if ln.get_marker() not in ("", None):
                ln.set_markerfacecolor("none")
                ln.set_markeredgecolor(ln.get_color())
            _apply_bode_visibility()

        # ---------------- Fonts ----------------
        # Initial font sizes (grab from current artists)
        try:
            init_tick_fs = float(ax.get_xticklabels()[0].get_fontsize()) if ax.get_xticklabels() else 10.0
        except Exception:
            init_tick_fs = 10.0

        try:
            init_label_fs = float(ax.xaxis.label.get_fontsize() or 12.0)
        except Exception:
            init_label_fs = 12.0

        try:
            init_title_fs = float(ax.title.get_fontsize() or 14.0)
        except Exception:
            init_title_fs = 14.0
        try:
            init_legend_fs = float(ax.get_legend().get_texts()[0].get_fontsize()) if ax.get_legend() else float(font_defaults.legend)
        except Exception:
            init_legend_fs = float(font_defaults.legend)

        tick_fs_var = tk.DoubleVar(value=init_tick_fs)
        label_fs_var = tk.DoubleVar(value=init_label_fs)
        title_fs_var = tk.DoubleVar(value=init_title_fs)
        legend_fs_var = tk.DoubleVar(value=init_legend_fs)
        refresh_frequency_labels = None

        def apply_fonts():
            # Apply to all axes in the figure (safe even if later you add multi-axes figs)
            try:
                tfs = float(tick_fs_var.get())
                lfs = float(label_fs_var.get())
                hfs = float(title_fs_var.get())
                gfs = float(legend_fs_var.get())
            except Exception:
                return

            for ax_ in fig.axes:
                ax_.tick_params(labelsize=tfs)
                ax_.xaxis.label.set_fontsize(lfs)
                ax_.yaxis.label.set_fontsize(lfs)
                ax_.title.set_fontsize(hfs)
                leg = ax_.get_legend()
                if leg is not None:
                    for t in leg.get_texts():
                        t.set_fontsize(gfs)
                    make_legend_draggable(leg)

            if is_multi_series:
                _update_legend()
            if is_bode_plot:
                _apply_bode_visibility()
            if callable(refresh_frequency_labels):
                refresh_frequency_labels()

            # Layout may need refresh when fonts change
            try:
                fig.tight_layout()
            except Exception:
                pass

            if is_multi_series:
                _update_right_axis_spacing(fig, canvas, _right_spacing_axes())

            canvas.draw_idle()

        def reset_fonts():
            tick_fs_var.set(init_tick_fs)
            label_fs_var.set(init_label_fs)
            title_fs_var.set(init_title_fs)
            legend_fs_var.set(init_legend_fs)
            title_text_var.set(init_title_text)
            show_title_var.set(True)
            apply_fonts()
            apply_title_text()

        def pick_color():
            if line is None:
                return
            chosen = colorchooser.askcolor(title="Choose line color")
            if chosen and chosen[1]:
                color_var.set(chosen[1])
                apply_style()

        # --- Controls UI ---
        axes_box = ttk.LabelFrame(ctrl_frame, text="Límites de ejes", padding=8)
        axes_box.pack(fill="x", pady=(0, 10))

        def _row(parent, r, label, var):
            ttk.Label(parent, text=label, width=5).grid(row=r, column=0, sticky="w", padx=(0, 6), pady=2)
            e = ttk.Entry(parent, textvariable=var, width=12)
            e.grid(row=r, column=1, sticky="w", pady=2)
            return e

        axis_entries = []
        next_row = 0
        axis_entries.append(_row(axes_box, next_row, "Xmin", xmin_var))
        next_row += 1
        axis_entries.append(_row(axes_box, next_row, "Xmax", xmax_var))
        next_row += 1
        if is_multi_series:
            for key, label in (("I", "I"), ("V", "V"), ("T", "T")):
                if key not in pt_limit_vars:
                    continue
                axis_entries.append(_row(axes_box, next_row, f"{label}min", pt_limit_vars[key]["min"]))
                next_row += 1
                axis_entries.append(_row(axes_box, next_row, f"{label}max", pt_limit_vars[key]["max"]))
                next_row += 1
        elif is_bode_plot:
            for key, label in (("mod", "M"), ("phz", "P")):
                if key not in bode_limit_vars:
                    continue
                axis_entries.append(_row(axes_box, next_row, f"{label}min", bode_limit_vars[key]["min"]))
                next_row += 1
                axis_entries.append(_row(axes_box, next_row, f"{label}max", bode_limit_vars[key]["max"]))
                next_row += 1
        else:
            axis_entries.append(_row(axes_box, next_row, "Ymin", ymin_var))
            next_row += 1
            axis_entries.append(_row(axes_box, next_row, "Ymax", ymax_var))
            next_row += 1
        x_tick_count_spin = None
        if is_multi_series:
            ttk.Label(axes_box, text="X ticks", width=7).grid(row=next_row, column=0, sticky="w", padx=(0, 6), pady=2)
            x_tick_count_spin = ttk.Spinbox(axes_box, from_=2, to=20, increment=1, textvariable=x_tick_count_var, width=10)
            x_tick_count_spin.grid(row=next_row, column=1, sticky="w", pady=2)
            next_row += 1
            ttk.Label(axes_box, text="Y ticks", width=7).grid(row=next_row, column=0, sticky="w", padx=(0, 6), pady=2)
        else:
            ttk.Label(axes_box, text="Ticks", width=5).grid(row=next_row, column=0, sticky="w", padx=(0, 6), pady=2)
        tick_count_spin = ttk.Spinbox(axes_box, from_=2, to=20, increment=1, textvariable=tick_count_var, width=10)
        tick_count_spin.grid(row=next_row, column=1, sticky="w", pady=2)

        pending_axes = {"id": None}

        def _schedule_apply_axes(_evt=None):
            if pending_axes["id"] is not None:
                tab.after_cancel(pending_axes["id"])
            pending_axes["id"] = tab.after(900, lambda: apply_axes(live=True))


        for entry in axis_entries:
            entry.bind("<Return>", lambda ev: apply_axes(live=False))
            entry.bind("<FocusOut>", lambda ev: apply_axes(live=False))

        btns_axes = ttk.Frame(axes_box)
        btns_axes.grid(row=next_row + 1, column=0, columnspan=2, sticky="ew", pady=(8, 0))
        ttk.Button(btns_axes, text="Autoescala", command=autoscale_axes).pack(side="left", expand=True, fill="x", padx=(0, 6))
        ttk.Button(btns_axes, text="Restablecer", command=reset_axes).pack(side="left", expand=True, fill="x")
        tick_count_spin.configure(command=autoscale_axes)
        tick_count_spin.bind("<Return>", lambda ev: autoscale_axes())
        tick_count_spin.bind("<FocusOut>", lambda ev: autoscale_axes())
        if x_tick_count_spin is not None:
            x_tick_count_spin.configure(command=autoscale_axes)
            x_tick_count_spin.bind("<Return>", lambda ev: autoscale_axes())
            x_tick_count_spin.bind("<FocusOut>", lambda ev: autoscale_axes())

        if is_bode_plot:
            bode_box = ttk.LabelFrame(ctrl_frame, text="Bode", padding=8)
            bode_box.pack(fill="x", pady=(0, 10))
            if "mod" in bode_show_vars:
                ttk.Checkbutton(bode_box, text="Zmod", variable=bode_show_vars["mod"], command=lambda: (_apply_bode_visibility(), autoscale_axes())).pack(anchor="w")
            if "phz" in bode_show_vars:
                ttk.Checkbutton(bode_box, text="Zphz", variable=bode_show_vars["phz"], command=lambda: (_apply_bode_visibility(), autoscale_axes())).pack(anchor="w")
            ttk.Checkbutton(
                bode_box,
                text="y-axes a color",
                variable=bode_color_axes_var,
                command=_apply_bode_visibility,
            ).pack(anchor="w", pady=(6, 0))

            bode_style_box = ttk.LabelFrame(ctrl_frame, text="Estilo", padding=8)
            bode_style_box.pack(fill="x", pady=(0, 10))
            bode_style_nb = ttk.Notebook(bode_style_box)
            bode_style_nb.pack(fill="x")

            for key, title in (("mod", "Zmod"), ("phz", "Zphz")):
                if key not in bode_style_vars:
                    continue
                f = ttk.Frame(bode_style_nb, padding=6)
                bode_style_nb.add(f, text=title)
                vars_map = bode_style_vars[key]

                ttk.Label(f, text="Color").grid(row=0, column=0, sticky="w")
                color_entry_local = tk.Entry(f, textvariable=vars_map["color"], width=10)
                color_entry_local.grid(row=0, column=1, sticky="w", padx=(6, 0))

                def _update_bode_color_entry(entry, color_value: str):
                    c = (color_value or "").strip()
                    if c.startswith("#") and len(c) == 7:
                        entry.configure(bg=c, fg=_safe_fg_for_bg(c))
                    else:
                        entry.configure(bg="white", fg="black")

                _update_bode_color_entry(color_entry_local, vars_map["color"].get())
                vars_map["color"].trace_add(
                    "write",
                    lambda *_args, e=color_entry_local, v=vars_map["color"]: _update_bode_color_entry(e, v.get()),
                )

                def _pick_bode_color(kk=key):
                    chosen = colorchooser.askcolor(title=f"Choose color for {kk}")
                    if chosen and chosen[1]:
                        bode_style_vars[kk]["color"].set(chosen[1])
                        _apply_bode_style(kk)

                ttk.Button(f, text="Pick…", command=_pick_bode_color).grid(row=0, column=2, sticky="w", padx=(6, 0))

                ttk.Label(f, text="Line").grid(row=1, column=0, sticky="w", pady=(6, 0))
                cb_ls = ttk.Combobox(f, textvariable=vars_map["ls"], values=linestyle_opts, state="readonly", width=8)
                cb_ls.grid(row=1, column=1, sticky="w", padx=(6, 0), pady=(6, 0))

                ttk.Label(f, text="Marker").grid(row=2, column=0, sticky="w", pady=(6, 0))
                cb_mk = ttk.Combobox(f, textvariable=vars_map["mk"], values=marker_opts, state="readonly", width=8)
                cb_mk.grid(row=2, column=1, sticky="w", padx=(6, 0), pady=(6, 0))

                ttk.Label(f, text="LW").grid(row=3, column=0, sticky="w", pady=(6, 0))
                sp_lw = ttk.Spinbox(f, from_=0.0, to=10.0, increment=0.1, textvariable=vars_map["lw"], width=8)
                sp_lw.grid(row=3, column=1, sticky="w", padx=(6, 0), pady=(6, 0))

                ttk.Label(f, text="MS").grid(row=4, column=0, sticky="w", pady=(6, 0))
                sp_ms = ttk.Spinbox(f, from_=0.0, to=20.0, increment=0.5, textvariable=vars_map["ms"], width=8)
                sp_ms.grid(row=4, column=1, sticky="w", padx=(6, 0), pady=(6, 0))

                color_entry_local.bind("<Return>", lambda e, kk=key: _apply_bode_style(kk))
                color_entry_local.bind("<FocusOut>", lambda e, kk=key: _apply_bode_style(kk))
                cb_ls.bind("<<ComboboxSelected>>", lambda e, kk=key: _apply_bode_style(kk))
                cb_mk.bind("<<ComboboxSelected>>", lambda e, kk=key: _apply_bode_style(kk))
                sp_lw.configure(command=lambda kk=key: _apply_bode_style(kk))
                sp_lw.bind("<KeyRelease>", lambda e, kk=key: _apply_bode_style(kk))
                sp_ms.configure(command=lambda kk=key: _apply_bode_style(kk))
                sp_ms.bind("<KeyRelease>", lambda e, kk=key: _apply_bode_style(kk))

        style_box = ttk.LabelFrame(ctrl_frame, text="Estilo", padding=8)
        if not is_multi_series and not is_bode_plot:
            style_box.pack(fill="x", pady=(0, 10))
        # else: don’t pack it (it won’t appear, but code stays safe)

        fonts_box = ttk.LabelFrame(ctrl_frame, text="Fuentes", padding=8)
        fonts_box.pack(fill="x", pady=(0, 10))
        fonts_box.columnconfigure(1, weight=1)

        ttk.Label(fonts_box, text="Ticks").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=2)
        tick_spin = ttk.Spinbox(fonts_box, from_=6.0, to=30.0, increment=0.5,
                                textvariable=tick_fs_var, width=10)
        tick_spin.grid(row=0, column=1, sticky="w", pady=2)

        ttk.Label(fonts_box, text="Labels").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=2)
        label_spin = ttk.Spinbox(fonts_box, from_=6.0, to=40.0, increment=0.5,
                                textvariable=label_fs_var, width=10)
        label_spin.grid(row=1, column=1, sticky="w", pady=2)

        ttk.Label(fonts_box, text="Leyenda").grid(row=2, column=0, sticky="w", padx=(0, 6), pady=2)
        legend_spin = ttk.Spinbox(fonts_box, from_=6.0, to=40.0, increment=0.5,
                                textvariable=legend_fs_var, width=10)
        legend_spin.grid(row=2, column=1, sticky="w", pady=2)

        ttk.Label(fonts_box, text="Título").grid(row=3, column=0, sticky="w", padx=(0, 6), pady=2)
        title_spin = ttk.Spinbox(fonts_box, from_=6.0, to=50.0, increment=0.5,
                                textvariable=title_fs_var, width=10)
        title_spin.grid(row=3, column=1, sticky="w", pady=2)

        ttk.Label(fonts_box, text="Text").grid(row=4, column=0, sticky="w", padx=(0, 6), pady=2)
        title_entry = ttk.Entry(fonts_box, textvariable=title_text_var, width=18)
        title_entry.grid(row=4, column=1, sticky="w", pady=2)
        title_entry.grid(row=4, column=1, sticky="ew", pady=2)
        ttk.Checkbutton(fonts_box, text="Mostrar titulo", variable=show_title_var, command=apply_title_text).grid(
            row=6, column=0, columnspan=2, sticky="w", pady=(6, 0)
        )

        btns_fonts = ttk.Frame(fonts_box)
        btns_fonts.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(8, 0))
        ttk.Button(btns_fonts, text="Restablecer", command=reset_fonts).pack(side="left", expand=True, fill="x")

        ttk.Label(style_box, text="Color").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=2)

        color_entry = tk.Entry(style_box, textvariable=color_var, width=12)
        color_entry.grid(row=0, column=1, sticky="w", pady=2)
        _update_color_entry_bg()
        ttk.Button(style_box, text="Pick…", command=pick_color).grid(row=0, column=2, sticky="w", padx=(6, 0), pady=2)

        ttk.Label(style_box, text="Line").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=2)
        linestyle_cb = ttk.Combobox(style_box, textvariable=linestyle_var, values=linestyle_opts, width=9, state="readonly")
        linestyle_cb.grid(row=1, column=1, sticky="w", pady=2)

        ttk.Label(style_box, text="Marker").grid(row=2, column=0, sticky="w", padx=(0, 6), pady=2)
        marker_cb = ttk.Combobox(style_box, textvariable=marker_var, values=marker_opts, width=9, state="readonly")
        marker_cb.grid(row=2, column=1, sticky="w", pady=2)

        ttk.Label(style_box, text="LW").grid(row=3, column=0, sticky="w", padx=(0, 6), pady=2)
        lw_spin = ttk.Spinbox(style_box, from_=0.0, to=10.0, increment=0.1, textvariable=lw_var, width=10)
        lw_spin.grid(row=3, column=1, sticky="w", pady=2)

        ttk.Label(style_box, text="MS").grid(row=4, column=0, sticky="w", padx=(0, 6), pady=2)
        ms_spin = ttk.Spinbox(style_box, from_=0.0, to=20.0, increment=0.5, textvariable=ms_var, width=10)
        ms_spin.grid(row=4, column=1, sticky="w", pady=2)

        # Apply on arrow clicks + typing
        tick_spin.configure(command=apply_fonts)
        label_spin.configure(command=apply_fonts)
        legend_spin.configure(command=apply_fonts)
        title_spin.configure(command=apply_fonts)

        tick_spin.bind("<KeyRelease>", lambda e: apply_fonts())
        label_spin.bind("<KeyRelease>", lambda e: apply_fonts())
        legend_spin.bind("<KeyRelease>", lambda e: apply_fonts())
        title_spin.bind("<KeyRelease>", lambda e: apply_fonts())

        # Comboboxes apply instantly on selection
        linestyle_cb.bind("<<ComboboxSelected>>", lambda e: apply_style())
        marker_cb.bind("<<ComboboxSelected>>", lambda e: apply_style())

        # Spinboxes: apply on arrow clicks + typing
        lw_spin.configure(command=apply_style)
        ms_spin.configure(command=apply_style)
        lw_spin.bind("<KeyRelease>", lambda e: apply_style())
        ms_spin.bind("<KeyRelease>", lambda e: apply_style())

        # Color entry: apply on Enter or leaving the field
        color_entry.bind("<Return>", lambda e: apply_style())
        color_entry.bind("<FocusOut>", lambda e: apply_style())

        btns_style = ttk.Frame(style_box)
        btns_style.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(8, 0))
        ttk.Button(btns_style, text="Restablecer", command=reset_style).pack(side="left", expand=True, fill="x")

        pending_title = {"id": None}

        def _schedule_title(_evt=None):
            if pending_title["id"] is not None:
                tab.after_cancel(pending_title["id"])
            pending_title["id"] = tab.after(250, apply_title_text)

        # ---------------- Frequency tools (Nyquist only) ----------------
        freqs = getattr(line, "_eis_freq", None) if line is not None else None

        def _fmt_freq_hz(v: float) -> str:
            # nice readable frequency formatting
            av = abs(v)
            if av >= 1e6:
                return f"{v/1e6:.3g} MHz"
            if av >= 1e3:
                return f"{v/1e3:.3g} kHz"
            return f"{v:.3g} Hz"

        # Only enable this panel for Nyquist plots that actually have freq data
        if is_nyquist and line is not None and isinstance(freqs, list) and len(freqs) == len(line.get_xdata()):
            freq_box = ttk.LabelFrame(ctrl_frame, text="Frequency", padding=8)
            freq_box.pack(fill="x", pady=(0, 10))

            # --- Hover tooltip ---
            hover_var = tk.BooleanVar(value=True)

            hover_annot = ax.annotate(
                "",
                xy=(0, 0),
                xytext=(10, 10),
                textcoords="offset points",
                bbox=dict(boxstyle="round", fc="white", alpha=0.9),
                arrowprops=dict(arrowstyle="->", alpha=0.7),
            )
            hover_annot.set_visible(False)

            def _toggle_hover():
                if not hover_var.get():
                    hover_annot.set_visible(False)
                    canvas.draw_idle()

            ttk.Checkbutton(freq_box, text="Hover shows frequency", variable=hover_var, command=_toggle_hover)\
                .grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 6))

            ttk.Label(
                freq_box,
                text="Click a point to add or remove its frequency label.",
                wraplength=220,
                justify="left",
            ).grid(row=1, column=0, columnspan=2, sticky="w", pady=(0, 6))

            def _nearest_freq_point_index(event) -> int | None:
                if event.inaxes != ax or event.x is None or event.y is None:
                    return None
                xdata = list(line.get_xdata(orig=False))
                ydata = list(line.get_ydata(orig=False))

                best_i = None
                best_d2 = 1e18
                # threshold ~12 px
                thresh2 = 12.0 * 12.0

                for i, (xv, yv) in enumerate(zip(xdata, ydata)):
                    xp, yp = ax.transData.transform((xv, yv))
                    d2 = (xp - event.x) ** 2 + (yp - event.y) ** 2
                    if d2 < best_d2:
                        best_d2 = d2
                        best_i = i

                if best_i is None or best_d2 > thresh2:
                    return None
                return best_i

            def _on_move(event):
                if not hover_var.get():
                    return

                best_i = _nearest_freq_point_index(event)
                if best_i is None:
                    if hover_annot.get_visible():
                        hover_annot.set_visible(False)
                        canvas.draw_idle()
                    return

                xdata = list(line.get_xdata(orig=False))
                ydata = list(line.get_ydata(orig=False))
                fval = freqs[best_i]
                hover_annot.xy = (xdata[best_i], ydata[best_i])
                hover_annot.set_text(f"f = {_fmt_freq_hz(float(fval))}")
                if not hover_annot.get_visible():
                    hover_annot.set_visible(True)
                canvas.draw_idle()

            canvas.mpl_connect("motion_notify_event", _on_move)

            # --- Static labels / click-to-pin labels ---
            ttk.Separator(freq_box).grid(row=2, column=0, columnspan=2, sticky="ew", pady=6)

            # Defaults: full range, labels OFF (N=0)
            fmin_default = min(freqs)
            fmax_default = max(freqs)

            freqmin_var = tk.StringVar(value=f"{fmin_default:.6g}")
            freqmax_var = tk.StringVar(value=f"{fmax_default:.6g}")
            nlabels_var = tk.IntVar(value=0)

            static_artists: list[object] = []
            auto_label_idxs: list[int] = []
            manual_label_idxs: list[int] = []
            auto_label_spec: dict[str, float | int] | None = None

            def _sync_freq_label_metadata(idxs: list[int], spec: dict[str, float | int] | None = None):
                try:
                    line._freq_label_idxs = list(idxs)  # type: ignore[attr-defined]
                    line._freq_label_spec = spec        # type: ignore[attr-defined]
                except Exception:
                    pass

            def _clear_label_artists():
                nonlocal static_artists
                for a in static_artists:
                    try:
                        a.remove()
                    except Exception:
                        pass
                static_artists = []

            def _redraw_frequency_labels():
                _clear_label_artists()

                idxs = sorted(set(auto_label_idxs + manual_label_idxs))
                if not idxs:
                    _sync_freq_label_metadata([], None)
                    canvas.draw_idle()
                    return

                xdata = list(line.get_xdata(orig=False))
                ydata = list(line.get_ydata(orig=False))
                label_fs = max(7.0, float(label_fs_var.get()))

                for i in idxs:
                    txt = _fmt_freq_hz(float(freqs[i]))
                    a = ax.annotate(
                        txt,
                        xy=(xdata[i], ydata[i]),
                        xytext=(6, 6),
                        textcoords="offset points",
                        fontsize=label_fs,
                        bbox=dict(boxstyle="round,pad=0.15", fc="white", alpha=0.7),
                    )
                    static_artists.append(a)

                spec = auto_label_spec if auto_label_idxs and not manual_label_idxs and idxs == sorted(auto_label_idxs) else None
                _sync_freq_label_metadata(idxs, spec)
                canvas.draw_idle()

            refresh_frequency_labels = _redraw_frequency_labels

            def _clear_all_frequency_labels():
                nonlocal auto_label_idxs, manual_label_idxs, auto_label_spec
                auto_label_idxs = []
                manual_label_idxs = []
                auto_label_spec = None
                _redraw_frequency_labels()

            def _parse_float_or_none(s: str) -> float | None:
                s = s.strip()
                if not s:
                    return None
                try:
                    return float(s)
                except ValueError:
                    return None

            def _apply_static_labels():
                nonlocal auto_label_idxs, auto_label_spec
                raws = [freqmin_var.get(), freqmax_var.get()]
                if any(_is_incomplete_number(r) for r in raws):
                    return

                n = int(nlabels_var.get())
                if n <= 0:
                    auto_label_idxs = []
                    auto_label_spec = None
                    _redraw_frequency_labels()
                    return

                fmin_in = _parse_float_or_none(freqmin_var.get())
                fmax_in = _parse_float_or_none(freqmax_var.get())

                # if blank/invalid, use full range
                fmin_use = fmin_default if fmin_in is None else fmin_in
                fmax_use = fmax_default if fmax_in is None else fmax_in

                lo = min(fmin_use, fmax_use)
                hi = max(fmin_use, fmax_use)

                # candidate indices in the requested freq window
                candidates = [i for i, fv in enumerate(freqs) if lo <= float(fv) <= hi]

                if not candidates:
                    # fallback: whole curve
                    candidates = list(range(len(freqs)))

                # Choose indices:
                if n == 1:
                    target = fmin_use  # spec: nearest to Freqmin
                    idx = min(candidates, key=lambda i: abs(float(freqs[i]) - target))
                    idxs = [idx]
                else:
                    # evenly spaced in point number *within* candidates
                    m = len(candidates)
                    if n > m:
                        n = m
                    if n == 1:
                        idxs = [candidates[0]]
                    else:
                        pos = [round(k * (m - 1) / (n - 1)) for k in range(n)]
                        idxs = sorted({candidates[p] for p in pos})

                auto_label_idxs = list(idxs)
                auto_label_spec = {
                    "freqmin": fmin_use,
                    "freqmax": fmax_use,
                    "n": len(auto_label_idxs),
                }
                _redraw_frequency_labels()

            def _on_click_freq_label(event):
                nonlocal manual_label_idxs
                if event.button != 1:
                    return
                if str(getattr(toolbar, "mode", "") or "").strip():
                    return

                best_i = _nearest_freq_point_index(event)
                if best_i is None:
                    return

                if hover_annot.get_visible():
                    hover_annot.set_visible(False)

                if best_i in manual_label_idxs:
                    manual_label_idxs = [idx for idx in manual_label_idxs if idx != best_i]
                else:
                    manual_label_idxs = sorted(manual_label_idxs + [best_i])
                _redraw_frequency_labels()

            canvas.mpl_connect("button_press_event", _on_click_freq_label)

            # UI widgets
            ttk.Label(freq_box, text="Freqmin").grid(row=3, column=0, sticky="w", padx=(0, 6), pady=2)
            fmin_entry = ttk.Entry(freq_box, textvariable=freqmin_var, width=12)
            fmin_entry.grid(row=3, column=1, sticky="w", pady=2)

            ttk.Label(freq_box, text="Freqmax").grid(row=4, column=0, sticky="w", padx=(0, 6), pady=2)
            fmax_entry = ttk.Entry(freq_box, textvariable=freqmax_var, width=12)
            fmax_entry.grid(row=4, column=1, sticky="w", pady=2)

            ttk.Label(freq_box, text="N labels").grid(row=5, column=0, sticky="w", padx=(0, 6), pady=2)
            n_spin = ttk.Spinbox(freq_box, from_=0, to=50, increment=1, textvariable=nlabels_var, width=10)
            n_spin.grid(row=5, column=1, sticky="w", pady=2)

            btns_f = ttk.Frame(freq_box)
            btns_f.grid(row=6, column=0, columnspan=2, sticky="ew", pady=(8, 0))
            ttk.Button(btns_f, text="Limpiar", command=lambda: (nlabels_var.set(0), _clear_all_frequency_labels()))\
                .pack(side="left", expand=True, fill="x")

            # Debounced auto-apply for static labels
            pending_freq = {"id": None}

            def _schedule_labels(_evt=None):
                if pending_freq["id"] is not None:
                    tab.after_cancel(pending_freq["id"])
                pending_freq["id"] = tab.after(300, _apply_static_labels)

            for e in (fmin_entry, fmax_entry):
                e.bind("<Return>", lambda ev: _apply_static_labels())
                e.bind("<FocusOut>", lambda ev: _apply_static_labels())
                e.bind("<KeyRelease>", _schedule_labels)

            n_spin.configure(command=_apply_static_labels)
            n_spin.bind("<KeyRelease>", lambda ev: _apply_static_labels())

        title_entry.bind("<Return>", lambda e: apply_title_text())
        title_entry.bind("<FocusOut>", lambda e: apply_title_text())
        title_entry.bind("<KeyRelease>", _schedule_title)

        if is_multi_series:
            _apply_visibility()

        if is_bode_plot:
            for key in bode_style_vars:
                _apply_bode_style(key)
            _apply_bode_visibility()
            autoscale_axes()

        if line is None and not is_bode_plot:
            for child in style_box.winfo_children():
                try:
                    child.configure(state="disabled")
                except Exception:
                    pass

        win._mpl_refs.append((canvas, toolbar, fig, ax, line))  # type: ignore[attr-defined]

    def open_composer_nyquist():

        if not nyquist_sources:
            return

        # If already open, bring to front
        existing = getattr(win, "_composer_win", None)
        if existing is not None and existing.winfo_exists():
            existing.lift()
            existing.focus_force()
            return

        def _src_label(key: str) -> str:
            # IMPORTANT: uses the *current* title of the source plot (user-editable)
            ax_src = nyquist_sources[key]["ax"]
            t = (ax_src.get_title() or "").strip()
            return t if t else key

        comp = tk.Toplevel(win)
        win._composer_win = comp  # type: ignore[attr-defined]
        comp.title("Composite (Nyquist)")
        comp.geometry("1250x780")

        ctrl_host, plot_frame = create_resizable_plot_layout(
            comp,
            sidebar_width=320,
            sidebar_side="right",
            plot_padding=0,
        )

        figc = _new_figure()
        axc = figc.add_subplot(111)

        def _reset_composite_axes():
            axc.set_aspect("equal", adjustable="box")
            axc.grid(True)
            axc.set_title("Composite - Nyquist")
            axc.set_xlabel("Zreal (ohm)")
            axc.set_ylabel("-Zimag (ohm)")

        axc.set_aspect("equal", adjustable="box")
        axc.grid(True)
        axc.set_title("Composite - Nyquist")
        axc.set_xlabel("Zreal (ohm)")
        axc.set_ylabel("-Zimag (ohm)")

        canvas = FigureCanvasTkAgg(figc, master=plot_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

        toolbar = NavigationToolbar2Tk(canvas, plot_frame)
        toolbar.update()

        # Right: controls (scrollable)
        _topbar, ctrl_scroll = _build_scrollable_controls(ctrl_host)
        ctrl = ttk.Frame(ctrl_scroll, padding=10)
        ctrl.pack(fill="both", expand=True)

        # ---- Sources list (show current title + key) ----
        src_box = ttk.LabelFrame(ctrl, text="Nyquist sources", padding=8)
        src_box.pack(fill="x", pady=(0, 10))

        lb = tk.Listbox(src_box, selectmode="extended", height=12, exportselection=False)
        lb.pack(fill="x", expand=False)

        # mapping listbox index -> key
        idx_to_key: list[str] = []

        def _rebuild_listbox():
            nonlocal idx_to_key
            lb.delete(0, "end")
            idx_to_key = []

            # sort by label for nicer UX
            keys = list(nyquist_sources.keys())
            keys.sort(key=lambda k: _src_label(k).lower())

            for k in keys:
                lb.insert("end", f"{_src_label(k)}   [{k}]")
                idx_to_key.append(k)

        _rebuild_listbox()

        def _selected_keys() -> list[str]:
            return [idx_to_key[i] for i in lb.curselection()]

        # ---- Composite line storage: key -> line ----
        comp_lines: dict[str, object] = {}
        comp_label_artists: dict[str, list[object]] = {}  # NEW: per-curve annotations

        def _fmt_freq_hz(v: float) -> str:
            av = abs(v)
            if av >= 1e6:
                return f"{v/1e6:.3g} MHz"
            if av >= 1e3:
                return f"{v/1e3:.3g} kHz"
            return f"{v:.3g} Hz"
        
        def _clear_comp_labels(key: str):
            arts = comp_label_artists.pop(key, [])
            for a in arts:
                try:
                    a.remove()
                except Exception:
                    pass

        def _apply_comp_labels_from_source(key: str, dst_line):
            # remove any old ones
            _clear_comp_labels(key)

            src_line = nyquist_sources[key]["line"]
            idxs = getattr(src_line, "_freq_label_idxs", None)
            freqs = getattr(src_line, "_eis_freq", None)

            if not idxs or freqs is None:
                return
            # defensive: keep only valid indices
            npts = len(list(dst_line.get_xdata(orig=False)))
            idxs = [int(i) for i in idxs if 0 <= int(i) < npts]
            if not idxs:
                return

            xdata = list(dst_line.get_xdata(orig=False))
            ydata = list(dst_line.get_ydata(orig=False))

            arts: list[object] = []
            for i in idxs:
                txt = _fmt_freq_hz(float(freqs[i]))
                a = axc.annotate(
                    txt,
                    xy=(xdata[i], ydata[i]),
                    xytext=(6, 6),
                    textcoords="offset points",
                    fontsize=8.5,
                    bbox=dict(boxstyle="round,pad=0.15", fc="white", alpha=0.7),
                )
                arts.append(a)

            comp_label_artists[key] = arts

        legend_var = tk.BooleanVar(value=True)
        title_text_var = tk.StringVar(value=axc.get_title() or "Composite - Nyquist")
        show_title_var = tk.BooleanVar(value=True)
        init_title_fs = float(font_defaults.title)
        init_label_fs = float(font_defaults.label)
        init_tick_fs = float(font_defaults.tick)
        title_fs_var = tk.DoubleVar(value=init_title_fs)
        label_fs_var = tk.DoubleVar(value=init_label_fs)
        legend_fs_var = tk.DoubleVar(value=float(font_defaults.legend))
        tick_fs_var = tk.DoubleVar(value=init_tick_fs)
        x_tick_count_var = tk.IntVar(value=6)
        y_tick_count_var = tk.IntVar(value=6)

        def _apply_legend(redraw: bool = True):
            # Always remove existing legend first (prevents stacking / stale legends)
            leg = axc.get_legend()
            if leg is not None:
                leg.remove()

            if not legend_var.get():
                if redraw:
                    canvas.draw_idle()
                return

            handles, labels = axc.get_legend_handles_labels()
            # Keep only meaningful labels (ignore '_' internal ones)
            pairs = [(h, l) for h, l in zip(handles, labels) if l and not l.startswith("_")]
            if not pairs:
                if redraw:
                    canvas.draw_idle()
                return

            h2, l2 = zip(*pairs)
            try:
                legend_fs = float(legend_fs_var.get())
            except (tk.TclError, ValueError):
                legend_fs = float(font_defaults.legend)
            make_legend_draggable(axc.legend(h2, l2, loc="best", fontsize=legend_fs))
            if redraw:
                canvas.draw_idle()

        def _apply_plot_settings(redraw: bool = True):
            try:
                title_fs = float(title_fs_var.get())
                label_fs = float(label_fs_var.get())
                tick_fs = float(tick_fs_var.get())
                x_tick_count = max(2, int(x_tick_count_var.get()))
                y_tick_count = max(2, int(y_tick_count_var.get()))
            except (tk.TclError, ValueError):
                return

            axc.set_title(title_text_var.get() if show_title_var.get() else "", fontsize=title_fs)
            axc.xaxis.label.set_fontsize(label_fs)
            axc.yaxis.label.set_fontsize(label_fs)
            axc.tick_params(axis="both", labelsize=tick_fs)
            apply_x_tick_label_padding(axc, tick_fs)
            axc.xaxis.set_major_locator(MaxNLocator(nbins=x_tick_count))
            axc.yaxis.set_major_locator(MaxNLocator(nbins=y_tick_count))
            axc.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
            axc.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
            _apply_legend(redraw=False)
            if redraw:
                canvas.draw_idle()

        pending_plot_settings = {"id": None}

        def _schedule_plot_settings(_evt=None):
            if pending_plot_settings["id"] is not None:
                comp.after_cancel(pending_plot_settings["id"])
            pending_plot_settings["id"] = comp.after(180, _apply_plot_settings)

        def _copy_style(src_line, dst_line):
            # Copy *current* formatting from source line
            dst_line.set_color(src_line.get_color())
            dst_line.set_linestyle(src_line.get_linestyle())
            dst_line.set_marker(src_line.get_marker())
            dst_line.set_linewidth(src_line.get_linewidth())
            dst_line.set_markersize(src_line.get_markersize())

            # Also copy marker fill/edge if present
            try:
                dst_line.set_markerfacecolor(src_line.get_markerfacecolor())
            except Exception:
                pass
            try:
                dst_line.set_markeredgecolor(src_line.get_markeredgecolor())
            except Exception:
                pass
            try:
                dst_line.set_alpha(src_line.get_alpha())
            except Exception:
                pass

        def _fit_all():
            if not comp_lines:
                return

            xs: list[float] = []
            ys: list[float] = []

            for ln in comp_lines.values():
                x = [float(v) for v in ln.get_xdata(orig=False)]
                y = [float(v) for v in ln.get_ydata(orig=False)]
                xs.extend(x)
                ys.extend(y)

            if not xs or not ys:
                return

            x0, x1 = min(xs), max(xs)
            y0, y1 = min(ys), max(ys)

            dx = (x1 - x0) if x1 != x0 else (abs(x0) * 0.1 + 1.0)
            dy = (y1 - y0) if y1 != y0 else (abs(y0) * 0.1 + 1.0)

            pad_x = 0.05 * dx
            pad_y = 0.05 * dy

            axc.set_xlim(x0 - pad_x, x1 + pad_x)
            axc.set_ylim(y0 - pad_y, y1 + pad_y)
            axc.set_aspect("equal", adjustable="box")

            _apply_plot_settings(redraw=False)
            canvas.draw_idle()
            _sync_limit_entries()

        def add_selected():
            for key in _selected_keys():
                if key in comp_lines:
                    continue

                src_line = nyquist_sources[key]["line"]
                x = list(src_line.get_xdata(orig=False))
                y = list(src_line.get_ydata(orig=False))

                # label must match the *source plot title*
                (ln,) = axc.plot(x, y, label=_src_label(key))
                _copy_style(src_line, ln)

                # keep freqs too (future hover freq on composite)
                freqs = getattr(src_line, "_eis_freq", None)
                if freqs is not None:
                    ln._eis_freq = freqs  # type: ignore[attr-defined]

                comp_lines[key] = ln
                _apply_comp_labels_from_source(key, ln)   # NEW

            axc.set_aspect("equal", adjustable="box")
            _apply_legend()
            _fit_all()

        def remove_selected():
            removed = False
            for key in _selected_keys():
                ln = comp_lines.pop(key, None)
                if ln is not None:
                    try:
                        _clear_comp_labels(key)   # NEW
                        ln.remove()
                    except Exception:
                        pass
                    removed = True
            if removed:
                _apply_legend()
                _fit_all()

        def clear_all():
            # Clear our bookkeeping first
            comp_lines.clear()

            # If you implemented composed freq labels:
            try:
                comp_label_artists.clear()  # type: ignore[name-defined]
            except Exception:
                pass

            # Clear the axes in one shot (fast, removes lines + annotations + legend)
            axc.cla()
            _reset_composite_axes()
            _apply_plot_settings(redraw=False)

            # Redraw + sync UI fields
            canvas.draw_idle()
            _sync_limit_entries()

            # Optional: clear selection so user doesn't accidentally "Remove" nothing
            try:
                lb.selection_clear(0, "end")
            except Exception:
                pass

            # Legend should now be empty (but keep checkbox state consistent)
            _apply_legend()

        def refresh_formatting():
            # Refresh BOTH style and legend labels from the current state of source tabs
            for key, ln in comp_lines.items():
                src = nyquist_sources.get(key)
                if not src:
                    continue
                src_line = src["line"]
                _copy_style(src_line, ln)
                ln.set_label(_src_label(key))  # <-- ensures legend matches edited titles
                _apply_comp_labels_from_source(key, ln)

            _rebuild_listbox()
            _apply_legend()
            canvas.draw_idle()

        btns = ttk.Frame(src_box)
        btns.pack(fill="x", pady=(8, 0))
        ttk.Button(btns, text="Añadir", command=add_selected).pack(side="left", expand=True, fill="x", padx=(0, 6))
        ttk.Button(btns, text="Remover", command=remove_selected).pack(side="left", expand=True, fill="x", padx=(0, 6))
        ttk.Button(btns, text="Limpiar", command=clear_all).pack(side="left", expand=True, fill="x")

        ttk.Button(src_box, text="Actualizar formato", command=refresh_formatting).pack(fill="x", pady=(8, 0))
        ttk.Checkbutton(ctrl, text="Leyenda", variable=legend_var, command=_apply_legend).pack(anchor="w", pady=(0, 10))

        plot_box = ttk.LabelFrame(ctrl, text="Gráfico", padding=8)
        plot_box.pack(fill="x", pady=(0, 10))
        plot_box.columnconfigure(1, weight=1)

        ttk.Label(plot_box, text="Título").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=2)
        title_entry = ttk.Entry(plot_box, textvariable=title_text_var, width=20)
        title_entry.grid(row=0, column=1, sticky="ew", pady=2)
        ttk.Checkbutton(plot_box, text="Mostrar titulo", variable=show_title_var, command=_apply_plot_settings).grid(
            row=7, column=0, columnspan=2, sticky="w", pady=(6, 0)
        )

        ttk.Label(plot_box, text="Tamaño del título").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=2)
        title_spin = ttk.Spinbox(plot_box, from_=6.0, to=50.0, increment=0.5, textvariable=title_fs_var, width=10)
        title_spin.grid(row=1, column=1, sticky="w", pady=2)

        ttk.Label(plot_box, text="Tamaño de etiquetas").grid(row=2, column=0, sticky="w", padx=(0, 6), pady=2)
        label_spin = ttk.Spinbox(plot_box, from_=6.0, to=40.0, increment=0.5, textvariable=label_fs_var, width=10)
        label_spin.grid(row=2, column=1, sticky="w", pady=2)

        ttk.Label(plot_box, text="Tamaño de leyenda").grid(row=3, column=0, sticky="w", padx=(0, 6), pady=2)
        legend_spin = ttk.Spinbox(plot_box, from_=6.0, to=40.0, increment=0.5, textvariable=legend_fs_var, width=10)
        legend_spin.grid(row=3, column=1, sticky="w", pady=2)

        ttk.Label(plot_box, text="Tamaño de ticks").grid(row=4, column=0, sticky="w", padx=(0, 6), pady=2)
        tick_spin = ttk.Spinbox(plot_box, from_=6.0, to=40.0, increment=0.5, textvariable=tick_fs_var, width=10)
        tick_spin.grid(row=4, column=1, sticky="w", pady=2)

        ttk.Label(plot_box, text="X ticks").grid(row=5, column=0, sticky="w", padx=(0, 6), pady=2)
        x_tick_spin = ttk.Spinbox(plot_box, from_=2, to=20, increment=1, textvariable=x_tick_count_var, width=10)
        x_tick_spin.grid(row=5, column=1, sticky="w", pady=2)

        ttk.Label(plot_box, text="Y ticks").grid(row=6, column=0, sticky="w", padx=(0, 6), pady=2)
        y_tick_spin = ttk.Spinbox(plot_box, from_=2, to=20, increment=1, textvariable=y_tick_count_var, width=10)
        y_tick_spin.grid(row=6, column=1, sticky="w", pady=2)

        # ---- Axis limits (independent in composite) ----
        lim_box = ttk.LabelFrame(ctrl, text="Límites de ejes", padding=8)
        lim_box.pack(fill="x", pady=(0, 10))

        def _fmt(v: float) -> str:
            return f"{v:.6g}"

        xmin_var = tk.StringVar()
        xmax_var = tk.StringVar()
        ymin_var = tk.StringVar()
        ymax_var = tk.StringVar()

        def _sync_limit_entries():
            x0, x1 = axc.get_xlim()
            y0, y1 = axc.get_ylim()
            xmin_var.set(_fmt(x0))
            xmax_var.set(_fmt(x1))
            ymin_var.set(_fmt(y0))
            ymax_var.set(_fmt(y1))

        _sync_limit_entries()

        def _parse_float(s: str) -> float | None:
            s = s.strip()
            if not s:
                return None
            try:
                return float(s)
            except ValueError:
                return None

        def apply_limits(live: bool = False):
            cx0, cx1 = axc.get_xlim()
            cy0, cy1 = axc.get_ylim()

            raws = [xmin_var.get(), xmax_var.get(), ymin_var.get(), ymax_var.get()]
            if any(_is_incomplete_number(r) for r in raws):
                return
            if live and any(not r.strip() for r in raws):
                return

            nx0 = _parse_float(xmin_var.get())
            nx1 = _parse_float(xmax_var.get())
            ny0 = _parse_float(ymin_var.get())
            ny1 = _parse_float(ymax_var.get())

            axc.set_xlim(cx0 if nx0 is None else nx0, cx1 if nx1 is None else nx1)
            axc.set_ylim(cy0 if ny0 is None else ny0, cy1 if ny1 is None else ny1)
            axc.set_aspect("equal", adjustable="box")
            _apply_plot_settings(redraw=False)
            canvas.draw_idle()
            _sync_limit_entries()

        def _row(parent, r, label, var):
            ttk.Label(parent, text=label, width=5).grid(row=r, column=0, sticky="w", padx=(0, 6), pady=2)
            e = ttk.Entry(parent, textvariable=var, width=12)
            e.grid(row=r, column=1, sticky="w", pady=2)
            return e

        exmin = _row(lim_box, 0, "Xmin", xmin_var)
        exmax = _row(lim_box, 1, "Xmax", xmax_var)
        eymin = _row(lim_box, 2, "Ymin", ymin_var)
        eymax = _row(lim_box, 3, "Ymax", ymax_var)

        pending = {"id": None}

        def _schedule(_evt=None):
            if pending["id"] is not None:
                comp.after_cancel(pending["id"])
            pending["id"] = comp.after(300, lambda: apply_limits(live=True))

        for e in (exmin, exmax, eymin, eymax):
            e.bind("<Return>", lambda ev: apply_limits(live=False))
            e.bind("<FocusOut>", lambda ev: apply_limits(live=False))

        title_entry.bind("<Return>", lambda ev: _apply_plot_settings())
        title_entry.bind("<FocusOut>", lambda ev: _apply_plot_settings())
        title_entry.bind("<KeyRelease>", _schedule_plot_settings)

        for spin in (title_spin, label_spin, legend_spin, tick_spin, x_tick_spin, y_tick_spin):
            spin.configure(command=_apply_plot_settings)
            spin.bind("<Return>", lambda ev: _apply_plot_settings())
            spin.bind("<FocusOut>", lambda ev: _apply_plot_settings())

        b2 = ttk.Frame(lim_box)
        b2.grid(row=4, column=0, columnspan=2, sticky="ew", pady=(8, 0))
        ttk.Button(b2, text="Ajustar todo", command=_fit_all).pack(side="left", expand=True, fill="x", padx=(0, 6))
        ttk.Button(b2, text="Refresh fields", command=_sync_limit_entries).pack(side="left", expand=True, fill="x")

        _apply_plot_settings(redraw=False)

        def _on_close():
            comp.destroy()
            try:
                delattr(win, "_composer_win")
            except Exception:
                pass

        comp.protocol("WM_DELETE_WINDOW", _on_close)

    def open_composer_bode():

        source_keys = sorted(
            set(bode_sources.get("zmod", {}).keys()) | set(bode_sources.get("zphz", {}).keys()),
            key=lambda k: k.lower(),
        )
        if not source_keys:
            return

        existing = getattr(win, "_composer_win_bode", None)
        if existing is not None and existing.winfo_exists():
            existing.lift()
            existing.focus_force()
            return

        def _src_label(key: str) -> str:
            src = bode_sources.get("zmod", {}).get(key) or bode_sources.get("zphz", {}).get(key)
            if not src:
                return key
            ax_src = src["ax"]
            t = (ax_src.get_title() or "").strip()
            return t if t else key

        any_key = source_keys[0]
        src_mod = bode_sources.get("zmod", {}).get(any_key)
        src_phz = bode_sources.get("zphz", {}).get(any_key)
        sample_src = src_mod or src_phz
        sample_ax = sample_src["ax"]
        default_xlabel = sample_ax.get_xlabel() or "Frecuencia"
        default_ylabel_mod = src_mod["ax"].get_ylabel() if src_mod else "Zmod"
        default_ylabel_phz = src_phz["ax"].get_ylabel() if src_phz else "Zphz"
        default_title = "Composite - Bode"

        comp = tk.Toplevel(win)
        win._composer_win_bode = comp  # type: ignore[attr-defined]
        comp.title("Composite (Bode)")
        comp.geometry("1280x800")

        ctrl_host, plot_frame = create_resizable_plot_layout(
            comp,
            sidebar_width=320,
            sidebar_side="right",
            plot_padding=0,
        )

        figc = _new_figure()
        axc_mod = figc.add_subplot(111)
        axc_phz = axc_mod.twinx()
        axc_phz.spines["left"].set_visible(False)
        axc_phz.yaxis.tick_right()
        axc_phz.yaxis.set_label_position("right")

        def _reset_axes():
            axc_mod.cla()
            axc_phz.cla()
            axc_mod.set_xscale("log")
            axc_mod.grid(True, which="both")
            axc_mod.set_title(default_title)
            axc_mod.set_xlabel(default_xlabel)
            axc_mod.set_ylabel(default_ylabel_mod)
            axc_phz.set_ylabel(default_ylabel_phz)
            axc_phz.spines["left"].set_visible(False)
            axc_phz.yaxis.tick_right()
            axc_phz.yaxis.set_label_position("right")

        _reset_axes()

        canvas = FigureCanvasTkAgg(figc, master=plot_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

        toolbar = NavigationToolbar2Tk(canvas, plot_frame)
        toolbar.update()

        _topbar, ctrl_scroll = _build_scrollable_controls(ctrl_host)
        ctrl = ttk.Frame(ctrl_scroll, padding=10)
        ctrl.pack(fill="both", expand=True)

        src_box = ttk.LabelFrame(ctrl, text="Bode sources", padding=8)
        src_box.pack(fill="x", pady=(0, 10))

        lb = tk.Listbox(src_box, selectmode="extended", height=12, exportselection=False)
        lb.pack(fill="x", expand=False)

        idx_to_key: list[str] = []

        def _rebuild_listbox():
            nonlocal idx_to_key
            lb.delete(0, "end")
            idx_to_key = []
            keys = sorted(source_keys, key=lambda k: _src_label(k).lower())
            for k in keys:
                lb.insert("end", f"{_src_label(k)}   [{k}]")
                idx_to_key.append(k)

        _rebuild_listbox()

        def _selected_keys() -> list[str]:
            return [idx_to_key[i] for i in lb.curselection()]

        comp_lines = {"zmod": {}, "zphz": {}}
        legend_var = tk.BooleanVar(value=True)
        show_mod_var = tk.BooleanVar(value=True)
        show_phz_var = tk.BooleanVar(value=True)
        title_text_var = tk.StringVar(value=axc_mod.get_title() or default_title)
        show_title_var = tk.BooleanVar(value=True)
        init_title_fs = float(font_defaults.title)
        init_label_fs = float(font_defaults.label)
        init_tick_fs = float(font_defaults.tick)
        title_fs_var = tk.DoubleVar(value=init_title_fs)
        label_fs_var = tk.DoubleVar(value=init_label_fs)
        legend_fs_var = tk.DoubleVar(value=float(font_defaults.legend))
        tick_fs_var = tk.DoubleVar(value=init_tick_fs)
        x_tick_count_var = tk.IntVar(value=6)
        y_tick_count_var = tk.IntVar(value=6)

        def _copy_style(src_line, dst_line):
            dst_line.set_color(src_line.get_color())
            dst_line.set_linestyle(src_line.get_linestyle())
            dst_line.set_marker(src_line.get_marker())
            dst_line.set_linewidth(src_line.get_linewidth())
            dst_line.set_markersize(src_line.get_markersize())
            try:
                dst_line.set_markerfacecolor(src_line.get_markerfacecolor())
            except Exception:
                pass
            try:
                dst_line.set_markeredgecolor(src_line.get_markeredgecolor())
            except Exception:
                pass
            try:
                dst_line.set_alpha(src_line.get_alpha())
            except Exception:
                pass

        def _apply_legend(redraw: bool = True):
            for axis in (axc_mod, axc_phz):
                leg = axis.get_legend()
                if leg is not None:
                    leg.remove()

            if not legend_var.get():
                if redraw:
                    canvas.draw_idle()
                return

            handles = []
            labels = []
            for kind in ("zmod", "zphz"):
                for key in sorted(comp_lines[kind].keys()):
                    ln = comp_lines[kind][key]
                    if ln.get_visible():
                        handles.append(ln)
                        labels.append(ln.get_label())

            if handles:
                try:
                    legend_fs = float(legend_fs_var.get())
                except (tk.TclError, ValueError):
                    legend_fs = float(font_defaults.legend)
                make_legend_draggable(axc_mod.legend(handles, labels, loc="best", fontsize=legend_fs))
            if redraw:
                canvas.draw_idle()

        def _apply_tick_settings():
            try:
                x_tick_count = max(2, int(x_tick_count_var.get()))
            except (tk.TclError, ValueError):
                x_tick_count = 6
            try:
                y_tick_count = max(2, int(y_tick_count_var.get()))
            except (tk.TclError, ValueError):
                y_tick_count = 6
            axc_mod.xaxis.set_major_locator(LogLocator(base=10.0, numticks=x_tick_count))
            for axis in (axc_mod, axc_phz):
                axis.yaxis.set_major_locator(LinearLocator(y_tick_count))
                axis.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
                axis.yaxis.get_offset_text().set_visible(False)

        def _apply_plot_settings(redraw: bool = True):
            try:
                title_fs = float(title_fs_var.get())
                label_fs = float(label_fs_var.get())
                tick_fs = float(tick_fs_var.get())
            except (tk.TclError, ValueError):
                return

            axc_mod.set_title(title_text_var.get() if show_title_var.get() else "", fontsize=title_fs)
            axc_mod.xaxis.label.set_fontsize(label_fs)
            axc_mod.yaxis.label.set_fontsize(label_fs)
            axc_phz.yaxis.label.set_fontsize(label_fs)
            axc_mod.tick_params(axis="both", labelsize=tick_fs)
            apply_x_tick_label_padding(axc_mod, tick_fs)
            axc_phz.tick_params(axis="y", labelsize=tick_fs)
            _apply_tick_settings()
            _apply_legend(redraw=False)
            if redraw:
                canvas.draw_idle()

        pending_plot_settings = {"id": None}

        def _schedule_plot_settings(_evt=None):
            if pending_plot_settings["id"] is not None:
                comp.after_cancel(pending_plot_settings["id"])
            pending_plot_settings["id"] = comp.after(180, _apply_plot_settings)

        def _sync_axis_visibility():
            mod_visible = bool(show_mod_var.get() and comp_lines["zmod"])
            phz_visible = bool(show_phz_var.get() and comp_lines["zphz"])

            for ln in comp_lines["zmod"].values():
                ln.set_visible(show_mod_var.get())
            for ln in comp_lines["zphz"].values():
                ln.set_visible(show_phz_var.get())

            axc_mod.yaxis.set_visible(mod_visible)
            axc_mod.spines["left"].set_visible(mod_visible)
            axc_mod.yaxis.label.set_visible(mod_visible)

            axc_phz.yaxis.set_visible(phz_visible)
            axc_phz.spines["right"].set_visible(phz_visible)
            axc_phz.yaxis.label.set_visible(phz_visible)

            _apply_tick_settings()
            _apply_legend()
            canvas.draw_idle()

        def _fit_all():
            active_mod = [ln for ln in comp_lines["zmod"].values() if ln.get_visible()]
            active_phz = [ln for ln in comp_lines["zphz"].values() if ln.get_visible()]
            if not active_mod and not active_phz:
                return

            xs: list[float] = []
            for lineset in (active_mod, active_phz):
                for ln in lineset:
                    xs.extend([float(v) for v in ln.get_xdata(orig=False) if float(v) > 0])

            if not xs:
                return

            x0, x1 = min(xs), max(xs)
            axc_mod.set_xlim(x0 / 1.2, x1 * 1.2)
            axc_phz.set_xlim(x0 / 1.2, x1 * 1.2)

            if active_mod:
                ys = []
                for ln in active_mod:
                    ys.extend([float(v) for v in ln.get_ydata(orig=False)])
                y0, y1 = min(ys), max(ys)
                dy = (y1 - y0) if y1 != y0 else (abs(y0) * 0.1 + 1.0)
                axc_mod.set_ylim(y0 - 0.05 * dy, y1 + 0.05 * dy)

            if active_phz:
                ys = []
                for ln in active_phz:
                    ys.extend([float(v) for v in ln.get_ydata(orig=False)])
                y0, y1 = min(ys), max(ys)
                dy = (y1 - y0) if y1 != y0 else (abs(y0) * 0.1 + 1.0)
                axc_phz.set_ylim(y0 - 0.05 * dy, y1 + 0.05 * dy)

            _apply_plot_settings(redraw=False)
            canvas.draw_idle()
            _sync_limit_entries()

        def add_selected():
            added = False
            for key in _selected_keys():
                src_mod_local = bode_sources.get("zmod", {}).get(key)
                if src_mod_local and key not in comp_lines["zmod"]:
                    src_line = src_mod_local["line"]
                    x = list(src_line.get_xdata(orig=False))
                    y = list(src_line.get_ydata(orig=False))
                    (ln,) = axc_mod.plot(x, y, label=f"{_src_label(key)} - Zmod")
                    _copy_style(src_line, ln)
                    comp_lines["zmod"][key] = ln
                    added = True

                src_phz_local = bode_sources.get("zphz", {}).get(key)
                if src_phz_local and key not in comp_lines["zphz"]:
                    src_line = src_phz_local["line"]
                    x = list(src_line.get_xdata(orig=False))
                    y = list(src_line.get_ydata(orig=False))
                    (ln,) = axc_phz.plot(x, y, label=f"{_src_label(key)} - Zphz")
                    _copy_style(src_line, ln)
                    comp_lines["zphz"][key] = ln
                    added = True

            if added:
                _sync_axis_visibility()
                _fit_all()

        def remove_selected():
            removed = False
            for key in _selected_keys():
                for kind in ("zmod", "zphz"):
                    ln = comp_lines[kind].pop(key, None)
                    if ln is not None:
                        try:
                            ln.remove()
                        except Exception:
                            pass
                        removed = True
            if removed:
                _sync_axis_visibility()
                _fit_all()

        def clear_all():
            comp_lines["zmod"].clear()
            comp_lines["zphz"].clear()
            _reset_axes()
            _apply_plot_settings(redraw=False)
            canvas.draw_idle()
            _sync_limit_entries()
            _apply_legend()

        def refresh_formatting():
            for key, ln in comp_lines["zmod"].items():
                src = bode_sources.get("zmod", {}).get(key)
                if src:
                    _copy_style(src["line"], ln)
                    ln.set_label(f"{_src_label(key)} - Zmod")
            for key, ln in comp_lines["zphz"].items():
                src = bode_sources.get("zphz", {}).get(key)
                if src:
                    _copy_style(src["line"], ln)
                    ln.set_label(f"{_src_label(key)} - Zphz")
            _rebuild_listbox()
            _apply_legend()
            canvas.draw_idle()

        btns = ttk.Frame(src_box)
        btns.pack(fill="x", pady=(8, 0))
        ttk.Button(btns, text="Añadir", command=add_selected).pack(side="left", expand=True, fill="x", padx=(0, 6))
        ttk.Button(btns, text="Remover", command=remove_selected).pack(side="left", expand=True, fill="x", padx=(0, 6))
        ttk.Button(btns, text="Limpiar", command=clear_all).pack(side="left", expand=True, fill="x")

        ttk.Button(src_box, text="Actualizar formato", command=refresh_formatting).pack(fill="x", pady=(8, 0))
        ttk.Checkbutton(ctrl, text="Leyenda", variable=legend_var, command=_apply_legend).pack(anchor="w", pady=(0, 4))
        ttk.Checkbutton(ctrl, text="Zmod", variable=show_mod_var, command=_sync_axis_visibility).pack(anchor="w")
        ttk.Checkbutton(ctrl, text="Zphz", variable=show_phz_var, command=_sync_axis_visibility).pack(anchor="w", pady=(0, 10))

        plot_box = ttk.LabelFrame(ctrl, text="Gráfico", padding=8)
        plot_box.pack(fill="x", pady=(0, 10))
        plot_box.columnconfigure(1, weight=1)

        ttk.Label(plot_box, text="Título").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=2)
        title_entry = ttk.Entry(plot_box, textvariable=title_text_var, width=20)
        title_entry.grid(row=0, column=1, sticky="ew", pady=2)

        ttk.Label(plot_box, text="Tamaño del título").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=2)
        title_spin = ttk.Spinbox(plot_box, from_=6.0, to=50.0, increment=0.5, textvariable=title_fs_var, width=10)
        title_spin.grid(row=1, column=1, sticky="w", pady=2)

        ttk.Label(plot_box, text="Tamaño de etiquetas").grid(row=2, column=0, sticky="w", padx=(0, 6), pady=2)
        label_spin = ttk.Spinbox(plot_box, from_=6.0, to=40.0, increment=0.5, textvariable=label_fs_var, width=10)
        label_spin.grid(row=2, column=1, sticky="w", pady=2)

        ttk.Label(plot_box, text="Tamaño de leyenda").grid(row=3, column=0, sticky="w", padx=(0, 6), pady=2)
        legend_spin = ttk.Spinbox(plot_box, from_=6.0, to=40.0, increment=0.5, textvariable=legend_fs_var, width=10)
        legend_spin.grid(row=3, column=1, sticky="w", pady=2)

        ttk.Label(plot_box, text="Tamaño de ticks").grid(row=4, column=0, sticky="w", padx=(0, 6), pady=2)
        tick_spin = ttk.Spinbox(plot_box, from_=6.0, to=40.0, increment=0.5, textvariable=tick_fs_var, width=10)
        tick_spin.grid(row=4, column=1, sticky="w", pady=2)

        ttk.Label(plot_box, text="X ticks").grid(row=5, column=0, sticky="w", padx=(0, 6), pady=2)
        x_tick_spin = ttk.Spinbox(plot_box, from_=2, to=20, increment=1, textvariable=x_tick_count_var, width=10)
        x_tick_spin.grid(row=5, column=1, sticky="w", pady=2)

        ttk.Label(plot_box, text="Y ticks").grid(row=6, column=0, sticky="w", padx=(0, 6), pady=2)
        y_tick_spin = ttk.Spinbox(plot_box, from_=2, to=20, increment=1, textvariable=y_tick_count_var, width=10)
        y_tick_spin.grid(row=6, column=1, sticky="w", pady=2)

        lim_box = ttk.LabelFrame(ctrl, text="Límites de ejes", padding=8)
        lim_box.pack(fill="x", pady=(0, 10))

        xmin_var = tk.StringVar()
        xmax_var = tk.StringVar()
        modmin_var = tk.StringVar()
        modmax_var = tk.StringVar()
        phzmin_var = tk.StringVar()
        phzmax_var = tk.StringVar()

        def _sync_limit_entries():
            x0, x1 = axc_mod.get_xlim()
            m0, m1 = axc_mod.get_ylim()
            p0, p1 = axc_phz.get_ylim()
            xmin_var.set(_fmt(x0))
            xmax_var.set(_fmt(x1))
            modmin_var.set(_fmt(m0))
            modmax_var.set(_fmt(m1))
            phzmin_var.set(_fmt(p0))
            phzmax_var.set(_fmt(p1))

        _sync_limit_entries()

        def apply_limits():
            raws = [xmin_var.get(), xmax_var.get(), modmin_var.get(), modmax_var.get(), phzmin_var.get(), phzmax_var.get()]
            if any(_is_incomplete_number(r) for r in raws):
                return

            cur_x0, cur_x1 = axc_mod.get_xlim()
            nx0 = _parse_float(xmin_var.get())
            nx1 = _parse_float(xmax_var.get())
            new_x0 = cur_x0 if nx0 is None else nx0
            new_x1 = cur_x1 if nx1 is None else nx1
            if new_x0 <= 0 or new_x1 <= 0:
                _sync_limit_entries()
                return

            axc_mod.set_xlim(new_x0, new_x1)
            axc_phz.set_xlim(new_x0, new_x1)

            nm0 = _parse_float(modmin_var.get())
            nm1 = _parse_float(modmax_var.get())
            np0 = _parse_float(phzmin_var.get())
            np1 = _parse_float(phzmax_var.get())

            cur_m0, cur_m1 = axc_mod.get_ylim()
            cur_p0, cur_p1 = axc_phz.get_ylim()
            axc_mod.set_ylim(cur_m0 if nm0 is None else nm0, cur_m1 if nm1 is None else nm1)
            axc_phz.set_ylim(cur_p0 if np0 is None else np0, cur_p1 if np1 is None else np1)
            _apply_plot_settings(redraw=False)
            canvas.draw_idle()
            _sync_limit_entries()

        def _row(parent, r, label, var):
            ttk.Label(parent, text=label, width=7).grid(row=r, column=0, sticky="w", padx=(0, 6), pady=2)
            e = ttk.Entry(parent, textvariable=var, width=12)
            e.grid(row=r, column=1, sticky="w", pady=2)
            return e

        entries = [
            _row(lim_box, 0, "Xmin", xmin_var),
            _row(lim_box, 1, "Xmax", xmax_var),
            _row(lim_box, 2, "Mmin", modmin_var),
            _row(lim_box, 3, "Mmax", modmax_var),
            _row(lim_box, 4, "Pmin", phzmin_var),
            _row(lim_box, 5, "Pmax", phzmax_var),
        ]

        for e in entries:
            e.bind("<Return>", lambda ev: apply_limits())
            e.bind("<FocusOut>", lambda ev: apply_limits())

        title_entry.bind("<Return>", lambda ev: _apply_plot_settings())
        title_entry.bind("<FocusOut>", lambda ev: _apply_plot_settings())
        title_entry.bind("<KeyRelease>", _schedule_plot_settings)

        for spin in (title_spin, label_spin, legend_spin, tick_spin, x_tick_spin, y_tick_spin):
            spin.configure(command=_apply_plot_settings)
            spin.bind("<Return>", lambda ev: _apply_plot_settings())
            spin.bind("<FocusOut>", lambda ev: _apply_plot_settings())

        b2 = ttk.Frame(lim_box)
        b2.grid(row=6, column=0, columnspan=2, sticky="ew", pady=(8, 0))
        ttk.Button(b2, text="Ajustar todo", command=_fit_all).pack(side="left", expand=True, fill="x", padx=(0, 6))
        ttk.Button(b2, text="Refresh fields", command=_sync_limit_entries).pack(side="left", expand=True, fill="x")

        _apply_plot_settings(redraw=False)

        def _on_close():
            comp.destroy()
            try:
                delattr(win, "_composer_win_bode")
            except Exception:
                pass

        comp.protocol("WM_DELETE_WINDOW", _on_close)

    def open_composer_for_current():
        if not composer_enabled:
            return

        key = plot_names_var.get().lower()

        if "nyquist" in key:
            open_composer_nyquist()
            return

        if "bode" in key:
            open_composer_bode()
            return

        if "series vs pt" in key or "series by pt" in key:
            _open_composer_pt(win, pt_sources, font_defaults)
            return

        if "pre-estabiliz" in key:
            _open_composer_pt(
                win,
                pre_stabilization_sources,
                font_defaults,
                window_title="Composite (Pre-estabilizacion)",
                default_title="Composite - Pre-estabilizacion",
                sources_title="Pre-estabilizacion sources",
            )
            return

        mb.showinfo("Componer", "Componer esta disponible para graficos Nyquist, Bode, Series vs Pt y Pre-estabilizacion.")

    compose_btn.configure(command=open_composer_for_current)

    for tab_title, fig in figures:
        _add_tab(tab_title, fig)

    if created_root:
        win.mainloop()
        


# ---------------------------------------------------------------------------
# Folder export
# ---------------------------------------------------------------------------

def find_eis_files(input_dir: Path) -> list[Path]:
    return sorted(
        [
            p for p in input_dir.iterdir()
            if p.is_file() and p.suffix.lower() == ".dta" and "EISPOT" in p.name
        ]
    )

def find_pre_stabilization_files(input_dir: Path) -> list[Path]:
    return sorted(
        [
            p for p in input_dir.iterdir()
            if p.is_file() and p.suffix.lower() == ".dta" and p.name.lower().startswith("est_eis")
        ]
    )


def _wants_pre_stabilization(selected_options: Iterable[str] | None) -> bool:
    return any("pre-estabiliz" in option.lower() for option in (selected_options or []))


def export_folder(
    input_dir: Path,
    output_dir: Path,
    *,
    include_pre_stabilization: bool = False,
) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)

    dta_files = find_eis_files(input_dir)
    pre_stabilization_files = find_pre_stabilization_files(input_dir) if include_pre_stabilization else []
    if not dta_files and not pre_stabilization_files:
        return []

    exported_xlsx: list[Path] = []

    for dta_file in dta_files:
        parsed = parse_gamry_dta(dta_file)
        xlsx_path = output_dir / f"{dta_file.stem}.xlsx"
        export_to_xlsx(parsed, xlsx_path)
        exported_xlsx.append(xlsx_path)

    for dta_file in pre_stabilization_files:
        parsed = parse_gamry_curve_dta(dta_file)
        xlsx_path = output_dir / f"{dta_file.stem}.xlsx"
        export_to_xlsx(
            parsed,
            xlsx_path,
            meta_fields=PRE_STAB_META_FIELDS,
            data_map=PRE_STAB_DATA_MAP,
            numeric_meta_keys={"ISTEP1", "TSTEP1", "SAMPLETIME", "AREA"},
        )
        exported_xlsx.append(xlsx_path)

    return exported_xlsx

def run_pipeline(
    input_dir: Path,
    output_dir: Path,
    selected_options: Iterable[str] | None = None,
    font_defaults: PlotFontDefaults | None = None,
) -> list[Path]:
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)

    chosen = set(selected_options or [])
    wants_pre_stabilization = _wants_pre_stabilization(chosen)

    exported_xlsx = export_folder(
        input_dir,
        output_dir,
        include_pre_stabilization=wants_pre_stabilization,
    )

    if not exported_xlsx:
        return []

    if not chosen:
        return exported_xlsx

    dta_files = find_eis_files(input_dir)
    pre_stabilization_files = find_pre_stabilization_files(input_dir) if wants_pre_stabilization else []
    plot_entries = _collect_eis_plot_entries(dta_files)

    option_order = [
        "Nyquist plot",
        "Bode plot",
        "Series by Pt",
        "Pre-estabilización",
        "Equivalent circuit fit",
        "MultiFit",
    ]

    for option in option_order:
        if option not in chosen:
            continue

        if option == "MultiFit":
            from pipelines.multi_eis_fit_pip import open_multifit_window

            open_multifit_window(eis_files=dta_files)
            continue

        option_figs: list[tuple[str, Figure]] = []
        if _wants_pre_stabilization([option]):
            pre_entries = _collect_pre_stabilization_entries(pre_stabilization_files)
            for entry in pre_entries:
                fig = fig_pre_stabilization(entry, font_defaults=font_defaults)
                if fig is not None:
                    option_figs.append((f"{entry.display_name} - Pre-estabilizacion", fig))
        else:
            for entry in plot_entries:
                option_figs.extend(build_figures(entry, [option], font_defaults=font_defaults))

        if option_figs:
            show_figures_tk(option_figs, window_title=f"EIS - {option}", font_defaults=font_defaults)

    return exported_xlsx


def main() -> None:
    """Manual standalone test."""
    input_dir = Path(r"C:\\path\\to\\your\\input")
    output_dir = Path(r"C:\\path\\to\\your\\output")

    exported = export_folder(input_dir, output_dir)
    print(f"Exported {len(exported)} file(s)")
    for path in exported:
        print(" -", path)


if __name__ == "__main__":
    main()
