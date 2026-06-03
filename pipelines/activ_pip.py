"""Activation (.DTA) -> Excel (.xlsx) exporter."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
import re

import tkinter as tk
from tkinter import ttk
from matplotlib import colors as mcolors
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
from matplotlib.ticker import LinearLocator, MaxNLocator, StrMethodFormatter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from plot_defaults import (
    PlotFontDefaults,
    apply_x_tick_label_padding,
    ensure_axis_bottom_margin,
    resolve_plot_font_defaults,
)
from ui_layout import create_resizable_plot_layout, create_scrollable_controls


META_ROWS_ORDER = [
    "Tecnica",
    "Fecha",
    "Hora",
    "Duracion del paso",
    "Rango I",
    "Paso I",
    "Tiempo de muestreo",
    "Area",
]

DATA_EXPORT = [
    ("Pt", "Pt", ""),
    ("T", "time", "s"),
    ("Vf", "Voltaje", "V"),
    ("Im", "Corriente", "A"),
    ("Sig", "Sig", "V"),
    ("Ach", "Ach", "V"),
    ("Temp", "Temperatura", "C"),
]

ACTIV_FILE_RE = re.compile(
    r"^Activacion_(?P<direction>Asc|Dsc)_(?P<label>.+?)_#(?P<cycle>\d+)_#(?P<file_index>\d+)\.DTA$",
    re.IGNORECASE,
)

LINESTYLE_OPTIONS = ["none", "-", "--", ":", "-."]
TIME_UNIT_OPTIONS = ["s", "min", "h"]
SECONDS_PER_MINUTE = 60.0
SECONDS_PER_HOUR = 3600.0
ACTIV_CYCLE_SELECTOR_HEIGHT = 176

ACTIV_PLOT_COLORS = {
    "voltage": "#1f5f99",
    "current": "#2f7d32",
    "temperature": "#b33a3a",
}

ACTIV_CYCLE_GRADIENTS = {
    "voltage": ("#5367ad", "#00b7ff"),
    "current": ("#067300", "#82d900"),
    "temperature": ("#700000", "#dbbe00"),
}


@dataclass(frozen=True)
class ActivationFile:
    path: Path
    direction: str
    label: str
    cycle: int
    file_index: int


@dataclass
class ParsedDTA:
    meta_values: dict[str, str]
    meta_units: dict[str, str]
    header: list[str]
    units: list[str]
    rows: list[list[str]]


@dataclass
class ActivationBundle:
    label: str
    asc_cycles: dict[int, list[ActivationFile]]
    dsc_cycles: dict[int, list[ActivationFile]]


@dataclass
class ActivationRamp:
    key: str
    direction: str
    cycle: int
    files: list[ActivationFile]


def to_float(val: str) -> float | None:
    s = val.strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _drop_leading_blank(parts: list[str]) -> list[str]:
    if parts and parts[0] == "":
        return parts[1:]
    return parts


def _extract_parenthesized_unit(text: str) -> str:
    matches = re.findall(r"\(([^()]*)\)", text)
    if matches:
        return matches[-1].strip()
    return ""


def _extract_meta_unit(key: str, description: str) -> str:
    unit = _extract_parenthesized_unit(description)
    if unit:
        return unit

    fallback_units = {
        "TITLE": "",
        "DATE": "",
        "TIME": "",
        "IPRESTEP": "A",
        "ISTEP1": "A",
        "ISTEP2": "A",
        "TPRESTEP": "s",
        "TSTEP1": "s",
        "TSTEP2": "s",
        "SAMPLETIME": "s",
        "AREA": "cm^2",
    }
    return fallback_units.get(key, "")


def parse_gamry_dta(path: Path) -> ParsedDTA:
    text = path.read_text(encoding="latin-1", errors="replace")
    lines = text.splitlines()

    meta_values: dict[str, str] = {}
    meta_units: dict[str, str] = {}
    header: list[str] | None = None
    units: list[str] = []
    rows: list[list[str]] = []
    table_started = False

    for line in lines:
        if not table_started:
            if line.startswith("CURVE") and "TABLE" in line:
                table_started = True
                continue

            if not line.strip():
                continue

            parts = line.split("\t")
            if len(parts) >= 3 and parts[0].strip():
                key = parts[0].strip()
                value = parts[2].strip()
                description = " ".join(p.strip() for p in parts[3:] if p.strip())
                meta_values[key] = value
                meta_units[key] = _extract_meta_unit(key, description)
            continue

        if not line.strip():
            continue

        parts = _drop_leading_blank([p.strip() for p in line.rstrip("\r\n").split("\t")])
        if not parts:
            continue

        first = parts[0]
        if header is None:
            if first == "Pt":
                header = parts
            continue

        if not units:
            if first == "#":
                units = parts
            continue

        if re.fullmatch(r"-?\d+", first):
            rows.append(parts)

    if header is None:
        raise ValueError(f"No data header found in {path.name} (expected 'Pt ...')")

    return ParsedDTA(
        meta_values=meta_values,
        meta_units=meta_units,
        header=header,
        units=units,
        rows=rows,
    )


def _column_index(parsed: ParsedDTA, column_name: str) -> int | None:
    try:
        return parsed.header.index(column_name)
    except ValueError:
        return None


def _required_float(row: list[str], idx: int | None) -> float:
    if idx is None or idx >= len(row):
        raise ValueError("Falta una columna requerida en la tabla CURVE.")
    num = to_float(row[idx])
    if num is None:
        raise ValueError(f"No se pudo convertir a numero: {row[idx]!r}")
    return num


def _extract_local_rows(parsed: ParsedDTA) -> list[dict[str, float]]:
    idx_map = {source: _column_index(parsed, source) for source, _, _ in DATA_EXPORT}
    missing = [source for source, idx in idx_map.items() if idx is None]
    if missing:
        raise ValueError("Faltan columnas requeridas en la tabla CURVE: " + ", ".join(missing))

    out: list[dict[str, float]] = []
    for raw_row in parsed.rows:
        record: dict[str, float] = {}
        for source, export_name, _unit in DATA_EXPORT:
            record[export_name] = _required_float(raw_row, idx_map[source])
        out.append(record)
    return out


def _parse_filename(path: Path) -> ActivationFile | None:
    match = ACTIV_FILE_RE.match(path.name)
    if not match:
        return None
    return ActivationFile(
        path=path,
        direction=match.group("direction").title(),
        label=match.group("label"),
        cycle=int(match.group("cycle")),
        file_index=int(match.group("file_index")),
    )


def discover_activation_bundles(input_dir: Path) -> list[ActivationBundle]:
    grouped: dict[str, dict[str, dict[int, list[ActivationFile]]]] = defaultdict(
        lambda: {"Asc": defaultdict(list), "Dsc": defaultdict(list)}
    )

    for path in sorted(input_dir.glob("*.DTA")):
        info = _parse_filename(path)
        if info is None:
            continue
        grouped[info.label][info.direction][info.cycle].append(info)

    bundles: list[ActivationBundle] = []
    for label, by_dir in sorted(grouped.items()):
        asc_cycles = {
            cycle: sorted(files, key=lambda item: item.file_index)
            for cycle, files in sorted(by_dir["Asc"].items())
        }
        dsc_cycles = {
            cycle: sorted(files, key=lambda item: item.file_index)
            for cycle, files in sorted(by_dir["Dsc"].items())
        }
        bundles.append(ActivationBundle(label=label, asc_cycles=asc_cycles, dsc_cycles=dsc_cycles))
    return bundles


def _step_delta_from_file(item: ActivationFile) -> float | None:
    parsed = parse_gamry_dta(item.path)
    i1 = to_float(parsed.meta_values.get("ISTEP1", ""))
    i2 = to_float(parsed.meta_values.get("ISTEP2", ""))
    if i1 is None or i2 is None:
        return None
    return abs(i2 - i1)


def infer_current_tolerance(files: list[ActivationFile]) -> float:
    for item in files:
        step_delta = _step_delta_from_file(item)
        if step_delta is not None and step_delta > 0:
            return max(min(step_delta * 0.1, 1e-3), 1e-5)
    return 1e-5


def build_activation_ramps(bundle: ActivationBundle) -> list[ActivationRamp]:
    ramps: list[ActivationRamp] = []
    cycle_ids = sorted(set(bundle.asc_cycles) | set(bundle.dsc_cycles))

    for cycle in cycle_ids:
        asc_files = bundle.asc_cycles.get(cycle, [])
        dsc_files = bundle.dsc_cycles.get(cycle, [])
        if asc_files:
            ramps.append(
                ActivationRamp(
                    key=f"Asc_{cycle}",
                    direction="Asc",
                    cycle=cycle,
                    files=asc_files,
                )
            )
        if dsc_files:
            ramps.append(
                ActivationRamp(
                    key=f"Dsc_{cycle}",
                    direction="Dsc",
                    cycle=cycle,
                    files=dsc_files,
                )
            )
    return ramps


def _fmt_range_value(i_min: float | None, i_max: float | None) -> str:
    if i_min is None or i_max is None:
        return ""
    return f"{i_min:g} a {i_max:g}"


def _collect_current_extremes(files: list[ActivationFile]) -> tuple[float | None, float | None]:
    currents: list[float] = []
    for item in files:
        parsed = parse_gamry_dta(item.path)
        for key in ("ISTEP1", "ISTEP2"):
            val = to_float(parsed.meta_values.get(key, ""))
            if val is not None:
                currents.append(val)
    if not currents:
        return None, None
    return min(currents), max(currents)


def build_metadata(bundle: ActivationBundle) -> list[tuple[str, object, str]]:
    all_files = [*sum(bundle.asc_cycles.values(), []), *sum(bundle.dsc_cycles.values(), [])]
    if not all_files:
        raise ValueError("No se encontraron archivos de activacion para exportar.")

    first_parsed = parse_gamry_dta(all_files[0].path)
    step_duration = to_float(first_parsed.meta_values.get("TSTEP1", ""))
    sample_time = to_float(first_parsed.meta_values.get("SAMPLETIME", ""))
    area = to_float(first_parsed.meta_values.get("AREA", ""))
    i1 = to_float(first_parsed.meta_values.get("ISTEP1", ""))
    i2 = to_float(first_parsed.meta_values.get("ISTEP2", ""))
    delta_i = abs(i2 - i1) if i1 is not None and i2 is not None else None
    i_min, i_max = _collect_current_extremes(all_files)

    metadata_map: dict[str, tuple[object, str]] = {
        "Tecnica": (first_parsed.meta_values.get("TITLE", ""), ""),
        "Fecha": (first_parsed.meta_values.get("DATE", ""), ""),
        "Hora": (first_parsed.meta_values.get("TIME", ""), ""),
        "Duracion del paso": (step_duration if step_duration is not None else "", "s"),
        "Rango I": (_fmt_range_value(i_min, i_max), "A"),
        "Paso I": (delta_i if delta_i is not None else "", "A"),
        "Tiempo de muestreo": (sample_time if sample_time is not None else "", "s"),
        "Area": (area if area is not None else "", "cm^2"),
    }

    return [(field, *metadata_map[field]) for field in META_ROWS_ORDER]


def concatenate_cycle_data(files: list[ActivationFile]) -> list[dict[str, float]]:
    all_rows: list[dict[str, float]] = []
    time_offset = 0.0
    global_pt = 0

    for item in files:
        parsed = parse_gamry_dta(item.path)
        local_rows = _extract_local_rows(parsed)
        if not local_rows:
            continue

        first_local_time = local_rows[0]["time"]
        for local in local_rows:
            row = {
                "Pt": float(global_pt),
                "time": local["time"] - first_local_time + time_offset,
                "Voltaje": local["Voltaje"],
                "Corriente": local["Corriente"],
                "Sig": local["Sig"],
                "Ach": local["Ach"],
                "Temperatura": local["Temperatura"],
            }
            all_rows.append(row)
            global_pt += 1

        sample_time = to_float(parsed.meta_values.get("SAMPLETIME", ""))
        if sample_time is None:
            if len(local_rows) >= 2:
                sample_time = local_rows[1]["time"] - local_rows[0]["time"]
            else:
                sample_time = 0.0

        time_offset = all_rows[-1]["time"] + sample_time

    return all_rows


def find_last_point_of_each_step(
    rows: list[dict[str, float]],
    current_tolerance: float,
) -> list[dict[str, float]]:
    if not rows:
        return []

    stable_rows: list[dict[str, float]] = []
    plateau_current = rows[0]["Corriente"]
    step_number = 1

    for idx in range(1, len(rows)):
        current = rows[idx]["Corriente"]
        if abs(current - plateau_current) > current_tolerance:
            last_row = dict(rows[idx - 1])
            last_row["Step"] = float(step_number)
            stable_rows.append(last_row)
            plateau_current = rows[idx]["Corriente"]
            step_number += 1

    last_row = dict(rows[-1])
    last_row["Step"] = float(step_number)
    stable_rows.append(last_row)
    return stable_rows


def build_v_vs_t_plot_data(bundle: ActivationBundle, time_unit: str = "s") -> dict[str, object]:
    ramps = build_activation_ramps(bundle)
    cycle_ids = sorted({ramp.cycle for ramp in ramps})
    time_scale = _time_unit_scale(time_unit)

    plotted_ramps: list[dict[str, object]] = []
    time_offset = 0.0

    for ramp in ramps:
        rows = concatenate_cycle_data(ramp.files)
        if not rows:
            continue

        plotted_rows: list[dict[str, float]] = []
        for row in rows:
            new_row = dict(row)
            new_row["plot_time"] = (row["time"] + time_offset) / time_scale
            plotted_rows.append(new_row)

        sample_time = 0.0
        if len(rows) >= 2:
            sample_time = max(0.0, rows[-1]["time"] - rows[-2]["time"])

        time_offset = rows[-1]["time"] + time_offset + sample_time
        plotted_ramps.append(
            {
                "key": ramp.key,
                "label": f"{ramp.direction} #{ramp.cycle}",
                "direction": ramp.direction,
                "cycle": ramp.cycle,
                "rows": plotted_rows,
            }
        )

    return {"ramps": plotted_ramps}


def build_visible_cycle_plot_data(
    bundle: ActivationBundle,
    visible_ramp_keys: set[str],
    time_unit: str = "s",
    local_cycle_time: bool = False,
) -> dict[str, object]:
    ramps = build_activation_ramps(bundle)
    visible_ramps = [ramp for ramp in ramps if ramp.key in visible_ramp_keys]
    cycle_ids = sorted({ramp.cycle for ramp in visible_ramps})
    time_scale = _time_unit_scale(time_unit)

    cycles: list[dict[str, object]] = []
    global_offset = 0.0

    for cycle in cycle_ids:
        cycle_ramps = [ramp for ramp in visible_ramps if ramp.cycle == cycle]
        cycle_ramps.sort(key=lambda item: 0 if item.direction == "Asc" else 1)

        cycle_rows: list[dict[str, float]] = []
        local_offset = 0.0
        for ramp in cycle_ramps:
            rows = concatenate_cycle_data(ramp.files)
            if not rows:
                continue

            ramp_start = rows[0]["time"]
            for row in rows:
                new_row = dict(row)
                if local_cycle_time:
                    new_row["plot_time"] = ((row["time"] - ramp_start) + local_offset) / time_scale
                else:
                    new_row["plot_time"] = ((row["time"] - ramp_start) + global_offset + local_offset) / time_scale
                cycle_rows.append(new_row)

            if len(rows) >= 2:
                sample_step = max(0.0, rows[-1]["time"] - rows[-2]["time"])
            else:
                sample_step = 0.0
            local_offset += (rows[-1]["time"] - ramp_start) + sample_step

        if cycle_rows:
            cycles.append({"cycle": cycle, "rows": cycle_rows})
            if not local_cycle_time:
                global_offset += local_offset

    return {"cycles": cycles}


def compute_default_v_vs_t_limits(bundle: ActivationBundle, time_unit: str = "s") -> dict[str, str]:
    plot_data = build_v_vs_t_plot_data(bundle, time_unit=time_unit)
    rows = [row for ramp in plot_data["ramps"] for row in ramp["rows"]]
    decimals = _time_unit_decimals(time_unit)

    if not rows:
        return {
            "t_min": "",
            "t_max": "",
            "v_min": "",
            "v_max": "",
            "i_min": "",
            "i_max": "",
            "temp_min": "",
            "temp_max": "",
        }

    t_min, t_max = _padded_limits([r["plot_time"] for r in rows], decimals=decimals)
    v_min, v_max = _padded_limits([r["Voltaje"] for r in rows])
    i_min, i_max = _padded_limits([r["Corriente"] for r in rows])
    temp_min, temp_max = _padded_limits([r["Temperatura"] for r in rows])

    return {
        "t_min": _format_limit_value(t_min, decimals),
        "t_max": _format_limit_value(t_max, decimals),
        "v_min": _format_limit_value(v_min),
        "v_max": _format_limit_value(v_max),
        "i_min": _format_limit_value(i_min),
        "i_max": _format_limit_value(i_max),
        "temp_min": _format_limit_value(temp_min),
        "temp_max": _format_limit_value(temp_max),
    }


def compute_autofit_v_vs_t_limits(
    bundle: ActivationBundle,
    visible_ramp_keys: set[str],
    show_voltage: bool,
    show_current: bool,
    show_temperature: bool,
    time_unit: str = "s",
    local_cycle_time: bool = False,
) -> dict[str, str]:
    if not visible_ramp_keys:
        raise ValueError("Debe seleccionar al menos una rampa para usar Autoescala.")
    if not (show_voltage or show_current or show_temperature):
        raise ValueError("Debe seleccionar al menos una magnitud para usar Autoescala.")

    plot_data = build_visible_cycle_plot_data(
        bundle,
        visible_ramp_keys=visible_ramp_keys,
        time_unit=time_unit,
        local_cycle_time=local_cycle_time,
    )
    rows = [row for cycle in plot_data["cycles"] for row in cycle["rows"]]
    if not rows:
        raise ValueError("No hay datos validos para ajustar los ejes.")

    decimals = _time_unit_decimals(time_unit)
    out = {
        "t_min": "",
        "t_max": "",
        "v_min": "",
        "v_max": "",
        "i_min": "",
        "i_max": "",
        "temp_min": "",
        "temp_max": "",
    }

    t_values = [r["plot_time"] for r in rows]
    if t_values:
        out["t_min"] = _format_limit_value(_round_down_dec(min(t_values), decimals), decimals)
        out["t_max"] = _format_limit_value(_round_up_dec(max(t_values), decimals), decimals)

    if show_voltage:
        v_values = [r["Voltaje"] for r in rows]
        if v_values:
            out["v_min"] = _format_limit_value(_round_down_dec(min(v_values)), 1)
            out["v_max"] = _format_limit_value(_round_up_dec(max(v_values)), 1)

    if show_current:
        i_values = [r["Corriente"] for r in rows]
        if i_values:
            out["i_min"] = _format_limit_value(_round_down_dec(min(i_values)), 1)
            out["i_max"] = _format_limit_value(_round_up_dec(max(i_values)), 1)

    if show_temperature:
        temp_values = [r["Temperatura"] for r in rows]
        if temp_values:
            out["temp_min"] = _format_limit_value(_round_down_dec(min(temp_values)), 1)
            out["temp_max"] = _format_limit_value(_round_up_dec(max(temp_values)), 1)

    return out


def build_visible_cycle_v_vs_i_data(
    bundle: ActivationBundle,
    visible_ramp_keys: set[str],
) -> dict[str, object]:
    ramps = build_activation_ramps(bundle)
    visible_ramps = [ramp for ramp in ramps if ramp.key in visible_ramp_keys]
    cycle_ids = sorted({ramp.cycle for ramp in visible_ramps})

    cycles: list[dict[str, object]] = []

    for cycle in cycle_ids:
        cycle_ramps = [ramp for ramp in visible_ramps if ramp.cycle == cycle]
        cycle_ramps.sort(key=lambda item: 0 if item.direction == "Asc" else 1)

        cycle_rows: list[dict[str, float]] = []
        for ramp in cycle_ramps:
            rows = concatenate_cycle_data(ramp.files)
            if rows:
                cycle_rows.extend(dict(row) for row in rows)

        if cycle_rows:
            cycles.append({"cycle": cycle, "rows": cycle_rows})

    return {"cycles": cycles}


def compute_default_v_vs_i_limits(bundle: ActivationBundle) -> dict[str, str]:
    ramps = build_activation_ramps(bundle)
    plot_data = build_visible_cycle_v_vs_i_data(bundle, {ramp.key for ramp in ramps})
    rows = [row for cycle in plot_data["cycles"] for row in cycle["rows"]]

    if not rows:
        return {
            "i_min": "",
            "i_max": "",
            "v_min": "",
            "v_max": "",
            "temp_min": "",
            "temp_max": "",
        }

    i_min, i_max = _padded_limits([r["Corriente"] for r in rows])
    v_min, v_max = _padded_limits([r["Voltaje"] for r in rows])
    temp_min, temp_max = _padded_limits([r["Temperatura"] for r in rows])

    return {
        "i_min": _format_limit_value(i_min),
        "i_max": _format_limit_value(i_max),
        "v_min": _format_limit_value(v_min),
        "v_max": _format_limit_value(v_max),
        "temp_min": _format_limit_value(temp_min),
        "temp_max": _format_limit_value(temp_max),
    }


def compute_autofit_v_vs_i_limits(
    bundle: ActivationBundle,
    visible_ramp_keys: set[str],
    show_voltage: bool,
    show_temperature: bool,
) -> dict[str, str]:
    if not visible_ramp_keys:
        raise ValueError("Debe seleccionar al menos una rampa para usar Autoescala.")
    if not (show_voltage or show_temperature):
        raise ValueError("Debe seleccionar al menos una magnitud para usar Autoescala.")

    plot_data = build_visible_cycle_v_vs_i_data(bundle, visible_ramp_keys=visible_ramp_keys)
    rows = [row for cycle in plot_data["cycles"] for row in cycle["rows"]]
    if not rows:
        raise ValueError("No hay datos validos para ajustar los ejes.")

    out = {
        "i_min": "",
        "i_max": "",
        "v_min": "",
        "v_max": "",
        "temp_min": "",
        "temp_max": "",
    }

    i_values = [r["Corriente"] for r in rows]
    if i_values:
        out["i_min"] = _format_limit_value(_round_down_dec(min(i_values)), 1)
        out["i_max"] = _format_limit_value(_round_up_dec(max(i_values)), 1)

    if show_voltage:
        v_values = [r["Voltaje"] for r in rows]
        if v_values:
            out["v_min"] = _format_limit_value(_round_down_dec(min(v_values)), 1)
            out["v_max"] = _format_limit_value(_round_up_dec(max(v_values)), 1)

    if show_temperature:
        temp_values = [r["Temperatura"] for r in rows]
        if temp_values:
            out["temp_min"] = _format_limit_value(_round_down_dec(min(temp_values)), 1)
            out["temp_max"] = _format_limit_value(_round_up_dec(max(temp_values)), 1)

    return out


def _write_metadata_sheet(ws, metadata_rows: list[tuple[str, object, str]]) -> None:
    ws.title = "Metadata"
    ws["A1"] = "Campo"
    ws["B1"] = "Valor"
    ws["C1"] = "Unidad"
    for ref in ("A1", "B1", "C1"):
        ws[ref].font = Font(bold=True)

    for row_idx, (field, value, unit) in enumerate(metadata_rows, start=2):
        ws.cell(row=row_idx, column=1, value=field)
        ws.cell(row=row_idx, column=2, value=value)
        ws.cell(row=row_idx, column=3, value=unit)

    ws.freeze_panes = "A2"


def _write_data_sheet(ws, rows: list[dict[str, float]], include_step: bool = False) -> None:
    headers = [label for _src, label, _unit in DATA_EXPORT]
    units = [unit for _src, _label, unit in DATA_EXPORT]

    if include_step:
        headers = ["Step"] + headers
        units = [""] + units

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for col_num, unit in enumerate(units, start=1):
        ws.cell(row=2, column=col_num, value=unit)

    for row_num, record in enumerate(rows, start=3):
        for col_num, header in enumerate(headers, start=1):
            value = record.get(header, "")
            if header in {"Pt", "Step"} and value != "":
                value = int(value)
            ws.cell(row=row_num, column=col_num, value=value)

    ws.freeze_panes = "A3"


def _mpl_linestyle(value: str) -> str:
    return "None" if value == "none" else value


def _optional_float(value: str | None) -> float | None:
    if value is None:
        return None
    text = value.strip()
    if not text:
        return None
    return float(text.replace(",", "."))


def _positive_float(text: str, name: str) -> float:
    value = text.strip().replace(",", ".")
    if not value:
        raise ValueError(f"{name} no puede estar vacio.")
    num = float(value)
    if num <= 0:
        raise ValueError(f"{name} debe ser mayor que 0.")
    return num


def _round_down_dec(value: float, decimals: int = 1) -> float:
    scale = 10 ** decimals
    return int(value * scale) / scale if value >= 0 else -int((-value * scale) + 0.9999999999) / scale


def _round_up_dec(value: float, decimals: int = 1) -> float:
    scale = 10 ** decimals
    return -int((-value * scale)) / scale if value <= 0 else int((value * scale) + 0.9999999999) / scale


def _format_limit_value(value: float | None, decimals: int = 1) -> str:
    if value is None:
        return ""
    return f"{value:.{decimals}f}"


def _padded_limits(
    values: list[float],
    rel_pad: float = 0.05,
    decimals: int = 1,
) -> tuple[float | None, float | None]:
    if not values:
        return None, None

    vmin = min(values)
    vmax = max(values)

    if vmin == vmax:
        pad = max(abs(vmin) * rel_pad, 1e-6)
    else:
        pad = (vmax - vmin) * rel_pad

    lo = _round_down_dec(vmin - pad, decimals)
    hi = _round_up_dec(vmax + pad, decimals)
    return lo, hi


def _time_unit_scale(time_unit: str) -> float:
    if time_unit == "min":
        return SECONDS_PER_MINUTE
    if time_unit == "h":
        return SECONDS_PER_HOUR
    return 1.0


def _time_unit_decimals(time_unit: str) -> int:
    if time_unit == "h":
        return 2
    if time_unit == "min":
        return 1
    return 1


def apply_x_edge_ticks(ax, x_min: float | None, x_max: float | None, tick_count: int) -> None:
    if x_min is None and x_max is None:
        return
    ax.xaxis.set_major_locator(LinearLocator(max(2, int(tick_count))))
    ax.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))


def apply_current_axis_scaling(
    ax_current,
    current_values: list[float],
    tick_count: int,
    i_min: float | None = None,
    i_max: float | None = None,
) -> None:
    if not current_values:
        return

    if i_min is not None or i_max is not None:
        current_lo, current_hi = ax_current.get_ylim()
        i_lo = current_lo if i_min is None else i_min
        i_hi = current_hi if i_max is None else i_max
    else:
        i_lo, i_hi = _padded_limits(current_values)
        if i_lo is None or i_hi is None:
            return

    if i_hi <= i_lo:
        i_hi = i_lo + 1.0
    tick_count = max(2, int(tick_count))
    ax_current.set_ylim(i_lo, i_hi)
    ax_current.yaxis.set_major_locator(LinearLocator(tick_count))
    ax_current.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))


def apply_temperature_axis_scaling(
    ax_temp,
    temp_values: list[float],
    tick_count: int,
    t_min: float | None = None,
    t_max: float | None = None,
) -> None:
    if not temp_values:
        return

    if t_min is not None or t_max is not None:
        current_lo, current_hi = ax_temp.get_ylim()
        y_lo = current_lo if t_min is None else t_min
        y_hi = current_hi if t_max is None else t_max
    else:
        y_lo, y_hi = _padded_limits(temp_values)
        if y_lo is None or y_hi is None:
            return

    if y_lo == y_hi:
        y_hi = y_lo + 1.0

    tick_count = max(2, int(tick_count))
    ax_temp.set_ylim(y_lo, y_hi)
    ax_temp.yaxis.set_major_locator(LinearLocator(tick_count))
    ax_temp.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))


def _cycle_gradient_color(color_range: tuple[str, str], index: int, count: int) -> tuple[float, float, float]:
    start_rgb = mcolors.to_rgb(color_range[0])
    end_rgb = mcolors.to_rgb(color_range[1])
    if count <= 1:
        t = 0.0
    else:
        t = index / (count - 1)
    return tuple(start + (end - start) * t for start, end in zip(start_rgb, end_rgb))


def _draw_cycle_scale_bars(
    fig: Figure,
    cycle_ids: list[int],
    show_voltage: bool,
    show_current: bool,
    show_temperature: bool,
    legend_fontsize: float,
    legend_scale: float,
) -> None:
    series_specs = []
    if show_voltage:
        series_specs.append(("V", "voltage"))
    if show_current:
        series_specs.append(("I", "current"))
    if show_temperature:
        series_specs.append(("T", "temperature"))

    if not series_specs or not cycle_ids:
        return

    legend_scale = max(0.5, legend_scale)
    count = len(series_specs)
    label_width = min(0.06, 0.025 + 0.01 * legend_scale)
    total_width = min(0.72, 0.22 + 0.12 * max(1, len(cycle_ids)) * legend_scale)
    bar_height = min(0.055, 0.022 * legend_scale)
    gap = min(0.03, 0.012 * legend_scale)
    total_height = count * bar_height + (count - 1) * gap
    left = 0.5 - (label_width + total_width) / 2
    bottom = max(0.03, 0.05 - 0.01 * (legend_scale - 1))

    for idx, (label, series_key) in enumerate(series_specs):
        y = bottom + total_height - (idx + 1) * bar_height - idx * gap
        label_ax = fig.add_axes([left, y, label_width, bar_height])
        label_ax.axis("off")
        label_ax.text(
            0.95,
            0.5,
            label,
            ha="right",
            va="center",
            fontsize=legend_fontsize,
            fontweight="bold",
        )

        bar_ax = fig.add_axes([left + label_width, y, total_width, bar_height])
        colors = [_cycle_gradient_color(ACTIV_CYCLE_GRADIENTS[series_key], color_idx, len(cycle_ids)) for color_idx in range(len(cycle_ids))]
        bar_ax.imshow([colors], aspect="auto")
        bar_ax.set_yticks([])
        if idx == count - 1:
            bar_ax.set_xticks(range(len(cycle_ids)))
            bar_ax.set_xticklabels([str(cycle) for cycle in cycle_ids], fontsize=max(6, legend_fontsize - 1))
            bar_ax.tick_params(axis="x", length=0, pad=1)
        else:
            bar_ax.set_xticks([])
        for spine in bar_ax.spines.values():
            spine.set_visible(False)


def _build_scrollable_controls(parent) -> ttk.Frame:
    _outer, inner = create_scrollable_controls(parent, outer_padding=0, inner_padding=(10, 0, 10, 0))
    return inner


def _build_scrollable_cycle_selector(parent, *, height: int = ACTIV_CYCLE_SELECTOR_HEIGHT):
    outer = ttk.Frame(parent)
    outer.pack(fill="x")
    outer.columnconfigure(0, weight=1)

    style = ttk.Style(parent)
    canvas_bg = style.lookup("App.TFrame", "background") or style.lookup("TFrame", "background") or "#10161d"

    canvas = tk.Canvas(
        outer,
        height=1,
        width=1,
        highlightthickness=0,
        borderwidth=0,
        bg=canvas_bg,
        bd=0,
        relief="flat",
    )
    scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
    inner = ttk.Frame(canvas)
    canvas.configure(yscrollcommand=scrollbar.set)
    window_id = canvas.create_window((0, 0), window=inner, anchor="nw")

    def _update_scrollregion(_event=None):
        bbox = canvas.bbox("all")
        if bbox is None:
            return
        canvas.configure(scrollregion=bbox)
        content_height = max(1, bbox[3] - bbox[1])
        target_height = min(height, content_height)
        if int(float(canvas.cget("height"))) != target_height:
            canvas.configure(height=target_height)

    def _sync_width(event):
        canvas.itemconfigure(window_id, width=event.width)

    def _on_mousewheel(event):
        if event.delta:
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        else:
            step = -1 if event.num == 4 else 1
            canvas.yview_scroll(step, "units")
        return "break"

    def _bind_mousewheel(widget):
        widget.bind("<MouseWheel>", _on_mousewheel, add="+")
        widget.bind("<Button-4>", _on_mousewheel, add="+")
        widget.bind("<Button-5>", _on_mousewheel, add="+")

    inner.bind("<Configure>", _update_scrollregion)
    canvas.bind("<Configure>", _sync_width)

    canvas.grid(row=0, column=0, sticky="nsew")
    scrollbar.grid(row=0, column=1, sticky="ns")

    _bind_mousewheel(canvas)
    _bind_mousewheel(inner)
    outer.after_idle(lambda: canvas.yview_moveto(0))

    return inner, _bind_mousewheel


def _axis_right_footprint_px(fig: Figure, ax) -> float:
    renderer = fig.canvas.get_renderer()
    axis_bbox = ax.yaxis.get_tightbbox(renderer)
    axes_bbox = ax.get_window_extent(renderer)
    if axis_bbox is None:
        return 0.0
    return max(0.0, axis_bbox.x1 - axes_bbox.x1)


def _pixels_to_points(fig: Figure, pixels: float) -> float:
    return pixels * 72.0 / fig.dpi


def _auto_format_sheet(ws) -> None:
    for col_num in range(1, ws.max_column + 1):
        max_len = 0
        for row_num in range(1, min(ws.max_row, 100) + 1):
            value = ws.cell(row=row_num, column=col_num).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_num)].width = min(max(10, max_len + 2), 45)

    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top")


def draw_v_vs_t_on_figure(
    fig: Figure,
    bundle: ActivationBundle,
    visible_ramp_keys: set[str],
    show_voltage: bool,
    show_current: bool,
    show_temperature: bool,
    voltage_linestyle: str,
    current_linestyle: str,
    temperature_linestyle: str,
    time_unit: str = "s",
    local_cycle_time: bool = False,
    x_tick_count: int = 6,
    y_tick_count: int = 6,
    t_min: float | None = None,
    t_max: float | None = None,
    v_min: float | None = None,
    v_max: float | None = None,
    temp_min: float | None = None,
    temp_max: float | None = None,
    current_min: float | None = None,
    current_max: float | None = None,
    plot_title: str = "",
    show_title: bool = True,
    title_fontsize: float = 14,
    tick_fontsize: float = 10,
    label_fontsize: float = 11,
    legend_fontsize: float = 10,
    legend_scale: float = 1.0,
    color_axes_by_magnitude: bool = False,
    line_width: float = 1.5,
) -> bool:
    fig.clear()

    if not visible_ramp_keys:
        return False
    if not (show_voltage or show_current or show_temperature):
        return False

    plot_data = build_visible_cycle_plot_data(
        bundle,
        visible_ramp_keys=visible_ramp_keys,
        time_unit=time_unit,
        local_cycle_time=local_cycle_time,
    )
    visible_cycles = plot_data["cycles"]
    if not visible_cycles:
        return False

    cycle_ids = [cycle_data["cycle"] for cycle_data in visible_cycles]

    x_tick_count = max(2, int(x_tick_count))
    y_tick_count = max(2, int(y_tick_count))

    ax_main = fig.add_subplot(111)
    ax_current = None
    ax_temp = None

    if show_voltage:
        if show_current:
            ax_current = ax_main.twinx()
            ax_current.spines["left"].set_visible(False)
            ax_current.yaxis.tick_right()
            ax_current.yaxis.set_label_position("right")
        if show_temperature:
            ax_temp = ax_main.twinx()
            ax_temp.spines["left"].set_visible(False)
            ax_temp.yaxis.tick_right()
            ax_temp.yaxis.set_label_position("right")
    else:
        if show_current:
            ax_current = ax_main
        if show_temperature:
            if show_current:
                ax_temp = ax_main.twinx()
                ax_temp.spines["left"].set_visible(False)
                ax_temp.yaxis.tick_right()
                ax_temp.yaxis.set_label_position("right")
            else:
                ax_temp = ax_main

    voltage_ls = _mpl_linestyle(voltage_linestyle)
    current_ls = _mpl_linestyle(current_linestyle)
    temp_ls = _mpl_linestyle(temperature_linestyle)

    axis_colors = {
        "voltage": ACTIV_PLOT_COLORS["voltage"],
        "current": ACTIV_PLOT_COLORS["current"],
        "temperature": ACTIV_PLOT_COLORS["temperature"],
    }

    for color_idx, cycle_data in enumerate(visible_cycles):
        grouped_rows = cycle_data["rows"]
        x_vals = [row["plot_time"] for row in grouped_rows]
        voltage_color = _cycle_gradient_color(ACTIV_CYCLE_GRADIENTS["voltage"], color_idx, len(cycle_ids))
        current_color = _cycle_gradient_color(ACTIV_CYCLE_GRADIENTS["current"], color_idx, len(cycle_ids))
        temp_color = _cycle_gradient_color(ACTIV_CYCLE_GRADIENTS["temperature"], color_idx, len(cycle_ids))

        if show_voltage and voltage_ls != "None":
            ax_main.plot(
                x_vals,
                [row["Voltaje"] for row in grouped_rows],
                color=voltage_color,
                linestyle=voltage_ls,
                linewidth=line_width,
            )

        if show_current and ax_current is not None and current_ls != "None":
            ax_current.plot(
                x_vals,
                [row["Corriente"] for row in grouped_rows],
                color=current_color,
                linestyle=current_ls,
                linewidth=line_width,
            )

        if show_temperature and ax_temp is not None and temp_ls != "None":
            ax_temp.plot(
                x_vals,
                [row["Temperatura"] for row in grouped_rows],
                color=temp_color,
                linestyle=temp_ls,
                linewidth=line_width,
            )

    handles, labels = [], []

    if ax_current is ax_main:
        current_values = [row["Corriente"] for cycle_data in visible_cycles for row in cycle_data["rows"]]
        ax_main.set_ylabel("Corriente (A)", fontsize=label_fontsize)
        apply_current_axis_scaling(ax_main, current_values, y_tick_count, current_min, current_max)

    if not (show_voltage and voltage_ls != "None") and not (show_current and current_ls != "None") and not (show_temperature and temp_ls != "None"):
        fig.clear()
        return False

    default_title = "V vs local t - Activacion " + bundle.label if local_cycle_time else f"V vs t - Activacion {bundle.label}"
    final_title = plot_title.strip() if plot_title.strip() else default_title

    ax_main.set_xlabel(f"Tiempo ({time_unit})", fontsize=label_fontsize)
    ax_main.set_title(final_title if show_title else "", fontsize=title_fontsize)
    ax_main.grid(True)
    ax_main.xaxis.set_major_locator(MaxNLocator(nbins=x_tick_count))
    ax_main.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
    ax_main.tick_params(axis="both", labelsize=tick_fontsize)
    apply_x_tick_label_padding(ax_main, tick_fontsize)

    if t_min is not None or t_max is not None:
        ax_main.set_xlim(left=t_min, right=t_max)
        apply_x_edge_ticks(ax_main, t_min, t_max, x_tick_count)

    if show_voltage:
        ax_main.set_ylabel("Voltaje (V)", fontsize=label_fontsize)
        if v_min is not None or v_max is not None:
            ax_main.set_ylim(bottom=v_min, top=v_max)
            ax_main.yaxis.set_major_locator(LinearLocator(y_tick_count))
        else:
            ax_main.yaxis.set_major_locator(MaxNLocator(nbins=y_tick_count))
        ax_main.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
        if color_axes_by_magnitude:
            ax_main.yaxis.label.set_color(axis_colors["voltage"])
            ax_main.tick_params(axis="y", colors=axis_colors["voltage"], labelsize=tick_fontsize)
            ax_main.spines["left"].set_color(axis_colors["voltage"])
    elif ax_temp is ax_main:
        temp_values = [row["Temperatura"] for cycle_data in visible_cycles for row in cycle_data["rows"]]
        ax_main.set_ylabel("Temperatura (C)", fontsize=label_fontsize)
        apply_temperature_axis_scaling(ax_main, temp_values, y_tick_count, temp_min, temp_max)
        if color_axes_by_magnitude:
            ax_main.yaxis.label.set_color(axis_colors["temperature"])
            ax_main.tick_params(axis="y", colors=axis_colors["temperature"], labelsize=tick_fontsize)
            ax_main.spines["left"].set_color(axis_colors["temperature"])

    if ax_current is not None and ax_current is not ax_main:
        current_values = [row["Corriente"] for cycle_data in visible_cycles for row in cycle_data["rows"]]
        ax_current.tick_params(axis="y", labelsize=tick_fontsize)
        ax_current.set_ylabel("Corriente (A)", fontsize=label_fontsize)
        apply_current_axis_scaling(ax_current, current_values, y_tick_count, current_min, current_max)
        if color_axes_by_magnitude:
            ax_current.yaxis.label.set_color(axis_colors["current"])
            ax_current.tick_params(axis="y", colors=axis_colors["current"], labelsize=tick_fontsize)
            ax_current.spines["right"].set_color(axis_colors["current"])

    if ax_temp is not None and ax_temp is not ax_main:
        temp_values = [row["Temperatura"] for cycle_data in visible_cycles for row in cycle_data["rows"]]
        ax_temp.tick_params(axis="y", labelsize=tick_fontsize)
        ax_temp.set_ylabel("Temperatura (C)", fontsize=label_fontsize)
        apply_temperature_axis_scaling(ax_temp, temp_values, y_tick_count, temp_min, temp_max)
        if color_axes_by_magnitude:
            ax_temp.yaxis.label.set_color(axis_colors["temperature"])
            ax_temp.tick_params(axis="y", colors=axis_colors["temperature"], labelsize=tick_fontsize)
            ax_temp.spines["right"].set_color(axis_colors["temperature"])

    if ax_current is ax_main and color_axes_by_magnitude:
        ax_main.yaxis.label.set_color(axis_colors["current"])
        ax_main.tick_params(axis="y", colors=axis_colors["current"], labelsize=tick_fontsize)
        ax_main.spines["left"].set_color(axis_colors["current"])

    if show_voltage or show_current or show_temperature:
        legend_rows = sum([
            1 if show_voltage and voltage_ls != "None" else 0,
            1 if show_current and current_ls != "None" else 0,
            1 if show_temperature and temp_ls != "None" else 0,
        ])
        bottom_margin = min(0.32, 0.12 + legend_rows * (0.04 * max(0.5, legend_scale)))
        fig.subplots_adjust(right=0.90, bottom=bottom_margin)
        bottom_margin = ensure_axis_bottom_margin(
            fig,
            ax_main,
            bottom_margin,
            tick_fontsize,
            max_bottom_margin=0.38,
        )
        fig.subplots_adjust(right=0.90, bottom=bottom_margin)

        if ax_current is not None or ax_temp is not None:
            fig.canvas.draw()

            right_gap_px = max(12.0, tick_fontsize * 1.2)
            outer_pad_px = max(10.0, tick_fontsize)
            current_axis_offset_px = 0.0
            total_right_px = 0.0

            if ax_current is not None and ax_current is not ax_main:
                total_right_px = _axis_right_footprint_px(fig, ax_current)

            if ax_temp is not None and ax_temp is not ax_main:
                if ax_current is not None and ax_current is not ax_main:
                    current_axis_offset_px = _axis_right_footprint_px(fig, ax_current) + right_gap_px
                    ax_temp.spines["right"].set_position(
                        ("outward", _pixels_to_points(fig, current_axis_offset_px))
                    )
                    fig.canvas.draw()
                temp_axis_width_px = _axis_right_footprint_px(fig, ax_temp)
                total_right_px = max(total_right_px, current_axis_offset_px + temp_axis_width_px)

            fig_width_px = fig.get_size_inches()[0] * fig.dpi
            dynamic_right_margin = 1.0 - ((total_right_px + outer_pad_px) / fig_width_px)
            fig.subplots_adjust(right=min(0.96, max(0.55, dynamic_right_margin)), bottom=bottom_margin)

        _draw_cycle_scale_bars(
            fig,
            cycle_ids,
            show_voltage and voltage_ls != "None",
            show_current and current_ls != "None",
            show_temperature and temp_ls != "None",
            legend_fontsize,
            legend_scale,
        )

    return True


def draw_v_vs_i_on_figure(
    fig: Figure,
    bundle: ActivationBundle,
    visible_ramp_keys: set[str],
    show_voltage: bool,
    show_temperature: bool,
    voltage_linestyle: str,
    temperature_linestyle: str,
    x_tick_count: int = 6,
    y_tick_count: int = 6,
    i_min: float | None = None,
    i_max: float | None = None,
    v_min: float | None = None,
    v_max: float | None = None,
    temp_min: float | None = None,
    temp_max: float | None = None,
    plot_title: str = "",
    show_title: bool = True,
    title_fontsize: float = 14,
    tick_fontsize: float = 10,
    label_fontsize: float = 11,
    legend_fontsize: float = 10,
    legend_scale: float = 1.0,
    color_axes_by_magnitude: bool = False,
    line_width: float = 1.5,
) -> bool:
    fig.clear()

    if not visible_ramp_keys:
        return False
    if not (show_voltage or show_temperature):
        return False

    plot_data = build_visible_cycle_v_vs_i_data(bundle, visible_ramp_keys=visible_ramp_keys)
    visible_cycles = plot_data["cycles"]
    if not visible_cycles:
        return False

    cycle_ids = [cycle_data["cycle"] for cycle_data in visible_cycles]
    x_tick_count = max(2, int(x_tick_count))
    y_tick_count = max(2, int(y_tick_count))

    ax_main = fig.add_subplot(111)
    ax_temp = None

    if show_voltage and show_temperature:
        ax_temp = ax_main.twinx()
        ax_temp.spines["left"].set_visible(False)
        ax_temp.yaxis.tick_right()
        ax_temp.yaxis.set_label_position("right")

    voltage_ls = _mpl_linestyle(voltage_linestyle)
    temp_ls = _mpl_linestyle(temperature_linestyle)

    axis_colors = {
        "voltage": ACTIV_PLOT_COLORS["voltage"],
        "temperature": ACTIV_PLOT_COLORS["temperature"],
    }

    for color_idx, cycle_data in enumerate(visible_cycles):
        grouped_rows = cycle_data["rows"]
        x_vals = [row["Corriente"] for row in grouped_rows]
        voltage_color = _cycle_gradient_color(ACTIV_CYCLE_GRADIENTS["voltage"], color_idx, len(cycle_ids))
        temp_color = _cycle_gradient_color(ACTIV_CYCLE_GRADIENTS["temperature"], color_idx, len(cycle_ids))

        if show_voltage and voltage_ls != "None":
            ax_main.plot(
                x_vals,
                [row["Voltaje"] for row in grouped_rows],
                color=voltage_color,
                linestyle=voltage_ls,
                linewidth=line_width,
            )

        if show_temperature:
            target_ax = ax_temp if ax_temp is not None else ax_main
            if temp_ls != "None":
                target_ax.plot(
                    x_vals,
                    [row["Temperatura"] for row in grouped_rows],
                    color=temp_color,
                    linestyle=temp_ls,
                    linewidth=line_width,
                )

    if not (show_voltage and voltage_ls != "None") and not (show_temperature and temp_ls != "None"):
        fig.clear()
        return False

    default_title = f"V vs I - Activacion {bundle.label}"
    final_title = plot_title.strip() if plot_title.strip() else default_title

    ax_main.set_xlabel("Corriente (A)", fontsize=label_fontsize)
    ax_main.set_title(final_title if show_title else "", fontsize=title_fontsize)
    ax_main.grid(True)
    ax_main.xaxis.set_major_locator(MaxNLocator(nbins=x_tick_count))
    ax_main.xaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
    ax_main.tick_params(axis="both", labelsize=tick_fontsize)
    apply_x_tick_label_padding(ax_main, tick_fontsize)

    if i_min is not None or i_max is not None:
        ax_main.set_xlim(left=i_min, right=i_max)
        apply_x_edge_ticks(ax_main, i_min, i_max, x_tick_count)

    if show_voltage:
        ax_main.set_ylabel("Voltaje (V)", fontsize=label_fontsize)
        if v_min is not None or v_max is not None:
            ax_main.set_ylim(bottom=v_min, top=v_max)
            ax_main.yaxis.set_major_locator(LinearLocator(y_tick_count))
        else:
            ax_main.yaxis.set_major_locator(MaxNLocator(nbins=y_tick_count))
        ax_main.yaxis.set_major_formatter(StrMethodFormatter("{x:g}"))
        if color_axes_by_magnitude:
            ax_main.yaxis.label.set_color(axis_colors["voltage"])
            ax_main.tick_params(axis="y", colors=axis_colors["voltage"], labelsize=tick_fontsize)
            ax_main.spines["left"].set_color(axis_colors["voltage"])
    else:
        temp_values = [row["Temperatura"] for cycle_data in visible_cycles for row in cycle_data["rows"]]
        ax_main.set_ylabel("Temperatura (C)", fontsize=label_fontsize)
        apply_temperature_axis_scaling(ax_main, temp_values, y_tick_count, temp_min, temp_max)
        if color_axes_by_magnitude:
            ax_main.yaxis.label.set_color(axis_colors["temperature"])
            ax_main.tick_params(axis="y", colors=axis_colors["temperature"], labelsize=tick_fontsize)
            ax_main.spines["left"].set_color(axis_colors["temperature"])

    if ax_temp is not None:
        temp_values = [row["Temperatura"] for cycle_data in visible_cycles for row in cycle_data["rows"]]
        ax_temp.tick_params(axis="y", labelsize=tick_fontsize)
        ax_temp.set_ylabel("Temperatura (C)", fontsize=label_fontsize)
        apply_temperature_axis_scaling(ax_temp, temp_values, y_tick_count, temp_min, temp_max)
        if color_axes_by_magnitude:
            ax_temp.yaxis.label.set_color(axis_colors["temperature"])
            ax_temp.tick_params(axis="y", colors=axis_colors["temperature"], labelsize=tick_fontsize)
            ax_temp.spines["right"].set_color(axis_colors["temperature"])

    if show_voltage or show_temperature:
        legend_rows = sum([
            1 if show_voltage and voltage_ls != "None" else 0,
            1 if show_temperature and temp_ls != "None" else 0,
        ])
        bottom_margin = min(0.32, 0.12 + legend_rows * (0.04 * max(0.5, legend_scale)))
        fig.subplots_adjust(right=0.90, bottom=bottom_margin)
        bottom_margin = ensure_axis_bottom_margin(
            fig,
            ax_main,
            bottom_margin,
            tick_fontsize,
            max_bottom_margin=0.38,
        )
        fig.subplots_adjust(right=0.90, bottom=bottom_margin)

        if ax_temp is not None:
            fig.canvas.draw()
            total_right_px = _axis_right_footprint_px(fig, ax_temp)
            outer_pad_px = max(10.0, tick_fontsize)
            fig_width_px = fig.get_size_inches()[0] * fig.dpi
            dynamic_right_margin = 1.0 - ((total_right_px + outer_pad_px) / fig_width_px)
            fig.subplots_adjust(right=min(0.96, max(0.55, dynamic_right_margin)), bottom=bottom_margin)

        _draw_cycle_scale_bars(
            fig,
            cycle_ids,
            show_voltage and voltage_ls != "None",
            False,
            show_temperature and temp_ls != "None",
            legend_fontsize,
            legend_scale,
        )

    return True


def export_activation_bundle(bundle: ActivationBundle, out_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws_meta = wb.create_sheet("Metadata")
    _write_metadata_sheet(ws_meta, build_metadata(bundle))

    all_sheets = [ws_meta]

    for cycle, files in sorted(bundle.asc_cycles.items()):
        cycle_rows = concatenate_cycle_data(files)
        ws = wb.create_sheet(f"Asc_#{cycle}")
        _write_data_sheet(ws, cycle_rows)
        all_sheets.append(ws)

        last_rows = find_last_point_of_each_step(cycle_rows, infer_current_tolerance(files))
        ws_last = wb.create_sheet(f"Asc_#{cycle}_last")
        _write_data_sheet(ws_last, last_rows, include_step=True)
        all_sheets.append(ws_last)

    for cycle, files in sorted(bundle.dsc_cycles.items()):
        cycle_rows = concatenate_cycle_data(files)
        ws = wb.create_sheet(f"Dsc_#{cycle}")
        _write_data_sheet(ws, cycle_rows)
        all_sheets.append(ws)

        last_rows = find_last_point_of_each_step(cycle_rows, infer_current_tolerance(files))
        ws_last = wb.create_sheet(f"Dsc_#{cycle}_last")
        _write_data_sheet(ws_last, last_rows, include_step=True)
        all_sheets.append(ws_last)

    for ws in all_sheets:
        _auto_format_sheet(ws)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def open_v_vs_t_window(input_dir: Path, font_defaults: PlotFontDefaults | None = None) -> None:
    bundles = discover_activation_bundles(Path(input_dir))
    if not bundles:
        raise ValueError("No se encontraron archivos de activacion validos.")

    font_defaults = resolve_plot_font_defaults(font_defaults)
    font_default_values = font_defaults.as_strings()
    bundle = bundles[0]
    default_limits = compute_default_v_vs_t_limits(bundle, time_unit="s")
    ramps = build_activation_ramps(bundle)
    cycle_ids = sorted({ramp.cycle for ramp in ramps})

    win = tk.Toplevel()
    win.title(f"Activacion - V vs t - {bundle.label}")
    win.geometry("1200x720")

    controls_host, plot_outer = create_resizable_plot_layout(
        win,
        sidebar_width=420,
        min_sidebar_width=390,
    )
    controls_frame = _build_scrollable_controls(controls_host)

    toolbar_frame = ttk.Frame(plot_outer)
    toolbar_frame.pack(side="top", fill="x")

    canvas_frame = ttk.Frame(plot_outer)
    canvas_frame.pack(side="top", fill="both", expand=True)

    fig = Figure(figsize=(9, 5.5), dpi=100)
    canvas = FigureCanvasTkAgg(fig, master=canvas_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    toolbar = NavigationToolbar2Tk(canvas, toolbar_frame, pack_toolbar=False)
    toolbar.update()
    toolbar.pack(side="left", fill="x")

    status_var = tk.StringVar(value="Listo.")

    asc_var = tk.BooleanVar(value=True)
    dsc_var = tk.BooleanVar(value=True)
    voltage_var = tk.BooleanVar(value=True)
    current_var = tk.BooleanVar(value=False)
    temperature_var = tk.BooleanVar(value=False)
    time_unit_var = tk.StringVar(value="s")
    local_cycle_time_var = tk.BooleanVar(value=False)
    color_axes_var = tk.BooleanVar(value=False)

    voltage_line_var = tk.StringVar(value="-")
    current_line_var = tk.StringVar(value="-.")
    temperature_line_var = tk.StringVar(value="--")

    t_min_var = tk.StringVar(value=default_limits["t_min"])
    t_max_var = tk.StringVar(value=default_limits["t_max"])
    v_min_var = tk.StringVar(value=default_limits["v_min"])
    v_max_var = tk.StringVar(value=default_limits["v_max"])
    i_min_var = tk.StringVar(value=default_limits["i_min"])
    i_max_var = tk.StringVar(value=default_limits["i_max"])
    temp_min_var = tk.StringVar(value=default_limits["temp_min"])
    temp_max_var = tk.StringVar(value=default_limits["temp_max"])

    x_tick_count_var = tk.IntVar(value=6)
    y_tick_count_var = tk.IntVar(value=6)

    plot_title_var = tk.StringVar(value="")
    show_title_var = tk.BooleanVar(value=True)
    title_fontsize_var = tk.StringVar(value=font_default_values["title"])
    tick_fontsize_var = tk.StringVar(value=font_default_values["tick"])
    label_fontsize_var = tk.StringVar(value=font_default_values["label"])
    legend_fontsize_var = tk.StringVar(value=font_default_values["legend"])
    legend_scale_var = tk.StringVar(value="1.0")
    line_width_var = tk.StringVar(value="1.5")

    cycle_vars = {cycle: tk.BooleanVar(value=True) for cycle in cycle_ids}

    initial_state = {
        "asc": True,
        "dsc": True,
        "voltage": True,
        "current": False,
        "temperature": False,
        "time_unit": "s",
        "local_cycle_time": False,
        "color_axes": False,
        "voltage_line": "-",
        "current_line": "-.",
        "temperature_line": "--",
        "t_min": default_limits["t_min"],
        "t_max": default_limits["t_max"],
        "v_min": default_limits["v_min"],
        "v_max": default_limits["v_max"],
        "i_min": default_limits["i_min"],
        "i_max": default_limits["i_max"],
        "temp_min": default_limits["temp_min"],
        "temp_max": default_limits["temp_max"],
        "x_tick_count": 6,
        "y_tick_count": 6,
        "plot_title": "",
        "show_title": True,
        "title_fontsize": font_default_values["title"],
        "tick_fontsize": font_default_values["tick"],
        "label_fontsize": font_default_values["label"],
        "legend_fontsize": font_default_values["legend"],
        "legend_scale": "1.0",
        "line_width": "1.5",
        "visible_cycles": {cycle: True for cycle in cycle_ids},
    }

    def _visible_ramp_keys() -> set[str]:
        visible_cycles = {cycle for cycle, var in cycle_vars.items() if var.get()}
        keys: set[str] = set()
        for ramp in ramps:
            if ramp.cycle not in visible_cycles:
                continue
            if ramp.direction == "Asc" and not asc_var.get():
                continue
            if ramp.direction == "Dsc" and not dsc_var.get():
                continue
            keys.add(ramp.key)
        return keys

    def _collect_limits():
        return dict(
            t_min=_optional_float(t_min_var.get()),
            t_max=_optional_float(t_max_var.get()),
            v_min=_optional_float(v_min_var.get()),
            v_max=_optional_float(v_max_var.get()),
            current_min=_optional_float(i_min_var.get()),
            current_max=_optional_float(i_max_var.get()),
            temp_min=_optional_float(temp_min_var.get()),
            temp_max=_optional_float(temp_max_var.get()),
        )

    plot_job = {"id": None}
    suspend_events = {"value": False}
    current_time_unit = {"value": time_unit_var.get()}

    def _schedule_plot(*_args):
        if suspend_events["value"]:
            return
        if plot_job["id"] is not None:
            win.after_cancel(plot_job["id"])
        plot_job["id"] = win.after(20, _plot)

    def _plot():
        plot_job["id"] = None
        try:
            has_plot = draw_v_vs_t_on_figure(
                fig=fig,
                bundle=bundle,
                visible_ramp_keys=_visible_ramp_keys(),
                show_voltage=voltage_var.get(),
                show_current=current_var.get(),
                show_temperature=temperature_var.get(),
                voltage_linestyle=voltage_line_var.get(),
                current_linestyle=current_line_var.get(),
                temperature_linestyle=temperature_line_var.get(),
                time_unit=time_unit_var.get(),
                local_cycle_time=local_cycle_time_var.get(),
                color_axes_by_magnitude=color_axes_var.get(),
                x_tick_count=x_tick_count_var.get(),
                y_tick_count=y_tick_count_var.get(),
                plot_title=plot_title_var.get(),
                show_title=show_title_var.get(),
                title_fontsize=_positive_float(title_fontsize_var.get(), "Tamaño del título"),
                tick_fontsize=_positive_float(tick_fontsize_var.get(), "Tamaño de ticks"),
                label_fontsize=_positive_float(label_fontsize_var.get(), "Tamaño de etiquetas"),
                legend_fontsize=_positive_float(legend_fontsize_var.get(), "Tamaño de leyenda"),
                legend_scale=_positive_float(legend_scale_var.get(), "Escala del gradiente"),
                line_width=_positive_float(line_width_var.get(), "Grosor de línea"),
                **_collect_limits(),
            )
        except ValueError as exc:
            fig.clear()
            canvas.draw_idle()
            status_var.set(f"Error: {exc}")
            return

        if not has_plot:
            fig.clear()
            canvas.draw_idle()
            status_var.set("No se muestra grafico: seleccione al menos una rampa y una magnitud.")
            return

        canvas.draw_idle()
        status_var.set("Grafico actualizado.")

    def _autofit():
        try:
            fitted = compute_autofit_v_vs_t_limits(
                bundle=bundle,
                visible_ramp_keys=_visible_ramp_keys(),
                show_voltage=voltage_var.get(),
                show_current=current_var.get(),
                show_temperature=temperature_var.get(),
                time_unit=time_unit_var.get(),
                local_cycle_time=local_cycle_time_var.get(),
            )
        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            return

        suspend_events["value"] = True
        try:
            t_min_var.set(fitted["t_min"])
            t_max_var.set(fitted["t_max"])
            if fitted["v_min"] != "" or fitted["v_max"] != "":
                v_min_var.set(fitted["v_min"])
                v_max_var.set(fitted["v_max"])
            if fitted["i_min"] != "" or fitted["i_max"] != "":
                i_min_var.set(fitted["i_min"])
                i_max_var.set(fitted["i_max"])
            if fitted["temp_min"] != "" or fitted["temp_max"] != "":
                temp_min_var.set(fitted["temp_min"])
                temp_max_var.set(fitted["temp_max"])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Autoescala aplicada.")

    def _on_local_cycle_time_changed():
        if suspend_events["value"]:
            return
        _autofit()

    def _convert_time_limit_text(value_text: str, from_unit: str, to_unit: str) -> str:
        value = _optional_float(value_text)
        if value is None or from_unit == to_unit:
            return value_text
        seconds = value * _time_unit_scale(from_unit)
        converted = seconds / _time_unit_scale(to_unit)
        return _format_limit_value(converted, _time_unit_decimals(to_unit))

    def _on_time_unit_changed(*_args):
        if suspend_events["value"]:
            return

        old_unit = current_time_unit["value"]
        new_unit = time_unit_var.get()

        suspend_events["value"] = True
        try:
            if t_min_var.get().strip():
                t_min_var.set(_convert_time_limit_text(t_min_var.get(), old_unit, new_unit))
            if t_max_var.get().strip():
                t_max_var.set(_convert_time_limit_text(t_max_var.get(), old_unit, new_unit))
            current_time_unit["value"] = new_unit
        finally:
            suspend_events["value"] = False

        _schedule_plot()

    def _reset():
        suspend_events["value"] = True
        try:
            asc_var.set(initial_state["asc"])
            dsc_var.set(initial_state["dsc"])
            voltage_var.set(initial_state["voltage"])
            current_var.set(initial_state["current"])
            temperature_var.set(initial_state["temperature"])
            time_unit_var.set(initial_state["time_unit"])
            local_cycle_time_var.set(initial_state["local_cycle_time"])
            color_axes_var.set(initial_state["color_axes"])
            current_time_unit["value"] = initial_state["time_unit"]
            voltage_line_var.set(initial_state["voltage_line"])
            current_line_var.set(initial_state["current_line"])
            temperature_line_var.set(initial_state["temperature_line"])
            t_min_var.set(initial_state["t_min"])
            t_max_var.set(initial_state["t_max"])
            v_min_var.set(initial_state["v_min"])
            v_max_var.set(initial_state["v_max"])
            i_min_var.set(initial_state["i_min"])
            i_max_var.set(initial_state["i_max"])
            temp_min_var.set(initial_state["temp_min"])
            temp_max_var.set(initial_state["temp_max"])
            x_tick_count_var.set(initial_state["x_tick_count"])
            y_tick_count_var.set(initial_state["y_tick_count"])
            plot_title_var.set(initial_state["plot_title"])
            show_title_var.set(initial_state["show_title"])
            title_fontsize_var.set(initial_state["title_fontsize"])
            tick_fontsize_var.set(initial_state["tick_fontsize"])
            label_fontsize_var.set(initial_state["label_fontsize"])
            legend_fontsize_var.set(initial_state["legend_fontsize"])
            legend_scale_var.set(initial_state["legend_scale"])
            line_width_var.set(initial_state["line_width"])
            for cycle, var in cycle_vars.items():
                var.set(initial_state["visible_cycles"][cycle])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Valores restaurados.")

    ttk.Label(
        controls_frame,
        text=f"Serie detectada: Activacion {bundle.label}",
        justify="left",
        wraplength=260,
    ).pack(anchor="w", pady=(0, 10))

    cycles_box = ttk.LabelFrame(controls_frame, text="Ciclos")
    cycles_box.pack(fill="x", pady=5)

    series_box = ttk.LabelFrame(controls_frame, text="Series")
    series_box.pack(fill="x", pady=5)

    style_box = ttk.LabelFrame(controls_frame, text="Estilo")
    style_box.pack(fill="x", pady=5)

    text_box = ttk.LabelFrame(controls_frame, text="Texto / tamaños")
    text_box.pack(fill="x", pady=5)

    limits_box = ttk.LabelFrame(controls_frame, text="Límites de ejes")
    limits_box.pack(fill="x", pady=5)

    cycles_inner, bind_cycle_scroll = _build_scrollable_cycle_selector(cycles_box)

    for index, cycle in enumerate(cycle_ids):
        cycle_toggle = ttk.Checkbutton(
            cycles_inner,
            text=f"Ciclo #{cycle}",
            variable=cycle_vars[cycle],
            command=_schedule_plot,
        )
        pady = (0, 2) if index < len(cycle_ids) - 1 else 0
        cycle_toggle.pack(anchor="w", padx=4, pady=pady)
        bind_cycle_scroll(cycle_toggle)

    ttk.Checkbutton(series_box, text="Ascendente", variable=asc_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="Descendente", variable=dsc_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="Voltaje", variable=voltage_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="Corriente", variable=current_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="Temperatura", variable=temperature_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="Tiempo local del ciclo", variable=local_cycle_time_var, command=_on_local_cycle_time_changed).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="y-axes a color", variable=color_axes_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Label(series_box, text="Unidad de tiempo").pack(anchor="w", padx=8, pady=(8, 2))
    time_unit_combo = ttk.Combobox(series_box, textvariable=time_unit_var, values=TIME_UNIT_OPTIONS, state="readonly", width=8)
    time_unit_combo.pack(anchor="w", padx=8, pady=(0, 4))

    ttk.Label(style_box, text="Línea de voltaje").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    voltage_line_combo = ttk.Combobox(style_box, textvariable=voltage_line_var, values=LINESTYLE_OPTIONS, state="readonly", width=10)
    voltage_line_combo.grid(row=0, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Línea de corriente").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    current_line_combo = ttk.Combobox(style_box, textvariable=current_line_var, values=LINESTYLE_OPTIONS, state="readonly", width=10)
    current_line_combo.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Línea de temperatura").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    temperature_line_combo = ttk.Combobox(style_box, textvariable=temperature_line_var, values=LINESTYLE_OPTIONS, state="readonly", width=10)
    temperature_line_combo.grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Título").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    title_entry = ttk.Entry(text_box, textvariable=plot_title_var, width=28)
    title_entry.grid(row=0, column=1, sticky="we", padx=8, pady=3)
    ttk.Checkbutton(text_box, text="Mostrar titulo", variable=show_title_var, command=_schedule_plot).grid(
        row=7, column=0, columnspan=2, sticky="w", padx=8, pady=3
    )

    ttk.Label(text_box, text="Tamaño del título").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    title_size_spin = tk.Spinbox(text_box, from_=6, to=30, increment=0.5, textvariable=title_fontsize_var, width=8)
    title_size_spin.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de ticks").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    tick_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=tick_fontsize_var, width=8)
    tick_size_spin.grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de etiquetas").grid(row=3, column=0, sticky="w", padx=8, pady=3)
    label_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=label_fontsize_var, width=8)
    label_size_spin.grid(row=3, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de leyenda").grid(row=4, column=0, sticky="w", padx=8, pady=3)
    legend_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=legend_fontsize_var, width=8)
    legend_size_spin.grid(row=4, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Escala del gradiente").grid(row=5, column=0, sticky="w", padx=8, pady=3)
    legend_scale_spin = tk.Spinbox(text_box, from_=0.5, to=3.0, increment=0.1, textvariable=legend_scale_var, width=8)
    legend_scale_spin.grid(row=5, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Grosor de línea").grid(row=6, column=0, sticky="w", padx=8, pady=3)
    line_width_spin = tk.Spinbox(text_box, from_=0.0, to=10.0, increment=0.1, textvariable=line_width_var, width=8)
    line_width_spin.grid(row=6, column=1, sticky="w", padx=8, pady=3)
    limit_specs = [
        ("t min", t_min_var),
        ("t max", t_max_var),
        ("V min", v_min_var),
        ("V max", v_max_var),
        ("I min", i_min_var),
        ("I max", i_max_var),
        ("T min", temp_min_var),
        ("T max", temp_max_var),
    ]
    for row_idx, (label, var) in enumerate(limit_specs):
        ttk.Label(limits_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        entry = ttk.Entry(limits_box, textvariable=var, width=12)
        entry.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        entry.bind("<Return>", _schedule_plot)
        entry.bind("<KP_Enter>", _schedule_plot)
        entry.bind("<FocusOut>", _schedule_plot)

    ttk.Label(limits_box, text="x-Ticks").grid(row=len(limit_specs), column=0, sticky="w", padx=8, pady=3)
    x_tick_spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=x_tick_count_var, width=8)
    x_tick_spin.grid(row=len(limit_specs), column=1, sticky="w", padx=8, pady=3)

    ttk.Label(limits_box, text="y-Ticks").grid(row=len(limit_specs) + 1, column=0, sticky="w", padx=8, pady=3)
    y_tick_spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=y_tick_count_var, width=8)
    y_tick_spin.grid(row=len(limit_specs) + 1, column=1, sticky="w", padx=8, pady=3)

    for combo in (voltage_line_combo, current_line_combo, temperature_line_combo):
        combo.bind("<<ComboboxSelected>>", _schedule_plot)
    time_unit_combo.bind("<<ComboboxSelected>>", _on_time_unit_changed)

    for widget in (
        title_entry,
        title_size_spin,
        tick_size_spin,
        label_size_spin,
        legend_size_spin,
        legend_scale_spin,
        line_width_spin,
    ):
        widget.bind("<Return>", _schedule_plot)
        widget.bind("<KP_Enter>", _schedule_plot)
        widget.bind("<FocusOut>", _schedule_plot)

    for spin in (title_size_spin, tick_size_spin, label_size_spin, legend_size_spin, legend_scale_spin, line_width_spin, x_tick_spin, y_tick_spin):
        spin.config(command=_schedule_plot)
        spin.bind("<Return>", _schedule_plot)
        spin.bind("<FocusOut>", _schedule_plot)

    ttk.Label(
        controls_frame,
        textvariable=status_var,
        wraplength=240,
        justify="left",
    ).pack(anchor="w", fill="x", pady=(10, 10))

    buttons_frame = ttk.Frame(controls_frame)
    buttons_frame.pack(fill="x", pady=(5, 0))
    ttk.Button(buttons_frame, text="Restablecer", command=_reset).pack(side="left", padx=(0, 6))
    ttk.Button(buttons_frame, text="Autoescala", command=_autofit).pack(side="left")

    _plot()


def open_v_vs_i_window(input_dir: Path, font_defaults: PlotFontDefaults | None = None) -> None:
    bundles = discover_activation_bundles(Path(input_dir))
    if not bundles:
        raise ValueError("No se encontraron archivos de activacion validos.")

    font_defaults = resolve_plot_font_defaults(font_defaults)
    font_default_values = font_defaults.as_strings()
    bundle = bundles[0]
    default_limits = compute_default_v_vs_i_limits(bundle)
    ramps = build_activation_ramps(bundle)
    cycle_ids = sorted({ramp.cycle for ramp in ramps})

    win = tk.Toplevel()
    win.title(f"Activacion - V vs I - {bundle.label}")
    win.geometry("1200x720")

    controls_host, plot_outer = create_resizable_plot_layout(
        win,
        sidebar_width=420,
        min_sidebar_width=390,
    )
    controls_frame = _build_scrollable_controls(controls_host)

    toolbar_frame = ttk.Frame(plot_outer)
    toolbar_frame.pack(side="top", fill="x")

    canvas_frame = ttk.Frame(plot_outer)
    canvas_frame.pack(side="top", fill="both", expand=True)

    fig = Figure(figsize=(9, 5.5), dpi=100)
    canvas = FigureCanvasTkAgg(fig, master=canvas_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    toolbar = NavigationToolbar2Tk(canvas, toolbar_frame, pack_toolbar=False)
    toolbar.update()
    toolbar.pack(side="left", fill="x")

    status_var = tk.StringVar(value="Listo.")

    asc_var = tk.BooleanVar(value=True)
    dsc_var = tk.BooleanVar(value=True)
    voltage_var = tk.BooleanVar(value=True)
    temperature_var = tk.BooleanVar(value=False)
    color_axes_var = tk.BooleanVar(value=False)

    voltage_line_var = tk.StringVar(value="-")
    temperature_line_var = tk.StringVar(value="--")

    i_min_var = tk.StringVar(value=default_limits["i_min"])
    i_max_var = tk.StringVar(value=default_limits["i_max"])
    v_min_var = tk.StringVar(value=default_limits["v_min"])
    v_max_var = tk.StringVar(value=default_limits["v_max"])
    temp_min_var = tk.StringVar(value=default_limits["temp_min"])
    temp_max_var = tk.StringVar(value=default_limits["temp_max"])

    x_tick_count_var = tk.IntVar(value=6)
    y_tick_count_var = tk.IntVar(value=6)

    plot_title_var = tk.StringVar(value="")
    show_title_var = tk.BooleanVar(value=True)
    title_fontsize_var = tk.StringVar(value=font_default_values["title"])
    tick_fontsize_var = tk.StringVar(value=font_default_values["tick"])
    label_fontsize_var = tk.StringVar(value=font_default_values["label"])
    legend_fontsize_var = tk.StringVar(value=font_default_values["legend"])
    legend_scale_var = tk.StringVar(value="1.0")
    line_width_var = tk.StringVar(value="1.5")

    cycle_vars = {cycle: tk.BooleanVar(value=True) for cycle in cycle_ids}

    initial_state = {
        "asc": True,
        "dsc": True,
        "voltage": True,
        "temperature": False,
        "color_axes": False,
        "voltage_line": "-",
        "temperature_line": "--",
        "i_min": default_limits["i_min"],
        "i_max": default_limits["i_max"],
        "v_min": default_limits["v_min"],
        "v_max": default_limits["v_max"],
        "temp_min": default_limits["temp_min"],
        "temp_max": default_limits["temp_max"],
        "x_tick_count": 6,
        "y_tick_count": 6,
        "plot_title": "",
        "show_title": True,
        "title_fontsize": font_default_values["title"],
        "tick_fontsize": font_default_values["tick"],
        "label_fontsize": font_default_values["label"],
        "legend_fontsize": font_default_values["legend"],
        "legend_scale": "1.0",
        "line_width": "1.5",
        "visible_cycles": {cycle: True for cycle in cycle_ids},
    }

    def _visible_ramp_keys() -> set[str]:
        visible_cycles = {cycle for cycle, var in cycle_vars.items() if var.get()}
        keys: set[str] = set()
        for ramp in ramps:
            if ramp.cycle not in visible_cycles:
                continue
            if ramp.direction == "Asc" and not asc_var.get():
                continue
            if ramp.direction == "Dsc" and not dsc_var.get():
                continue
            keys.add(ramp.key)
        return keys

    def _collect_limits():
        return dict(
            i_min=_optional_float(i_min_var.get()),
            i_max=_optional_float(i_max_var.get()),
            v_min=_optional_float(v_min_var.get()),
            v_max=_optional_float(v_max_var.get()),
            temp_min=_optional_float(temp_min_var.get()),
            temp_max=_optional_float(temp_max_var.get()),
        )

    plot_job = {"id": None}
    suspend_events = {"value": False}

    def _schedule_plot(*_args):
        if suspend_events["value"]:
            return
        if plot_job["id"] is not None:
            win.after_cancel(plot_job["id"])
        plot_job["id"] = win.after(20, _plot)

    def _plot():
        plot_job["id"] = None
        try:
            has_plot = draw_v_vs_i_on_figure(
                fig=fig,
                bundle=bundle,
                visible_ramp_keys=_visible_ramp_keys(),
                show_voltage=voltage_var.get(),
                show_temperature=temperature_var.get(),
                voltage_linestyle=voltage_line_var.get(),
                temperature_linestyle=temperature_line_var.get(),
                color_axes_by_magnitude=color_axes_var.get(),
                x_tick_count=x_tick_count_var.get(),
                y_tick_count=y_tick_count_var.get(),
                plot_title=plot_title_var.get(),
                show_title=show_title_var.get(),
                title_fontsize=_positive_float(title_fontsize_var.get(), "Tamaño del título"),
                tick_fontsize=_positive_float(tick_fontsize_var.get(), "Tamaño de ticks"),
                label_fontsize=_positive_float(label_fontsize_var.get(), "Tamaño de etiquetas"),
                legend_fontsize=_positive_float(legend_fontsize_var.get(), "Tamaño de leyenda"),
                legend_scale=_positive_float(legend_scale_var.get(), "Escala del gradiente"),
                line_width=_positive_float(line_width_var.get(), "Grosor de línea"),
                **_collect_limits(),
            )
        except ValueError as exc:
            fig.clear()
            canvas.draw_idle()
            status_var.set(f"Error: {exc}")
            return

        if not has_plot:
            fig.clear()
            canvas.draw_idle()
            status_var.set("No se muestra grafico: seleccione al menos una rampa y una magnitud.")
            return

        canvas.draw_idle()
        status_var.set("Grafico actualizado.")

    def _autofit():
        try:
            fitted = compute_autofit_v_vs_i_limits(
                bundle=bundle,
                visible_ramp_keys=_visible_ramp_keys(),
                show_voltage=voltage_var.get(),
                show_temperature=temperature_var.get(),
            )
        except ValueError as exc:
            status_var.set(f"Error: {exc}")
            return

        suspend_events["value"] = True
        try:
            i_min_var.set(fitted["i_min"])
            i_max_var.set(fitted["i_max"])
            if fitted["v_min"] != "" or fitted["v_max"] != "":
                v_min_var.set(fitted["v_min"])
                v_max_var.set(fitted["v_max"])
            if fitted["temp_min"] != "" or fitted["temp_max"] != "":
                temp_min_var.set(fitted["temp_min"])
                temp_max_var.set(fitted["temp_max"])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Autoescala aplicada.")

    def _reset():
        suspend_events["value"] = True
        try:
            asc_var.set(initial_state["asc"])
            dsc_var.set(initial_state["dsc"])
            voltage_var.set(initial_state["voltage"])
            temperature_var.set(initial_state["temperature"])
            color_axes_var.set(initial_state["color_axes"])
            voltage_line_var.set(initial_state["voltage_line"])
            temperature_line_var.set(initial_state["temperature_line"])
            i_min_var.set(initial_state["i_min"])
            i_max_var.set(initial_state["i_max"])
            v_min_var.set(initial_state["v_min"])
            v_max_var.set(initial_state["v_max"])
            temp_min_var.set(initial_state["temp_min"])
            temp_max_var.set(initial_state["temp_max"])
            x_tick_count_var.set(initial_state["x_tick_count"])
            y_tick_count_var.set(initial_state["y_tick_count"])
            plot_title_var.set(initial_state["plot_title"])
            show_title_var.set(initial_state["show_title"])
            title_fontsize_var.set(initial_state["title_fontsize"])
            tick_fontsize_var.set(initial_state["tick_fontsize"])
            label_fontsize_var.set(initial_state["label_fontsize"])
            legend_fontsize_var.set(initial_state["legend_fontsize"])
            legend_scale_var.set(initial_state["legend_scale"])
            line_width_var.set(initial_state["line_width"])
            for cycle, var in cycle_vars.items():
                var.set(initial_state["visible_cycles"][cycle])
        finally:
            suspend_events["value"] = False

        _plot()
        status_var.set("Valores restaurados.")

    ttk.Label(
        controls_frame,
        text=f"Serie detectada: Activacion {bundle.label}",
        justify="left",
        wraplength=260,
    ).pack(anchor="w", pady=(0, 10))

    cycles_box = ttk.LabelFrame(controls_frame, text="Ciclos")
    cycles_box.pack(fill="x", pady=5)

    series_box = ttk.LabelFrame(controls_frame, text="Series")
    series_box.pack(fill="x", pady=5)

    style_box = ttk.LabelFrame(controls_frame, text="Estilo")
    style_box.pack(fill="x", pady=5)

    text_box = ttk.LabelFrame(controls_frame, text="Texto / tamaños")
    text_box.pack(fill="x", pady=5)

    limits_box = ttk.LabelFrame(controls_frame, text="Límites de ejes")
    limits_box.pack(fill="x", pady=5)

    cycles_inner, bind_cycle_scroll = _build_scrollable_cycle_selector(cycles_box)

    for index, cycle in enumerate(cycle_ids):
        cycle_toggle = ttk.Checkbutton(
            cycles_inner,
            text=f"Ciclo #{cycle}",
            variable=cycle_vars[cycle],
            command=_schedule_plot,
        )
        pady = (0, 2) if index < len(cycle_ids) - 1 else 0
        cycle_toggle.pack(anchor="w", padx=4, pady=pady)
        bind_cycle_scroll(cycle_toggle)

    ttk.Checkbutton(series_box, text="Ascendente", variable=asc_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="Descendente", variable=dsc_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="Voltaje", variable=voltage_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="Temperatura", variable=temperature_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)
    ttk.Checkbutton(series_box, text="y-axes a color", variable=color_axes_var, command=_schedule_plot).pack(anchor="w", padx=8, pady=2)

    ttk.Label(style_box, text="Línea de voltaje").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    voltage_line_combo = ttk.Combobox(style_box, textvariable=voltage_line_var, values=LINESTYLE_OPTIONS, state="readonly", width=10)
    voltage_line_combo.grid(row=0, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(style_box, text="Línea de temperatura").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    temperature_line_combo = ttk.Combobox(style_box, textvariable=temperature_line_var, values=LINESTYLE_OPTIONS, state="readonly", width=10)
    temperature_line_combo.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Título").grid(row=0, column=0, sticky="w", padx=8, pady=3)
    title_entry = ttk.Entry(text_box, textvariable=plot_title_var, width=28)
    title_entry.grid(row=0, column=1, sticky="we", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño del título").grid(row=1, column=0, sticky="w", padx=8, pady=3)
    title_size_spin = tk.Spinbox(text_box, from_=6, to=30, increment=0.5, textvariable=title_fontsize_var, width=8)
    title_size_spin.grid(row=1, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de ticks").grid(row=2, column=0, sticky="w", padx=8, pady=3)
    tick_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=tick_fontsize_var, width=8)
    tick_size_spin.grid(row=2, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de etiquetas").grid(row=3, column=0, sticky="w", padx=8, pady=3)
    label_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=label_fontsize_var, width=8)
    label_size_spin.grid(row=3, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Tamaño de leyenda").grid(row=4, column=0, sticky="w", padx=8, pady=3)
    legend_size_spin = tk.Spinbox(text_box, from_=6, to=24, increment=0.5, textvariable=legend_fontsize_var, width=8)
    legend_size_spin.grid(row=4, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Escala del gradiente").grid(row=5, column=0, sticky="w", padx=8, pady=3)
    legend_scale_spin = tk.Spinbox(text_box, from_=0.5, to=3.0, increment=0.1, textvariable=legend_scale_var, width=8)
    legend_scale_spin.grid(row=5, column=1, sticky="w", padx=8, pady=3)

    ttk.Label(text_box, text="Grosor de línea").grid(row=6, column=0, sticky="w", padx=8, pady=3)
    line_width_spin = tk.Spinbox(text_box, from_=0.0, to=10.0, increment=0.1, textvariable=line_width_var, width=8)
    line_width_spin.grid(row=6, column=1, sticky="w", padx=8, pady=3)
    ttk.Checkbutton(text_box, text="Mostrar titulo", variable=show_title_var, command=_schedule_plot).grid(
        row=7, column=0, columnspan=2, sticky="w", padx=8, pady=3
    )

    limit_specs = [
        ("I min", i_min_var),
        ("I max", i_max_var),
        ("V min", v_min_var),
        ("V max", v_max_var),
        ("T min", temp_min_var),
        ("T max", temp_max_var),
    ]
    for row_idx, (label, var) in enumerate(limit_specs):
        ttk.Label(limits_box, text=label).grid(row=row_idx, column=0, sticky="w", padx=8, pady=3)
        entry = ttk.Entry(limits_box, textvariable=var, width=12)
        entry.grid(row=row_idx, column=1, sticky="w", padx=8, pady=3)
        entry.bind("<Return>", _schedule_plot)
        entry.bind("<KP_Enter>", _schedule_plot)
        entry.bind("<FocusOut>", _schedule_plot)

    ttk.Label(limits_box, text="x-Ticks").grid(row=len(limit_specs), column=0, sticky="w", padx=8, pady=3)
    x_tick_spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=x_tick_count_var, width=8)
    x_tick_spin.grid(row=len(limit_specs), column=1, sticky="w", padx=8, pady=3)

    ttk.Label(limits_box, text="y-Ticks").grid(row=len(limit_specs) + 1, column=0, sticky="w", padx=8, pady=3)
    y_tick_spin = tk.Spinbox(limits_box, from_=2, to=10, textvariable=y_tick_count_var, width=8)
    y_tick_spin.grid(row=len(limit_specs) + 1, column=1, sticky="w", padx=8, pady=3)

    for combo in (voltage_line_combo, temperature_line_combo):
        combo.bind("<<ComboboxSelected>>", _schedule_plot)

    for widget in (
        title_entry,
        title_size_spin,
        tick_size_spin,
        label_size_spin,
        legend_size_spin,
        legend_scale_spin,
        line_width_spin,
    ):
        widget.bind("<Return>", _schedule_plot)
        widget.bind("<KP_Enter>", _schedule_plot)
        widget.bind("<FocusOut>", _schedule_plot)

    for spin in (title_size_spin, tick_size_spin, label_size_spin, legend_size_spin, legend_scale_spin, line_width_spin, x_tick_spin, y_tick_spin):
        spin.config(command=_schedule_plot)
        spin.bind("<Return>", _schedule_plot)
        spin.bind("<FocusOut>", _schedule_plot)

    ttk.Label(
        controls_frame,
        textvariable=status_var,
        wraplength=240,
        justify="left",
    ).pack(anchor="w", fill="x", pady=(10, 10))

    buttons_frame = ttk.Frame(controls_frame)
    buttons_frame.pack(fill="x", pady=(5, 0))
    ttk.Button(buttons_frame, text="Restablecer", command=_reset).pack(side="left", padx=(0, 6))
    ttk.Button(buttons_frame, text="Autoescala", command=_autofit).pack(side="left")

    _plot()


def export_folder(input_dir: Path, output_dir: Path) -> list[Path]:
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)

    bundles = discover_activation_bundles(input_dir)
    exported_files: list[Path] = []

    for bundle in bundles:
        out_name = f"Activacion_{bundle.label}.xlsx"
        out_path = output_dir / out_name
        export_activation_bundle(bundle, out_path)
        exported_files.append(out_path)

    return exported_files


def run_pipeline(
    input_dir: Path,
    output_dir: Path,
    selected_options: list[str] | None = None,
    font_defaults: PlotFontDefaults | None = None,
    language: str = "es",
) -> list[Path]:
    exported_files = export_folder(input_dir, output_dir)
    if not exported_files:
        return []

    chosen = set(selected_options or [])
    if "V vs t" in chosen:
        open_v_vs_t_window(input_dir, font_defaults=font_defaults)
    if "V vs I" in chosen:
        open_v_vs_i_window(input_dir, font_defaults=font_defaults)

    return exported_files
